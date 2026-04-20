"""
Grid Drawing Generator
----------------------
Generates a Foundation/Grid Layout Plan DXF from gridline_coordinates and node_coordinates.
Matches the reference drawing style: colors, labels, dimensions, perimeter, title block, notes.

Usage: python grid_drawing_generator.py
No arguments needed - auto-finds latest files in STD ANL model folder.

Pipeline: Run AFTER gridline_calculator.py.
"""

import os
import glob
import math
import openpyxl
import ezdxf
from ezdxf.enums import TextEntityAlignment
from datetime import datetime


# ── Placement Registry (collision avoidance for dimension text) ──────
class PlacementRegistry:
    """Tracks bounding boxes of placed text and lines.
    Provides collision-aware text placement with Nudge -> Stagger -> Leader strategies."""

    def __init__(self, msp):
        self.msp = msp
        self._boxes = []      # list of (x1, y1, x2, y2) bounding boxes
        self._lines = []      # list of (x1, y1, x2, y2) line segments

    def register_box(self, x1, y1, x2, y2):
        """Register a rectangular occupied area."""
        bx1, bx2 = min(x1, x2), max(x1, x2)
        by1, by2 = min(y1, y2), max(y1, y2)
        self._boxes.append((bx1, by1, bx2, by2))

    def register_line(self, x1, y1, x2, y2, thickness=0):
        """Register a line segment as an obstacle. Thickness expands the line box."""
        lx1, lx2 = min(x1, x2), max(x1, x2)
        ly1, ly2 = min(y1, y2), max(y1, y2)
        ht = thickness / 2
        self._lines.append((lx1 - ht, ly1 - ht, lx2 + ht, ly2 + ht))

    @staticmethod
    def _rendered_len(text):
        """Count rendered character length, accounting for DXF codes.
        %%189 renders as single char (1/2), %%nnn is 1 rendered char."""
        import re
        # Replace %%nnn codes with single placeholder char
        rendered = re.sub(r'%%\d+', 'X', text)
        return len(rendered)

    def _text_bbox(self, cx, cy, text, height, rotation=0, align=None):
        """Estimate bounding box accounting for text alignment.
        For BOTTOM_CENTER: (cx,cy) is at bottom-center, text extends upward.
        For TOP_CENTER: text extends downward. For MIDDLE_CENTER: centered."""
        tw = self._rendered_len(text) * height * 0.65
        th = height * 1.4
        hw = tw / 2

        # Determine vertical offset based on alignment
        if align in (TextEntityAlignment.BOTTOM_CENTER, TextEntityAlignment.BOTTOM_LEFT,
                     TextEntityAlignment.BOTTOM_RIGHT):
            y_lo, y_hi = cy, cy + th
        elif align in (TextEntityAlignment.TOP_CENTER, TextEntityAlignment.TOP_LEFT,
                       TextEntityAlignment.TOP_RIGHT):
            y_lo, y_hi = cy - th, cy
        else:  # MIDDLE_CENTER or default
            y_lo, y_hi = cy - th / 2, cy + th / 2

        if rotation == 90 or rotation == -90:
            # Rotated 90: width becomes vertical, height becomes horizontal
            return (cx - th / 2, cy - hw, cx + th / 2, cy + hw)
        return (cx - hw, y_lo, cx + hw, y_hi)

    def _collides(self, bbox, margin=60):
        """Check if bbox overlaps any registered box or line."""
        bx1, by1, bx2, by2 = bbox
        bx1 -= margin
        by1 -= margin
        bx2 += margin
        by2 += margin

        for ox1, oy1, ox2, oy2 in self._boxes:
            if bx1 < ox2 and bx2 > ox1 and by1 < oy2 and by2 > oy1:
                return True
        for lx1, ly1, lx2, ly2 in self._lines:
            if bx1 < lx2 and bx2 > lx1 and by1 < ly2 and by2 > ly1:
                return True
        return False

    def place_text(self, text, height, cx, cy, layer="DIM_TEXT",
                   rotation=0, align=TextEntityAlignment.BOTTOM_CENTER,
                   dim_axis="h", dim_span=0):
        """Place dimension text with collision avoidance.
        Tries: original -> grid search (nudge+stagger combos) -> leader line.

        dim_axis: 'h' for horizontal dim, 'v' for vertical dim
        dim_span: length of the dimension line (for nudge bounds)
        """
        def _try(tx, ty):
            bb = self._text_bbox(tx, ty, text, height, rotation, align)
            return bb if not self._collides(bb) else None

        # Strategy 1: Original position
        bbox = _try(cx, cy)
        if bbox:
            self._place(text, height, cx, cy, layer, rotation, align)
            self.register_box(*bbox)
            return

        # Strategy 2: Grid search — try nudge, stagger, and diagonal combos
        # along_offsets: movement along the dim line
        # perp_offsets: movement perpendicular to dim line
        unit = max(height * 2.5, 200)
        along_steps = [0, 1, -1, 2, -2, 3, -3, 4, -4, 5, -5]
        perp_steps = [0, 1, -1, 2, -2, 3, -3, 4, -4, 5, -5, 6, -6, 7, -7]

        best = None
        best_dist = float('inf')

        for a in along_steps:
            for p in perp_steps:
                if a == 0 and p == 0:
                    continue  # already tried original
                if dim_axis == "h":
                    tx, ty = cx + a * unit, cy + p * unit
                else:
                    tx, ty = cx + p * unit, cy + a * unit
                bb = _try(tx, ty)
                if bb:
                    d = math.sqrt((tx - cx) ** 2 + (ty - cy) ** 2)
                    if d < best_dist:
                        best = (tx, ty, bb)
                        best_dist = d
                        if d <= unit * 2.5:
                            break
            if best and best_dist <= unit * 2.5:
                break

        if best:
            tx, ty, bbox = best
            # Draw leader line if text moved significantly from original
            if best_dist > unit * 1.5:
                self.msp.add_line((cx, cy), (tx, ty),
                                  dxfattribs={"layer": layer})
            self._place(text, height, tx, ty, layer, rotation, align)
            self.register_box(*bbox)
            return

        # Strategy 3: Leader line — search 8 directions at increasing distances
        directions = [(0, 1), (0, -1), (1, 0), (-1, 0),
                      (1, 1), (1, -1), (-1, 1), (-1, -1)]
        for dist_mult in [8, 12, 16, 20]:
            for dx, dy in directions:
                lx = cx + dx * height * dist_mult
                ly = cy + dy * height * dist_mult
                bbox = _try(lx, ly)
                if bbox:
                    self.msp.add_line((cx, cy), (lx, ly),
                                      dxfattribs={"layer": layer})
                    self._place(text, height, lx, ly, layer, rotation, align)
                    self.register_box(*bbox)
                    return

        # Fallback: place at original position (rare worst case)
        self._place(text, height, cx, cy, layer, rotation, align)
        bbox = self._text_bbox(cx, cy, text, height, rotation, align)
        self.register_box(*bbox)

    def _place(self, text, height, x, y, layer, rotation, align):
        """Actually add text entity to modelspace."""
        attribs = {"layer": layer}
        if rotation:
            self.msp.add_text(text, height=height, rotation=rotation,
                              dxfattribs=attribs).set_placement(
                (x, y), align=align)
        else:
            self.msp.add_text(text, height=height,
                              dxfattribs=attribs).set_placement(
                (x, y), align=align)

    def register_grid_obstacles(self, grid_x, grid_y, columns):
        """Register grid lines, column outlines, perimeter as obstacles."""
        min_x = min(grid_x) - GRID_LINE_END
        max_x = max(grid_x) + GRID_LINE_END
        min_y = min(grid_y) - GRID_LINE_END
        max_y = max(grid_y) + GRID_LINE_END

        # Grid lines (avoid text sitting on grid lines)
        for x in grid_x:
            self.register_line(x, min_y, x, max_y, thickness=40)
        for y in grid_y:
            self.register_line(min_x, y, max_x, y, thickness=40)

        # Column rectangles (expanded slightly to keep text off edges)
        for col in columns:
            pad = 30
            self.register_box(col["x1"] - pad, col["y1"] - pad,
                              col["x2"] + pad, col["y2"] + pad)

        # Building perimeter line (thick blue)
        bx1 = min(c["x1"] for c in columns)
        bx2 = max(c["x2"] for c in columns)
        by1 = min(c["y1"] for c in columns)
        by2 = max(c["y2"] for c in columns)
        self.register_line(bx1, by1, bx2, by1, thickness=30)
        self.register_line(bx1, by2, bx2, by2, thickness=30)
        self.register_line(bx1, by1, bx1, by2, thickness=30)
        self.register_line(bx2, by1, bx2, by2, thickness=30)

# ── Paths ──────────────────────────────────────────────────────────────
STD_ANL_FOLDER = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model"
OUTPUT_FOLDER = os.path.join(STD_ANL_FOLDER, "final")

# ── Drawing constants (all in mm) ─────────────────────────────────────
SCALE = 1000                # metres to mm

# Layout zones (distance from outermost grid edge, stacked outward)
DIM_OFFSET_1 = 800          # individual span dimensions
DIM_OFFSET_2 = 1800         # overall span dimensions
GRID_LABEL_OFFSET = 2800    # center of grid label circles
GRID_LINE_END = 3250        # grid lines terminate here

# Grid
CIRCLE_RADIUS = 350         # mm for grid label circles
GRID_LABEL_HEIGHT = 250     # mm text height inside circles

# Columns
COL_FILL_COLOR = 31         # salmon/pink for column rectangle fill
COL_LABEL_COLOR = 30        # orange for column name labels
COL_LABEL_HEIGHT = 180      # mm text height for C1, C2 etc.
COL_LABEL_CIRCLE_RADIUS = 220  # mm for column label circle

# Dimensions
DIM_TEXT_HEIGHT = 120        # mm
DIM_TICK_SIZE = 80           # mm tick mark half-size

# Footing annotation
FOOTING_LINE_EXTEND = 450   # mm, cyan lines extending from column face
FOOTING_DIM_HEIGHT = 100    # mm text height for column dimensions

# Title block
TB_WIDTH = 5000
TB_ROW_H = 300              # row height in title block

# Notes panel
NOTES_WIDTH = 5500
NOTES_TEXT_H = 100           # notes body text height
NOTES_HEADER_H = 140         # notes header text height

# Colors (AutoCAD Color Index)
CLR_WHITE = 7
CLR_RED = 1
CLR_YELLOW = 2
CLR_CYAN = 4
CLR_BLUE = 5
CLR_MAGENTA = 6
CLR_ORANGE = 30
CLR_SALMON = 31

# Sheet border
BORDER_MARGIN = 600          # mm clearance from content to magenta border


def find_latest_file(pattern):
    """Find the latest file matching a glob pattern."""
    files = glob.glob(os.path.join(STD_ANL_FOLDER, pattern))
    if not files:
        raise FileNotFoundError(f"No files matching {pattern} in {STD_ANL_FOLDER}")
    return max(files, key=os.path.getmtime)


def read_gridline_coordinates(filepath):
    """Read gridline_coordinates Excel. Returns list of column dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    columns = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        columns.append({
            "no": row[0],
            "name": row[1],
            "location": row[2],
            "anchor_x": row[3] * SCALE,
            "anchor_y": row[4] * SCALE,
            "beam_x_loc": row[5],
            "beam_y_loc": row[6],
            "orientation": row[7],
            "yd": row[8],
            "zd": row[9],
            "x1": row[10] * SCALE,
            "x2": row[11] * SCALE,
            "y1": row[12] * SCALE,
            "y2": row[13] * SCALE,
        })
    wb.close()
    return columns


def read_unique_grid_values(filepath):
    """Read node_coordinates and extract unique grid X/Y values in mm."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Node coordinates"]
    xs, ys = set(), set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        xs.add(round(row[21] * SCALE, 1))  # Node X (col 22, index 21)
        ys.add(round(row[22] * SCALE, 1))  # Node Y (col 23, index 22)
    wb.close()
    return sorted(xs), sorted(ys)


def setup_document():
    """Create DXF document with layers matching the reference drawing."""
    doc = ezdxf.new("R2010")

    # Override the default "Standard" text style to use simplex.shx (engineering font)
    style = doc.styles.get("Standard")
    style.dxf.font = "simplex.shx"

    doc.layers.add("GRID", color=CLR_WHITE)
    doc.layers.add("GRID_LABELS", color=CLR_WHITE)
    doc.layers.add("COLUMNS", color=CLR_RED)
    doc.layers.add("COL_FILL", color=9)           # grey fill
    doc.layers.add("COL_LABELS", color=CLR_ORANGE)     # circle + leader line
    doc.layers.add("COL_LABEL_TEXT", color=CLR_WHITE)  # text inside circle
    doc.layers.add("DIMENSIONS", color=CLR_CYAN)
    doc.layers.add("FOOTING_DIM", color=CLR_CYAN)
    doc.layers.add("DIM_TEXT", color=CLR_YELLOW)    # dimension numbers only
    doc.layers.add("PERIMETER", color=CLR_BLUE)
    doc.layers.add("PLOT_BOUNDARY", color=CLR_RED)
    doc.layers.add("TITLE", color=CLR_CYAN)
    doc.layers.add("TITLE_BLOCK", color=CLR_WHITE)
    doc.layers.add("NOTES", color=CLR_YELLOW)
    doc.layers.add("NOTES_HEADER", color=CLR_RED)
    doc.layers.add("BORDER", color=CLR_MAGENTA)

    return doc


# ── 1. Grid Lines ─────────────────────────────────────────────────────

def draw_grid_lines(msp, grid_x, grid_y):
    """Draw grid lines extending to GRID_LINE_END beyond building boundary."""
    min_x = min(grid_x) - GRID_LINE_END
    max_x = max(grid_x) + GRID_LINE_END
    min_y = min(grid_y) - GRID_LINE_END
    max_y = max(grid_y) + GRID_LINE_END

    for x in grid_x:
        msp.add_line((x, min_y), (x, max_y), dxfattribs={"layer": "GRID"})

    for y in grid_y:
        msp.add_line((min_x, y), (max_x, y), dxfattribs={"layer": "GRID"})


# ── 2. Grid Labels (circles with numbers/letters) ────────────────────

def _offset_close_labels(values, min_gap):
    """For grid values that are too close, offset their label positions.
    Returns list of (value, y_offset) where y_offset staggers close labels."""
    offsets = [0.0] * len(values)
    for i in range(len(values) - 1):
        gap = values[i + 1] - values[i]
        if gap < min_gap:
            # Stagger: push previous up, next down (or left/right for Y)
            offsets[i] = CIRCLE_RADIUS * 1.5
            offsets[i + 1] = -CIRCLE_RADIUS * 1.5
    return offsets


def draw_grid_labels(msp, grid_x, grid_y):
    """Draw numbered/lettered circle labels at both ends of grid lines.
    Staggers labels for grid lines that are too close together."""
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)

    min_gap = CIRCLE_RADIUS * 2.5  # minimum gap before staggering

    # X grid labels (numbers 1, 2, 3...) at bottom and top
    x_offsets = _offset_close_labels(grid_x, min_gap)
    for i, x in enumerate(grid_x):
        label = str(i + 1)
        stagger = x_offsets[i]  # vertical stagger for close labels
        for y_base in [min_y - GRID_LABEL_OFFSET, max_y + GRID_LABEL_OFFSET]:
            # At bottom, stagger downward; at top, stagger upward
            if y_base < min_y:
                y_pos = y_base - stagger
            else:
                y_pos = y_base + stagger
            msp.add_circle((x, y_pos), CIRCLE_RADIUS,
                           dxfattribs={"layer": "GRID_LABELS"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "GRID_LABELS"}).set_placement(
                (x, y_pos), align=TextEntityAlignment.MIDDLE_CENTER)

    # Y grid labels (letters A, B, C...) at left and right
    y_offsets = _offset_close_labels(grid_y, min_gap)
    for i, y in enumerate(grid_y):
        label = chr(65 + i)  # A, B, C, D, E...
        stagger = y_offsets[i]  # horizontal stagger for close labels
        for x_base in [min_x - GRID_LABEL_OFFSET, max_x + GRID_LABEL_OFFSET]:
            if x_base < min_x:
                x_pos = x_base - stagger
            else:
                x_pos = x_base + stagger
            msp.add_circle((x_pos, y), CIRCLE_RADIUS,
                           dxfattribs={"layer": "GRID_LABELS"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "GRID_LABELS"}).set_placement(
                (x_pos, y), align=TextEntityAlignment.MIDDLE_CENTER)


# ── 3. Column Rectangles (filled) ────────────────────────────────────

def draw_column_rectangles(msp, columns):
    """Draw filled column footprint rectangles with outline."""
    for col in columns:
        x1, y1, x2, y2 = col["x1"], col["y1"], col["x2"], col["y2"]

        # Solid fill
        hatch = msp.add_hatch(color=9, dxfattribs={"layer": "COL_FILL"})
        hatch.paths.add_polyline_path([
            (x1, y1), (x2, y1), (x2, y2), (x1, y2), (x1, y1)
        ], is_closed=True)

        # Outline
        msp.add_lwpolyline(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)],
            dxfattribs={"layer": "COLUMNS"}, close=True)


# ── 4. Column Labels (C1, C2...) ─────────────────────────────────────

def _find_best_label_pos(col, columns, radius):
    """Find the best label position that avoids other columns and labels."""
    cx = (col["x1"] + col["x2"]) / 2
    cy = (col["y1"] + col["y2"]) / 2
    col_w = abs(col["x2"] - col["x1"])
    col_h = abs(col["y2"] - col["y1"])

    # Try 8 directions: prioritize diagonal corners, then cardinals
    offset = radius * 2.5
    candidates = [
        (col["x2"] + offset, col["y2"] + offset),     # top-right
        (col["x1"] - offset, col["y2"] + offset),     # top-left
        (col["x2"] + offset, col["y1"] - offset),     # bottom-right
        (col["x1"] - offset, col["y1"] - offset),     # bottom-left
        (cx, col["y2"] + offset * 1.2),                # top
        (cx, col["y1"] - offset * 1.2),                # bottom
        (col["x2"] + offset * 1.2, cy),                # right
        (col["x1"] - offset * 1.2, cy),                # left
    ]

    min_clearance = radius * 3  # minimum distance from other column centers

    for lx, ly in candidates:
        clear = True
        for other in columns:
            if other["no"] == col["no"]:
                continue
            # Check distance from other column center
            ocx = (other["x1"] + other["x2"]) / 2
            ocy = (other["y1"] + other["y2"]) / 2
            dist = math.sqrt((lx - ocx) ** 2 + (ly - ocy) ** 2)
            if dist < min_clearance:
                clear = False
                break
            # Check if label overlaps other column rectangle
            if (other["x1"] - radius < lx < other["x2"] + radius and
                    other["y1"] - radius < ly < other["y2"] + radius):
                clear = False
                break
        if clear:
            return lx, ly

    # Fallback: top-right with larger offset
    return col["x2"] + offset * 1.5, col["y2"] + offset * 1.5


def draw_column_labels(msp, columns):
    """Draw column name labels with orange circle, smart placement to avoid overlap."""
    placed = []  # track placed label positions

    for col in columns:
        lx, ly = _find_best_label_pos(col, columns, COL_LABEL_CIRCLE_RADIUS)

        # Also check against already-placed labels
        retry = 0
        while retry < 3:
            collision = False
            for px, py in placed:
                if math.sqrt((lx - px) ** 2 + (ly - py) ** 2) < COL_LABEL_CIRCLE_RADIUS * 4:
                    collision = True
                    break
            if not collision:
                break
            # Shift further away
            cx = (col["x1"] + col["x2"]) / 2
            cy = (col["y1"] + col["y2"]) / 2
            lx += (lx - cx) * 0.5
            ly += (ly - cy) * 0.5
            retry += 1

        placed.append((lx, ly))

        msp.add_circle((lx, ly), COL_LABEL_CIRCLE_RADIUS,
                        dxfattribs={"layer": "COL_LABELS"})
        msp.add_text(col["name"], height=COL_LABEL_HEIGHT,
                     dxfattribs={"layer": "COL_LABEL_TEXT"}).set_placement(
            (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)

        # Leader line from circle edge to column center
        ccx = (col["x1"] + col["x2"]) / 2
        ccy = (col["y1"] + col["y2"]) / 2
        dx = ccx - lx
        dy = ccy - ly
        dist = math.sqrt(dx ** 2 + dy ** 2)
        if dist > COL_LABEL_CIRCLE_RADIUS * 1.2:
            # Start from circle edge, end at column center
            ratio = COL_LABEL_CIRCLE_RADIUS / dist
            start_x = lx + dx * ratio
            start_y = ly + dy * ratio
            msp.add_line((start_x, start_y), (ccx, ccy),
                         dxfattribs={"layer": "COL_LABELS"})


# ── 5. Footing Detail Annotations ────────────────────────────────────

def _check_clearance(col, columns, direction, min_dist):
    """Check if a column face has enough clearance from other columns in given direction.
    direction: 'bottom', 'top', 'left', 'right'
    Returns True if clear."""
    x1, y1, x2, y2 = col["x1"], col["y1"], col["x2"], col["y2"]

    for other in columns:
        if other["no"] == col["no"]:
            continue
        ox1, oy1, ox2, oy2 = other["x1"], other["y1"], other["x2"], other["y2"]

        # Check X overlap (columns are in same horizontal band)
        x_overlap = not (ox2 < x1 - 200 or ox1 > x2 + 200)
        # Check Y overlap (columns are in same vertical band)
        y_overlap = not (oy2 < y1 - 200 or oy1 > y2 + 200)

        if direction == "bottom" and x_overlap:
            if 0 < y1 - oy2 < min_dist:
                return False
        elif direction == "top" and x_overlap:
            if 0 < oy1 - y2 < min_dist:
                return False
        elif direction == "left" and y_overlap:
            if 0 < x1 - ox2 < min_dist:
                return False
        elif direction == "right" and y_overlap:
            if 0 < ox1 - x2 < min_dist:
                return False
    return True



def _draw_footing_cross_h(msp, col, y_line, th, text_above, ext, registry=None):
    """Draw horizontal footing cross-dim showing segment breakdown.
    Shows: left_offset | YD (column width) | right_offset as 3 segments,
    with extension lines from the outer ends and through column edges."""
    layer = "FOOTING_DIM"
    x1, x2 = col["x1"], col["x2"]
    ax = col["anchor_x"]
    yd = col["yd"]
    tick = DIM_TICK_SIZE * 0.6

    # Segment positions: anchor divides the column width
    beam_x = col["beam_x_loc"]
    if beam_x == "Left":
        left_ext_x = x1 - ext * 0.3
        right_ext_x = x2 + ext * 0.3
    elif beam_x == "Right":
        left_ext_x = x1 - ext * 0.3
        right_ext_x = x2 + ext * 0.3
    else:  # Centre
        left_ext_x = x1 - ext * 0.3
        right_ext_x = x2 + ext * 0.3

    # Extension lines from column edges down/up to dim line
    gap = 40
    if text_above:
        for ex in [x1, x2]:
            msp.add_line((ex, col["y2"] + gap), (ex, y_line + tick),
                         dxfattribs={"layer": layer})
            if registry:
                registry.register_line(ex, col["y2"] + gap, ex, y_line + tick, thickness=10)
        msp.add_line((left_ext_x, y_line), (right_ext_x, y_line),
                     dxfattribs={"layer": layer})
    else:
        for ex in [x1, x2]:
            msp.add_line((ex, col["y1"] - gap), (ex, y_line - tick),
                         dxfattribs={"layer": layer})
            if registry:
                registry.register_line(ex, col["y1"] - gap, ex, y_line - tick, thickness=10)
        msp.add_line((left_ext_x, y_line), (right_ext_x, y_line),
                     dxfattribs={"layer": layer})

    # Register dim line as obstacle
    if registry:
        registry.register_line(left_ext_x, y_line, right_ext_x, y_line, thickness=10)

    # Ticks at column edges
    for tx in [x1, x2]:
        msp.add_line((tx - tick * 0.5, y_line - tick * 0.5),
                     (tx + tick * 0.5, y_line + tick * 0.5),
                     dxfattribs={"layer": layer})

    # Dimension text: total YD across column width
    cx = (x1 + x2) / 2
    y_text = y_line + th * 0.5 if text_above else y_line - th * 0.5
    align = TextEntityAlignment.BOTTOM_CENTER if text_above else TextEntityAlignment.TOP_CENTER
    text = _dim_text(yd)
    if registry:
        registry.place_text(text, th, cx, y_text, rotation=0, align=align,
                            dim_axis="h", dim_span=abs(x2 - x1))
    else:
        msp.add_text(text, height=th,
                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
            (cx, y_text), align=align)


def _draw_footing_cross_v(msp, col, x_line, th, text_right, ext, registry=None):
    """Draw vertical footing cross-dim showing ZD segment."""
    layer = "FOOTING_DIM"
    y1, y2 = col["y1"], col["y2"]
    zd = col["zd"]
    tick = DIM_TICK_SIZE * 0.6

    top_ext_y = y2 + ext * 0.3
    bot_ext_y = y1 - ext * 0.3

    # Extension lines from column edges to dim line
    gap = 40
    if text_right:
        for ey in [y1, y2]:
            msp.add_line((col["x2"] + gap, ey), (x_line + tick, ey),
                         dxfattribs={"layer": layer})
            if registry:
                registry.register_line(col["x2"] + gap, ey, x_line + tick, ey, thickness=10)
        msp.add_line((x_line, bot_ext_y), (x_line, top_ext_y),
                     dxfattribs={"layer": layer})
    else:
        for ey in [y1, y2]:
            msp.add_line((col["x1"] - gap, ey), (x_line - tick, ey),
                         dxfattribs={"layer": layer})
            if registry:
                registry.register_line(col["x1"] - gap, ey, x_line - tick, ey, thickness=10)
        msp.add_line((x_line, bot_ext_y), (x_line, top_ext_y),
                     dxfattribs={"layer": layer})

    # Register dim line as obstacle
    if registry:
        registry.register_line(x_line, bot_ext_y, x_line, top_ext_y, thickness=10)

    # Ticks at column edges
    for ty in [y1, y2]:
        msp.add_line((x_line - tick * 0.5, ty - tick * 0.5),
                     (x_line + tick * 0.5, ty + tick * 0.5),
                     dxfattribs={"layer": layer})

    # Dimension text: total ZD
    cy = (y1 + y2) / 2
    x_text = x_line + th * 0.5 if text_right else x_line - th * 0.5
    text = _dim_text(zd)
    if registry:
        registry.place_text(text, th, x_text, cy, rotation=90,
                            align=TextEntityAlignment.MIDDLE_CENTER,
                            dim_axis="v", dim_span=abs(y2 - y1))
    else:
        msp.add_text(text, height=th, rotation=90,
                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
            (x_text, cy), align=TextEntityAlignment.MIDDLE_CENTER)


def draw_footing_annotations(msp, columns, registry=None):
    """Draw full cross-shaped dimension annotations around each column footprint.
    Always draws all 4 faces. Uses compact mode for tight spaces."""
    ext = FOOTING_LINE_EXTEND
    th = FOOTING_DIM_HEIGHT
    min_clearance = ext * 2.0
    layer = "FOOTING_DIM"

    for col in columns:
        x1, y1, x2, y2 = col["x1"], col["y1"], col["x2"], col["y2"]

        # Check clearance for full vs compact dims
        can_bottom = _check_clearance(col, columns, "bottom", min_clearance)
        can_top = _check_clearance(col, columns, "top", min_clearance)
        can_right = _check_clearance(col, columns, "right", min_clearance)
        can_left = _check_clearance(col, columns, "left", min_clearance)

        # Bottom face
        if can_bottom:
            _draw_footing_cross_h(msp, col, y1 - ext, th, text_above=False, ext=ext, registry=registry)
        else:
            text = _dim_text(col["yd"])
            tcx, tcy = (x1 + x2) / 2, y1 - th * 1.5
            if registry:
                registry.place_text(text, th * 0.7, tcx, tcy, rotation=0,
                                    align=TextEntityAlignment.TOP_CENTER,
                                    dim_axis="h", dim_span=abs(x2 - x1))
            else:
                msp.add_text(text, height=th * 0.7,
                             dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                    (tcx, tcy), align=TextEntityAlignment.TOP_CENTER)

        # Top face
        if can_top:
            _draw_footing_cross_h(msp, col, y2 + ext, th, text_above=True, ext=ext, registry=registry)
        else:
            text = _dim_text(col["yd"])
            tcx, tcy = (x1 + x2) / 2, y2 + th * 0.8
            if registry:
                registry.place_text(text, th * 0.7, tcx, tcy, rotation=0,
                                    align=TextEntityAlignment.BOTTOM_CENTER,
                                    dim_axis="h", dim_span=abs(x2 - x1))
            else:
                msp.add_text(text, height=th * 0.7,
                             dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                    (tcx, tcy), align=TextEntityAlignment.BOTTOM_CENTER)

        # Right face
        if can_right:
            _draw_footing_cross_v(msp, col, x2 + ext, th, text_right=True, ext=ext, registry=registry)
        else:
            text = _dim_text(col["zd"])
            tcx, tcy = x2 + th * 1.0, (y1 + y2) / 2
            if registry:
                registry.place_text(text, th * 0.7, tcx, tcy, rotation=90,
                                    align=TextEntityAlignment.MIDDLE_CENTER,
                                    dim_axis="v", dim_span=abs(y2 - y1))
            else:
                msp.add_text(text, height=th * 0.7, rotation=90,
                             dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                    (tcx, tcy), align=TextEntityAlignment.MIDDLE_CENTER)

        # Left face
        if can_left:
            _draw_footing_cross_v(msp, col, x1 - ext, th, text_right=False, ext=ext, registry=registry)
        else:
            text = _dim_text(col["zd"])
            tcx, tcy = x1 - th * 1.0, (y1 + y2) / 2
            if registry:
                registry.place_text(text, th * 0.7, tcx, tcy, rotation=90,
                                    align=TextEntityAlignment.MIDDLE_CENTER,
                                    dim_axis="v", dim_span=abs(y2 - y1))
            else:
                msp.add_text(text, height=th * 0.7, rotation=90,
                             dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                    (tcx, tcy), align=TextEntityAlignment.MIDDLE_CENTER)


# ── 6. Grid Spacing Dimensions ───────────────────────────────────────

def _dim_text(value_mm):
    """Format dimension in mm as feet-inches with half-inch rounding.
    Examples: 150mm -> 6", 300mm -> 1'-0", 450mm -> 1'-5%%189",
              3921mm -> 12'-10%%189", 600mm -> 1'-11%%189"
    Uses %%189 for the ½ symbol in DXF (AutoCAD special char)."""
    total_inches = value_mm / 25.4
    # Round to nearest half inch
    half_inches = round(total_inches * 2)
    total_half = half_inches  # in half-inch units

    feet = int(total_half // 24)  # 24 half-inches per foot
    remaining_half = int(total_half % 24)
    whole_inches = remaining_half // 2
    has_half = remaining_half % 2 == 1

    if feet == 0:
        # Less than 1 foot — just inches
        if has_half:
            if whole_inches == 0:
                return '%%189"'
            return f'{whole_inches}%%189"'
        else:
            return f'{whole_inches}"'
    else:
        # Feet and inches
        if whole_inches == 0 and not has_half:
            return f"{feet}'-0\""
        elif has_half:
            if whole_inches == 0:
                return f"{feet}'-%%189\""
            return f"{feet}'-{whole_inches}%%189\""
        else:
            return f"{feet}'-{whole_inches}\""


def _draw_dim_line(msp, p1, p2, offset_dir, offset_dist, registry=None):
    """Draw a dimension line between two points, offset in given direction.
    Extension lines start from a short gap past the grid edge (not from grid origin)."""
    x1, y1 = p1
    x2, y2 = p2

    # Offset positions
    if offset_dir == "down":
        ox1, oy1 = x1, y1 - offset_dist
        ox2, oy2 = x2, y2 - offset_dist
    elif offset_dir == "up":
        ox1, oy1 = x1, y1 + offset_dist
        ox2, oy2 = x2, y2 + offset_dist
    elif offset_dir == "left":
        ox1, oy1 = x1 - offset_dist, y1
        ox2, oy2 = x2 - offset_dist, y2
    elif offset_dir == "right":
        ox1, oy1 = x1 + offset_dist, y1
        ox2, oy2 = x2 + offset_dist, y2

    # Extension lines: start from a short distance past grid edge, end past dim line
    ext_past = DIM_TICK_SIZE * 1.5  # how far extension line goes past dim line
    ext_start = 200  # start 200mm past the grid edge (not from grid origin)
    if offset_dir == "down":
        msp.add_line((x1, y1 - ext_start), (ox1, oy1 - ext_past), dxfattribs={"layer": "DIMENSIONS"})
        msp.add_line((x2, y2 - ext_start), (ox2, oy2 - ext_past), dxfattribs={"layer": "DIMENSIONS"})
    elif offset_dir == "up":
        msp.add_line((x1, y1 + ext_start), (ox1, oy1 + ext_past), dxfattribs={"layer": "DIMENSIONS"})
        msp.add_line((x2, y2 + ext_start), (ox2, oy2 + ext_past), dxfattribs={"layer": "DIMENSIONS"})
    elif offset_dir == "left":
        msp.add_line((x1 - ext_start, y1), (ox1 - ext_past, oy1), dxfattribs={"layer": "DIMENSIONS"})
        msp.add_line((x2 - ext_start, y2), (ox2 - ext_past, oy2), dxfattribs={"layer": "DIMENSIONS"})
    elif offset_dir == "right":
        msp.add_line((x1 + ext_start, y1), (ox1 + ext_past, oy1), dxfattribs={"layer": "DIMENSIONS"})
        msp.add_line((x2 + ext_start, y2), (ox2 + ext_past, oy2), dxfattribs={"layer": "DIMENSIONS"})

    # Dimension line
    msp.add_line((ox1, oy1), (ox2, oy2), dxfattribs={"layer": "DIMENSIONS"})

    # Diagonal tick marks (slash style, like reference)
    tick = DIM_TICK_SIZE
    if offset_dir in ("down", "up"):
        for tx, ty_val in [(ox1, oy1), (ox2, oy2)]:
            msp.add_line((tx - tick * 0.5, ty_val - tick * 0.5),
                         (tx + tick * 0.5, ty_val + tick * 0.5),
                         dxfattribs={"layer": "DIMENSIONS"})
    else:
        for pt in [(ox1, oy1), (ox2, oy2)]:
            msp.add_line((pt[0] - tick * 0.5, pt[1] - tick * 0.5),
                         (pt[0] + tick * 0.5, pt[1] + tick * 0.5),
                         dxfattribs={"layer": "DIMENSIONS"})

    # Register dimension line as obstacle
    if registry:
        registry.register_line(ox1, oy1, ox2, oy2, thickness=10)

    # Text
    dist = math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
    text = _dim_text(dist)
    mid_x = (ox1 + ox2) / 2
    mid_y = (oy1 + oy2) / 2

    # Scale text size down if span is very small (avoid text wider than dim line)
    text_h = DIM_TEXT_HEIGHT
    text_width_est = len(text) * text_h * 0.6
    if offset_dir in ("down", "up") and text_width_est > dist * 0.9:
        text_h = max(60, dist * 0.9 / (len(text) * 0.6))
    elif offset_dir in ("left", "right") and text_width_est > dist * 0.9:
        text_h = max(60, dist * 0.9 / (len(text) * 0.6))

    if offset_dir in ("down", "up"):
        tcx, tcy = mid_x, mid_y + text_h * 0.6
        if registry:
            registry.place_text(text, text_h, tcx, tcy, rotation=0,
                                align=TextEntityAlignment.BOTTOM_CENTER,
                                dim_axis="h", dim_span=dist)
        else:
            msp.add_text(text, height=text_h,
                         dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                (tcx, tcy), align=TextEntityAlignment.BOTTOM_CENTER)
    else:
        tcx, tcy = mid_x - text_h * 0.6, mid_y
        if registry:
            registry.place_text(text, text_h, tcx, tcy, rotation=90,
                                align=TextEntityAlignment.BOTTOM_CENTER,
                                dim_axis="v", dim_span=dist)
        else:
            msp.add_text(text, height=text_h, rotation=90,
                         dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                (tcx, tcy), align=TextEntityAlignment.BOTTOM_CENTER)


def draw_dimensions(msp, grid_x, grid_y, registry=None):
    """Draw dimension lines on all 4 sides + overall spans."""
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)

    # X individual dims at bottom only
    for i in range(len(grid_x) - 1):
        _draw_dim_line(msp, (grid_x[i], min_y), (grid_x[i + 1], min_y),
                       "down", DIM_OFFSET_1, registry=registry)

    # X overall dim at bottom
    if len(grid_x) > 2:
        _draw_dim_line(msp, (grid_x[0], min_y), (grid_x[-1], min_y),
                       "down", DIM_OFFSET_2, registry=registry)

    # Y individual dims at left only
    for i in range(len(grid_y) - 1):
        _draw_dim_line(msp, (min_x, grid_y[i]), (min_x, grid_y[i + 1]),
                       "left", DIM_OFFSET_1, registry=registry)

    # Y overall dim at left
    if len(grid_y) > 2:
        _draw_dim_line(msp, (min_x, grid_y[0]), (min_x, grid_y[-1]),
                       "left", DIM_OFFSET_2, registry=registry)


# ── 7. Building Perimeter (thick blue) ───────────────────────────────

def draw_perimeter(msp, columns):
    """Draw building perimeter through outermost column faces (thick blue)."""
    # Find the actual building extents from column footprints
    all_x1 = [c["x1"] for c in columns]
    all_x2 = [c["x2"] for c in columns]
    all_y1 = [c["y1"] for c in columns]
    all_y2 = [c["y2"] for c in columns]

    bx1 = min(all_x1)
    bx2 = max(all_x2)
    by1 = min(all_y1)
    by2 = max(all_y2)

    msp.add_lwpolyline(
        [(bx1, by1), (bx2, by1), (bx2, by2), (bx1, by2)],
        dxfattribs={"layer": "PERIMETER", "const_width": 18},
        close=True)


# ── 8. Plot Boundary (red inner line) ────────────────────────────────

def draw_plot_boundary(msp, grid_x, grid_y):
    """Draw red plot boundary line slightly inside the perimeter."""
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)

    # Plot boundary passes through grid line centers
    msp.add_lwpolyline(
        [(min_x, min_y), (max_x, min_y), (max_x, max_y), (min_x, max_y)],
        dxfattribs={"layer": "PLOT_BOUNDARY", "const_width": 6},
        close=True)


# ── 9. Title Block ───────────────────────────────────────────────────

def draw_title_block(msp, tb_x, tb_y):
    """Draw title block at bottom-right with PROJECT, ARCHITECTS, TITLE, etc."""
    layer = "TITLE_BLOCK"
    x = tb_x
    y = tb_y
    w = TB_WIDTH
    rh = TB_ROW_H
    th = NOTES_TEXT_H  # text height for entries

    # Rows listed top-to-bottom, but drawn bottom-to-top, so reverse
    rows_top_down = [
        ("PROJECT:-", rh * 1.2),
        ("", rh * 1.5),           # blank project name
        ("ARCHITECTS", rh * 1.0),
        ("", rh * 1.5),           # blank architect name
        ("TITLE:-", rh * 1.0),
        ("GRID LAYOUT PLAN", rh * 1.8),
        ("SCALE_ROW", rh * 1.5),
        ("STRUCT_ROW", rh * 1.5),
    ]
    # Reverse so first drawn row (at y) is STRUCT_ROW (bottom)
    rows = [(label, "", height) for label, height in reversed(rows_top_down)]

    # Draw outer border
    total_h = sum(r[2] for r in rows)
    msp.add_lwpolyline(
        [(x, y), (x + w, y), (x + w, y + total_h), (x, y + total_h)],
        dxfattribs={"layer": layer}, close=True)

    # Draw rows
    cy = y
    for i, (label, _, height) in enumerate(rows):
        # Horizontal separator
        if i > 0:
            msp.add_line((x, cy), (x + w, cy), dxfattribs={"layer": layer})

        text_y = cy + height / 2

        if label == "PROJECT:-":
            msp.add_text(label, height=NOTES_HEADER_H,
                         dxfattribs={"layer": "NOTES_HEADER"}).set_placement(
                (x + 100, text_y), align=TextEntityAlignment.MIDDLE_LEFT)
        elif label == "TITLE:-":
            msp.add_text(label, height=NOTES_HEADER_H,
                         dxfattribs={"layer": "NOTES_HEADER"}).set_placement(
                (x + 100, text_y), align=TextEntityAlignment.MIDDLE_LEFT)
        elif label == "GRID LAYOUT PLAN":
            msp.add_text(label, height=NOTES_HEADER_H * 1.1,
                         dxfattribs={"layer": "TITLE"}).set_placement(
                (x + w / 2, text_y), align=TextEntityAlignment.MIDDLE_CENTER)
        elif label == "ARCHITECTS":
            msp.add_text(label, height=th,
                         dxfattribs={"layer": layer}).set_placement(
                (x + 100, text_y), align=TextEntityAlignment.MIDDLE_LEFT)
        elif label == "SCALE_ROW":
            # Split into SCALE | SHEET NO | REV NO | DATE
            col_w = w / 4
            mid_line = cy + height / 2
            # Horizontal mid-line splitting label/value
            msp.add_line((x, mid_line), (x + w, mid_line), dxfattribs={"layer": layer})
            # Vertical separators
            for j in range(1, 4):
                msp.add_line((x + col_w * j, cy), (x + col_w * j, cy + height),
                             dxfattribs={"layer": layer})
            # Top half: labels
            top_y = mid_line + (height / 2) * 0.5
            msp.add_text("SCALE", height=th * 0.85,
                         dxfattribs={"layer": layer}).set_placement(
                (x + col_w * 0.5, top_y), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_text("SHEET NO.", height=th * 0.8,
                         dxfattribs={"layer": layer}).set_placement(
                (x + col_w * 1.5, top_y), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_text("REV. NO.", height=th * 0.8,
                         dxfattribs={"layer": layer}).set_placement(
                (x + col_w * 2.5, top_y), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_text("DATE", height=th * 0.8,
                         dxfattribs={"layer": layer}).set_placement(
                (x + col_w * 3.5, top_y), align=TextEntityAlignment.MIDDLE_CENTER)
            # Bottom half: values
            bot_y = cy + (height / 2) * 0.5
            msp.add_text("N.T.S.", height=th * 0.85,
                         dxfattribs={"layer": layer}).set_placement(
                (x + col_w * 0.5, bot_y), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_text("01", height=th * 0.85,
                         dxfattribs={"layer": layer}).set_placement(
                (x + col_w * 1.5, bot_y), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_text("R00", height=th * 0.85,
                         dxfattribs={"layer": layer}).set_placement(
                (x + col_w * 2.5, bot_y), align=TextEntityAlignment.MIDDLE_CENTER)
        elif label == "STRUCT_ROW":
            mid = w / 2
            msp.add_line((x + mid, cy), (x + mid, cy + height),
                         dxfattribs={"layer": layer})
            msp.add_text("STRUCTURAL", height=th * 0.9,
                         dxfattribs={"layer": layer}).set_placement(
                (x + mid / 2, text_y + th * 0.7), align=TextEntityAlignment.BOTTOM_CENTER)
            msp.add_text("DRAWING", height=th * 0.9,
                         dxfattribs={"layer": layer}).set_placement(
                (x + mid / 2, text_y - th * 0.7), align=TextEntityAlignment.TOP_CENTER)
            msp.add_text("DRAWING NO.:-ST/01", height=th * 0.85,
                         dxfattribs={"layer": layer}).set_placement(
                (x + mid + mid / 2, text_y), align=TextEntityAlignment.MIDDLE_CENTER)

        cy += height

    return total_h


# ── 10. Revision / Print Tables ──────────────────────────────────────

def draw_revision_tables(msp, tb_x, tb_y_top):
    """Draw 'PRINT ISSUED TO' and 'REMARK' tables above title block."""
    layer = "TITLE_BLOCK"
    x = tb_x
    w = TB_WIDTH
    rh = 250
    th = NOTES_TEXT_H * 0.85

    # REMARK table (4 rows)
    y = tb_y_top
    cols = [800, 500, w - 1300]  # DATE, NO, REMARK, (BY implied)
    header_labels = ["DATE", "NO", "REMARK", "BY"]
    col_positions = [0, 800, 1300, w - 500]
    col_widths_x = [0, 800, 1300, w - 500, w]

    # Header row
    msp.add_lwpolyline(
        [(x, y), (x + w, y), (x + w, y + rh), (x, y + rh)],
        dxfattribs={"layer": layer}, close=True)
    for j, lbl in enumerate(header_labels):
        cx = x + (col_widths_x[j] + col_widths_x[j + 1]) / 2 if j < len(col_widths_x) - 1 else x + col_widths_x[j]
        msp.add_text(lbl, height=th,
                     dxfattribs={"layer": layer}).set_placement(
            (cx, y + rh / 2), align=TextEntityAlignment.MIDDLE_CENTER)
    # Vertical dividers
    for cx_val in col_widths_x[1:-1]:
        msp.add_line((x + cx_val, y), (x + cx_val, y + rh),
                     dxfattribs={"layer": layer})

    # 4 empty rows
    for i in range(4):
        ry = y + rh + i * rh
        msp.add_lwpolyline(
            [(x, ry), (x + w, ry), (x + w, ry + rh), (x, ry + rh)],
            dxfattribs={"layer": layer}, close=True)
        for cx_val in col_widths_x[1:-1]:
            msp.add_line((x + cx_val, ry), (x + cx_val, ry + rh),
                         dxfattribs={"layer": layer})

    remark_top = y + rh * 5

    # PRINT ISSUED TO table (same structure, above remark)
    y2 = remark_top + 200  # gap
    header_labels2 = ["DATE", "NO", "PRINT ISSUED TO", "BY"]

    msp.add_lwpolyline(
        [(x, y2), (x + w, y2), (x + w, y2 + rh), (x, y2 + rh)],
        dxfattribs={"layer": layer}, close=True)
    for j, lbl in enumerate(header_labels2):
        cx = x + (col_widths_x[j] + col_widths_x[j + 1]) / 2 if j < len(col_widths_x) - 1 else x + col_widths_x[j]
        msp.add_text(lbl, height=th,
                     dxfattribs={"layer": layer}).set_placement(
            (cx, y2 + rh / 2), align=TextEntityAlignment.MIDDLE_CENTER)
    for cx_val in col_widths_x[1:-1]:
        msp.add_line((x + cx_val, y2), (x + cx_val, y2 + rh),
                     dxfattribs={"layer": layer})

    for i in range(4):
        ry = y2 + rh + i * rh
        msp.add_lwpolyline(
            [(x, ry), (x + w, ry), (x + w, ry + rh), (x, ry + rh)],
            dxfattribs={"layer": layer}, close=True)
        for cx_val in col_widths_x[1:-1]:
            msp.add_line((x + cx_val, ry), (x + cx_val, ry + rh),
                         dxfattribs={"layer": layer})

    return y2 + rh * 5  # top of print table


# ── 11. General Notes Panel ──────────────────────────────────────────

def draw_notes_panel(msp, notes_x, notes_y, notes_h):
    """Draw General Notes panel on the right side of the drawing."""
    layer_body = "NOTES"
    layer_hdr = "NOTES_HEADER"
    x = notes_x
    y = notes_y
    w = NOTES_WIDTH
    h = notes_h
    th = NOTES_TEXT_H
    hh = NOTES_HEADER_H
    line_sp = th * 1.8  # line spacing

    # Outer border
    msp.add_lwpolyline(
        [(x, y), (x + w, y), (x + w, y + h), (x, y + h)],
        dxfattribs={"layer": layer_body}, close=True)

    # Content (top to bottom)
    cy = y + h - 200  # start from top with margin

    # GENERAL NOTES header
    msp.add_text("GENERAL NOTES:-", height=hh, dxfattribs={"layer": layer_body}).set_placement(
        (x + 150, cy), align=TextEntityAlignment.TOP_LEFT)
    cy -= hh + 50
    msp.add_line((x, cy), (x + w, cy), dxfattribs={"layer": layer_body})
    cy -= line_sp

    notes = [
        "G1. DO NOT SCALE THE DRAWING FOLLOW ONLY FIGURED",
        "    DIMENSIONS.",
        "G2. ALL STRUCTURAL DRAWINGS SHOULD BE READ IN",
        "    CONJUNCTION WITH RELEVANT ARCHITECTURAL DRAWINGS.",
        "    ANY DISCREPANCY OR AMBIGUITY IN EITHER SHOULD",
        "    BE BROUGHT TO THE NOTICE OF THE ARCHITECT",
        "G3. ALL DIMENSIONS ARE IN MM",
    ]
    for line in notes:
        msp.add_text(line, height=th, dxfattribs={"layer": layer_body}).set_placement(
            (x + 150, cy), align=TextEntityAlignment.TOP_LEFT)
        cy -= line_sp

    cy -= line_sp * 0.5
    msp.add_line((x, cy), (x + w, cy), dxfattribs={"layer": layer_body})
    cy -= line_sp

    # CONCRETE section
    msp.add_text("CONCRETE:-", height=hh, dxfattribs={"layer": layer_hdr}).set_placement(
        (x + 150, cy), align=TextEntityAlignment.TOP_LEFT)
    cy -= hh + line_sp
    msp.add_text("ALL R.C.C WORK SHALL BE IN MIX M25", height=th,
                 dxfattribs={"layer": layer_body}).set_placement(
        (x + 150, cy), align=TextEntityAlignment.TOP_LEFT)

    cy -= line_sp * 2
    msp.add_line((x, cy), (x + w, cy), dxfattribs={"layer": layer_body})
    cy -= line_sp

    # REINFORCING STEEL section
    msp.add_text("REINFORCING STEEL:-", height=hh, dxfattribs={"layer": layer_hdr}).set_placement(
        (x + 150, cy), align=TextEntityAlignment.TOP_LEFT)
    cy -= hh + line_sp

    steel_notes = [
        "1. ALL REINFORCING STEEL WILL BE OF TESTED QUALITY",
        "   CONFORMING TO IS:1786 LATEST.",
        "2. REFER TO HIGH YIELD STRENGTH DEFORMED BARS",
        "   WITH CHARACTERISTIC STRENGTH OF 500 N/sq. mm.",
        "3. CLEAR COVER TO MAIN REINFORCEMENT SHALL BE",
        "   * FOUNDATION    50 mm. ALL AROUND",
        "   * COLUMNS       40 mm. ALL AROUND",
        "   * BEAMS         30 mm. ALL AROUND",
        "   * SLABS         20 mm. TOP & BOTTOM",
        "   * WALLS         25 mm. EARTH FACE",
        "   * WALLS         20 mm. INNER FACE",
        "4. LAP LENGTH TO BE 50xDIA OF BAR MINIMUM.",
        "5. SLAB BARS IN SHORTER DIRECTION, SHALL BE",
        "   BELOW BARS FOR THE LONGER DIRECTION",
        "6. IN BEAMS, FIRST STIRRUP SHALL BE AT NO MORE THAN",
        "   40 mm FROM FACE OF THE SUPPORTING MEMBER.",
        "7. IN BEAMS TOP BARS ARE NOT TO BE SPLICED IN THE",
        "   END QUARTERS OF THE SPAN, AND THE BOTTOM BARS",
        "   ARE NOT TO BE SPLICED AT MIDDLE HALF OF THE SPAN.",
        "8. SAFE BEARING CAPACITY OF SOIL HAS BEEN TAKEN AS",
        "   16 TON/SQM BELOW 10'-6\" FROM NGL(ASSUMED)",
        "9. FOUNDATION HAS BEEN DESIGNED FOR BASEMENT, STILT,",
        "   GROUND, FIRST & SECOND FLOORS ONLY.",
        "10. DO NOT SCALE FOLLOW WRITTEN DIMENSION ONLY.",
    ]
    for line in steel_notes:
        msp.add_text(line, height=th, dxfattribs={"layer": layer_body}).set_placement(
            (x + 150, cy), align=TextEntityAlignment.TOP_LEFT)
        cy -= line_sp


# ── 14. Clear Span Dimensions (between column faces) ─────────────────

CLEAR_DIM_OFFSET = 350  # mm offset from grid line for clear span dim line

def draw_clear_span_dims(msp, columns, grid_x, grid_y, registry=None):
    """Draw clear span dimensions between adjacent column faces
    along each grid row (horizontal) and grid column (vertical)."""
    tol = 300  # mm tolerance for matching column center to grid line
    layer = "FOOTING_DIM"
    th = FOOTING_DIM_HEIGHT * 0.85
    tick = DIM_TICK_SIZE * 0.6

    def nearest(val, grid_vals):
        best = min(grid_vals, key=lambda g: abs(g - val))
        return best if abs(best - val) < tol else None

    # ── Horizontal clear spans (along grid Y rows) ──
    y_groups = {}
    for col in columns:
        cy = (col["y1"] + col["y2"]) / 2
        gy = nearest(cy, grid_y)
        if gy is not None:
            y_groups.setdefault(gy, []).append(col)

    for gy, group in y_groups.items():
        group.sort(key=lambda c: c["x1"])
        for i in range(len(group) - 1):
            span_x1 = group[i]["x2"]       # right face of left column
            span_x2 = group[i + 1]["x1"]   # left face of right column
            clear = span_x2 - span_x1
            if clear < 50:
                continue

            y_dim = gy - CLEAR_DIM_OFFSET
            # Dim line
            msp.add_line((span_x1, y_dim), (span_x2, y_dim),
                         dxfattribs={"layer": layer})
            # Ticks
            for tx in [span_x1, span_x2]:
                msp.add_line((tx - tick * 0.5, y_dim - tick * 0.5),
                             (tx + tick * 0.5, y_dim + tick * 0.5),
                             dxfattribs={"layer": layer})
            # Extension lines from column face to dim line
            for tx in [span_x1, span_x2]:
                msp.add_line((tx, gy - 40), (tx, y_dim - tick),
                             dxfattribs={"layer": layer})
            # Register dim line
            if registry:
                registry.register_line(span_x1, y_dim, span_x2, y_dim, thickness=10)
            # Text
            cx = (span_x1 + span_x2) / 2
            text = _dim_text(clear)
            text_h = th
            text_w_est = len(text) * th * 0.6
            if text_w_est > clear * 0.9:
                text_h = max(50, clear * 0.9 / (len(text) * 0.6))
            tcx, tcy = cx, y_dim + text_h * 0.3
            if registry:
                registry.place_text(text, text_h, tcx, tcy, rotation=0,
                                    align=TextEntityAlignment.BOTTOM_CENTER,
                                    dim_axis="h", dim_span=clear)
            else:
                msp.add_text(text, height=text_h,
                             dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                    (tcx, tcy), align=TextEntityAlignment.BOTTOM_CENTER)

    # ── Vertical clear spans (along grid X columns) ──
    x_groups = {}
    for col in columns:
        ccx = (col["x1"] + col["x2"]) / 2
        gx = nearest(ccx, grid_x)
        if gx is not None:
            x_groups.setdefault(gx, []).append(col)

    for gx, group in x_groups.items():
        group.sort(key=lambda c: c["y1"])
        for i in range(len(group) - 1):
            span_y1 = group[i]["y2"]       # top face of bottom column
            span_y2 = group[i + 1]["y1"]   # bottom face of top column
            clear = span_y2 - span_y1
            if clear < 50:
                continue

            x_dim = gx - CLEAR_DIM_OFFSET
            msp.add_line((x_dim, span_y1), (x_dim, span_y2),
                         dxfattribs={"layer": layer})
            for ty in [span_y1, span_y2]:
                msp.add_line((x_dim - tick * 0.5, ty - tick * 0.5),
                             (x_dim + tick * 0.5, ty + tick * 0.5),
                             dxfattribs={"layer": layer})
            for ty in [span_y1, span_y2]:
                msp.add_line((gx - 40, ty), (x_dim - tick, ty),
                             dxfattribs={"layer": layer})
            # Register dim line
            if registry:
                registry.register_line(x_dim, span_y1, x_dim, span_y2, thickness=10)
            cy = (span_y1 + span_y2) / 2
            text = _dim_text(clear)
            text_h = th
            text_w_est = len(text) * th * 0.6
            if text_w_est > clear * 0.9:
                text_h = max(50, clear * 0.9 / (len(text) * 0.6))
            tcx, tcy = x_dim - text_h * 0.3, cy
            if registry:
                registry.place_text(text, text_h, tcx, tcy, rotation=90,
                                    align=TextEntityAlignment.BOTTOM_CENTER,
                                    dim_axis="v", dim_span=clear)
            else:
                msp.add_text(text, height=text_h, rotation=90,
                             dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                    (tcx, tcy), align=TextEntityAlignment.BOTTOM_CENTER)


# ── 15. Edge Offset Dimensions (grid to building perimeter) ──────────

def draw_edge_offset_dims(msp, columns, grid_x, grid_y, registry=None):
    """Draw offset dimensions from grid lines to building perimeter
    (outermost column faces) on all 4 edges."""
    layer = "FOOTING_DIM"
    th = FOOTING_DIM_HEIGHT * 0.85
    tick = DIM_TICK_SIZE * 0.6

    # Building perimeter extents
    bx1 = min(c["x1"] for c in columns)
    bx2 = max(c["x2"] for c in columns)
    by1 = min(c["y1"] for c in columns)
    by2 = max(c["y2"] for c in columns)

    min_gx, max_gx = min(grid_x), max(grid_x)
    min_gy, max_gy = min(grid_y), max(grid_y)

    offset_y = 500  # vertical offset for horizontal edge dims
    offset_x = 500  # horizontal offset for vertical edge dims

    def _draw_h_edge(x_from, x_to, y_base, y_off, above):
        """Draw a horizontal edge offset dim."""
        dist = abs(x_to - x_from)
        if dist < 10:
            return
        y_dim = y_base + y_off if above else y_base - y_off
        msp.add_line((x_from, y_dim), (x_to, y_dim),
                     dxfattribs={"layer": layer})
        if registry:
            registry.register_line(x_from, y_dim, x_to, y_dim, thickness=10)
        for tx in [x_from, x_to]:
            msp.add_line((tx - tick * 0.5, y_dim - tick * 0.5),
                         (tx + tick * 0.5, y_dim + tick * 0.5),
                         dxfattribs={"layer": layer})
            ext_y1 = y_base + 40 if above else y_base - 40
            ext_y2 = y_dim + tick if above else y_dim - tick
            msp.add_line((tx, ext_y1), (tx, ext_y2),
                         dxfattribs={"layer": layer})
        cx = (x_from + x_to) / 2
        text = _dim_text(dist)
        ty = y_dim + th * 0.3 if above else y_dim - th * 0.3
        align = TextEntityAlignment.BOTTOM_CENTER if above else TextEntityAlignment.TOP_CENTER
        if registry:
            registry.place_text(text, th, cx, ty, rotation=0, align=align,
                                dim_axis="h", dim_span=dist)
        else:
            msp.add_text(text, height=th,
                         dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                (cx, ty), align=align)

    def _draw_v_edge(y_from, y_to, x_base, x_off, right):
        """Draw a vertical edge offset dim."""
        dist = abs(y_to - y_from)
        if dist < 10:
            return
        x_dim = x_base + x_off if right else x_base - x_off
        msp.add_line((x_dim, y_from), (x_dim, y_to),
                     dxfattribs={"layer": layer})
        if registry:
            registry.register_line(x_dim, y_from, x_dim, y_to, thickness=10)
        for ty in [y_from, y_to]:
            msp.add_line((x_dim - tick * 0.5, ty - tick * 0.5),
                         (x_dim + tick * 0.5, ty + tick * 0.5),
                         dxfattribs={"layer": layer})
            ext_x1 = x_base + 40 if right else x_base - 40
            ext_x2 = x_dim + tick if right else x_dim - tick
            msp.add_line((ext_x1, ty), (ext_x2, ty),
                         dxfattribs={"layer": layer})
        cy = (y_from + y_to) / 2
        text = _dim_text(dist)
        tx = x_dim + th * 0.3 if right else x_dim - th * 0.3
        if registry:
            registry.place_text(text, th, tx, cy, rotation=90,
                                align=TextEntityAlignment.MIDDLE_CENTER,
                                dim_axis="v", dim_span=dist)
        else:
            msp.add_text(text, height=th, rotation=90,
                         dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                (tx, cy), align=TextEntityAlignment.MIDDLE_CENTER)

    # Bottom edge: building perimeter y1 to grid y[0]
    _draw_h_edge(bx1, min_gx, min_gy, offset_y, above=False)
    # Top edge: grid y[-1] to building perimeter y2
    _draw_h_edge(max_gx, bx2, max_gy, offset_y, above=True)
    # Left edge: building perimeter x1 to grid x[0]
    _draw_v_edge(by1, min_gy, min_gx, offset_x, right=False)
    # Right edge: grid x[-1] to building perimeter x2
    _draw_v_edge(max_gy, by2, max_gx, offset_x, right=True)


# ── 16. Grid-to-Face Offset Dimensions ──────────────────────────────

def draw_grid_face_offsets(msp, columns, grid_x, grid_y, registry=None):
    """Draw offset dimensions from grid line to each column face.
    For Centre columns, shows YD/2 on each side of grid line.
    For Left/Right, shows single offset (grid to far face)."""
    tol = 300
    layer = "FOOTING_DIM"
    th = FOOTING_DIM_HEIGHT * 0.7
    tick = DIM_TICK_SIZE * 0.5

    def nearest(val, grid_vals):
        best = min(grid_vals, key=lambda g: abs(g - val))
        return best if abs(best - val) < tol else None

    for col in columns:
        cx = (col["x1"] + col["x2"]) / 2
        cy = (col["y1"] + col["y2"]) / 2
        x1, y1, x2, y2 = col["x1"], col["y1"], col["x2"], col["y2"]

        gx = nearest(cx, grid_x)
        gy = nearest(cy, grid_y)

        # X-direction: grid line to column faces
        if gx is not None:
            left_off = abs(gx - x1)
            right_off = abs(x2 - gx)

            # Only show offsets if the grid line is INSIDE the column (Centre)
            if left_off > 10 and right_off > 10:
                dim_y = y2 + FOOTING_LINE_EXTEND * 1.5  # past footing annotation
                col_w = abs(x2 - x1)

                # Draw dim line across full column width
                msp.add_line((x1, dim_y), (x2, dim_y),
                             dxfattribs={"layer": layer})
                if registry:
                    registry.register_line(x1, dim_y, x2, dim_y, thickness=10)

                # Tick at grid line and faces
                msp.add_line((gx - tick * 0.5, dim_y - tick * 0.5),
                             (gx + tick * 0.5, dim_y + tick * 0.5),
                             dxfattribs={"layer": layer})
                for tx in [x1, x2]:
                    msp.add_line((tx - tick * 0.5, dim_y - tick * 0.5),
                                 (tx + tick * 0.5, dim_y + tick * 0.5),
                                 dxfattribs={"layer": layer})

                if col_w < 500 or abs(left_off - right_off) < 20:
                    # Small column or equal halves: single centered text
                    text_c = _dim_text(left_off) + "|" + _dim_text(right_off) if abs(left_off - right_off) >= 20 else _dim_text(left_off)
                    tcx_c, tcy_c = (x1 + x2) / 2, dim_y + th * 0.3
                    if registry:
                        registry.place_text(text_c, th, tcx_c, tcy_c, rotation=0,
                                            align=TextEntityAlignment.BOTTOM_CENTER,
                                            dim_axis="h", dim_span=col_w)
                    else:
                        msp.add_text(text_c, height=th,
                                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                            (tcx_c, tcy_c), align=TextEntityAlignment.BOTTOM_CENTER)
                else:
                    # Wide column: two separate texts
                    text_l = _dim_text(left_off)
                    tcx_l, tcy_l = (x1 + gx) / 2, dim_y + th * 0.3
                    if registry:
                        registry.place_text(text_l, th, tcx_l, tcy_l, rotation=0,
                                            align=TextEntityAlignment.BOTTOM_CENTER,
                                            dim_axis="h", dim_span=left_off)
                    else:
                        msp.add_text(text_l, height=th,
                                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                            (tcx_l, tcy_l), align=TextEntityAlignment.BOTTOM_CENTER)
                    text_r = _dim_text(right_off)
                    tcx_r, tcy_r = (gx + x2) / 2, dim_y + th * 0.3
                    if registry:
                        registry.place_text(text_r, th, tcx_r, tcy_r, rotation=0,
                                            align=TextEntityAlignment.BOTTOM_CENTER,
                                            dim_axis="h", dim_span=right_off)
                    else:
                        msp.add_text(text_r, height=th,
                                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                            (tcx_r, tcy_r), align=TextEntityAlignment.BOTTOM_CENTER)

        # Y-direction: grid line to column faces
        if gy is not None:
            front_off = abs(gy - y1)
            back_off = abs(y2 - gy)

            if front_off > 10 and back_off > 10:
                # Centre Y: show both halves
                dim_x = x2 + FOOTING_LINE_EXTEND * 1.5  # past footing annotation
                col_h = abs(y2 - y1)

                # Draw dim line across full column height
                msp.add_line((dim_x, y1), (dim_x, y2),
                             dxfattribs={"layer": layer})
                if registry:
                    registry.register_line(dim_x, y1, dim_x, y2, thickness=10)

                # Tick at grid line and faces
                msp.add_line((dim_x - tick * 0.5, gy - tick * 0.5),
                             (dim_x + tick * 0.5, gy + tick * 0.5),
                             dxfattribs={"layer": layer})
                for ty in [y1, y2]:
                    msp.add_line((dim_x - tick * 0.5, ty - tick * 0.5),
                                 (dim_x + tick * 0.5, ty + tick * 0.5),
                                 dxfattribs={"layer": layer})

                if col_h < 500 or abs(front_off - back_off) < 20:
                    # Small column or equal halves: single centered text
                    text_c = _dim_text(front_off) + "|" + _dim_text(back_off) if abs(front_off - back_off) >= 20 else _dim_text(front_off)
                    tcx_c, tcy_c = dim_x + th * 0.3, (y1 + y2) / 2
                    if registry:
                        registry.place_text(text_c, th, tcx_c, tcy_c, rotation=90,
                                            align=TextEntityAlignment.MIDDLE_CENTER,
                                            dim_axis="v", dim_span=col_h)
                    else:
                        msp.add_text(text_c, height=th, rotation=90,
                                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                            (tcx_c, tcy_c), align=TextEntityAlignment.MIDDLE_CENTER)
                else:
                    # Tall column: two separate texts
                    text_f = _dim_text(front_off)
                    tcx_f, tcy_f = dim_x + th * 0.3, (y1 + gy) / 2
                    if registry:
                        registry.place_text(text_f, th, tcx_f, tcy_f, rotation=90,
                                            align=TextEntityAlignment.MIDDLE_CENTER,
                                            dim_axis="v", dim_span=front_off)
                    else:
                        msp.add_text(text_f, height=th, rotation=90,
                                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                            (tcx_f, tcy_f), align=TextEntityAlignment.MIDDLE_CENTER)
                    text_b = _dim_text(back_off)
                    tcx_b, tcy_b = dim_x + th * 0.3, (gy + y2) / 2
                    if registry:
                        registry.place_text(text_b, th, tcx_b, tcy_b, rotation=90,
                                            align=TextEntityAlignment.MIDDLE_CENTER,
                                            dim_axis="v", dim_span=back_off)
                    else:
                        msp.add_text(text_b, height=th, rotation=90,
                                     dxfattribs={"layer": "DIM_TEXT"}).set_placement(
                            (tcx_b, tcy_b), align=TextEntityAlignment.MIDDLE_CENTER)


# ── 12. Foundation Plan Title (large, below drawing) ─────────────────

def draw_main_title(msp, grid_x, grid_y):
    """Draw 'GRID LAYOUT PLAN' title centered below the grid label circles."""
    cx = (min(grid_x) + max(grid_x)) / 2
    # Place below the bottom grid label circles (GRID_LABEL_OFFSET + CIRCLE_RADIUS + gap)
    ty = min(grid_y) - GRID_LABEL_OFFSET - CIRCLE_RADIUS - 800

    msp.add_text("GRID LAYOUT PLAN", height=500,
                 dxfattribs={"layer": "TITLE"}).set_placement(
        (cx, ty), align=TextEntityAlignment.MIDDLE_CENTER)


# ── 13. Magenta Sheet Border ─────────────────────────────────────────

def draw_sheet_border(msp, bounds):
    """Draw sheet border: 3 nested rectangles (magenta-cyan-magenta) + corner marks."""
    bx1, by1, bx2, by2 = bounds
    margin = BORDER_MARGIN
    gap1 = 150   # outer magenta to cyan
    gap2 = 150   # cyan to inner magenta

    # Layer 1: Outer magenta
    ox1, oy1 = bx1 - margin, by1 - margin
    ox2, oy2 = bx2 + margin, by2 + margin
    msp.add_lwpolyline(
        [(ox1, oy1), (ox2, oy1), (ox2, oy2), (ox1, oy2)],
        dxfattribs={"layer": "BORDER"},
        close=True)

    # Layer 2: Middle cyan
    cx1, cy1 = ox1 + gap1, oy1 + gap1
    cx2, cy2 = ox2 - gap1, oy2 - gap1
    msp.add_lwpolyline(
        [(cx1, cy1), (cx2, cy1), (cx2, cy2), (cx1, cy2)],
        dxfattribs={"layer": "DIMENSIONS"},  # cyan
        close=True)

    # Layer 3: Inner magenta
    ix1, iy1 = cx1 + gap2, cy1 + gap2
    ix2, iy2 = cx2 - gap2, cy2 - gap2
    msp.add_lwpolyline(
        [(ix1, iy1), (ix2, iy1), (ix2, iy2), (ix1, iy2)],
        dxfattribs={"layer": "BORDER"},
        close=True)

    # Corner marks — small L-shaped magenta marks at the outermost corners
    corner_len = 500  # length of corner mark lines
    corner_offset = 200  # distance outside outer border
    corners = [
        (ox1 - corner_offset, oy1 - corner_offset),  # bottom-left
        (ox2 + corner_offset, oy1 - corner_offset),  # bottom-right
        (ox2 + corner_offset, oy2 + corner_offset),  # top-right
        (ox1 - corner_offset, oy2 + corner_offset),  # top-left
    ]
    # L-shape directions for each corner: (dx_horiz, dy_vert)
    dirs = [
        (1, 1),    # bottom-left: right + up
        (-1, 1),   # bottom-right: left + up
        (-1, -1),  # top-right: left + down
        (1, -1),   # top-left: right + down
    ]
    for (cx, cy), (dx, dy) in zip(corners, dirs):
        # Horizontal arm
        msp.add_line((cx, cy), (cx + corner_len * dx, cy),
                     dxfattribs={"layer": "BORDER"})
        # Vertical arm
        msp.add_line((cx, cy), (cx, cy + corner_len * dy),
                     dxfattribs={"layer": "BORDER"})


# ── Main ─────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  GRID DRAWING GENERATOR")
    print("=" * 60)
    print()

    # Find latest input files
    gridline_path = find_latest_file("gridline_coordinates_*.xlsx")
    node_path = find_latest_file("node_coordinates_*.xlsx")
    print(f"Gridline coordinates: {os.path.basename(gridline_path)}")
    print(f"Node coordinates: {os.path.basename(node_path)}")

    # Read data
    columns = read_gridline_coordinates(gridline_path)
    grid_x, grid_y = read_unique_grid_values(node_path)
    print(f"\n{len(columns)} columns")
    print(f"Grid X ({len(grid_x)}): {grid_x}")
    print(f"Grid Y ({len(grid_y)}): {grid_y}")

    # Create DXF document
    doc = setup_document()
    msp = doc.modelspace()

    # Create placement registry for collision avoidance
    registry = PlacementRegistry(msp)

    # Calculate layout bounds
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)

    # Right panel X position (after grid labels on right side)
    panel_x = max_x + GRID_LABEL_OFFSET + CIRCLE_RADIUS + 500

    print("\nDrawing elements...")

    print("  1. Grid lines")
    draw_grid_lines(msp, grid_x, grid_y)

    print("  2. Grid labels")
    draw_grid_labels(msp, grid_x, grid_y)

    print("  3. Column rectangles")
    draw_column_rectangles(msp, columns)

    # Register grid lines and column rectangles as obstacles
    registry.register_grid_obstacles(grid_x, grid_y, columns)

    print("  4. Column labels")
    draw_column_labels(msp, columns)

    print("  5. Footing annotations")
    draw_footing_annotations(msp, columns, registry=registry)

    print("  6. Grid dimensions")
    draw_dimensions(msp, grid_x, grid_y, registry=registry)

    print("  6a. Clear span dimensions")
    draw_clear_span_dims(msp, columns, grid_x, grid_y, registry=registry)

    print("  6b. Edge offset dimensions")
    draw_edge_offset_dims(msp, columns, grid_x, grid_y, registry=registry)

    print("  6c. Grid-to-face offsets")
    draw_grid_face_offsets(msp, columns, grid_x, grid_y, registry=registry)

    print("  7. Building perimeter")
    draw_perimeter(msp, columns)

    print("  8. Plot boundary")
    draw_plot_boundary(msp, grid_x, grid_y)

    print("  9. Title block")
    # Title block at bottom-right of panel area
    tb_x = panel_x
    tb_y = min_y - GRID_LINE_END  # align with bottom of drawing
    tb_h = draw_title_block(msp, tb_x, tb_y)

    print("  10. Revision tables")
    rev_top = draw_revision_tables(msp, tb_x, tb_y + tb_h + 200)

    print("  11. General notes")
    notes_y = rev_top + 200
    notes_h = max_y + GRID_LINE_END - notes_y
    if notes_h < 5000:
        notes_h = 5000
    draw_notes_panel(msp, panel_x, notes_y, notes_h)

    print("  12. Main title")
    draw_main_title(msp, grid_x, grid_y)

    print("  13. Sheet border")
    # Calculate total bounds for border — include main title below
    main_title_y = min_y - GRID_LABEL_OFFSET - CIRCLE_RADIUS - 800 - 500
    border_left = min_x - GRID_LINE_END
    border_right = panel_x + NOTES_WIDTH
    border_bottom = min(tb_y - 200, main_title_y - 300)
    border_top = max_y + GRID_LINE_END
    draw_sheet_border(msp, (border_left, border_bottom, border_right, border_top))

    # Save
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outpath = os.path.join(OUTPUT_FOLDER, f"grid_layout_{timestamp}.dxf")
    doc.saveas(outpath)
    print(f"\nSaved: {outpath}")


if __name__ == "__main__":
    main()
