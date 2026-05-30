"""
Framing Plan Generator V2 (Hybrid)
-----------------------------------
Generates ALL FLOOR LEVEL FRAMING PLAN DXF.

Beam data:  Databeametabsupdated.xlsx (explicit beam segments with start/end coords)
Old data:   other_coordinates_*.xlsx (lift, staircase, mumty rectangles)

X in beam sheet = X in drawing, Z in beam sheet = Y in drawing.

Usage: python framing_plan_v2.py
       python framing_plan_v2.py path_to_beam_excel.xlsx
"""

import os
import sys
import glob
import openpyxl
import ezdxf
from ezdxf.enums import TextEntityAlignment
from datetime import datetime

from dxf_export import save_dxf_and_dwg

# ── Paths ──────────────────────────────────────────────────────────────
DEFAULT_INPUT = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\Databeametabsupdated.xlsx"
RECT_GEOM_PATH = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\Rectangle_geometry_filled.xlsx"
STD_ANL_FOLDER = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model"
OUTPUT_FOLDER = os.path.join(STD_ANL_FOLDER, "final dxf")

# Notes panel
NOTES_WIDTH = 5500

# ── Drawing constants (all in mm) ─────────────────────────────────────
SCALE = 1000  # metres to mm

# Grid
GRID_EXTEND = 2500
CIRCLE_RADIUS = 350
GRID_LABEL_HEIGHT = 250

# Dimensions
DIM_OFFSET_1 = 800
DIM_OFFSET_2 = 1800
DIM_TEXT_HEIGHT = 120
DIM_TICK_SIZE = 80

# Layout
FLOOR_GAP = 5000
MUMTY_GAP = 4000

# Title
TITLE_HEIGHT = 300
TITLE_GAP = 800

# Colors (AutoCAD Color Index)
CLR_WHITE = 7
CLR_RED = 1
CLR_YELLOW = 2
CLR_GREEN = 3
CLR_CYAN = 4
CLR_BLUE = 5
CLR_MAGENTA = 6
CLR_GREY = 253


# ══════════════════════════════════════════════════════════════════════
#  DATA READING
# ══════════════════════════════════════════════════════════════════════

def read_beam_data(filepath):
    """Read beam data from Excel. Returns list of beam dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Databeam"]
    beams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        beams.append({
            "member_no": row[0],
            "start_node": row[1],
            "end_node": row[2],
            "start_x": row[3],   # metres
            "start_z": row[5],   # metres (Z in sheet = Y in drawing)
            "end_x": row[6],
            "end_z": row[8],
            "length": row[9],
            "beam_type": row[10],
            "floor": row[11],
            "yd": row[12] or 300,     # beam depth (mm)
            "zd": row[13] or 230,     # beam width (mm)
            "wt": row[14] or 115,     # wall thickness (mm)
        })
    wb.close()
    return beams


def classify_beams(beams):
    """Classify beams as horizontal or vertical based on coordinates."""
    for b in beams:
        dz = abs(b["end_z"] - b["start_z"])
        dx = abs(b["end_x"] - b["start_x"])
        if dz < 0.001:
            b["direction"] = "horizontal"
        elif dx < 0.001:
            b["direction"] = "vertical"
        else:
            b["direction"] = "diagonal"
    return beams


def get_floor_beams(beams, floor_key):
    """Filter beams for a specific floor."""
    return [b for b in beams if b["floor"] == floor_key]


# ══════════════════════════════════════════════════════════════════════
#  OLD DATA READERS (for lift, staircase, mumty from other_coordinates)
# ══════════════════════════════════════════════════════════════════════

def find_latest_file(pattern):
    """Find the latest file matching a glob pattern in STD ANL folder."""
    files = glob.glob(os.path.join(STD_ANL_FOLDER, pattern))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def read_rectangles(filepath):
    """Read rectangle coordinates from other_coordinates Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    rects = []
    if "Rectangle coordinates" in wb.sheetnames:
        ws = wb["Rectangle coordinates"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rects.append({
                "no": row[0], "type": row[1],
                "x1": row[2], "y1": row[3], "x2": row[4], "y2": row[5],
                "location": row[6],
                "snapped_x1": row[13] if len(row) > 13 and row[13] else row[2],
                "snapped_y1": row[14] if len(row) > 14 and row[14] else row[3],
                "snapped_x2": row[15] if len(row) > 15 and row[15] else row[4],
                "snapped_y2": row[16] if len(row) > 16 and row[16] else row[5],
            })
    wb.close()
    return rects


GRID_SNAP_TOL = 300


def read_slab_panels(filepath):
    """Read slab panel geometry. Returns dict: floor_type -> list of panels."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Sheet1"]
    panels = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        floor_type = row[12]
        zone = row[11]
        if zone and str(zone).strip() in ("None", ""):
            zone = None
        panel = {
            "rect_no": row[1],
            "x_start": row[7],
            "x_end": row[8],
            "z_start": row[9],
            "z_end": row[10],
            "zone": zone,
        }
        panels.setdefault(floor_type, []).append(panel)
    wb.close()
    return panels


def read_unique_grid_values(filepath):
    """Read node_coordinates and extract unique grid X/Y values in mm."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Node coordinates"]
    xs, ys = set(), set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        xs.add(round(row[21] * SCALE, 1))
        ys.add(round(row[22] * SCALE, 1))
    wb.close()
    return sorted(xs), sorted(ys)


def read_gridline_coordinates(filepath):
    """Read gridline_coordinates Excel (new format with Floor Type and Y).
    Returns dict: floor_type -> list of column dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    all_columns = {}  # floor_type -> [columns]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # New format: No, Name, FloorType, Y, Location, AnchorX, AnchorY,
        #             BeamXloc, BeamYloc, Orientation, YD, ZD, X1, X2, Y1, Y2
        floor_type = row[2]
        raw_x1 = row[12] * SCALE
        raw_x2 = row[13] * SCALE
        raw_y1 = row[14] * SCALE
        raw_y2 = row[15] * SCALE
        col = {
            "no": row[0], "name": row[1],
            "floor_type": floor_type, "y_value": row[3],
            "location": row[4],
            "anchor_x": row[5] * SCALE, "anchor_y": row[6] * SCALE,
            "beam_x_loc": row[7], "beam_y_loc": row[8],
            "orientation": row[9], "yd": row[10], "zd": row[11],
            "x1": min(raw_x1, raw_x2), "x2": max(raw_x1, raw_x2),
            "y1": min(raw_y1, raw_y2), "y2": max(raw_y1, raw_y2),
        }
        all_columns.setdefault(floor_type, []).append(col)
    wb.close()
    return all_columns


def get_columns_for_floor(all_columns, beam_floor):
    """Get columns for a specific beam floor.
    Maps beam floor names to gridline floor types."""
    # # ── Old floor map (kept for reference) ──
    # floor_map = {
    #     "Stilt": "Stilt floor",
    #     "Typical": "Stilt floor roof",
    #     "Terrace": "Terrace",
    #     "Mumty": "Mumty roof",
    # }
    floor_map = {
        "Stilt": "Plinth",
        "Typical": "Stilt roof",
        "Terrace": "Stilt roof",     # Terrace no longer separate; use Stilt roof
        "Mumty": "Mumty",
    }
    floor_type = floor_map.get(beam_floor, beam_floor)
    cols = all_columns.get(floor_type, [])
    if not cols:
        # Fallback: try Plinth if specific floor not found
        cols = all_columns.get("Plinth", [])
    return cols


# ══════════════════════════════════════════════════════════════════════
#  DXF DOCUMENT SETUP
# ══════════════════════════════════════════════════════════════════════

def setup_document():
    """Create DXF document with layers and linetypes."""
    doc = ezdxf.new("R2010")

    style = doc.styles.get("Standard")
    style.dxf.font = "romans.shx"

    doc.linetypes.add("CENTER", pattern="A,1250,-250,250,-250",
                       description="Center ____ _ ____ _ ____")

    doc.layers.add("S-GRID", color=CLR_RED, linetype="CENTER")
    doc.layers.add("S-COL.", color=CLR_GREEN)
    doc.layers.add("S-COL-HATCH", color=CLR_GREY)
    doc.layers.add("S-BEAM", color=CLR_YELLOW)
    doc.layers.add("S-LINE", color=CLR_RED)
    doc.layers.add("S-DIM", color=CLR_RED)
    doc.layers.add("S-DIM-TEXT", color=CLR_MAGENTA)
    doc.layers.add("TEXT", color=CLR_GREEN)
    doc.layers.add("S-BEAM-NO", color=CLR_WHITE)
    doc.layers.add("NOTES", color=CLR_WHITE)
    doc.layers.add("TITLE_BLOCK", color=CLR_WHITE)
    doc.layers.add("BORDER", color=CLR_MAGENTA)

    return doc


# ══════════════════════════════════════════════════════════════════════
#  DIMENSION TEXT
# ══════════════════════════════════════════════════════════════════════

def _dim_text(value_mm):
    """Format dimension in mm as feet-inches with half-inch rounding."""
    total_inches = value_mm / 25.4
    half_inches = round(total_inches * 2)
    total_half = half_inches

    feet = int(total_half // 24)
    remaining_half = int(total_half % 24)
    whole_inches = remaining_half // 2
    has_half = remaining_half % 2 == 1

    if feet == 0:
        if has_half:
            if whole_inches == 0:
                return '%%189"'
            return f'{whole_inches}%%189"'
        else:
            return f'{whole_inches}"'
    else:
        if whole_inches == 0 and not has_half:
            return f"{feet}'-0\""
        elif has_half:
            if whole_inches == 0:
                return f"{feet}'-%%189\""
            return f"{feet}'-{whole_inches}%%189\""
        else:
            return f"{feet}'-{whole_inches}\""


# ══════════════════════════════════════════════════════════════════════
#  DRAWING FUNCTIONS
# ══════════════════════════════════════════════════════════════════════

def draw_grid_stubs(msp, grid_x, grid_y, ox, oy):
    """Draw grid line stubs from circles to building edge."""
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)
    label_offset = GRID_EXTEND + CIRCLE_RADIUS + 200
    extension = 400

    attribs = {"layer": "S-GRID", "ltscale": 50.0}

    for x in grid_x:
        msp.add_line((ox + x, oy + max_y - extension),
                     (ox + x, oy + max_y + label_offset - CIRCLE_RADIUS),
                     dxfattribs=attribs)

    for y in grid_y:
        msp.add_line((ox + min_x + extension, oy + y),
                     (ox + min_x - label_offset + CIRCLE_RADIUS, oy + y),
                     dxfattribs=attribs)


def draw_grid_labels(msp, grid_x, grid_y, ox, oy):
    """Draw grid labels: numbers at left (Y), letters at top (X)."""
    min_x = min(grid_x)
    max_y = max(grid_y)
    label_offset = GRID_EXTEND + CIRCLE_RADIUS + 200

    # Y labels (1, 2, 3...) at left
    min_gap = CIRCLE_RADIUS * 2
    stagger_offset = 800
    y_staggered = set()
    for i in range(len(grid_y) - 1):
        if grid_y[i + 1] - grid_y[i] < min_gap:
            y_staggered.add(i + 1)

    for i, y in enumerate(grid_y):
        label = str(i + 1)
        base_lx = ox + min_x - label_offset
        ly = oy + y

        if i in y_staggered:
            lx = base_lx - stagger_offset
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_line((lx + CIRCLE_RADIUS, ly),
                         (base_lx - CIRCLE_RADIUS, ly),
                         dxfattribs={"layer": "S-GRID", "ltscale": 50.0})
        else:
            lx = base_lx
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)

    # X labels (A, B, C...) at top
    x_staggered = set()
    for i in range(len(grid_x) - 1):
        if grid_x[i + 1] - grid_x[i] < min_gap:
            x_staggered.add(i + 1)

    for i, x in enumerate(grid_x):
        # Support beyond Z (AA, AB, etc.)
        if i < 26:
            label = chr(65 + i)
        else:
            label = chr(65 + (i // 26) - 1) + chr(65 + (i % 26))

        lx = ox + x
        base_ly = oy + max_y + label_offset

        if i in x_staggered:
            ly = base_ly + stagger_offset
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_line((lx, ly - CIRCLE_RADIUS),
                         (lx, base_ly + CIRCLE_RADIUS),
                         dxfattribs={"layer": "S-GRID", "ltscale": 50.0})
        else:
            ly = base_ly
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)


def draw_beams(msp, floor_beams, columns, ox, oy):
    """Draw each beam as a pair of parallel lines showing beam width (ZD).
    At terminus endpoints (where no other beam of the same direction continues),
    retract inward by half_zd so edge lines meet perpendicular beam edges."""

    # Count how many H and V beams meet at each node
    node_counts = {}
    for b in floor_beams:
        sx = round(b["start_x"] * SCALE, 1)
        sy = round(b["start_z"] * SCALE, 1)
        ex = round(b["end_x"] * SCALE, 1)
        ey = round(b["end_z"] * SCALE, 1)
        for nx, ny in [(sx, sy), (ex, ey)]:
            if (nx, ny) not in node_counts:
                node_counts[(nx, ny)] = {"h": 0, "v": 0}
            if b["direction"] == "horizontal":
                node_counts[(nx, ny)]["h"] += 1
            elif b["direction"] == "vertical":
                node_counts[(nx, ny)]["v"] += 1

    retracted = 0
    for b in floor_beams:
        sx = b["start_x"] * SCALE
        sy = b["start_z"] * SCALE
        ex = b["end_x"] * SCALE
        ey = b["end_z"] * SCALE
        half_zd = b["zd"] / 2

        if b["direction"] == "horizontal":
            left_x = min(sx, ex)
            right_x = max(sx, ex)
            left_node = (round(left_x, 1), round(sy, 1))
            right_node = (round(right_x, 1), round(sy, 1))

            # Retract left end only if this is a terminus (no other H beam continues left)
            if node_counts.get(left_node, {}).get("h", 0) <= 1:
                left_x += half_zd
                retracted += 1
            # Retract right end only if this is a terminus
            if node_counts.get(right_node, {}).get("h", 0) <= 1:
                right_x -= half_zd
                retracted += 1

            msp.add_line((ox + left_x, oy + sy - half_zd),
                         (ox + right_x, oy + sy - half_zd),
                         dxfattribs={"layer": "S-BEAM"})
            msp.add_line((ox + left_x, oy + sy + half_zd),
                         (ox + right_x, oy + sy + half_zd),
                         dxfattribs={"layer": "S-BEAM"})
            # Close balcony beam ends
            if b["beam_type"] == "BALCONY_BEAM":
                msp.add_line((ox + left_x, oy + sy - half_zd),
                             (ox + left_x, oy + sy + half_zd),
                             dxfattribs={"layer": "S-BEAM"})
                msp.add_line((ox + right_x, oy + sy - half_zd),
                             (ox + right_x, oy + sy + half_zd),
                             dxfattribs={"layer": "S-BEAM"})

        elif b["direction"] == "vertical":
            bot_y = min(sy, ey)
            top_y = max(sy, ey)
            bot_node = (round(sx, 1), round(bot_y, 1))
            top_node = (round(sx, 1), round(top_y, 1))

            # Retract bottom only if terminus
            if node_counts.get(bot_node, {}).get("v", 0) <= 1:
                bot_y += half_zd
                retracted += 1
            # Retract top only if terminus
            if node_counts.get(top_node, {}).get("v", 0) <= 1:
                top_y -= half_zd
                retracted += 1

            msp.add_line((ox + sx - half_zd, oy + bot_y),
                         (ox + ex - half_zd, oy + top_y),
                         dxfattribs={"layer": "S-BEAM"})
            msp.add_line((ox + sx + half_zd, oy + bot_y),
                         (ox + ex + half_zd, oy + top_y),
                         dxfattribs={"layer": "S-BEAM"})
            # Close balcony beam ends
            if b["beam_type"] == "BALCONY_BEAM":
                msp.add_line((ox + sx - half_zd, oy + bot_y),
                             (ox + sx + half_zd, oy + bot_y),
                             dxfattribs={"layer": "S-BEAM"})
                msp.add_line((ox + ex - half_zd, oy + top_y),
                             (ox + ex + half_zd, oy + top_y),
                             dxfattribs={"layer": "S-BEAM"})

        elif b["direction"] == "diagonal":
            msp.add_line((ox + sx, oy + sy),
                         (ox + ex, oy + ey),
                         dxfattribs={"layer": "S-BEAM"})

    print(f"      ({retracted} terminus endpoints retracted)")

    # ── Balcony closing lines (outer edge connecting all balcony beams) ──
    col_y_min = min(c["y1"] for c in columns)
    col_y_max = max(c["y2"] for c in columns)
    front_bal_beams = [b for b in floor_beams if b["beam_type"] == "BALCONY_BEAM"
                       and min(b["start_z"], b["end_z"]) * SCALE < col_y_min - 100]
    back_bal_beams = [b for b in floor_beams if b["beam_type"] == "BALCONY_BEAM"
                      and max(b["start_z"], b["end_z"]) * SCALE > col_y_max + 100]

    for bal_group, is_front in [(front_bal_beams, True), (back_bal_beams, False)]:
        if len(bal_group) < 2:
            continue
        half_zd = bal_group[0]["zd"] / 2
        if is_front:
            outer_y = min(min(b["start_z"], b["end_z"]) * SCALE for b in bal_group) + half_zd
        else:
            outer_y = max(max(b["start_z"], b["end_z"]) * SCALE for b in bal_group) - half_zd
        all_left = []
        all_right = []
        for b in bal_group:
            sx = b["start_x"] * SCALE
            all_left.append(sx - half_zd)
            all_right.append(sx + half_zd)
        left_x = min(all_left)
        right_x = max(all_right)
        msp.add_line((ox + left_x, oy + outer_y),
                     (ox + right_x, oy + outer_y),
                     dxfattribs={"layer": "S-BEAM"})
        side = "Front" if is_front else "Back"
        print(f"      {side} balcony closing line: X={left_x:.0f}-{right_x:.0f}, Y={outer_y:.0f}")


def draw_columns(msp, columns, ox, oy):
    """Draw column rectangles with opaque fill + green outline.
    Drawn LAST so opaque fill hides beam overlaps at junctions."""
    for col in columns:
        x1, y1 = ox + col["x1"], oy + col["y1"]
        x2, y2 = ox + col["x2"], oy + col["y2"]

        # Opaque 2D SOLID fill
        msp.add_solid(
            [(x1, y1), (x2, y1), (x1, y2), (x2, y2)],
            dxfattribs={"layer": "S-COL-HATCH", "color": CLR_GREY})

        # Green outline
        msp.add_lwpolyline(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)],
            dxfattribs={"layer": "S-COL."}, close=True)


def draw_dimensions(msp, grid_x, grid_y, ox, oy):
    """Draw grid spacing dimensions at top (X) and left (Y)."""
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)
    tick = DIM_TICK_SIZE
    th = DIM_TEXT_HEIGHT

    # X dims at top
    for i in range(len(grid_x) - 1):
        x1, x2 = grid_x[i], grid_x[i + 1]
        y_dim = max_y + DIM_OFFSET_1
        msp.add_line((ox + x1, oy + y_dim), (ox + x2, oy + y_dim),
                     dxfattribs={"layer": "S-DIM"})
        for tx in [x1, x2]:
            msp.add_line((ox + tx - tick * 0.5, oy + y_dim - tick * 0.5),
                         (ox + tx + tick * 0.5, oy + y_dim + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + tx, oy + max_y + 200),
                         (ox + tx, oy + y_dim - tick),
                         dxfattribs={"layer": "S-DIM"})
        dist = x2 - x1
        msp.add_text(_dim_text(dist), height=th,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + (x1 + x2) / 2, oy + y_dim + th * 0.5),
            align=TextEntityAlignment.BOTTOM_CENTER)

    # X overall
    if len(grid_x) > 2:
        y_dim = max_y + DIM_OFFSET_2
        msp.add_line((ox + grid_x[0], oy + y_dim),
                     (ox + grid_x[-1], oy + y_dim),
                     dxfattribs={"layer": "S-DIM"})
        for tx in [grid_x[0], grid_x[-1]]:
            msp.add_line((ox + tx - tick * 0.5, oy + y_dim - tick * 0.5),
                         (ox + tx + tick * 0.5, oy + y_dim + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + tx, oy + max_y + 200),
                         (ox + tx, oy + y_dim - tick),
                         dxfattribs={"layer": "S-DIM"})
        dist = grid_x[-1] - grid_x[0]
        msp.add_text(_dim_text(dist), height=th,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + (grid_x[0] + grid_x[-1]) / 2, oy + y_dim + th * 0.5),
            align=TextEntityAlignment.BOTTOM_CENTER)

    # Y dims at left
    for i in range(len(grid_y) - 1):
        y1, y2 = grid_y[i], grid_y[i + 1]
        x_dim = min_x - DIM_OFFSET_1
        msp.add_line((ox + x_dim, oy + y1), (ox + x_dim, oy + y2),
                     dxfattribs={"layer": "S-DIM"})
        for ty in [y1, y2]:
            msp.add_line((ox + x_dim - tick * 0.5, oy + ty - tick * 0.5),
                         (ox + x_dim + tick * 0.5, oy + ty + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + min_x - 200, oy + ty),
                         (ox + x_dim - tick, oy + ty),
                         dxfattribs={"layer": "S-DIM"})
        dist = y2 - y1
        msp.add_text(_dim_text(dist), height=th, rotation=90,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + x_dim - th * 0.5, oy + (y1 + y2) / 2),
            align=TextEntityAlignment.BOTTOM_CENTER)

    # Y overall
    if len(grid_y) > 2:
        x_dim = min_x - DIM_OFFSET_2
        msp.add_line((ox + x_dim, oy + grid_y[0]),
                     (ox + x_dim, oy + grid_y[-1]),
                     dxfattribs={"layer": "S-DIM"})
        for ty in [grid_y[0], grid_y[-1]]:
            msp.add_line((ox + x_dim - tick * 0.5, oy + ty - tick * 0.5),
                         (ox + x_dim + tick * 0.5, oy + ty + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + min_x - 200, oy + ty),
                         (ox + x_dim - tick, oy + ty),
                         dxfattribs={"layer": "S-DIM"})
        dist = grid_y[-1] - grid_y[0]
        msp.add_text(_dim_text(dist), height=th, rotation=90,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + x_dim - th * 0.5, oy + (grid_y[0] + grid_y[-1]) / 2),
            align=TextEntityAlignment.BOTTOM_CENTER)


def draw_slab_labels(msp, panels, ox, oy):
    """Draw S1, S2, S3... labels on slab panels. Skip Lift/Stair/Shaft."""
    if not panels:
        return
    counter = 1
    for p in panels:
        zone = p.get("zone")
        if zone and str(zone).strip().lower() in ("lift", "stair", "shaft"):
            continue
        x1 = min(p["x_start"], p["x_end"]) * SCALE
        x2 = max(p["x_start"], p["x_end"]) * SCALE
        y1 = min(p["z_start"], p["z_end"]) * SCALE
        y2 = max(p["z_start"], p["z_end"]) * SCALE
        cx, cy = (x1 + x2) / 2, (y1 + y2) / 2
        label = f"S{counter}"
        r = 150 + len(label) * 30  # tight circle based on text length
        msp.add_circle((ox + cx, oy + cy), r,
                      dxfattribs={"layer": "TEXT", "lineweight": 50})
        msp.add_text(label, height=150,
                     dxfattribs={"layer": "TEXT", "lineweight": 50}).set_placement(
            (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        counter += 1


def draw_floor_title(msp, title, grid_x, grid_y, ox, oy):
    """Draw underlined floor plan title below the plan."""
    cx = (min(grid_x) + max(grid_x)) / 2
    ty = min(grid_y) - DIM_OFFSET_2 - TITLE_GAP
    msp.add_text("%%U" + title, height=TITLE_HEIGHT,
                 dxfattribs={"layer": "TEXT"}).set_placement(
        (ox + cx, oy + ty), align=TextEntityAlignment.MIDDLE_CENTER)


# ══════════════════════════════════════════════════════════════════════
#  STAIRCASE, LIFT, NOTES, TITLE BLOCK (from old code)
# ══════════════════════════════════════════════════════════════════════

def draw_staircase(msp, stair_rects, ox, oy):
    """Draw unified U-shaped staircase."""
    if not stair_rects:
        return
    all_x, all_y = [], []
    for rect in stair_rects:
        all_x.extend([rect["x1"] * SCALE, rect["x2"] * SCALE])
        all_y.extend([rect["y1"] * SCALE, rect["y2"] * SCALE])
    x1, x2 = min(all_x), max(all_x)
    y1, y2 = min(all_y), max(all_y)
    w, h = x2 - x1, y2 - y1
    if min(w, h) < 1200:
        return
    msp.add_lwpolyline(
        [(ox+x1,oy+y1),(ox+x2,oy+y1),(ox+x2,oy+y2),(ox+x1,oy+y2)],
        dxfattribs={"layer":"S-LINE"}, close=True)
    if h >= w:
        fw, ld = w*0.42, h*0.25
        vx1, vx2 = x1+fw, x2-fw
        vy1, vy2 = y1+ld, y2-ld
        msp.add_lwpolyline(
            [(ox+vx1,oy+vy1),(ox+vx2,oy+vy1),(ox+vx2,oy+vy2),(ox+vx1,oy+vy2)],
            dxfattribs={"layer":"S-LINE"}, close=True)
        flight_h = vy2 - vy1
        nt = max(4, round(flight_h/250))
        sp = flight_h/nt
        for t in range(1, nt):
            ty = vy1 + t*sp
            msp.add_line((ox+x1,oy+ty),(ox+vx1,oy+ty), dxfattribs={"layer":"S-LINE"})
            msp.add_line((ox+vx2,oy+ty),(ox+x2,oy+ty), dxfattribs={"layer":"S-LINE"})
        ax = vx2 + (x2-vx2)/2
        msp.add_line((ox+ax,oy+vy1+sp),(ox+ax,oy+vy2-sp), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+ax-75,oy+vy2-sp-150),(ox+ax,oy+vy2-sp), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+ax+75,oy+vy2-sp-150),(ox+ax,oy+vy2-sp), dxfattribs={"layer":"S-LINE"})
        zy = vy2 - 2*sp
        zx = x1 + fw/2
        zw = fw*0.3
        msp.add_lwpolyline(
            [(ox+x1-50,oy+zy),(ox+zx-zw/2,oy+zy),(ox+zx-zw/4,oy+zy+150),
             (ox+zx+zw/4,oy+zy-150),(ox+zx+zw/2,oy+zy),(ox+vx1+50,oy+zy)],
            dxfattribs={"layer":"S-LINE"})
    else:
        fw, ld = h*0.42, w*0.25
        vx1, vx2 = x1+ld, x2-ld
        vy1, vy2 = y1+fw, y2-fw
        msp.add_lwpolyline(
            [(ox+vx1,oy+vy1),(ox+vx2,oy+vy1),(ox+vx2,oy+vy2),(ox+vx1,oy+vy2)],
            dxfattribs={"layer":"S-LINE"}, close=True)
        flight_w = vx2 - vx1
        nt = max(4, round(flight_w/250))
        sp = flight_w/nt
        for t in range(1, nt):
            tx = vx1 + t*sp
            msp.add_line((ox+tx,oy+y1),(ox+tx,oy+vy1), dxfattribs={"layer":"S-LINE"})
            msp.add_line((ox+tx,oy+vy2),(ox+tx,oy+y2), dxfattribs={"layer":"S-LINE"})
        ay = y1 + fw/2
        msp.add_line((ox+vx1+sp,oy+ay),(ox+vx2-sp,oy+ay), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+vx2-sp-150,oy+ay-75),(ox+vx2-sp,oy+ay), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+vx2-sp-150,oy+ay+75),(ox+vx2-sp,oy+ay), dxfattribs={"layer":"S-LINE"})
        zx = vx2 - 2*sp
        zy = y2 - fw/2
        zw = fw*0.3
        msp.add_lwpolyline(
            [(ox+zx,oy+y2+50),(ox+zx,oy+zy+zw/2),(ox+zx+150,oy+zy+zw/4),
             (ox+zx-150,oy+zy-zw/4),(ox+zx,oy+zy-zw/2),(ox+zx,oy+vy2-50)],
            dxfattribs={"layer":"S-LINE"})


def draw_lift(msp, rect, ox, oy):
    """Draw lift X-cross + LIFT text."""
    x1 = rect.get("snapped_x1", rect["x1"]) * SCALE
    y1 = rect.get("snapped_y1", rect["y1"]) * SCALE
    x2 = rect.get("snapped_x2", rect["x2"]) * SCALE
    y2 = rect.get("snapped_y2", rect["y2"]) * SCALE
    if x1 > x2: x1, x2 = x2, x1
    if y1 > y2: y1, y2 = y2, y1
    msp.add_line((ox+x1,oy+y1),(ox+x2,oy+y2), dxfattribs={"layer":"S-LINE"})
    msp.add_line((ox+x2,oy+y1),(ox+x1,oy+y2), dxfattribs={"layer":"S-LINE"})
    cx, cy = (x1+x2)/2, (y1+y2)/2
    msp.add_text("LIFT", height=200,
                 dxfattribs={"layer":"S-BEAM-NO"}).set_placement(
        (ox+cx, oy+cy), align=TextEntityAlignment.MIDDLE_CENTER)


def draw_shaft(msp, rect, ox, oy):
    """Draw shaft markings: X-cross + SHAFT text (same style as lift)."""
    x1 = rect.get("snapped_x1", rect["x1"]) * SCALE
    y1 = rect.get("snapped_y1", rect["y1"]) * SCALE
    x2 = rect.get("snapped_x2", rect["x2"]) * SCALE
    y2 = rect.get("snapped_y2", rect["y2"]) * SCALE
    if x1 > x2: x1, x2 = x2, x1
    if y1 > y2: y1, y2 = y2, y1
    msp.add_line((ox+x1,oy+y1),(ox+x2,oy+y2), dxfattribs={"layer":"S-LINE"})
    msp.add_line((ox+x2,oy+y1),(ox+x1,oy+y2), dxfattribs={"layer":"S-LINE"})
    cx, cy = (x1+x2)/2, (y1+y2)/2
    msp.add_text("SHAFT", height=200,
                 dxfattribs={"layer":"S-BEAM-NO"}).set_placement(
        (ox+cx, oy+cy), align=TextEntityAlignment.MIDDLE_CENTER)


def _group_nearby_rects(rects, gap_threshold=1.0):
    """Group rectangles by spatial proximity using union-find."""
    if not rects:
        return []
    boxes = []
    for r in rects:
        boxes.append((min(r["x1"],r["x2"]), min(r["y1"],r["y2"]),
                       max(r["x1"],r["x2"]), max(r["y1"],r["y2"])))
    parent = list(range(len(rects)))
    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i
    def union(i, j):
        pi, pj = find(i), find(j)
        if pi != pj: parent[pi] = pj
    for i in range(len(rects)):
        for j in range(i+1, len(rects)):
            x_gap = max(0, max(boxes[i][0]-boxes[j][2], boxes[j][0]-boxes[i][2]))
            y_gap = max(0, max(boxes[i][1]-boxes[j][3], boxes[j][1]-boxes[i][3]))
            if x_gap < gap_threshold and y_gap < gap_threshold:
                union(i, j)
    groups = {}
    for i in range(len(rects)):
        groups.setdefault(find(i), []).append(rects[i])
    return list(groups.values())


def filter_mumty_columns(columns, rectangles):
    """Filter columns to only those within the staircase/lift zone for Mumty."""
    if not rectangles:
        return columns
    all_x, all_y = [], []
    for r in rectangles:
        all_x.extend([r["snapped_x1"] * SCALE, r["snapped_x2"] * SCALE])
        all_y.extend([r["snapped_y1"] * SCALE, r["snapped_y2"] * SCALE])
    zone_x1, zone_x2 = min(all_x) - 500, max(all_x) + 500
    zone_y1, zone_y2 = min(all_y) - 500, max(all_y) + 500
    return [col for col in columns
            if zone_x1 <= (col["x1"]+col["x2"])/2 <= zone_x2
            and zone_y1 <= (col["y1"]+col["y2"])/2 <= zone_y2]


def filter_mumty_grids(grid_x, grid_y, mumty_columns):
    """Filter grid values to only those that have mumty columns on them."""
    if not mumty_columns:
        return grid_x, grid_y
    col_xs, col_ys = set(), set()
    for col in mumty_columns:
        cx = (col["x1"] + col["x2"]) / 2
        cy = (col["y1"] + col["y2"]) / 2
        for gx in grid_x:
            if abs(gx - cx) < GRID_SNAP_TOL: col_xs.add(gx)
        for gy in grid_y:
            if abs(gy - cy) < GRID_SNAP_TOL: col_ys.add(gy)
    return sorted(col_xs), sorted(col_ys)


def _fill_mumty_corners(mumty_cols, mumty_gx, mumty_gy):
    """Add 300x600mm columns at any empty corners of mumty bounding box."""
    corners = [
        (min(mumty_gx), min(mumty_gy)),  # bottom-left
        (max(mumty_gx), min(mumty_gy)),  # bottom-right
        (min(mumty_gx), max(mumty_gy)),  # top-left
        (max(mumty_gx), max(mumty_gy)),  # top-right
    ]
    for cx, cy in corners:
        # Check if a column already exists near this corner
        has_col = False
        for col in mumty_cols:
            col_cx = (col["x1"] + col["x2"]) / 2
            col_cy = (col["y1"] + col["y2"]) / 2
            if abs(col_cx - cx) < GRID_SNAP_TOL and abs(col_cy - cy) < GRID_SNAP_TOL:
                has_col = True
                break
        if not has_col:
            # Add a 300x600mm column centered on this corner
            w, h = 300, 600
            mumty_cols.append({
                "no": f"M_{cx}_{cy}",
                "name": f"MC",
                "x1": cx - w / 2,
                "x2": cx + w / 2,
                "y1": cy - h / 2,
                "y2": cy + h / 2,
            })
            print(f"    Added corner column at ({cx}, {cy})")
    return mumty_cols


def filter_columns_by_beams(columns, beams):
    """Keep only columns that have at least one beam endpoint near them."""
    if not beams:
        return columns
    beam_pts = set()
    for b in beams:
        beam_pts.add((round(b["start_x"] * SCALE, 1), round(b["start_z"] * SCALE, 1)))
        beam_pts.add((round(b["end_x"] * SCALE, 1), round(b["end_z"] * SCALE, 1)))
    result = []
    for col in columns:
        cx = (col["x1"] + col["x2"]) / 2
        cy = (col["y1"] + col["y2"]) / 2
        for bx, by in beam_pts:
            if abs(bx - cx) < GRID_SNAP_TOL and abs(by - cy) < GRID_SNAP_TOL:
                result.append(col)
                break
    return result


def filter_rects_to_mumty_zone(rectangles):
    """Filter rectangles to only those within the mumty zone."""
    # Find mumty bounding box from the Mumty rectangle
    mumty_rects = [r for r in rectangles
                   if "mumty" in str(r.get("location", "")).lower()]
    if not mumty_rects:
        return []
    all_x = [v * SCALE for r in mumty_rects for v in [r["x1"], r["x2"]]]
    all_y = [v * SCALE for r in mumty_rects for v in [r["y1"], r["y2"]]]
    zx1, zx2 = min(all_x) - 500, max(all_x) + 500
    zy1, zy2 = min(all_y) - 500, max(all_y) + 500
    # Keep rectangles whose center falls within the mumty zone
    result = []
    for r in rectangles:
        cx = (r["x1"] + r["x2"]) / 2 * SCALE
        cy = (r["y1"] + r["y2"]) / 2 * SCALE
        if zx1 <= cx <= zx2 and zy1 <= cy <= zy2:
            result.append(r)
    return result


def draw_rectangles_on_floor(msp, rectangles, floor_key, ox, oy):
    """Draw lift and staircase markings from rectangle data."""
    if not rectangles:
        return
    lift_rect = None
    for rect in rectangles:
        if "lift" in str(rect.get("location", "")).lower():
            lift_rect = rect
            break
    lift_y_min = min(lift_rect["y1"], lift_rect["y2"]) if lift_rect else None
    lift_y_max = max(lift_rect["y1"], lift_rect["y2"]) if lift_rect else None

    stair_rects = []
    for rect in rectangles:
        rloc = str(rect.get("location", "")).lower()
        if "lift" in rloc:
            print(f"    Lift: {rect['no']} ({rect['type']})")
            draw_lift(msp, rect, ox, oy)
        elif "shaft" in rloc:
            print(f"    Shaft: {rect['no']} ({rect['type']})")
            draw_shaft(msp, rect, ox, oy)
        elif "stair" in rloc:
            if lift_rect is not None:
                ry_min = min(rect["y1"], rect["y2"])
                ry_max = max(rect["y1"], rect["y2"])
                overlap = max(0, min(ry_max, lift_y_max) - max(ry_min, lift_y_min))
                ry_range = ry_max - ry_min
                rx_min = min(rect["x1"], rect["x2"])
                rx_max = max(rect["x1"], rect["x2"])
                lx_min = min(lift_rect["x1"], lift_rect["x2"])
                lx_max = max(lift_rect["x1"], lift_rect["x2"])
                x_gap = max(0, max(rx_min - lx_max, lx_min - rx_max))
                if ry_range > 0 and overlap / ry_range > 0.5 and x_gap < 0.5:
                    print(f"    Skip shaft-adjacent: {rect['no']} ({rect['type']})")
                    continue
            print(f"    Staircase rect: {rect['no']} ({rect['type']})")
            stair_rects.append(rect)

    stair_groups = _group_nearby_rects(stair_rects, gap_threshold=1.0)
    for group in stair_groups:
        names = ", ".join(f"{r['no']}({r['type']})" for r in group)
        print(f"    Drawing staircase group: [{names}]")
        draw_staircase(msp, group, ox, oy)


def draw_notes_panel(msp, notes_x, notes_y, notes_h):
    """Draw General Notes and Legend."""
    x, y = notes_x, notes_y + notes_h
    th, hh = 80, 110
    line_sp, indent, wrap_indent = th*2.0, 250, 400
    cy = y - 100
    msp.add_text("%%UGENERAL  NOTES:-", height=hh,
                 dxfattribs={"layer":"NOTES"}).set_placement(
        (x, cy), align=TextEntityAlignment.TOP_LEFT)
    cy -= hh + line_sp*0.5
    notes = [
        ("1.","STRUCTURAL DRAWINGS SHALL BE READ IN CONJUNCTION WITH",
              "RELEVANT ARCHITECTURAL AND SERVICES DRAWINGS."),
        ("2.","DO NOT SCALE.  FOLLOW WRITTEN DIMENSIONS ONLY.",None),
        ("3.","ALL DIMENSIONS ARE IN MM. UNLESS OTHERWISE SPECIFIED.",None),
        ("4.","UNLESS SPECIFIED OTHERWISE, ALL LEVELS SHOWN IN STRUCTURAL",
              "DRAWINGS ARE STRUCTURAL LEVELS ONLY."),
        ("5.","COVER",None), (None,"a) BEAM                : 1\"",None),
        ("6.","CONCRETE MIX",None), (None,"a) BEAM                : M25",None),
        ("7.","REINFORCEMENT   : REINF. STEEL SHALL BE TMT BARS OF",
              "GRADE Fe500D CONFORMING TO IS 1786-2008."),
        ("8.","ALL LEVELS ARE TO BE TAKEN FROM ARCHITECTURAL DRAWINGS.",None),
        ("9.","CONFIRM LOCATION OF BEAM FOR WALLS WITH RELEVANT","ARCH. DRGS."),
        ("10.","ALL R.C.C. TO BE MACHINE MIXED, VIBRATED AND CURED",
               "THOROUGHLY AS PER I.S 456-LATEST."),
        ("11.","ALL DIMENSIONS MUST BE CHECKED WITH ARCHITECT'S DRGS.& IN",
               "CASE OF ANY DISCREPANCY ARCHITECTS DRGS. SHALL PREVAIL."),
        ("12.","ALL CONSTRUCTION JOINTS SHALL BE APPROVED BY CONSULTANT",
               "ON THE BASIS OF SCHEME PREPARED BY CONTRACTOR."),
        ("13.","TOP AND BOTTOM EXTRA BARS IN BEAMS TO EXTEND BEYOND THE",
               "FACE OF SUPPORT AS SHOWN IN DRG UNLESS OTHERWISE SHOWN."),
        ("14.","THE FIRST STIRRUPS IN BEAMS SHALL BE AT A DISTANCE OF",
               "50MM FROM THE JOINT FACE."),
        ("15.","All ANGLES ARE RIGHT ANGLES UNLESS OTHERWISE SPECIFIED.",None),
    ]
    for item in notes:
        num = item[0]
        lines = [l for l in item[1:] if l is not None]
        if num:
            msp.add_text(f"{num}  {lines[0]}", height=th,
                         dxfattribs={"layer":"NOTES"}).set_placement(
                (x+indent, cy), align=TextEntityAlignment.TOP_LEFT)
            cy -= line_sp
            for cont in lines[1:]:
                msp.add_text(cont, height=th,
                             dxfattribs={"layer":"NOTES"}).set_placement(
                    (x+wrap_indent, cy), align=TextEntityAlignment.TOP_LEFT)
                cy -= line_sp
        else:
            msp.add_text(f"     {lines[0]}", height=th,
                         dxfattribs={"layer":"NOTES"}).set_placement(
                (x+wrap_indent, cy), align=TextEntityAlignment.TOP_LEFT)
            cy -= line_sp
    cy -= line_sp*0.8
    msp.add_text("%%ULEGEND:-", height=hh,
                 dxfattribs={"layer":"NOTES"}).set_placement(
        (x, cy), align=TextEntityAlignment.TOP_LEFT)
    cy -= hh + line_sp*0.5
    for line in ["TOB     = TOP OF BEAM","T.O      = TOP OF",
                 "NGL     = NATURAL GROUND LEVEL","FFL      = FLOOR FINISH LEVEL",
                 "TOS     = TOP OF SLAB"]:
        msp.add_text(line, height=th,
                     dxfattribs={"layer":"NOTES"}).set_placement(
            (x+indent, cy), align=TextEntityAlignment.TOP_LEFT)
        cy -= line_sp


def draw_title_block(msp, x, y, w):
    """Draw title block."""
    layer = "NOTES"
    th, th_med, th_big, th_label, pad = 70, 90, 130, 55, 100
    sp = 170
    def box(cy, h):
        msp.add_lwpolyline([(x,cy),(x+w,cy),(x+w,cy-h),(x,cy-h)],
                           dxfattribs={"layer":layer}, close=True)
        return cy - h
    def label(text, cy):
        msp.add_text(text, height=th_label,
                     dxfattribs={"layer":layer,"color":1}).set_placement(
            (x+pad, cy-40), align=TextEntityAlignment.TOP_LEFT)
    def center_text(text, cy, h):
        msp.add_text(text, height=h,
                     dxfattribs={"layer":layer}).set_placement(
            (x+w/2, cy), align=TextEntityAlignment.MIDDLE_CENTER)
    cy = y
    cy = box(cy, 350)  # FOR REVIEW
    msp.add_text("FOR REVIEW", height=th_big,
                 dxfattribs={"layer":layer}).set_placement(
        (x+pad, cy+175), align=TextEntityAlignment.MIDDLE_LEFT)
    for _ in range(4): cy = box(cy, 160)  # revision rows
    rh = 180; cy_top = cy; cy = box(cy, rh)  # rev header
    c1,c2,c3,c4 = x+w*0.08, x+w*0.18, x+w*0.65, x+w*0.82
    for cx_div in [c1,c2,c3,c4]:
        msp.add_line((cx_div,cy_top),(cx_div,cy), dxfattribs={"layer":layer})
    for lbl,lx,rx in [("Rev",x,c1),("Date",c1,c2),("Description",c2,c3),
                       ("Drn",c3,c4),("Chk",c4,x+w)]:
        msp.add_text(lbl, height=th_label,
                     dxfattribs={"layer":layer,"color":1}).set_placement(
            ((lx+rx)/2, cy+rh/2), align=TextEntityAlignment.MIDDLE_CENTER)
    h=600; cy_top=cy; cy=box(cy,h); label("CLIENT:",cy_top)
    center_text("-------------------", cy+h/2-50, th_med)
    h=1400; cy_top=cy; cy=box(cy,h); label("ARCHITECTS:",cy_top)
    ty=cy_top-250; center_text("PLACEHOLDER & ASSOCIATES",ty,th_big)
    h=1200; cy_top=cy; cy=box(cy,h); label("STRUCTURAL CONSULTANT:",cy_top)
    ty=cy_top-250; center_text("Placeholder Consultants",ty,th_big)
    h=900; cy_top=cy; cy=box(cy,h); label("PROJECT:",cy_top)
    ty=cy_top-280; center_text("PROPOSED FLOOR PLAN FOR",ty,th_big)
    h=800; cy_top=cy; cy=box(cy,h); label("Drawing Title:",cy_top)
    ty=cy_top-300; center_text("ALL FLOOR LEVEL",ty,th_big*1.2)
    ty-=sp*1.5; center_text("FRAMING PLAN",ty,th_big*1.2)
    h=280; cy_top=cy; cy=box(cy,h)
    c1,c2,c3 = x+w*0.22, x+w*0.68, x+w*0.85
    for cx_div in [c1,c2,c3]:
        msp.add_line((cx_div,cy_top),(cx_div,cy), dxfattribs={"layer":layer})


def draw_north_arrow(msp, cx, cy):
    """Draw north arrow symbol."""
    r = 400
    msp.add_circle((cx,cy), r, dxfattribs={"layer":"S-LINE"})
    msp.add_line((cx,cy+r*0.3),(cx,cy+r*1.3), dxfattribs={"layer":"S-LINE"})
    msp.add_line((cx,cy+r*1.3),(cx-r*0.2,cy+r*1.0), dxfattribs={"layer":"S-LINE"})
    msp.add_line((cx,cy+r*1.3),(cx+r*0.2,cy+r*1.0), dxfattribs={"layer":"S-LINE"})
    msp.add_text("N", height=250,
                 dxfattribs={"layer":"TEXT"}).set_placement(
        (cx, cy+r*1.6), align=TextEntityAlignment.BOTTOM_CENTER)


def draw_sheet_border(msp, bounds):
    """Draw sheet border."""
    bx1,by1,bx2,by2 = bounds
    m = 600
    msp.add_lwpolyline(
        [(bx1-m,by1-m),(bx2+m,by1-m),(bx2+m,by2+m),(bx1-m,by2+m)],
        dxfattribs={"layer":"BORDER"}, close=True)


def draw_right_panel(msp, grid_x, grid_y, typ_ox, draw_w,
                     border_top, border_bottom):
    """Draw right-side panel with KEY PLAN, NOTES, TITLE BLOCK."""
    tb_width = NOTES_WIDTH
    panel_x = typ_ox + draw_w + 2000
    panel_right = panel_x + tb_width
    panel_top, panel_bottom = border_top, border_bottom

    msp.add_lwpolyline(
        [(panel_x,panel_bottom),(panel_right,panel_bottom),
         (panel_right,panel_top),(panel_x,panel_top)],
        dxfattribs={"layer":"NOTES"}, close=True)

    # KEY PLAN
    keyplan_h = 3500
    keyplan_bottom = panel_top - keyplan_h
    msp.add_line((panel_x,keyplan_bottom),(panel_right,keyplan_bottom),
                 dxfattribs={"layer":"NOTES"})
    msp.add_text("%%UKEY PLAN", height=120,
                 dxfattribs={"layer":"NOTES"}).set_placement(
        (panel_x+tb_width/2, keyplan_bottom+100),
        align=TextEntityAlignment.BOTTOM_CENTER)
    draw_north_arrow(msp, panel_right-600, keyplan_bottom+keyplan_h/2+200)

    # Title block
    tb_total_h = 350+4*160+180+600+1400+1200+900+800+280
    tb_top_y = panel_bottom + tb_total_h
    draw_title_block(msp, panel_x, tb_top_y, tb_width)
    msp.add_line((panel_x,tb_top_y),(panel_right,tb_top_y),
                 dxfattribs={"layer":"NOTES"})

    # Notes
    notes_bottom = tb_top_y + 100
    notes_top = keyplan_bottom - 100
    notes_h = max(notes_top - notes_bottom, 8000)
    draw_notes_panel(msp, panel_x, notes_bottom, notes_h)

    return panel_right + 500  # border_right


# ══════════════════════════════════════════════════════════════════════
#  FLOOR PLAN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════

def draw_floor_plan(msp, title, floor_beams, grid_x, grid_y, columns,
                    rectangles, floor_key, ox, oy, slab_panels=None):
    """Draw a complete floor plan at offset (ox, oy)."""
    print(f"    Grid stubs")
    draw_grid_stubs(msp, grid_x, grid_y, ox, oy)

    print(f"    Grid labels")
    draw_grid_labels(msp, grid_x, grid_y, ox, oy)

    print(f"    Beams ({len(floor_beams)})")
    draw_beams(msp, floor_beams, columns, ox, oy)

    # Staircase and lift from old rectangle data
    if rectangles:
        draw_rectangles_on_floor(msp, rectangles, floor_key, ox, oy)

    # Slab panel labels — disabled for framing plan
    # if slab_panels:
    #     print(f"    Slab labels ({len(slab_panels)})")
    #     draw_slab_labels(msp, slab_panels, ox, oy)

    # Columns drawn LAST (opaque fill hides beam overlaps)
    print(f"    Columns ({len(columns)})")
    draw_columns(msp, columns, ox, oy)

    print(f"    Dimensions")
    draw_dimensions(msp, grid_x, grid_y, ox, oy)

    print(f"    Title: {title}")
    draw_floor_title(msp, title, grid_x, grid_y, ox, oy)


# ══════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  FRAMING PLAN GENERATOR V2")
    print("=" * 60)
    print()

    # Input file
    input_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}")
        return
    print(f"Input: {input_path}")

    # Read and classify beams
    all_beams = read_beam_data(input_path)
    all_beams = classify_beams(all_beams)

    # Summary
    h_count = sum(1 for b in all_beams if b["direction"] == "horizontal")
    v_count = sum(1 for b in all_beams if b["direction"] == "vertical")
    d_count = sum(1 for b in all_beams if b["direction"] == "diagonal")
    floors = sorted(set(b["floor"] for b in all_beams))
    types = sorted(set(b["beam_type"] for b in all_beams))
    print(f"Total beams: {len(all_beams)} ({h_count} H, {v_count} V, {d_count} diagonal)")
    print(f"Floors: {floors}")
    print(f"Types: {types}")

    # ── Read OLD data (grids, columns, rectangles) ──
    node_path = find_latest_file("node_coordinates_*.xlsx")
    gridline_path = find_latest_file("gridline_coordinates_*.xlsx")
    other_path = find_latest_file("other_coordinates_*.xlsx")

    if not node_path or not gridline_path:
        print("ERROR: node_coordinates or gridline_coordinates not found in STD ANL folder")
        return

    print(f"Node coordinates: {os.path.basename(node_path)}")
    print(f"Gridline coordinates: {os.path.basename(gridline_path)}")

    # Grids from old node_coordinates (5 Y values, 6 X values)
    grid_x, grid_y = read_unique_grid_values(node_path)
    print(f"Grid X ({len(grid_x)}): {grid_x}")
    print(f"Grid Y ({len(grid_y)}): {grid_y}")

    # Columns from gridline_coordinates (per-floor)
    all_columns = read_gridline_coordinates(gridline_path)
    for ft, cols in all_columns.items():
        print(f"  {ft}: {len(cols)} columns")

    # Rectangles for lift/staircase
    rectangles = []
    if other_path:
        print(f"Other coordinates: {os.path.basename(other_path)}")
        rectangles = read_rectangles(other_path)
        print(f"{len(rectangles)} rectangles (lift/staircase)")

    # Slab panels for labels
    all_slab_panels = read_slab_panels(RECT_GEOM_PATH)
    for ft, pnls in all_slab_panels.items():
        print(f"  Slab panels {ft}: {len(pnls)}")

    # Create DXF
    doc = setup_document()
    msp = doc.modelspace()

    # Drawing extent
    draw_w = max(grid_x) - min(grid_x) + GRID_EXTEND * 2 + DIM_OFFSET_2 * 2
    draw_h = max(grid_y) - min(grid_y) + GRID_EXTEND * 2 + DIM_OFFSET_2 * 2

    # # ── Old 4-floor layout (kept for reference) ──
    # # Floor 1: STILT, Floor 2: TYPICAL, Floor 3: TERRACE, Floor 4: MUMTY
    # # (see git history for full old code)

    # ── Floor 1: PLINTH ──
    print("\n[PLINTH LEVEL FRAMING PLAN]")
    plinth_beams = get_floor_beams(all_beams, "Stilt")
    plinth_cols = get_columns_for_floor(all_columns, "Stilt")
    plinth_ox, plinth_oy = 0, 0
    # Plinth: no shaft — filter out shaft rectangles
    plinth_rects = [r for r in rectangles
                    if "shaft" not in str(r.get("location", "")).lower()
                    ] if rectangles else []
    draw_floor_plan(msp, "PLINTH LEVEL FRAMING PLAN",
                    plinth_beams, grid_x, grid_y, plinth_cols,
                    plinth_rects, "Plinth",
                    plinth_ox, plinth_oy,
                    slab_panels=all_slab_panels.get("Stilt"))

    # ── Floor 2: STILT ROOF ──
    print("\n[STILT ROOF FRAMING PLAN]")
    stilt_roof_beams = get_floor_beams(all_beams, "Typical")
    stilt_roof_cols = get_columns_for_floor(all_columns, "Typical")
    stilt_roof_ox = plinth_ox + draw_w + FLOOR_GAP
    stilt_roof_oy = 0
    draw_floor_plan(msp, "STILT ROOF FRAMING PLAN",
                    stilt_roof_beams, grid_x, grid_y, stilt_roof_cols,
                    rectangles, "Stilt roof",
                    stilt_roof_ox, stilt_roof_oy,
                    slab_panels=all_slab_panels.get("Typical"))

    # ── Floor 3: TYPICAL (shared for all numbered floor roofs) ──
    print("\n[TYP FLOOR LEVEL FRAMING PLAN]")
    typ_beams = get_floor_beams(all_beams, "Terrace")
    typ_cols = get_columns_for_floor(all_columns, "Terrace")
    typ_ox = stilt_roof_ox
    typ_oy = -(draw_h + FLOOR_GAP)
    draw_floor_plan(msp, "TYP FLOOR LEVEL FRAMING PLAN",
                    typ_beams, grid_x, grid_y, typ_cols,
                    rectangles, "Typical floor",
                    typ_ox, typ_oy,
                    slab_panels=all_slab_panels.get("Terrace"))

    # ── Floor 4: MUMTY ──
    print("\n[MUMTY LEVEL FRAMING PLAN]")
    mumty_beams = get_floor_beams(all_beams, "Mumty")
    mumty_cols = get_columns_for_floor(all_columns, "Mumty")
    if not mumty_cols:
        mumty_cols = filter_mumty_columns(
            get_columns_for_floor(all_columns, "Typical"), rectangles)
    mumty_gx, mumty_gy = filter_mumty_grids(grid_x, grid_y, mumty_cols)
    if mumty_gx and mumty_gy:
        mumty_cols = _fill_mumty_corners(mumty_cols, mumty_gx, mumty_gy)
    mumty_ox = plinth_ox
    mumty_oy = -(draw_h + MUMTY_GAP)
    print(f"  Mumty columns: {len(mumty_cols)}")
    print(f"  Mumty grids X: {mumty_gx}")
    print(f"  Mumty grids Y: {mumty_gy}")
    if mumty_cols and mumty_gx and mumty_gy:
        draw_floor_plan(msp, "MUMTY LEVEL FRAMING PLAN",
                        mumty_beams, mumty_gx, mumty_gy, mumty_cols,
                        [], "Mumty",
                        mumty_ox, mumty_oy,
                        slab_panels=all_slab_panels.get("Mumty"))
    else:
        print("  No mumty columns found -- skipping")

    # ── Right-side panel (notes, title block, north arrow) ──
    print("\nDrawing right panel...")
    border_top = max(grid_y) + GRID_EXTEND + 3500
    # Bottom border: lowest of mumty, typical, or plinth
    lowest_y = plinth_oy + min(grid_y)
    if typ_oy + min(grid_y) < lowest_y:
        lowest_y = typ_oy + min(grid_y)
    if mumty_gy and mumty_oy + min(mumty_gy) < lowest_y:
        lowest_y = mumty_oy + min(mumty_gy)
    border_bottom = lowest_y - GRID_EXTEND - DIM_OFFSET_2 - TITLE_GAP - TITLE_HEIGHT - 500
    border_left = min(plinth_ox + min(grid_x),
                      mumty_ox + min(mumty_gx) if mumty_gx else 0) - GRID_EXTEND - DIM_OFFSET_2 - 500

    border_right = draw_right_panel(msp, grid_x, grid_y,
                                     stilt_roof_ox, draw_w, border_top, border_bottom)

    # ── Sheet border ──
    print("Drawing sheet border...")
    draw_sheet_border(msp, (border_left, border_bottom, border_right, border_top))

    # Save
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outpath = os.path.join(OUTPUT_FOLDER, f"framing_plan_v2_{timestamp}.dxf")
    _, dwg_path = save_dxf_and_dwg(doc, outpath)
    print(f"\nSaved: {outpath}" + (f"\n       {dwg_path}" if dwg_path else ""))


if __name__ == "__main__":
    main()
