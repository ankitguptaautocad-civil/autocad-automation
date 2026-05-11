from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from statistics import median

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


DEFAULT_INPUT_PATH = None
DEFAULT_WALL_INPUT_PATH = None
DEFAULT_OUTPUT_PATH = None
DEFAULT_GROUP_TOLERANCE_M = 0.75
DEFAULT_BOUNDARY_TOLERANCE_M = 0.75
DEFAULT_WALL_ALIGNMENT_TOLERANCE_M = 0.65
DEFAULT_EDGE_WALL_COVERAGE_THRESHOLD_PCT = 10.0
DEFAULT_INTERIOR_WALL_COVERAGE_THRESHOLD_PCT = 30.0
DEFAULT_THICKNESS_TOLERANCE_MM = 40.0
FLOOR_SEQUENCE = ("Plinth", "Stilt roof", "Typical floor roof", "Terrace")
FLOOR_TO_WALL_SOURCE = {
    "Plinth": "plinth",
    "Stilt roof": "typical",
    "Typical floor roof": "typical",
    "Terrace": "terrace",
}
PLINTH_FLOORS = ("Plinth",)
NONPLINTH_FLOORS = ("Stilt roof", "Typical floor roof", "Terrace")
SECONDARY_MIN_SPAN_M = 0.8
SECONDARY_DUPLICATE_MINOR_TOL_M = 0.15
SECONDARY_DUPLICATE_ENDPOINT_TOL_M = 0.20
SECONDARY_PARALLEL_SPACING_MIN_M = 0.90
SECONDARY_WALL_MERGE_GAP_M = 0.30
SECONDARY_R3_MIN_SHORT_SIDE_M = 2.0
SECONDARY_R3_SPAN_TRIGGER_M = 6.0
SECONDARY_R3_ASPECT_RATIO_TRIGGER = 2.5
SECONDARY_R4_MIN_WALL_LENGTH_M = 0.8
SECONDARY_R4_NONPLINTH_MIN_WALL_LENGTH_M = 1.5
SECONDARY_MAX_GENERATION_PASSES = 2
SECONDARY_WALL_LINE_TOL_M = 0.15
SECONDARY_R3_WALL_ALIGN_TOL_M = 0.50
SECONDARY_SUPPORT_TOL_M = 0.15
SECONDARY_ZONE_INTERIOR_TOL_M = 0.15
SECONDARY_R1_MAX_SPAN_M = 2.5
SECONDARY_R1_ZONE_PROXIMITY_M = 1.5
SECONDARY_R4_END_EXTENSION_M = 0.45
SECONDARY_R4_STRONG_WALL_LENGTH_M = 2.0
SECONDARY_R5_SUPPORT_TOL_M = 0.05
SECONDARY_R5_EDGE_ALIGN_TOL_M = 0.20
SECONDARY_PLINTH_WIDTH_MM = 230
SECONDARY_PLINTH_DEPTH_MM = 150
RAW_PLINTH_SHEET = "Secondary beam coordinates_plin"
RAW_NONPLINTH_SHEET = "Secondary beam coordinates_nonp"
SECONDARY_RULE_PRIORITY = {
    "R2": 1,
    "R1": 2,
    "R5": 3,
    "R4": 4,
    "R3": 5,
}


@dataclass(frozen=True)
class SourceColumn:
    idx: int
    type_name: str
    xmin: float
    xmax: float
    ymin: float
    ymax: float
    width: float
    height: float
    left_right: str
    front_back: str
    source_location: str | None


@dataclass(frozen=True)
class ColumnRecord:
    idx: int
    type_name: str
    xmin: float
    xmax: float
    ymin: float
    ymax: float
    width: float
    height: float
    left_right: str
    front_back: str
    source_location: str | None
    location: str
    anchor_location: str
    anchor_x: float
    anchor_y: float
    orientation: str


@dataclass(frozen=True)
class WallSegment:
    wall_no: str
    start_x: float
    start_y: float
    end_x: float
    end_y: float
    thickness_m: float
    thickness_class_mm: int | None
    orientation: str
    source: str | None
    dxf_source: str | None


@dataclass(frozen=True)
class BeamPair:
    beam_no: str
    start: ColumnRecord
    end: ColumnRecord
    direction: str
    beam_class: str
    beam_start_x: float
    beam_start_y: float
    beam_end_x: float
    beam_end_y: float
    span_length_m: float
    group_coordinate_m: float


@dataclass(frozen=True)
class ZoneRectangle:
    rect_no: str
    type_name: str
    x1: float
    y1: float
    x2: float
    y2: float
    location: str
    floors: tuple[str, ...]


@dataclass(frozen=True)
class SupportSegment:
    axis: str
    fixed: float
    start: float
    end: float
    source_kind: str
    source_name: str


@dataclass(frozen=True)
class SecondaryBeamCandidate:
    axis: str
    x1: float
    y1: float
    x2: float
    y2: float
    beam_location: str
    floor_group: str
    floors: tuple[str, ...]
    rule_code: str
    beam_class: str
    detail: str
    score: float = 0.0


@dataclass(frozen=True)
class SecondaryBeam:
    no: int
    type_name: str
    axis: str
    x1: float
    y1: float
    x2: float
    y2: float
    beam_location: str
    floor_group: str
    floors: tuple[str, ...]
    rule_code: str
    beam_class: str
    detail: str


@dataclass(frozen=True)
class WallChain:
    axis: str
    fixed: float
    start: float
    end: float
    length: float
    thickness_mm: int
    source: str


@dataclass(frozen=True)
class ClosedRectangle:
    x1: float
    x2: float
    y1: float
    y2: float


def normalize_header(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


def infer_dxf_stem_from_column_workbook(path: Path) -> str:
    stem = path.stem
    suffixes = (
        "_col_rectangles_m_v2_wall_assisted",
        "_col_rectangles_m_v2",
        "_col_rectangles_m",
    )
    for suffix in suffixes:
        if stem.endswith(suffix):
            return stem[: -len(suffix)]
    return stem


def candidate_search_dirs() -> list[Path]:
    dirs = []
    for candidate in ((Path.cwd() / "STD ANL model").resolve(), Path.cwd().resolve()):
        if candidate.exists() and candidate not in dirs:
            dirs.append(candidate)
    return dirs


def is_excel_lockfile(path: Path) -> bool:
    return path.name.startswith("~$")


def discover_single_workbook(patterns: tuple[str, ...], label: str) -> Path:
    matches: list[Path] = []
    for folder in candidate_search_dirs():
        for pattern in patterns:
            matches.extend(path.resolve() for path in folder.glob(pattern) if not is_excel_lockfile(path))
    unique_matches = sorted(set(matches))
    if not unique_matches:
        raise SystemExit(f"Could not find the {label} workbook. Pass it explicitly.")
    if len(unique_matches) == 1:
        return unique_matches[0]
    return max(unique_matches, key=lambda path: (path.stat().st_mtime_ns, path.name.lower()))


def resolve_input_workbook(explicit_input: Path | None) -> Path:
    if explicit_input is not None:
        path = explicit_input.resolve()
        if not path.exists():
            raise SystemExit(f"Input workbook not found: {path}")
        return path
    return discover_single_workbook(
        (
            "*typical*col_rectangles_m_v2_wall_assisted.xlsx",
            "*typical*col_rectangles_m_v2.xlsx",
            "*typical*col_rectangles_m.xlsx",
        ),
        "column",
    )


def resolve_wall_workbook(explicit_walls: Path | None, input_workbook: Path) -> Path:
    if explicit_walls is not None:
        path = explicit_walls.resolve()
        if not path.exists():
            raise SystemExit(f"Wall workbook not found: {path}")
        return path

    dxf_stem = infer_dxf_stem_from_column_workbook(input_workbook)
    candidates = [
        input_workbook.with_name(f"{dxf_stem}_walls_m_v2.xlsx"),
        input_workbook.with_name(f"{dxf_stem}_walls.xlsx"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()

    raise SystemExit(
        "Could not find the wall workbook beside the column workbook. "
        "Pass --walls explicitly or generate the walls workbook first."
    )


def resolve_rectangle_workbook(input_workbook: Path) -> Path | None:
    candidates: list[Path] = []
    for folder in candidate_search_dirs():
        for pattern in ("*secondary_coordinates_nonplinth*.xlsx", "*secondary_coordinates_nonp*.xlsx"):
            candidates.extend(path.resolve() for path in folder.glob(pattern) if not is_excel_lockfile(path))
    unique_candidates = sorted(set(candidates), key=lambda path: (path.stat().st_mtime, path.name), reverse=True)
    return unique_candidates[0] if unique_candidates else None


def normalize_lr(tag: str | None) -> str:
    value = (tag or "").strip().lower()
    if value in {"left", "right", "centre", "center"}:
        return "Centre" if value in {"centre", "center"} else value.title()
    raise SystemExit(f"Invalid Left/Right tag: {tag!r}")


def normalize_fb(tag: str | None) -> str:
    value = (tag or "").strip().lower()
    if value in {"front", "back", "centre", "center"}:
        return "Centre" if value in {"centre", "center"} else value.title()
    raise SystemExit(f"Invalid Front/Back tag: {tag!r}")


def select_coordinate(min_value: float, max_value: float, tag: str) -> float:
    if tag in {"Left", "Front"}:
        return round(min_value, 3)
    if tag in {"Right", "Back"}:
        return round(max_value, 3)
    if tag == "Centre":
        return round((min_value + max_value) / 2.0, 3)
    raise SystemExit(f"Unsupported tag for coordinate selection: {tag!r}")


def build_anchor_location(front_back: str, left_right: str) -> str:
    return f"{front_back} {left_right}"


def parse_column_number(raw_value: object, fallback_idx: int) -> int:
    text = str(raw_value or "").strip()
    if text.upper().startswith("C") and text[1:].isdigit():
        return int(text[1:])
    if text.isdigit():
        return int(text)
    return fallback_idx


def snap_thickness_mm(raw_mm: float, tolerance_mm: float) -> int | None:
    candidates = []
    for target in (115.0, 230.0):
        delta = abs(raw_mm - target)
        if delta <= tolerance_mm:
            candidates.append((delta, int(target)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]))
    return candidates[0][1]


def read_source_columns(path: Path) -> list[SourceColumn]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {normalize_header(str(value or "")): idx for idx, value in enumerate(header_row)}

    required = {
        "columnno": "Column No",
        "xminm": "Xmin (m)",
        "xmaxm": "Xmax (m)",
        "yminm": "Ymin (m)",
        "ymaxm": "Ymax (m)",
        "widthm": "Width (m)",
        "heightm": "Height (m)",
        "leftright": "Left/Right",
        "frontback": "Front/Back",
        "location": "Location",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Missing required columns in input workbook: {', '.join(missing)}")

    columns: list[SourceColumn] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue
        try:
            left_right = normalize_lr(row[header_map["leftright"]])
            front_back = normalize_fb(row[header_map["frontback"]])
        except SystemExit as exc:
            raise SystemExit(f"Row {row_idx}: {exc}") from exc

        idx = parse_column_number(row[header_map["columnno"]], len(columns) + 1)

        columns.append(
            SourceColumn(
                idx=idx,
                type_name=f"C{idx}",
                xmin=round(float(row[header_map["xminm"]]), 3),
                xmax=round(float(row[header_map["xmaxm"]]), 3),
                ymin=round(float(row[header_map["yminm"]]), 3),
                ymax=round(float(row[header_map["ymaxm"]]), 3),
                width=round(float(row[header_map["widthm"]]), 3),
                height=round(float(row[header_map["heightm"]]), 3),
                left_right=left_right,
                front_back=front_back,
                source_location=str(row[header_map["location"]]).strip() if row[header_map["location"]] not in (None, "") else None,
            )
        )

    if not columns:
        raise SystemExit("No column rows found in the input workbook.")
    return sorted(columns, key=lambda col: col.idx)


def derive_columns(columns: list[SourceColumn], boundary_tolerance_m: float) -> list[ColumnRecord]:
    provisional: list[tuple[SourceColumn, float, float]] = []
    for column in columns:
        anchor_x = select_coordinate(column.xmin, column.xmax, column.left_right)
        anchor_y = select_coordinate(column.ymin, column.ymax, column.front_back)
        provisional.append((column, anchor_x, anchor_y))

    provisional_records = [
        ColumnRecord(
            idx=column.idx,
            type_name=column.type_name,
            xmin=column.xmin,
            xmax=column.xmax,
            ymin=column.ymin,
            ymax=column.ymax,
            width=column.width,
            height=column.height,
            left_right=column.left_right,
            front_back=column.front_back,
            source_location=column.source_location,
            location="Interior",
            anchor_location=build_anchor_location(column.front_back, column.left_right),
            anchor_x=anchor_x,
            anchor_y=anchor_y,
            orientation="Vertical" if column.height >= column.width else "Horizontal",
        )
        for column, anchor_x, anchor_y in provisional
    ]

    global_min_x = min(anchor_x for _, anchor_x, _ in provisional)
    global_max_x = max(anchor_x for _, anchor_x, _ in provisional)
    global_min_y = min(anchor_y for _, _, anchor_y in provisional)
    global_max_y = max(anchor_y for _, _, anchor_y in provisional)

    x_families = cluster_columns(provisional_records, lambda col: col.anchor_x, boundary_tolerance_m)

    left_boundary: set[int] = {column.idx for column, anchor_x, _ in provisional if abs(anchor_x - global_min_x) <= boundary_tolerance_m}
    right_boundary: set[int] = {column.idx for column, anchor_x, _ in provisional if abs(anchor_x - global_max_x) <= boundary_tolerance_m}
    front_boundary: set[int] = set()
    back_boundary: set[int] = set()

    for _, family in x_families:
        if len(family) >= 2:
            front_boundary.add(min(family, key=lambda col: col.anchor_y).idx)
            back_boundary.add(max(family, key=lambda col: col.anchor_y).idx)
            continue
        column = family[0]
        if abs(column.anchor_y - global_min_y) <= abs(column.anchor_y - global_max_y):
            front_boundary.add(column.idx)
        else:
            back_boundary.add(column.idx)

    corners = (
        (left_boundary & front_boundary)
        | (left_boundary & back_boundary)
        | (right_boundary & front_boundary)
        | (right_boundary & back_boundary)
    )

    derived: list[ColumnRecord] = []
    for column, anchor_x, anchor_y in provisional:
        is_left = column.idx in left_boundary
        is_right = column.idx in right_boundary
        is_front = column.idx in front_boundary
        is_back = column.idx in back_boundary

        geometric_location = "Interior"
        if column.idx in corners:
            geometric_location = "Corner"
        elif is_front:
            geometric_location = "Front edge"
        elif is_back:
            geometric_location = "Back edge"
        elif is_left:
            geometric_location = "Left edge"
        elif is_right:
            geometric_location = "Right edge"

        location = geometric_location
        if geometric_location != "Corner" and (column.source_location or "").strip().lower() == "lift":
            location = "Lift"
        derived.append(
            ColumnRecord(
                idx=column.idx,
                type_name=column.type_name,
                xmin=column.xmin,
                xmax=column.xmax,
                ymin=column.ymin,
                ymax=column.ymax,
                width=column.width,
                height=column.height,
                left_right=column.left_right,
                front_back=column.front_back,
                source_location=column.source_location,
                location=location,
                anchor_location=build_anchor_location(column.front_back, column.left_right),
                anchor_x=anchor_x,
                anchor_y=anchor_y,
                orientation="Vertical" if column.height >= column.width else "Horizontal",
            )
        )
    return derived


def read_walls(path: Path, thickness_tolerance_mm: float) -> list[WallSegment]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {normalize_header(str(value or "")): idx for idx, value in enumerate(header_row)}

    required = {
        "wallno": "Wall No",
        "startxm": "Start X (m)",
        "startym": "Start Y (m)",
        "endxm": "End X (m)",
        "endym": "End Y (m)",
        "thicknessm": "Thickness (m)",
        "orientation": "Orientation",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Missing required columns in wall workbook: {', '.join(missing)}")

    walls: list[WallSegment] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        thickness_m = float(row[header_map["thicknessm"]])
        walls.append(
            WallSegment(
                wall_no=str(row[header_map["wallno"]]),
                start_x=round(float(row[header_map["startxm"]]), 3),
                start_y=round(float(row[header_map["startym"]]), 3),
                end_x=round(float(row[header_map["endxm"]]), 3),
                end_y=round(float(row[header_map["endym"]]), 3),
                thickness_m=round(thickness_m, 3),
                thickness_class_mm=snap_thickness_mm(thickness_m * 1000.0, thickness_tolerance_mm),
                orientation=str(row[header_map["orientation"]]).strip(),
                source=str(row[header_map["source"]]).strip() if "source" in header_map and row[header_map["source"]] not in (None, "") else None,
                dxf_source=str(row[header_map["dxfsource"]]).strip().lower() if "dxfsource" in header_map and row[header_map["dxfsource"]] not in (None, "") else None,
            )
        )
    return walls


def read_zone_rectangles(path: Path | None) -> tuple[list[ZoneRectangle], dict[str, list[tuple[object, ...]]]]:
    if path is None or not path.exists():
        return [], {}

    wb = load_workbook(path, data_only=True)
    zones: list[ZoneRectangle] = []
    extra_sheets: dict[str, list[tuple[object, ...]]] = {}

    if "Rectangle coordinates" in wb.sheetnames:
        ws = wb["Rectangle coordinates"]
        extra_sheets["Rectangle coordinates"] = [tuple(cell.value for cell in row) for row in ws.iter_rows(values_only=False)]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_map = {normalize_header(str(value or "")): idx for idx, value in enumerate(header_row)}
        required = {"no": "No.", "type": "Type", "location": "Location"}
        missing = [label for key, label in required.items() if key not in header_map]
        has_snapped_coords = all(k in header_map for k in ("snappedx1m", "snappedy1m", "snappedx2m", "snappedy2m"))
        has_raw_coords = all(k in header_map for k in ("coordinatex1m", "coordinatey1m", "coordinatex2m", "coordinatey2m"))
        if missing:
            raise SystemExit(f"Rectangle coordinates sheet is missing required headers: {', '.join(missing)}")
        if not has_snapped_coords and not has_raw_coords:
            raise SystemExit("Rectangle coordinates sheet missing coordinate headers (need Snapped X1/Y1/X2/Y2 or Coordinate X1/Y1/X2/Y2)")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(value in (None, "") for value in row):
                continue
            x1_key = "snappedx1m" if "snappedx1m" in header_map else "coordinatex1m"
            y1_key = "snappedy1m" if "snappedy1m" in header_map else "coordinatey1m"
            x2_key = "snappedx2m" if "snappedx2m" in header_map else "coordinatex2m"
            y2_key = "snappedy2m" if "snappedy2m" in header_map else "coordinatey2m"
            zones.append(
                ZoneRectangle(
                    rect_no=str(row[header_map["no"]]).strip(),
                    type_name=str(row[header_map["type"]]).strip(),
                    x1=round(float(row[header_map[x1_key]]), 3),
                    y1=round(float(row[header_map[y1_key]]), 3),
                    x2=round(float(row[header_map[x2_key]]), 3),
                    y2=round(float(row[header_map[y2_key]]), 3),
                    location=str(row[header_map["location"]]).strip(),
                    floors=NONPLINTH_FLOORS,
                )
            )

    for optional_name in ("Staircase details", "Balcony coordinates"):
        if optional_name in wb.sheetnames:
            ws = wb[optional_name]
            extra_sheets[optional_name] = [tuple(cell.value for cell in row) for row in ws.iter_rows(values_only=False)]

    wb.close()
    return zones, extra_sheets


def cluster_columns(columns: list[ColumnRecord], value_fn, tolerance_m: float) -> list[tuple[float, list[ColumnRecord]]]:
    if not columns:
        return []

    sorted_cols = sorted(columns, key=value_fn)
    groups: list[list[ColumnRecord]] = [[sorted_cols[0]]]
    running_values: list[list[float]] = [[value_fn(sorted_cols[0])]]

    for column in sorted_cols[1:]:
        value = value_fn(column)
        current_values = running_values[-1]
        current_center = sum(current_values) / len(current_values)
        if abs(value - current_center) <= tolerance_m:
            groups[-1].append(column)
            current_values.append(value)
        else:
            groups.append([column])
            running_values.append([value])

    clustered: list[tuple[float, list[ColumnRecord]]] = []
    for group in groups:
        representative = round(float(median(value_fn(column) for column in group)), 3)
        clustered.append((representative, group))
    return clustered


def build_beam_class(start: ColumnRecord, end: ColumnRecord, direction: str) -> str:
    edge_x_locations = {"Front edge", "Back edge", "Corner"}
    edge_y_locations = {"Left edge", "Right edge", "Corner"}
    if (
        direction == "X"
        and start.front_back == end.front_back
        and start.front_back in {"Front", "Back"}
        and start.location in edge_x_locations
        and end.location in edge_x_locations
    ):
        return "Edge"
    if (
        direction == "Y"
        and start.left_right == end.left_right
        and start.left_right in {"Left", "Right"}
        and start.location in edge_y_locations
        and end.location in edge_y_locations
    ):
        return "Edge"
    return "Interior"


def generate_primary_beams(columns: list[ColumnRecord], group_tolerance_m: float) -> list[BeamPair]:
    beam_pairs: list[BeamPair] = []

    x_groups = cluster_columns(columns, lambda col: col.anchor_y, group_tolerance_m)
    for group_y, members in x_groups:
        ordered = sorted(members, key=lambda col: col.anchor_x)
        for start, end in zip(ordered, ordered[1:]):
            span = round(abs(end.anchor_x - start.anchor_x), 3)
            if span <= 0:
                continue
            beam_pairs.append(
                BeamPair(
                    beam_no="",
                    start=start,
                    end=end,
                    direction="X",
                    beam_class=build_beam_class(start, end, "X"),
                    beam_start_x=round(start.anchor_x, 3),
                    beam_start_y=group_y,
                    beam_end_x=round(end.anchor_x, 3),
                    beam_end_y=group_y,
                    span_length_m=span,
                    group_coordinate_m=group_y,
                )
            )

    y_groups = cluster_columns(columns, lambda col: col.anchor_x, group_tolerance_m)
    for group_x, members in y_groups:
        ordered = sorted(members, key=lambda col: col.anchor_y)
        for start, end in zip(ordered, ordered[1:]):
            span = round(abs(end.anchor_y - start.anchor_y), 3)
            if span <= 0:
                continue
            beam_pairs.append(
                BeamPair(
                    beam_no="",
                    start=start,
                    end=end,
                    direction="Y",
                    beam_class=build_beam_class(start, end, "Y"),
                    beam_start_x=group_x,
                    beam_start_y=round(start.anchor_y, 3),
                    beam_end_x=group_x,
                    beam_end_y=round(end.anchor_y, 3),
                    span_length_m=span,
                    group_coordinate_m=group_x,
                )
            )

    def sort_key(beam: BeamPair) -> tuple:
        if beam.direction == "X":
            return (0, beam.group_coordinate_m, beam.beam_start_x, beam.beam_end_x)
        return (1, beam.group_coordinate_m, beam.beam_start_y, beam.beam_end_y)

    ordered_pairs = sorted(beam_pairs, key=sort_key)
    return [
        BeamPair(
            beam_no=f"B{idx}",
            start=beam.start,
            end=beam.end,
            direction=beam.direction,
            beam_class=beam.beam_class,
            beam_start_x=beam.beam_start_x,
            beam_start_y=beam.beam_start_y,
            beam_end_x=beam.beam_end_x,
            beam_end_y=beam.beam_end_y,
            span_length_m=beam.span_length_m,
            group_coordinate_m=beam.group_coordinate_m,
        )
        for idx, beam in enumerate(ordered_pairs, start=1)
    ]


def overlap_1d(a1: float, a2: float, b1: float, b2: float) -> tuple[float, float] | None:
    lo = max(min(a1, a2), min(b1, b2))
    hi = min(max(a1, a2), max(b1, b2))
    if hi <= lo:
        return None
    return (round(lo, 3), round(hi, 3))


def merge_intervals(intervals: list[tuple[float, float]]) -> list[tuple[float, float]]:
    if not intervals:
        return []
    ordered = sorted(intervals)
    merged = [ordered[0]]
    for lo, hi in ordered[1:]:
        prev_lo, prev_hi = merged[-1]
        if lo <= prev_hi:
            merged[-1] = (prev_lo, max(prev_hi, hi))
        else:
            merged.append((lo, hi))
    return merged


def compute_wall_assignment(
    beam: BeamPair,
    walls: list[WallSegment],
    wall_alignment_tolerance_m: float,
    edge_wall_coverage_threshold_pct: float,
    interior_wall_coverage_threshold_pct: float,
) -> tuple[int, float]:
    if beam.span_length_m <= 0:
        return 0, 0.0

    coverage_by_class: dict[int, list[tuple[float, float]]] = {115: [], 230: []}
    all_intervals: list[tuple[float, float]] = []
    alignment_tolerance_m = wall_alignment_tolerance_m * (1.5 if beam.beam_class == "Edge" else 1.0)
    for wall in walls:
        if wall.thickness_class_mm not in coverage_by_class:
            continue

        if beam.direction == "X":
            if wall.orientation != "Horizontal":
                continue
            wall_y = (wall.start_y + wall.end_y) / 2.0
            if abs(wall_y - beam.beam_start_y) > alignment_tolerance_m:
                continue
            overlap = overlap_1d(beam.beam_start_x, beam.beam_end_x, wall.start_x, wall.end_x)
        else:
            if wall.orientation != "Vertical":
                continue
            wall_x = (wall.start_x + wall.end_x) / 2.0
            if abs(wall_x - beam.beam_start_x) > alignment_tolerance_m:
                continue
            overlap = overlap_1d(beam.beam_start_y, beam.beam_end_y, wall.start_y, wall.end_y)

        if overlap is not None:
            coverage_by_class[wall.thickness_class_mm].append(overlap)
            all_intervals.append(overlap)

    coverage_pct: dict[int, float] = {}
    for thickness_mm, intervals in coverage_by_class.items():
        merged = merge_intervals(intervals)
        total_length = sum(hi - lo for lo, hi in merged)
        coverage_pct[thickness_mm] = round((total_length / beam.span_length_m) * 100.0, 1)

    raw_best_coverage_pct = 0.0
    if all_intervals:
        merged_all = merge_intervals(all_intervals)
        total_all = sum(hi - lo for lo, hi in merged_all)
        raw_best_coverage_pct = round((total_all / beam.span_length_m) * 100.0, 1)

    threshold_pct = edge_wall_coverage_threshold_pct if beam.beam_class == "Edge" else interior_wall_coverage_threshold_pct
    qualified = [
        (pct, thickness_mm)
        for thickness_mm, pct in coverage_pct.items()
        if pct >= threshold_pct
    ]
    if not qualified:
        return 0, raw_best_coverage_pct

    qualified.sort(key=lambda item: (item[1], item[0]))
    _, winning_thickness = qualified[-1]
    return winning_thickness, raw_best_coverage_pct


def axis_major_values(axis: str, x1: float, y1: float, x2: float, y2: float) -> tuple[float, float, float]:
    if axis == "X":
        return round(y1, 3), round(min(x1, x2), 3), round(max(x1, x2), 3)
    return round(x1, 3), round(min(y1, y2), 3), round(max(y1, y2), 3)


def make_support_segment(axis: str, fixed: float, start: float, end: float, source_kind: str, source_name: str) -> SupportSegment:
    return SupportSegment(
        axis=axis,
        fixed=round(fixed, 3),
        start=round(min(start, end), 3),
        end=round(max(start, end), 3),
        source_kind=source_kind,
        source_name=source_name,
    )


def beam_to_support_segment(beam: BeamPair) -> SupportSegment:
    fixed, start, end = axis_major_values(beam.direction, beam.beam_start_x, beam.beam_start_y, beam.beam_end_x, beam.beam_end_y)
    return make_support_segment(beam.direction, fixed, start, end, "primary", beam.beam_no)


def secondary_to_support_segment(beam: SecondaryBeamCandidate | SecondaryBeam) -> SupportSegment:
    fixed, start, end = axis_major_values(beam.axis, beam.x1, beam.y1, beam.x2, beam.y2)
    name = getattr(beam, "type_name", None) or getattr(beam, "detail", "secondary")
    return make_support_segment(beam.axis, fixed, start, end, "secondary", str(name))


def zone_edge_segments(rectangles: list[ZoneRectangle]) -> list[SupportSegment]:
    segments: list[SupportSegment] = []
    for rect in rectangles:
        x1, x2 = sorted((rect.x1, rect.x2))
        y1, y2 = sorted((rect.y1, rect.y2))
        segments.extend(
            [
                make_support_segment("X", y1, x1, x2, "zone", f"{rect.type_name}:bottom"),
                make_support_segment("X", y2, x1, x2, "zone", f"{rect.type_name}:top"),
                make_support_segment("Y", x1, y1, y2, "zone", f"{rect.type_name}:left"),
                make_support_segment("Y", x2, y1, y2, "zone", f"{rect.type_name}:right"),
            ]
        )
    return segments


def merge_support_segments(segments: list[SupportSegment]) -> list[SupportSegment]:
    grouped: dict[tuple[str, float], list[SupportSegment]] = {}
    for segment in segments:
        grouped.setdefault((segment.axis, segment.fixed), []).append(segment)

    merged: list[SupportSegment] = []
    for (axis, fixed), members in grouped.items():
        ordered = sorted(members, key=lambda seg: (seg.start, seg.end))
        current_start = ordered[0].start
        current_end = ordered[0].end
        current_name = ordered[0].source_name
        current_kind = ordered[0].source_kind
        for segment in ordered[1:]:
            if segment.start <= current_end + 0.001:
                current_end = max(current_end, segment.end)
            else:
                merged.append(make_support_segment(axis, fixed, current_start, current_end, current_kind, current_name))
                current_start = segment.start
                current_end = segment.end
                current_name = segment.source_name
                current_kind = segment.source_kind
        merged.append(make_support_segment(axis, fixed, current_start, current_end, current_kind, current_name))
    return sorted(merged, key=lambda seg: (seg.axis, seg.fixed, seg.start, seg.end))


def build_support_segments(primary_beams: list[BeamPair], accepted: list[SecondaryBeamCandidate], rectangles: list[ZoneRectangle]) -> list[SupportSegment]:
    segments = [beam_to_support_segment(beam) for beam in primary_beams]
    segments.extend(secondary_to_support_segment(beam) for beam in accepted)
    return merge_support_segments(segments + zone_edge_segments(rectangles))


def build_support_segments_from_segments(primary_segments: list[SupportSegment], accepted: list[SecondaryBeamCandidate], rectangles: list[ZoneRectangle]) -> list[SupportSegment]:
    segments = list(primary_segments)
    segments.extend(secondary_to_support_segment(beam) for beam in accepted)
    return merge_support_segments(segments + zone_edge_segments(rectangles))


def point_is_support(
    x: float,
    y: float,
    column_points: list[tuple[str, float, float]],
    support_segments: list[SupportSegment],
    tolerance_m: float = SECONDARY_SUPPORT_TOL_M,
) -> bool:
    for _, col_x, col_y in column_points:
        if abs(col_x - x) <= tolerance_m and abs(col_y - y) <= tolerance_m:
            return True
    for segment in support_segments:
        if segment.axis == "X":
            if abs(segment.fixed - y) <= tolerance_m and segment.start - tolerance_m <= x <= segment.end + tolerance_m:
                return True
        else:
            if abs(segment.fixed - x) <= tolerance_m and segment.start - tolerance_m <= y <= segment.end + tolerance_m:
                return True
    return False


def has_support_in_direction(
    x: float,
    y: float,
    axis: str,
    sign: int,
    support_segments: list[SupportSegment],
    min_span_m: float = SECONDARY_MIN_SPAN_M,
    tolerance_m: float = SECONDARY_SUPPORT_TOL_M,
) -> bool:
    for segment in support_segments:
        if segment.axis != axis:
            continue
        if axis == "X":
            if abs(segment.fixed - y) > tolerance_m or not (segment.start - tolerance_m <= x <= segment.end + tolerance_m):
                continue
            extension = segment.end - x if sign > 0 else x - segment.start
        else:
            if abs(segment.fixed - x) > tolerance_m or not (segment.start - tolerance_m <= y <= segment.end + tolerance_m):
                continue
            extension = segment.end - y if sign > 0 else y - segment.start
        if extension >= min_span_m:
            return True
    return False


def first_support_intersection(
    x: float,
    y: float,
    axis: str,
    sign: int,
    column_points: list[tuple[str, float, float]],
    support_segments: list[SupportSegment],
    min_span_m: float = SECONDARY_MIN_SPAN_M,
    tolerance_m: float = SECONDARY_SUPPORT_TOL_M,
) -> tuple[float, float] | None:
    candidates: list[tuple[float, float, float]] = []
    if axis == "X":
        for _, col_x, col_y in column_points:
            if abs(col_y - y) > tolerance_m:
                continue
            delta = col_x - x
            if sign * delta >= min_span_m:
                candidates.append((abs(delta), round(col_x, 3), round(y, 3)))
        for segment in support_segments:
            if segment.axis != "Y" or not (segment.start - tolerance_m <= y <= segment.end + tolerance_m):
                continue
            delta = segment.fixed - x
            if sign * delta >= min_span_m:
                candidates.append((abs(delta), round(segment.fixed, 3), round(y, 3)))
    else:
        for _, col_x, col_y in column_points:
            if abs(col_x - x) > tolerance_m:
                continue
            delta = col_y - y
            if sign * delta >= min_span_m:
                candidates.append((abs(delta), round(x, 3), round(col_y, 3)))
        for segment in support_segments:
            if segment.axis != "X" or not (segment.start - tolerance_m <= x <= segment.end + tolerance_m):
                continue
            delta = segment.fixed - y
            if sign * delta >= min_span_m:
                candidates.append((abs(delta), round(x, 3), round(segment.fixed, 3)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0])
    return candidates[0][1], candidates[0][2]


def segment_covered_by_supports(
    axis: str,
    fixed: float,
    start: float,
    end: float,
    support_segments: list[SupportSegment],
    tolerance_m: float = SECONDARY_DUPLICATE_MINOR_TOL_M,
) -> bool:
    for segment in support_segments:
        if segment.axis != axis or abs(segment.fixed - fixed) > tolerance_m:
            continue
        if segment.start <= start + SECONDARY_DUPLICATE_ENDPOINT_TOL_M and segment.end >= end - SECONDARY_DUPLICATE_ENDPOINT_TOL_M:
            return True
    return False


def build_secondary_beam_class(
    axis: str,
    fixed: float,
    columns: list[ColumnRecord],
    boundary_tolerance_m: float,
) -> str:
    min_x = min(column.anchor_x for column in columns)
    max_x = max(column.anchor_x for column in columns)
    min_y = min(column.anchor_y for column in columns)
    max_y = max(column.anchor_y for column in columns)
    if axis == "X" and (abs(fixed - min_y) <= boundary_tolerance_m or abs(fixed - max_y) <= boundary_tolerance_m):
        return "Edge"
    if axis == "Y" and (abs(fixed - min_x) <= boundary_tolerance_m or abs(fixed - max_x) <= boundary_tolerance_m):
        return "Edge"
    return "Interior"


def compute_segment_wall_assignment(
    axis: str,
    fixed: float,
    start: float,
    end: float,
    beam_class: str,
    walls: list[WallSegment],
    wall_alignment_tolerance_m: float,
    edge_wall_coverage_threshold_pct: float,
    interior_wall_coverage_threshold_pct: float,
) -> tuple[int, float]:
    span_length_m = round(abs(end - start), 3)
    if span_length_m <= 0:
        return 0, 0.0

    coverage_by_class: dict[int, list[tuple[float, float]]] = {115: [], 230: []}
    all_intervals: list[tuple[float, float]] = []
    for wall in walls:
        if wall.thickness_class_mm not in coverage_by_class:
            continue
        if axis == "X":
            if wall.orientation != "Horizontal":
                continue
            wall_fixed = (wall.start_y + wall.end_y) / 2.0
            if abs(wall_fixed - fixed) > wall_alignment_tolerance_m:
                continue
            overlap = overlap_1d(start, end, wall.start_x, wall.end_x)
        else:
            if wall.orientation != "Vertical":
                continue
            wall_fixed = (wall.start_x + wall.end_x) / 2.0
            if abs(wall_fixed - fixed) > wall_alignment_tolerance_m:
                continue
            overlap = overlap_1d(start, end, wall.start_y, wall.end_y)
        if overlap is not None:
            coverage_by_class[wall.thickness_class_mm].append(overlap)
            all_intervals.append(overlap)

    coverage_pct: dict[int, float] = {}
    for thickness_mm, intervals in coverage_by_class.items():
        merged = merge_intervals(intervals)
        total_length = sum(hi - lo for lo, hi in merged)
        coverage_pct[thickness_mm] = round((total_length / span_length_m) * 100.0, 1)

    raw_best_coverage_pct = 0.0
    if all_intervals:
        merged_all = merge_intervals(all_intervals)
        total_all = sum(hi - lo for lo, hi in merged_all)
        raw_best_coverage_pct = round((total_all / span_length_m) * 100.0, 1)

    threshold_pct = edge_wall_coverage_threshold_pct if beam_class == "Edge" else interior_wall_coverage_threshold_pct
    qualified = [(pct, thickness_mm) for thickness_mm, pct in coverage_pct.items() if pct >= threshold_pct]
    if not qualified:
        return 0, raw_best_coverage_pct
    qualified.sort(key=lambda item: (item[1], item[0]))
    _, winning_thickness = qualified[-1]
    return winning_thickness, raw_best_coverage_pct


def merge_wall_chains(walls: list[WallSegment]) -> list[WallChain]:
    grouped: dict[tuple[str, int, float], list[tuple[float, float]]] = {}
    for wall in walls:
        if wall.thickness_class_mm not in {115, 230}:
            continue
        if wall.orientation == "Horizontal":
            fixed = round((wall.start_y + wall.end_y) / 2.0, 3)
            start = round(min(wall.start_x, wall.end_x), 3)
            end = round(max(wall.start_x, wall.end_x), 3)
            axis = "X"
        elif wall.orientation == "Vertical":
            fixed = round((wall.start_x + wall.end_x) / 2.0, 3)
            start = round(min(wall.start_y, wall.end_y), 3)
            end = round(max(wall.start_y, wall.end_y), 3)
            axis = "Y"
        else:
            continue
        grouped.setdefault((axis, wall.thickness_class_mm, fixed), []).append((start, end))

    chains: list[WallChain] = []
    for (axis, thickness_mm, fixed), intervals in grouped.items():
        intervals.sort()
        current_start, current_end = intervals[0]
        for start, end in intervals[1:]:
            if start <= current_end + SECONDARY_WALL_MERGE_GAP_M:
                current_end = max(current_end, end)
            else:
                chains.append(
                    WallChain(
                        axis=axis,
                        fixed=fixed,
                        start=current_start,
                        end=current_end,
                        length=round(current_end - current_start, 3),
                        thickness_mm=thickness_mm,
                        source="merged",
                    )
                )
                current_start, current_end = start, end
        chains.append(
            WallChain(
                axis=axis,
                fixed=fixed,
                start=current_start,
                end=current_end,
                length=round(current_end - current_start, 3),
                thickness_mm=thickness_mm,
                source="merged",
            )
        )
    return sorted(chains, key=lambda chain: (chain.axis, chain.fixed, chain.start, chain.end))


def find_wall_bracket_supports(
    axis: str,
    fixed: float,
    start: float,
    end: float,
    support_segments: list[SupportSegment],
    column_points: list[tuple[str, float, float]],
    tolerance_m: float = SECONDARY_SUPPORT_TOL_M,
) -> tuple[float, float] | None:
    """Find the nearest perpendicular structural supports that bracket a wall chain.
    Searches for crossing beams and column nodes, not just parallel beams."""
    perp_axis = "Y" if axis == "X" else "X"
    left_vals: list[float] = []
    right_vals: list[float] = []
    for segment in support_segments:
        if segment.axis == perp_axis:
            if not (segment.start - tolerance_m <= fixed <= segment.end + tolerance_m):
                continue
            if segment.fixed <= start + tolerance_m:
                left_vals.append(segment.fixed)
            if segment.fixed >= end - tolerance_m:
                right_vals.append(segment.fixed)
        elif segment.axis == axis and abs(segment.fixed - fixed) <= tolerance_m:
            if segment.start - tolerance_m <= start <= segment.end + tolerance_m:
                left_vals.append(start)
            if segment.start - tolerance_m <= end <= segment.end + tolerance_m:
                right_vals.append(end)
    for _, cx, cy in column_points:
        coord, perp = (cx, cy) if axis == "X" else (cy, cx)
        if abs(perp - fixed) <= tolerance_m:
            if coord <= start + tolerance_m:
                left_vals.append(coord)
            if coord >= end - tolerance_m:
                right_vals.append(coord)
    if not left_vals or not right_vals:
        return None
    left_bracket = max(left_vals)
    right_bracket = min(right_vals)
    if right_bracket - left_bracket < SECONDARY_MIN_SPAN_M:
        return None
    return round(left_bracket, 3), round(right_bracket, 3)


def find_bracketing_supports(
    axis: str,
    fixed: float,
    start: float,
    end: float,
    column_points: list[tuple[str, float, float]],
    support_segments: list[SupportSegment],
) -> tuple[float, float] | None:
    if axis == "X":
        left_point = (start, fixed)
        right_point = (end, fixed)
    else:
        left_point = (fixed, start)
        right_point = (fixed, end)

    left_support = start if point_is_support(*left_point, column_points, support_segments) else None
    right_support = end if point_is_support(*right_point, column_points, support_segments) else None
    if left_support is None:
        left_hit = first_support_intersection(left_point[0], left_point[1], axis, -1, column_points, support_segments)
        if left_hit is not None:
            left_support = left_hit[0] if axis == "X" else left_hit[1]
    if right_support is None:
        right_hit = first_support_intersection(right_point[0], right_point[1], axis, +1, column_points, support_segments)
        if right_hit is not None:
            right_support = right_hit[0] if axis == "X" else right_hit[1]
    if left_support is None or right_support is None or right_support - left_support < SECONDARY_MIN_SPAN_M:
        return None
    return round(left_support, 3), round(right_support, 3)


def clip_wall_support_extent(
    axis: str,
    fixed: float,
    start: float,
    end: float,
    column_points: list[tuple[str, float, float]],
    support_segments: list[SupportSegment],
    max_extension_m: float = SECONDARY_R4_END_EXTENSION_M,
) -> tuple[float, float]:
    clipped_start = start
    clipped_end = end

    if axis == "X":
        start_point = (start, fixed)
        end_point = (end, fixed)
    else:
        start_point = (fixed, start)
        end_point = (fixed, end)

    if not point_is_support(*start_point, column_points, support_segments):
        hit = first_support_intersection(start_point[0], start_point[1], axis, -1, column_points, support_segments, min_span_m=0.001)
        if hit is not None:
            target = hit[0] if axis == "X" else hit[1]
            if start - target <= max_extension_m:
                clipped_start = round(target, 3)

    if not point_is_support(*end_point, column_points, support_segments):
        hit = first_support_intersection(end_point[0], end_point[1], axis, +1, column_points, support_segments, min_span_m=0.001)
        if hit is not None:
            target = hit[0] if axis == "X" else hit[1]
            if target - end <= max_extension_m:
                clipped_end = round(target, 3)

    return clipped_start, clipped_end


def cell_is_near_zone(x1: float, x2: float, y1: float, y2: float, rectangles: list[ZoneRectangle], proximity_m: float = SECONDARY_R1_ZONE_PROXIMITY_M) -> bool:
    for rect in rectangles:
        rx1, rx2 = sorted((rect.x1, rect.x2))
        ry1, ry2 = sorted((rect.y1, rect.y2))
        dx = max(rx1 - x2, x1 - rx2, 0.0)
        dy = max(ry1 - y2, y1 - ry2, 0.0)
        if dx <= proximity_m and dy <= proximity_m:
            return True
    return False


def nearest_outer_grid_line(values: list[float], current: float, direction: int) -> float | None:
    if direction > 0:
        candidates = [value for value in values if value > current + 0.001]
        return min(candidates) if candidates else None
    candidates = [value for value in values if value < current - 0.001]
    return max(candidates) if candidates else None


def segment_has_exact_structural_support(
    axis: str,
    fixed: float,
    start: float,
    end: float,
    support_segments: list[SupportSegment],
    tolerance_m: float = SECONDARY_R5_SUPPORT_TOL_M,
) -> bool:
    for segment in support_segments:
        if segment.axis != axis or abs(segment.fixed - fixed) > tolerance_m:
            continue
        if segment.start <= start + tolerance_m and segment.end >= end - tolerance_m:
            return True
    return False


def chain_aligns_with_zone_edge(
    chain: WallChain,
    rectangles: list[ZoneRectangle],
    tolerance_m: float = SECONDARY_R5_EDGE_ALIGN_TOL_M,
) -> bool:
    for rect in rectangles:
        x1, x2 = sorted((rect.x1, rect.x2))
        y1, y2 = sorted((rect.y1, rect.y2))
        if chain.axis == "X":
            if abs(chain.fixed - y1) > tolerance_m and abs(chain.fixed - y2) > tolerance_m:
                continue
            overlap = overlap_1d(chain.start, chain.end, x1, x2)
        else:
            if abs(chain.fixed - x1) > tolerance_m and abs(chain.fixed - x2) > tolerance_m:
                continue
            overlap = overlap_1d(chain.start, chain.end, y1, y2)
        if overlap is not None and (overlap[1] - overlap[0]) >= min(chain.length, 0.8):
            return True
    return False


def chain_near_zone_interior_projection(
    chain: WallChain,
    rectangles: list[ZoneRectangle],
    tolerance_m: float = SECONDARY_R5_EDGE_ALIGN_TOL_M,
) -> bool:
    for rect in rectangles:
        x1, x2 = sorted((rect.x1, rect.x2))
        y1, y2 = sorted((rect.y1, rect.y2))
        if chain.axis == "Y":
            if not (x1 + tolerance_m < chain.fixed < x2 - tolerance_m):
                continue
            if abs(chain.start - y1) <= tolerance_m or abs(chain.start - y2) <= tolerance_m or abs(chain.end - y1) <= tolerance_m or abs(chain.end - y2) <= tolerance_m:
                return True
        else:
            if not (y1 + tolerance_m < chain.fixed < y2 - tolerance_m):
                continue
            if abs(chain.start - x1) <= tolerance_m or abs(chain.start - x2) <= tolerance_m or abs(chain.end - x1) <= tolerance_m or abs(chain.end - x2) <= tolerance_m:
                return True
    return False


def nearest_rectangle(columns_or_point_x: float, point_y: float, rectangles: list[ZoneRectangle]) -> ZoneRectangle | None:
    best_rect: ZoneRectangle | None = None
    best_dist: float | None = None
    for rect in rectangles:
        x1, x2 = sorted((rect.x1, rect.x2))
        y1, y2 = sorted((rect.y1, rect.y2))
        dx = max(x1 - columns_or_point_x, columns_or_point_x - x2, 0.0)
        dy = max(y1 - point_y, point_y - y2, 0.0)
        dist = dx + dy
        if best_dist is None or dist < best_dist:
            best_dist = dist
            best_rect = rect
    return best_rect if best_dist is not None and best_dist <= SECONDARY_R1_ZONE_PROXIMITY_M * 2.0 else None


def candidate_duplicate_key(candidate: SecondaryBeamCandidate) -> tuple[str, float, float, float]:
    fixed, start, end = axis_major_values(candidate.axis, candidate.x1, candidate.y1, candidate.x2, candidate.y2)
    return (candidate.axis, fixed, start, end)


def candidate_is_duplicate(candidate: SecondaryBeamCandidate, segments: list[SupportSegment]) -> bool:
    fixed, start, end = axis_major_values(candidate.axis, candidate.x1, candidate.y1, candidate.x2, candidate.y2)
    return segment_covered_by_supports(candidate.axis, fixed, start, end, segments)


def candidate_parallel_conflict(candidate: SecondaryBeamCandidate, segments: list[SupportSegment]) -> bool:
    if candidate.rule_code in {"R2", "R5"}:
        return False
    fixed, start, end = axis_major_values(candidate.axis, candidate.x1, candidate.y1, candidate.x2, candidate.y2)
    candidate_span = end - start
    for segment in segments:
        if segment.axis != candidate.axis:
            continue
        if abs(segment.fixed - fixed) >= SECONDARY_PARALLEL_SPACING_MIN_M:
            continue
        overlap = overlap_1d(start, end, segment.start, segment.end)
        if overlap is None:
            continue
        overlap_ratio = (overlap[1] - overlap[0]) / candidate_span if candidate_span else 0.0
        if overlap_ratio >= 0.5:
            return True
    return False


def candidate_crosses_zone(candidate: SecondaryBeamCandidate, rectangles: list[ZoneRectangle]) -> bool:
    if candidate.rule_code == "R5":
        return False
    fixed, start, end = axis_major_values(candidate.axis, candidate.x1, candidate.y1, candidate.x2, candidate.y2)
    for rect in rectangles:
        if str(rect.location).strip().lower() == "mumty":
            continue
        x1, x2 = sorted((rect.x1, rect.x2))
        y1, y2 = sorted((rect.y1, rect.y2))
        if candidate.axis == "X":
            if not (y1 + SECONDARY_ZONE_INTERIOR_TOL_M < fixed < y2 - SECONDARY_ZONE_INTERIOR_TOL_M):
                continue
            overlap = overlap_1d(start, end, x1, x2)
        else:
            if not (x1 + SECONDARY_ZONE_INTERIOR_TOL_M < fixed < x2 - SECONDARY_ZONE_INTERIOR_TOL_M):
                continue
            overlap = overlap_1d(start, end, y1, y2)
        if overlap is not None and (overlap[1] - overlap[0]) > SECONDARY_ZONE_INTERIOR_TOL_M:
            return True
    return False


def support_corner_points(
    support_segments: list[SupportSegment],
    column_points: list[tuple[str, float, float]],
    tolerance_m: float = 0.001,
) -> set[tuple[float, float]]:
    points: set[tuple[float, float]] = set()
    for _, x, y in column_points:
        points.add((round(x, 3), round(y, 3)))
    horizontals = [segment for segment in support_segments if segment.axis == "X"]
    verticals = [segment for segment in support_segments if segment.axis == "Y"]
    for segment in support_segments:
        if segment.axis == "X":
            points.add((round(segment.start, 3), round(segment.fixed, 3)))
            points.add((round(segment.end, 3), round(segment.fixed, 3)))
        else:
            points.add((round(segment.fixed, 3), round(segment.start, 3)))
            points.add((round(segment.fixed, 3), round(segment.end, 3)))
    for h_seg in horizontals:
        for v_seg in verticals:
            if v_seg.start - tolerance_m <= h_seg.fixed <= v_seg.end + tolerance_m and h_seg.start - tolerance_m <= v_seg.fixed <= h_seg.end + tolerance_m:
                points.add((round(v_seg.fixed, 3), round(h_seg.fixed, 3)))
    return points


def enumerate_closed_rectangles(
    support_segments: list[SupportSegment],
    column_points: list[tuple[str, float, float]],
) -> list[ClosedRectangle]:
    corner_points = support_corner_points(support_segments, column_points)
    xs = sorted({point[0] for point in corner_points})
    ys = sorted({point[1] for point in corner_points})
    seen: set[tuple[float, float, float, float]] = set()
    rectangles: list[ClosedRectangle] = []

    for i0, x1 in enumerate(xs[:-1]):
        for x2 in xs[i0 + 1 :]:
            for j0, y1 in enumerate(ys[:-1]):
                for y2 in ys[j0 + 1 :]:
                    if (
                        (x1, y1) not in corner_points
                        or (x1, y2) not in corner_points
                        or (x2, y1) not in corner_points
                        or (x2, y2) not in corner_points
                    ):
                        continue
                    if not support_side_exists("X", y1, x1, x2, support_segments):
                        continue
                    if not support_side_exists("X", y2, x1, x2, support_segments):
                        continue
                    if not support_side_exists("Y", x1, y1, y2, support_segments):
                        continue
                    if not support_side_exists("Y", x2, y1, y2, support_segments):
                        continue

                    has_split = False
                    for x_mid in xs:
                        if x1 + 0.001 < x_mid < x2 - 0.001 and support_side_exists("Y", x_mid, y1, y2, support_segments):
                            has_split = True
                            break
                    if has_split:
                        continue
                    for y_mid in ys:
                        if y1 + 0.001 < y_mid < y2 - 0.001 and support_side_exists("X", y_mid, x1, x2, support_segments):
                            has_split = True
                            break
                    if has_split:
                        continue

                    key = (round(x1, 3), round(x2, 3), round(y1, 3), round(y2, 3))
                    if key in seen:
                        continue
                    seen.add(key)
                    rectangles.append(ClosedRectangle(x1=key[0], x2=key[1], y1=key[2], y2=key[3]))
    return rectangles


def candidate_is_internal_split(
    candidate: SecondaryBeamCandidate,
    closed_rectangles: list[ClosedRectangle],
    tolerance_m: float = SECONDARY_SUPPORT_TOL_M,
) -> bool:
    fixed, start, end = axis_major_values(candidate.axis, candidate.x1, candidate.y1, candidate.x2, candidate.y2)
    for rect in closed_rectangles:
        if candidate.axis == "Y":
            if abs(start - rect.y1) <= tolerance_m and abs(end - rect.y2) <= tolerance_m and rect.x1 + tolerance_m < fixed < rect.x2 - tolerance_m:
                return True
        else:
            if abs(start - rect.x1) <= tolerance_m and abs(end - rect.x2) <= tolerance_m and rect.y1 + tolerance_m < fixed < rect.y2 - tolerance_m:
                return True
    return False


def generate_r2_edge_perpendicular(
    columns: list[ColumnRecord],
    column_points: list[tuple[str, float, float]],
    support_segments: list[SupportSegment],
    floor_group: str,
    floors: tuple[str, ...],
    boundary_tolerance_m: float,
) -> list[SecondaryBeamCandidate]:
    candidates: list[SecondaryBeamCandidate] = []
    point_lookup = {name: (cx, cy) for name, cx, cy in column_points}
    min_x = min(cx for _, cx, _ in column_points)
    max_x = max(cx for _, cx, _ in column_points)
    min_y = min(cy for _, _, cy in column_points)
    max_y = max(cy for _, _, cy in column_points)

    def neighbors(name: str) -> tuple[bool, bool, bool, bool]:
        x0, y0 = point_lookup[name]
        has_left = any(abs(other_y - y0) <= boundary_tolerance_m and other_x < x0 - 0.001 for other_name, other_x, other_y in column_points if other_name != name)
        has_right = any(abs(other_y - y0) <= boundary_tolerance_m and other_x > x0 + 0.001 for other_name, other_x, other_y in column_points if other_name != name)
        has_front = any(abs(other_x - x0) <= boundary_tolerance_m and other_y < y0 - 0.001 for other_name, other_x, other_y in column_points if other_name != name)
        has_back = any(abs(other_x - x0) <= boundary_tolerance_m and other_y > y0 + 0.001 for other_name, other_x, other_y in column_points if other_name != name)
        return has_left, has_right, has_front, has_back

    def add_candidate(column: ColumnRecord, axis: str, sign: int, reason: str) -> None:
        x0, y0 = column_points[[name for name, _, _ in column_points].index(column.type_name)][1:]
        if has_support_in_direction(x0, y0, axis, sign, support_segments):
            return
        end_point = first_support_intersection(x0, y0, axis, sign, column_points, support_segments)
        if end_point is None:
            return
        x2, y2 = end_point
        fixed, start, end = axis_major_values(axis, x0, y0, x2, y2)
        if end - start < SECONDARY_MIN_SPAN_M:
            return
        candidates.append(
            SecondaryBeamCandidate(
                axis=axis,
                x1=round(x0, 3),
                y1=round(y0, 3),
                x2=round(x2, 3),
                y2=round(y2, 3),
                beam_location="Centre",
                floor_group=floor_group,
                floors=floors,
                rule_code="R2",
                beam_class=build_secondary_beam_class(axis, fixed, columns, boundary_tolerance_m),
                detail=f"{column.type_name}:{reason}",
                score=100.0,
            )
        )

    for column in columns:
        has_left, has_right, has_front, has_back = neighbors(column.type_name)
        location = (column.location or "").strip()

        if location == "Front edge":
            add_candidate(column, "Y", +1, "front_inward")
            continue
        if location == "Back edge":
            add_candidate(column, "Y", -1, "back_inward")
            continue
        if location == "Left edge":
            add_candidate(column, "X", +1, "left_inward")
            continue
        if location == "Right edge":
            add_candidate(column, "X", -1, "right_inward")
            continue
        if location == "Corner":
            col_x, col_y = point_lookup[column.type_name]
            if not has_front:
                add_candidate(column, "Y", +1, "front_inward")
            elif not has_back:
                add_candidate(column, "Y", -1, "back_inward")
            else:
                if abs(col_y - min_y) <= abs(col_y - max_y):
                    add_candidate(column, "Y", +1, "front_inward")
                else:
                    add_candidate(column, "Y", -1, "back_inward")

            if not has_left:
                add_candidate(column, "X", +1, "left_inward")
            elif not has_right:
                add_candidate(column, "X", -1, "right_inward")
            else:
                if abs(col_x - min_x) <= abs(col_x - max_x):
                    add_candidate(column, "X", +1, "left_inward")
                else:
                    add_candidate(column, "X", -1, "right_inward")
    return candidates


def generate_r5_zone_perimeter(
    columns: list[ColumnRecord],
    rectangles: list[ZoneRectangle],
    structural_supports: list[SupportSegment],
    floor_group: str,
    floors: tuple[str, ...],
    boundary_tolerance_m: float,
) -> list[SecondaryBeamCandidate]:
    candidates: list[SecondaryBeamCandidate] = []
    for rect in rectangles:
        if str(rect.location).strip().lower() == "mumty":
            continue
        x1, x2 = sorted((rect.x1, rect.x2))
        y1, y2 = sorted((rect.y1, rect.y2))
        edges = [
            ("X", y1, x1, x2, x1, y1, x2, y1, "bottom"),
            ("X", y2, x1, x2, x1, y2, x2, y2, "top"),
            ("Y", x1, y1, y2, x1, y1, x1, y2, "left"),
            ("Y", x2, y1, y2, x2, y1, x2, y2, "right"),
        ]
        for axis, fixed, start, end, sx, sy, ex, ey, side in edges:
            if segment_has_exact_structural_support(axis, fixed, start, end, structural_supports):
                continue
            candidates.append(
                SecondaryBeamCandidate(
                    axis=axis,
                    x1=round(sx, 3),
                    y1=round(sy, 3),
                    x2=round(ex, 3),
                    y2=round(ey, 3),
                    beam_location="Centre",
                    floor_group=floor_group,
                    floors=floors,
                    rule_code="R5",
                    beam_class=build_secondary_beam_class(axis, fixed, columns, boundary_tolerance_m),
                    detail=f"{rect.location or rect.type_name}:{side}",
                    score=90.0,
                )
            )
    return candidates


def support_side_exists(axis: str, fixed: float, start: float, end: float, support_segments: list[SupportSegment]) -> bool:
    return segment_covered_by_supports(axis, fixed, start, end, support_segments, tolerance_m=0.001)


def side_coverage_intervals(
    axis: str,
    fixed: float,
    start: float,
    end: float,
    support_segments: list[SupportSegment],
    tolerance_m: float = 0.001,
) -> list[tuple[float, float]]:
    intervals: list[tuple[float, float]] = []
    for segment in support_segments:
        if segment.axis != axis or abs(segment.fixed - fixed) > tolerance_m:
            continue
        overlap = overlap_1d(start, end, segment.start, segment.end)
        if overlap is not None:
            intervals.append(overlap)
    return merge_intervals(intervals)


def side_missing_intervals(start: float, end: float, covered: list[tuple[float, float]], tolerance_m: float = 0.001) -> list[tuple[float, float]]:
    missing: list[tuple[float, float]] = []
    cursor = round(min(start, end), 3)
    target_end = round(max(start, end), 3)
    for lo, hi in covered:
        if lo - cursor > tolerance_m:
            missing.append((round(cursor, 3), round(lo, 3)))
        cursor = max(cursor, hi)
    if target_end - cursor > tolerance_m:
        missing.append((round(cursor, 3), round(target_end, 3)))
    return missing


def generate_r1_shape_closure(
    columns: list[ColumnRecord],
    column_points: list[tuple[str, float, float]],
    support_segments: list[SupportSegment],
    rectangles: list[ZoneRectangle],
    floor_group: str,
    floors: tuple[str, ...],
    boundary_tolerance_m: float,
) -> list[SecondaryBeamCandidate]:
    closed_rectangles = enumerate_closed_rectangles(support_segments, column_points)
    xs = sorted(
        {
            round(segment.fixed, 3)
            for segment in support_segments
            if segment.axis == "Y"
        }
        | {
            round(value, 3)
            for segment in support_segments
            if segment.axis == "X"
            for value in (segment.start, segment.end)
        }
    )
    ys = sorted(
        {
            round(segment.fixed, 3)
            for segment in support_segments
            if segment.axis == "X"
        }
        | {
            round(value, 3)
            for segment in support_segments
            if segment.axis == "Y"
            for value in (segment.start, segment.end)
        }
    )
    candidates: list[SecondaryBeamCandidate] = []
    seen_keys: set[tuple[str, float, float, float]] = set()

    def add_candidate(axis: str, sx: float, sy: float, ex: float, ey: float, detail: str) -> None:
        fixed, start, end = axis_major_values(axis, sx, sy, ex, ey)
        span = round(end - start, 3)
        if not (SECONDARY_MIN_SPAN_M <= span <= SECONDARY_R1_MAX_SPAN_M):
            return
        candidate = SecondaryBeamCandidate(
            axis=axis,
            x1=round(sx, 3),
            y1=round(sy, 3),
            x2=round(ex, 3),
            y2=round(ey, 3),
            beam_location="Centre",
            floor_group=floor_group,
            floors=floors,
            rule_code="R1",
            beam_class=build_secondary_beam_class(axis, fixed, columns, boundary_tolerance_m),
            detail=detail,
            score=100.0 - span,
        )
        if candidate_crosses_zone(candidate, rectangles):
            return
        if candidate_is_internal_split(candidate, closed_rectangles):
            return
        # Fallback: catch internal splits that enumerate_closed_rectangles misses due to
        # corner-registration edge cases. If parallel supports exist on both sides of this
        # candidate covering the same span, AND the perpendicular caps close the rectangle,
        # this candidate is an unnecessary internal split of an already-framed panel.
        _ep = SECONDARY_DUPLICATE_ENDPOINT_TOL_M
        _wall_axis = axis
        _cap_axis = "X" if axis == "Y" else "Y"
        _left_walls = [
            seg.fixed for seg in support_segments
            if seg.axis == _wall_axis and seg.fixed < fixed - SECONDARY_SUPPORT_TOL_M
            and seg.start <= start + _ep and seg.end >= end - _ep
        ]
        _right_walls = [
            seg.fixed for seg in support_segments
            if seg.axis == _wall_axis and seg.fixed > fixed + SECONDARY_SUPPORT_TOL_M
            and seg.start <= start + _ep and seg.end >= end - _ep
        ]
        if _left_walls and _right_walls:
            _lb, _rb = max(_left_walls), min(_right_walls)
            if (support_side_exists(_cap_axis, start, _lb, _rb, support_segments) and
                    support_side_exists(_cap_axis, end, _lb, _rb, support_segments)):
                return
        key = candidate_duplicate_key(candidate)
        if key in seen_keys:
            return
        seen_keys.add(key)
        candidates.append(candidate)

    def three_other_sides_exist(x1: float, x2: float, y1: float, y2: float, missing_side_name: str) -> bool:
        checks = {
            "left": (
                support_side_exists("Y", x2, y1, y2, support_segments),
                support_side_exists("X", y1, x1, x2, support_segments),
                support_side_exists("X", y2, x1, x2, support_segments),
            ),
            "right": (
                support_side_exists("Y", x1, y1, y2, support_segments),
                support_side_exists("X", y1, x1, x2, support_segments),
                support_side_exists("X", y2, x1, x2, support_segments),
            ),
            "bottom": (
                support_side_exists("X", y2, x1, x2, support_segments),
                support_side_exists("Y", x1, y1, y2, support_segments),
                support_side_exists("Y", x2, y1, y2, support_segments),
            ),
            "top": (
                support_side_exists("X", y1, x1, x2, support_segments),
                support_side_exists("Y", x1, y1, y2, support_segments),
                support_side_exists("Y", x2, y1, y2, support_segments),
            ),
        }
        return all(checks[missing_side_name])

    def side_gaps(side_name: str, x1: float, x2: float, y1: float, y2: float) -> list[tuple[float, float]]:
        if side_name == "left":
            covered = side_coverage_intervals("Y", x1, y1, y2, support_segments)
            return side_missing_intervals(y1, y2, covered)
        if side_name == "right":
            covered = side_coverage_intervals("Y", x2, y1, y2, support_segments)
            return side_missing_intervals(y1, y2, covered)
        if side_name == "bottom":
            covered = side_coverage_intervals("X", y1, x1, x2, support_segments)
            return side_missing_intervals(x1, x2, covered)
        covered = side_coverage_intervals("X", y2, x1, x2, support_segments)
        return side_missing_intervals(x1, x2, covered)

    def maybe_add_extended(side_name: str, x1: float, x2: float, y1: float, y2: float, gap_start: float, gap_end: float) -> None:
        if side_name in {"left", "right"}:
            fixed = x1 if side_name == "left" else x2
            if abs(gap_start - y1) <= SECONDARY_SUPPORT_TOL_M:
                hit = first_support_intersection(fixed, y1, "Y", -1, column_points, support_segments, min_span_m=0.001)
                if hit is not None and hit[1] < y1 - SECONDARY_SUPPORT_TOL_M:
                    new_y1 = round(hit[1], 3)
                    if three_other_sides_exist(x1, x2, new_y1, y2, side_name):
                        gaps = side_gaps(side_name, x1, x2, new_y1, y2)
                        if len(gaps) == 1 and abs(gaps[0][0] - new_y1) <= SECONDARY_SUPPORT_TOL_M:
                            add_candidate("Y", fixed, new_y1, fixed, gaps[0][1], f"rect({x1},{x2},{new_y1},{y2}) extend {side_name}")
            if abs(gap_end - y2) <= SECONDARY_SUPPORT_TOL_M:
                hit = first_support_intersection(fixed, y2, "Y", +1, column_points, support_segments, min_span_m=0.001)
                if hit is not None and hit[1] > y2 + SECONDARY_SUPPORT_TOL_M:
                    new_y2 = round(hit[1], 3)
                    if three_other_sides_exist(x1, x2, y1, new_y2, side_name):
                        gaps = side_gaps(side_name, x1, x2, y1, new_y2)
                        if len(gaps) == 1 and abs(gaps[0][1] - new_y2) <= SECONDARY_SUPPORT_TOL_M:
                            add_candidate("Y", fixed, gaps[0][0], fixed, new_y2, f"rect({x1},{x2},{y1},{new_y2}) extend {side_name}")
        else:
            fixed = y1 if side_name == "bottom" else y2
            if abs(gap_start - x1) <= SECONDARY_SUPPORT_TOL_M:
                hit = first_support_intersection(x1, fixed, "X", -1, column_points, support_segments, min_span_m=0.001)
                if hit is not None and hit[0] < x1 - SECONDARY_SUPPORT_TOL_M:
                    new_x1 = round(hit[0], 3)
                    if three_other_sides_exist(new_x1, x2, y1, y2, side_name):
                        gaps = side_gaps(side_name, new_x1, x2, y1, y2)
                        if len(gaps) == 1 and abs(gaps[0][0] - new_x1) <= SECONDARY_SUPPORT_TOL_M:
                            add_candidate("X", new_x1, fixed, gaps[0][1], fixed, f"rect({new_x1},{x2},{y1},{y2}) extend {side_name}")
            if abs(gap_end - x2) <= SECONDARY_SUPPORT_TOL_M:
                hit = first_support_intersection(x2, fixed, "X", +1, column_points, support_segments, min_span_m=0.001)
                if hit is not None and hit[0] > x2 + SECONDARY_SUPPORT_TOL_M:
                    new_x2 = round(hit[0], 3)
                    if three_other_sides_exist(x1, new_x2, y1, y2, side_name):
                        gaps = side_gaps(side_name, x1, new_x2, y1, y2)
                        if len(gaps) == 1 and abs(gaps[0][1] - new_x2) <= SECONDARY_SUPPORT_TOL_M:
                            add_candidate("X", gaps[0][0], fixed, new_x2, fixed, f"rect({x1},{new_x2},{y1},{y2}) extend {side_name}")

    for x_idx, x1 in enumerate(xs[:-1]):
        for x2 in xs[x_idx + 1 :]:
            for y_idx, y1 in enumerate(ys[:-1]):
                for y2 in ys[y_idx + 1 :]:
                    # Reject rectangles already subdivided by an internal full split.
                    has_internal_split = False
                    for x_mid in xs:
                        if x1 + 0.001 < x_mid < x2 - 0.001 and support_side_exists("Y", x_mid, y1, y2, support_segments):
                            has_internal_split = True
                            break
                    if has_internal_split:
                        continue
                    for y_mid in ys:
                        if y1 + 0.001 < y_mid < y2 - 0.001 and support_side_exists("X", y_mid, x1, x2, support_segments):
                            has_internal_split = True
                            break
                    if has_internal_split:
                        continue

                    sides = {
                        "left": side_coverage_intervals("Y", x1, y1, y2, support_segments),
                        "right": side_coverage_intervals("Y", x2, y1, y2, support_segments),
                        "bottom": side_coverage_intervals("X", y1, x1, x2, support_segments),
                        "top": side_coverage_intervals("X", y2, x1, x2, support_segments),
                    }
                    gapped_sides: list[tuple[str, list[tuple[float, float]]]] = []
                    for side_name, intervals in sides.items():
                        side_start, side_end = (y1, y2) if side_name in {"left", "right"} else (x1, x2)
                        gaps = side_missing_intervals(side_start, side_end, intervals)
                        if gaps:
                            gapped_sides.append((side_name, gaps))
                    if len(gapped_sides) != 1:
                        continue
                    missing_side_name, missing_intervals = gapped_sides[0]
                    if len(missing_intervals) != 1:
                        continue
                    gap_start, gap_end = missing_intervals[0]
                    if not three_other_sides_exist(x1, x2, y1, y2, missing_side_name):
                        continue

                    if missing_side_name in {"left", "right"}:
                        fixed = x1 if missing_side_name == "left" else x2
                        add_candidate("Y", fixed, gap_start, fixed, gap_end, f"rect({x1},{x2},{y1},{y2}) close {missing_side_name} gap({gap_start},{gap_end})")
                    else:
                        fixed = y1 if missing_side_name == "bottom" else y2
                        add_candidate("X", gap_start, fixed, gap_end, fixed, f"rect({x1},{x2},{y1},{y2}) close {missing_side_name} gap({gap_start},{gap_end})")

                    maybe_add_extended(missing_side_name, x1, x2, y1, y2, gap_start, gap_end)
    return candidates


def generate_r4_wall_support(
    columns: list[ColumnRecord],
    column_points: list[tuple[str, float, float]],
    structural_supports: list[SupportSegment],
    wall_chains: list[WallChain],
    rectangles: list[ZoneRectangle],
    floor_group: str,
    floors: tuple[str, ...],
    boundary_tolerance_m: float,
) -> list[SecondaryBeamCandidate]:
    candidates: list[SecondaryBeamCandidate] = []
    min_length_m = SECONDARY_R4_NONPLINTH_MIN_WALL_LENGTH_M
    for chain in wall_chains:
        if chain.length < min_length_m:
            continue
        if segment_covered_by_supports(chain.axis, chain.fixed, chain.start, chain.end, structural_supports):
            continue
        if floor_group != "plinth" and chain.length < SECONDARY_R4_STRONG_WALL_LENGTH_M and not chain_aligns_with_zone_edge(chain, rectangles):
            continue
        if chain_near_zone_interior_projection(chain, rectangles):
            continue
        if floor_group == "plinth":
            extents = find_wall_bracket_supports(chain.axis, chain.fixed, chain.start, chain.end, structural_supports, column_points)
        else:
            extents = find_bracketing_supports(chain.axis, chain.fixed, chain.start, chain.end, column_points, structural_supports)
        if extents is None:
            continue
        support_start, support_end = extents
        ext_left = round(chain.start - support_start, 3)
        ext_right = round(support_end - chain.end, 3)
        total_span = round(support_end - support_start, 3)
        if max(ext_left, ext_right) > 4.0:
            continue
        if total_span > (chain.length * 2.5):
            continue
        if support_end - support_start < min_length_m:
            continue
        if chain.axis == "X":
            x1, y1, x2, y2 = support_start, chain.fixed, support_end, chain.fixed
        else:
            x1, y1, x2, y2 = chain.fixed, support_start, chain.fixed, support_end
        candidates.append(
            SecondaryBeamCandidate(
                axis=chain.axis,
                x1=round(x1, 3),
                y1=round(y1, 3),
                x2=round(x2, 3),
                y2=round(y2, 3),
                beam_location="Centre",
                floor_group=floor_group,
                floors=floors,
                rule_code="R4",
                beam_class=build_secondary_beam_class(chain.axis, chain.fixed, columns, boundary_tolerance_m),
                detail=f"wall_chain({chain.axis},{chain.fixed},{chain.start},{chain.end})",
                score=35.0,
            )
        )
    return candidates


def closed_cells(support_segments: list[SupportSegment], column_points: list[tuple[str, float, float]]) -> list[tuple[float, float, float, float]]:
    xs = sorted(
        {
            round(segment.fixed, 3)
            for segment in support_segments
            if segment.axis == "Y"
        }
        | {
            round(value, 3)
            for segment in support_segments
            if segment.axis == "X"
            for value in (segment.start, segment.end)
        }
    )
    ys = sorted(
        {
            round(segment.fixed, 3)
            for segment in support_segments
            if segment.axis == "X"
        }
        | {
            round(value, 3)
            for segment in support_segments
            if segment.axis == "Y"
            for value in (segment.start, segment.end)
        }
    )
    cells: list[tuple[float, float, float, float]] = []
    for x1, x2 in zip(xs, xs[1:]):
        for y1, y2 in zip(ys, ys[1:]):
            if (
                support_side_exists("Y", x1, y1, y2, support_segments)
                and support_side_exists("Y", x2, y1, y2, support_segments)
                and support_side_exists("X", y1, x1, x2, support_segments)
                and support_side_exists("X", y2, x1, x2, support_segments)
            ):
                cells.append((x1, x2, y1, y2))
    return cells


def generate_r3_large_panel_dividers(
    columns: list[ColumnRecord],
    column_points: list[tuple[str, float, float]],
    support_segments: list[SupportSegment],
    wall_chains: list[WallChain],
    rectangles: list[ZoneRectangle],
    floor_group: str,
    floors: tuple[str, ...],
    boundary_tolerance_m: float,
) -> list[SecondaryBeamCandidate]:
    candidates: list[SecondaryBeamCandidate] = []
    for rect in enumerate_closed_rectangles(support_segments, column_points):
        x1, x2, y1, y2 = rect.x1, rect.x2, rect.y1, rect.y2
        dx = round(x2 - x1, 3)
        dy = round(y2 - y1, 3)
        long_span = max(dx, dy)
        short_span = min(dx, dy)
        if short_span < SECONDARY_R3_MIN_SHORT_SIDE_M:
            continue
        if not (long_span >= SECONDARY_R3_SPAN_TRIGGER_M or long_span / short_span >= SECONDARY_R3_ASPECT_RATIO_TRIGGER):
            continue

        if dx >= dy:
            midpoint = round((x1 + x2) / 2.0, 3)
            aligned = [
                chain.fixed
                for chain in wall_chains
                if chain.axis == "Y"
                and abs(chain.fixed - midpoint) <= SECONDARY_R3_WALL_ALIGN_TOL_M
                and overlap_1d(chain.start, chain.end, y1, y2) is not None
            ]
            fixed = aligned[0] if aligned else midpoint
            candidate = SecondaryBeamCandidate(
                axis="Y",
                x1=round(fixed, 3),
                y1=round(y1, 3),
                x2=round(fixed, 3),
                y2=round(y2, 3),
                beam_location="Centre",
                floor_group=floor_group,
                floors=floors,
                rule_code="R3",
                beam_class=build_secondary_beam_class("Y", fixed, columns, boundary_tolerance_m),
                detail=f"panel({x1},{x2},{y1},{y2}) split_y",
                score=25.0,
            )
        else:
            midpoint = round((y1 + y2) / 2.0, 3)
            aligned = [
                chain.fixed
                for chain in wall_chains
                if chain.axis == "X"
                and abs(chain.fixed - midpoint) <= SECONDARY_R3_WALL_ALIGN_TOL_M
                and overlap_1d(chain.start, chain.end, x1, x2) is not None
            ]
            fixed = aligned[0] if aligned else midpoint
            candidate = SecondaryBeamCandidate(
                axis="X",
                x1=round(x1, 3),
                y1=round(fixed, 3),
                x2=round(x2, 3),
                y2=round(fixed, 3),
                beam_location="Centre",
                floor_group=floor_group,
                floors=floors,
                rule_code="R3",
                beam_class=build_secondary_beam_class("X", fixed, columns, boundary_tolerance_m),
                detail=f"panel({x1},{x2},{y1},{y2}) split_x",
                score=25.0,
            )
        if not candidate_crosses_zone(candidate, rectangles):
            candidates.append(candidate)
    return candidates


def accept_secondary_candidates(
    candidates: list[SecondaryBeamCandidate],
    structural_segments: list[SupportSegment],
    rectangles: list[ZoneRectangle],
) -> list[SecondaryBeamCandidate]:
    accepted: list[SecondaryBeamCandidate] = []
    supports = list(structural_segments)
    ordered = sorted(candidates, key=lambda cand: (SECONDARY_RULE_PRIORITY[cand.rule_code], -cand.score, cand.detail))
    for candidate in ordered:
        fixed, start, end = axis_major_values(candidate.axis, candidate.x1, candidate.y1, candidate.x2, candidate.y2)
        if candidate.rule_code != "R1" and end - start < SECONDARY_MIN_SPAN_M:
            continue
        if candidate_is_duplicate(candidate, supports):
            continue
        if candidate_parallel_conflict(candidate, supports):
            continue
        if candidate_crosses_zone(candidate, rectangles):
            continue
        accepted.append(candidate)
        supports = merge_support_segments(supports + [secondary_to_support_segment(candidate)])
    return accepted


def assign_secondary_numbers(candidates: list[SecondaryBeamCandidate], prefix: str) -> list[SecondaryBeam]:
    result: list[SecondaryBeam] = []
    for idx, candidate in enumerate(candidates, start=1):
        result.append(
            SecondaryBeam(
                no=idx,
                type_name=f"{prefix}{idx}",
                axis=candidate.axis,
                x1=candidate.x1,
                y1=candidate.y1,
                x2=candidate.x2,
                y2=candidate.y2,
                beam_location=candidate.beam_location,
                floor_group=candidate.floor_group,
                floors=candidate.floors,
                rule_code=candidate.rule_code,
                beam_class=candidate.beam_class,
                detail=candidate.detail,
            )
        )
    return result


def generate_secondary_group(
    columns: list[ColumnRecord],
    primary_supports: list[SupportSegment],
    node_map: dict[str, tuple[float, float]],
    walls: list[WallSegment],
    rectangles: list[ZoneRectangle],
    floor_group: str,
    floors: tuple[str, ...],
    boundary_tolerance_m: float,
    enabled_rules: set[str] | None = None,
    preaccepted: list[SecondaryBeamCandidate] | None = None,
) -> tuple[list[SecondaryBeam], list[dict[str, object]]]:
    column_points = [(column.type_name, round(node_map[column.type_name][0], 3), round(node_map[column.type_name][1], 3)) for column in columns]
    _bnd_min_x = min(cx for _, cx, _ in column_points)
    _bnd_max_x = max(cx for _, cx, _ in column_points)
    _bnd_min_y = min(cy for _, _, cy in column_points)
    _bnd_max_y = max(cy for _, _, cy in column_points)
    active_rules = enabled_rules or {"R1", "R2", "R3", "R4", "R5"}
    accepted: list[SecondaryBeamCandidate] = list(preaccepted or [])
    audit_rows: list[dict[str, object]] = []
    wall_chains = merge_wall_chains(walls)

    for pass_idx in range(1, SECONDARY_MAX_GENERATION_PASSES + 1):
        base_supports = build_support_segments_from_segments(primary_supports, accepted, rectangles)
        structural_supports = build_support_segments_from_segments(primary_supports, accepted, [])
        staged: list[SecondaryBeamCandidate] = []
        if pass_idx == 1:
            if "R2" in active_rules:
                staged.extend(generate_r2_edge_perpendicular(columns, column_points, base_supports, floor_group, floors, boundary_tolerance_m))
            if "R1" in active_rules:
                r1_supports = build_support_segments_from_segments(primary_supports, accepted + staged, rectangles)
                staged.extend(generate_r1_shape_closure(columns, column_points, r1_supports, rectangles, floor_group, floors, boundary_tolerance_m))
            if "R5" in active_rules:
                r5_structural_supports = build_support_segments_from_segments(primary_supports, accepted + staged, [])
                staged.extend(generate_r5_zone_perimeter(columns, rectangles, r5_structural_supports, floor_group, floors, boundary_tolerance_m))
            if "R4" in active_rules:
                refreshed_structural_supports = build_support_segments_from_segments(primary_supports, accepted + staged, [])
                staged.extend(generate_r4_wall_support(columns, column_points, refreshed_structural_supports, wall_chains, rectangles, floor_group, floors, boundary_tolerance_m))
        else:
            if "R1" in active_rules:
                staged.extend(generate_r1_shape_closure(columns, column_points, base_supports, rectangles, floor_group, floors, boundary_tolerance_m))
        refreshed_supports = build_support_segments_from_segments(primary_supports, accepted + staged, rectangles)
        if "R3" in active_rules and floor_group != "plinth":
            r3_floors = tuple(floor for floor in floors if floor in {"Typical floor roof", "Terrace"})
            if r3_floors:
                staged.extend(generate_r3_large_panel_dividers(columns, column_points, refreshed_supports, wall_chains, rectangles, floor_group, r3_floors, boundary_tolerance_m))
        _tol = SECONDARY_SUPPORT_TOL_M
        staged = [
            c for c in staged
            if c.x1 >= _bnd_min_x - _tol and c.x2 >= _bnd_min_x - _tol
            and c.x1 <= _bnd_max_x + _tol and c.x2 <= _bnd_max_x + _tol
            and c.y1 >= _bnd_min_y - _tol and c.y2 >= _bnd_min_y - _tol
            and c.y1 <= _bnd_max_y + _tol and c.y2 <= _bnd_max_y + _tol
        ]
        newly_accepted = accept_secondary_candidates(staged, structural_supports, rectangles)
        existing_keys = {candidate_duplicate_key(beam) for beam in accepted}
        added = 0
        for candidate in newly_accepted:
            key = candidate_duplicate_key(candidate)
            if key in existing_keys:
                continue
            accepted.append(candidate)
            existing_keys.add(key)
            added += 1
            audit_rows.append(
                {
                    "Floor group": floor_group,
                    "Rule": candidate.rule_code,
                    "Axis": candidate.axis,
                    "X1": candidate.x1,
                    "Y1": candidate.y1,
                    "X2": candidate.x2,
                    "Y2": candidate.y2,
                    "Beam class": candidate.beam_class,
                    "Detail": candidate.detail,
                    "Pass": pass_idx,
                }
            )
        if added == 0:
            break
    prefix = "SP" if floor_group == "plinth" else "SO"
    return assign_secondary_numbers(accepted, prefix), audit_rows


def is_lift_to_lift_beam(beam: BeamPair) -> bool:
    return beam.start.location == "Lift" and beam.end.location == "Lift"


def compute_beam_defaults(beam_class: str, wall_thickness_mm: int) -> tuple[int, int]:
    if beam_class == "Edge":
        return 230, 450
    if wall_thickness_mm > 0:
        return 230, 375
    return 230, 375


def choose_node_width(primary: list[float] | None, fallback: list[float] | None) -> float:
    if primary:
        return max(primary)
    if fallback:
        return max(fallback)
    return 0.0


def calc_column_node(column: ColumnRecord, x_axis_width_mm: float, y_axis_width_mm: float) -> tuple[float, float]:
    if column.left_right == "Left":
        node_x = column.anchor_x + (x_axis_width_mm / 1000.0) / 2.0
    elif column.left_right == "Right":
        node_x = column.anchor_x - (x_axis_width_mm / 1000.0) / 2.0
    else:
        node_x = (column.xmin + column.xmax) / 2.0

    if column.front_back == "Front":
        node_y = column.anchor_y + (y_axis_width_mm / 1000.0) / 2.0
    elif column.front_back == "Back":
        node_y = column.anchor_y - (y_axis_width_mm / 1000.0) / 2.0
    else:
        node_y = (column.ymin + column.ymax) / 2.0
    return round(node_x, 3), round(node_y, 3)


def level_axis_values(values: list[float], tolerance_m: float) -> list[float]:
    result = list(values)
    used = [False] * len(values)
    for i, value in enumerate(values):
        if used[i]:
            continue
        group = [i]
        used[i] = True
        for j in range(i + 1, len(values)):
            if used[j]:
                continue
            if abs(values[j] - value) <= tolerance_m:
                group.append(j)
                used[j] = True
        if len(group) < 2:
            continue
        avg = round(sum(values[idx] for idx in group) / len(group), 3)
        for idx in group:
            result[idx] = avg
    return result


def build_provisional_node_map(
    columns: list[ColumnRecord],
    beams: list[BeamPair],
    walls_by_floor: dict[str, list[WallSegment]],
    wall_alignment_tolerance_m: float,
    edge_wall_coverage_threshold_pct: float,
    interior_wall_coverage_threshold_pct: float,
    node_level_tolerance_m: float,
) -> dict[str, tuple[float, float]]:
    x_widths: dict[str, list[float]] = {}
    y_widths: dict[str, list[float]] = {}
    typical_walls = walls_by_floor.get("typical", [])
    for beam in beams:
        wall_thickness_mm, _ = compute_wall_assignment(
            beam,
            typical_walls,
            wall_alignment_tolerance_m=wall_alignment_tolerance_m,
            edge_wall_coverage_threshold_pct=edge_wall_coverage_threshold_pct,
            interior_wall_coverage_threshold_pct=interior_wall_coverage_threshold_pct,
        )
        width_mm, _ = compute_beam_defaults(beam.beam_class, wall_thickness_mm)
        for column_name in (beam.start.type_name, beam.end.type_name):
            if beam.direction == "X":
                x_widths.setdefault(column_name, []).append(width_mm)
            else:
                y_widths.setdefault(column_name, []).append(width_mm)

    raw_nodes: dict[str, tuple[float, float]] = {}
    x_values: list[float] = []
    y_values: list[float] = []
    for column in columns:
        x_axis_width_mm = choose_node_width(primary=y_widths.get(column.type_name), fallback=x_widths.get(column.type_name))
        y_axis_width_mm = choose_node_width(primary=x_widths.get(column.type_name), fallback=y_widths.get(column.type_name))
        node_x, node_y = calc_column_node(column, x_axis_width_mm, y_axis_width_mm)
        raw_nodes[column.type_name] = (node_x, node_y)
        x_values.append(node_x)
        y_values.append(node_y)

    leveled_x = level_axis_values(x_values, node_level_tolerance_m)
    leveled_y = level_axis_values(y_values, node_level_tolerance_m)
    return {
        column.type_name: (round(node_x, 3), round(node_y, 3))
        for column, node_x, node_y in zip(columns, leveled_x, leveled_y)
    }


def build_node_primary_segments(primary_beams: list[BeamPair], node_map: dict[str, tuple[float, float]]) -> list[SupportSegment]:
    segments: list[SupportSegment] = []
    for beam in primary_beams:
        start_node = node_map[beam.start.type_name]
        end_node = node_map[beam.end.type_name]
        fixed, start, end = axis_major_values(beam.direction, start_node[0], start_node[1], end_node[0], end_node[1])
        segments.append(make_support_segment(beam.direction, fixed, start, end, "primary", beam.beam_no))
    return merge_support_segments(segments)


def autosize(ws) -> None:
    for col in ws.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in col]
        width = min(max(len(value) for value in values) + 2, 40)
        ws.column_dimensions[col[0].column_letter].width = width


def format_primary_beams_sheet(ws) -> None:
    header_map = {cell.value: cell.column_letter for cell in ws[1]}
    compact_headers = {
        "Beam No.": 7,
        "Direction": 7,
        "Beam class": 9,
        "Beam width (mm)": 10,
        "Beam depth (mm)": 10,
        "Wall thickness (mm)": 10,
        "Raw wall coverage (%)": 10,
        "Start Anchor X (m)": 9,
        "Start Anchor Y (m)": 9,
        "End Anchor X (m)": 9,
        "End Anchor Y (m)": 9,
    }
    for header, width in compact_headers.items():
        if header not in header_map:
            continue
        ws.column_dimensions[header_map[header]].width = width
        for cell in ws[header_map[header]]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def format_secondary_sheet(ws) -> None:
    compact_headers = {"No.": 7, "Type": 8, "Beam location": 11, "Floor": 18, "Present": 10, "Beam width (mm)": 10, "Beam depth (mm)": 10, "Wall thickness (mm)": 10}
    header_map = {cell.value: cell.column_letter for cell in ws[1]}
    for header, width in compact_headers.items():
        if header not in header_map:
            continue
        ws.column_dimensions[header_map[header]].width = width
        for cell in ws[header_map[header]]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def secondary_row_values(
    beam: SecondaryBeam,
    floor: str,
    walls_by_floor: dict[str, list[WallSegment]],
    wall_alignment_tolerance_m: float,
    edge_wall_coverage_threshold_pct: float,
    interior_wall_coverage_threshold_pct: float,
) -> list[object]:
    fixed, start, end = axis_major_values(beam.axis, beam.x1, beam.y1, beam.x2, beam.y2)
    floor_source = FLOOR_TO_WALL_SOURCE[floor]
    floor_walls = walls_by_floor.get(floor_source, [])
    typical_walls = walls_by_floor.get("typical", floor_walls)

    if beam.rule_code == "R4" and beam.floor_group == "plinth":
        typical_width_mm, typical_depth_mm = SECONDARY_PLINTH_WIDTH_MM, SECONDARY_PLINTH_DEPTH_MM
    else:
        typical_wall_mm, _ = compute_segment_wall_assignment(
            beam.axis,
            fixed,
            start,
            end,
            beam.beam_class,
            typical_walls,
            wall_alignment_tolerance_m,
            edge_wall_coverage_threshold_pct,
            interior_wall_coverage_threshold_pct,
        )
        typical_width_mm, typical_depth_mm = compute_beam_defaults(beam.beam_class, typical_wall_mm)

    wall_thickness_mm, _ = compute_segment_wall_assignment(
        beam.axis,
        fixed,
        start,
        end,
        beam.beam_class,
        floor_walls,
        wall_alignment_tolerance_m,
        edge_wall_coverage_threshold_pct,
        interior_wall_coverage_threshold_pct,
    )
    if beam.rule_code == "R4" and beam.floor_group == "plinth":
        beam_width_mm, beam_depth_mm = SECONDARY_PLINTH_WIDTH_MM, SECONDARY_PLINTH_DEPTH_MM
    else:
        beam_width_mm, beam_depth_mm = typical_width_mm, typical_depth_mm
    if floor == "Plinth":
        span_m = end - start
        if beam.beam_class == "Edge":
            beam_depth_mm = 375
        elif span_m <= 4.0:
            beam_depth_mm = 375
        elif span_m < 7.0:
            beam_depth_mm = 450
        else:
            beam_depth_mm = 525

    return [
        beam.no,
        beam.type_name,
        beam.x1,
        beam.y1,
        beam.x2,
        beam.y2,
        beam.beam_location,
        floor,
        "YES",
        beam_width_mm,
        beam_depth_mm,
        wall_thickness_mm,
    ]


_TERRACE_PARAPET_EXCLUDE_LOCATIONS: frozenset[str] = frozenset({"lift", "staircase", "shaft"})


def shaft_is_adjacent_to_lift_or_staircase(
    shaft_rect: ZoneRectangle,
    all_rects: list[ZoneRectangle],
    adjacency_tol_m: float = 0.5,
) -> bool:
    sx1, sx2 = sorted((shaft_rect.x1, shaft_rect.x2))
    sy1, sy2 = sorted((shaft_rect.y1, shaft_rect.y2))
    for rect in all_rects:
        if str(rect.location).strip().lower() not in {"lift", "staircase"}:
            continue
        rx1, rx2 = sorted((rect.x1, rect.x2))
        ry1, ry2 = sorted((rect.y1, rect.y2))
        x_gap = max(0.0, max(sx1, rx1) - min(sx2, rx2))
        y_gap = max(0.0, max(sy1, ry1) - min(sy2, ry2))
        if x_gap <= adjacency_tol_m and y_gap <= adjacency_tol_m:
            return True
    return False


def beam_overlaps_excluded_zone(
    beam: BeamPair,
    rectangles: list[ZoneRectangle],
    tol_m: float = 0.15,
) -> bool:
    for rect in rectangles:
        loc = str(rect.location).strip().lower()
        if loc not in _TERRACE_PARAPET_EXCLUDE_LOCATIONS:
            continue
        if loc == "shaft" and not shaft_is_adjacent_to_lift_or_staircase(rect, rectangles):
            continue
        rx1, rx2 = sorted((rect.x1, rect.x2))
        ry1, ry2 = sorted((rect.y1, rect.y2))
        if beam.direction == "X":
            if not (ry1 - tol_m <= beam.group_coordinate_m <= ry2 + tol_m):
                continue
            if overlap_1d(beam.beam_start_x, beam.beam_end_x, rx1, rx2) is None:
                continue
        else:
            if not (rx1 - tol_m <= beam.group_coordinate_m <= rx2 + tol_m):
                continue
            if overlap_1d(beam.beam_start_y, beam.beam_end_y, ry1, ry2) is None:
                continue
        return True
    return False


def write_output(
    path: Path,
    columns: list[ColumnRecord],
    beams: list[BeamPair],
    walls: list[WallSegment],
    plinth_secondaries: list[SecondaryBeam],
    nonplinth_secondaries: list[SecondaryBeam],
    secondary_audit_rows: list[dict[str, object]],
    rectangle_sheet_rows: dict[str, list[tuple[object, ...]]],
    wall_alignment_tolerance_m: float,
    edge_wall_coverage_threshold_pct: float,
    interior_wall_coverage_threshold_pct: float,
    rectangles: list[ZoneRectangle] | None = None,
) -> None:
    walls_by_floor: dict[str, list[WallSegment]] = {}
    for wall in walls:
        if wall.dxf_source:
            walls_by_floor.setdefault(wall.dxf_source, []).append(wall)

    wb = Workbook()
    ws_cols = wb.active
    ws_cols.title = "Columns"
    ws_cols.append(
        [
            "Column No.",
            "Type",
            "Left/Right",
            "Front/Back",
            "Location",
            "Anchor location",
        ]
    )

    for column in columns:
        ws_cols.append(
            [
                column.idx,
                column.type_name,
                column.left_right,
                column.front_back,
                column.location,
                column.anchor_location,
            ]
        )

    ws_beams = wb.create_sheet("Primary Beams")
    ws_beams.append(
        [
            "Beam No.",
            "StartC",
            "Start Location",
            "Start Anchor X (m)",
            "Start Anchor Y (m)",
            "EndC",
            "End Location",
            "End Anchor X (m)",
            "End Anchor Y (m)",
            "Direction",
            "Beam class",
            "Floor",
            "Beam width (mm)",
            "Beam depth (mm)",
            "Wall thickness (mm)",
            "Raw wall coverage (%)",
            "Beam Start X (m)",
            "Beam Start Y (m)",
            "Beam End X (m)",
            "Beam End Y (m)",
            "Span Length (m)",
        ]
    )

    _x_fixed = [b.group_coordinate_m for b in beams if b.direction == "Y"]
    _y_fixed = [b.group_coordinate_m for b in beams if b.direction == "X"]
    _geo_min_x = min(_x_fixed) if _x_fixed else 0.0
    _geo_max_x = max(_x_fixed) if _x_fixed else 0.0
    _geo_min_y = min(_y_fixed) if _y_fixed else 0.0
    _geo_max_y = max(_y_fixed) if _y_fixed else 0.0

    for beam in beams:
        typical_walls = walls_by_floor.get("typical", walls if not walls_by_floor else [])
        typical_wall_thickness_mm, _ = compute_wall_assignment(
            beam,
            typical_walls,
            wall_alignment_tolerance_m=wall_alignment_tolerance_m,
            edge_wall_coverage_threshold_pct=edge_wall_coverage_threshold_pct,
            interior_wall_coverage_threshold_pct=interior_wall_coverage_threshold_pct,
        )
        if is_lift_to_lift_beam(beam):
            typical_beam_width_mm, typical_beam_depth_mm = 230, 450
        else:
            typical_beam_width_mm, typical_beam_depth_mm = compute_beam_defaults(beam.beam_class, typical_wall_thickness_mm)

        for floor in FLOOR_SEQUENCE:
            floor_source = FLOOR_TO_WALL_SOURCE[floor]
            floor_walls = walls_by_floor.get(floor_source, walls if not walls_by_floor else [])
            wall_thickness_mm, raw_wall_coverage_pct = compute_wall_assignment(
                beam,
                floor_walls,
                wall_alignment_tolerance_m=wall_alignment_tolerance_m,
                edge_wall_coverage_threshold_pct=edge_wall_coverage_threshold_pct,
                interior_wall_coverage_threshold_pct=interior_wall_coverage_threshold_pct,
            )
            if is_lift_to_lift_beam(beam):
                beam_width_mm, beam_depth_mm = 230, 450
                wall_thickness_value = 230
            else:
                beam_width_mm, beam_depth_mm = typical_beam_width_mm, typical_beam_depth_mm
                wall_thickness_value = wall_thickness_mm
            if floor == "Plinth":
                if beam.beam_class == "Edge":
                    beam_depth_mm = 375
                elif beam.span_length_m <= 4.0:
                    beam_depth_mm = 375
                elif beam.span_length_m < 7.0:
                    beam_depth_mm = 450
                else:
                    beam_depth_mm = 525
            else:
                _tol = DEFAULT_BOUNDARY_TOLERANCE_M
                _on_boundary = (
                    beam.direction == "X" and (
                        abs(beam.group_coordinate_m - _geo_min_y) <= _tol
                        or abs(beam.group_coordinate_m - _geo_max_y) <= _tol
                    )
                ) or (
                    beam.direction == "Y" and (
                        abs(beam.group_coordinate_m - _geo_min_x) <= _tol
                        or abs(beam.group_coordinate_m - _geo_max_x) <= _tol
                    )
                )
                if _on_boundary:
                    beam_depth_mm = 450
                    if floor == "Terrace" and wall_thickness_value == 115:
                        if not beam_overlaps_excluded_zone(beam, rectangles or []):
                            wall_thickness_value = 40
            wall_coverage_value = raw_wall_coverage_pct
            if floor == "Plinth" and wall_thickness_value == 0:
                wall_thickness_value = 90

            ws_beams.append(
                [
                    beam.beam_no,
                    beam.start.type_name,
                    beam.start.location,
                    beam.start.anchor_x,
                    beam.start.anchor_y,
                    beam.end.type_name,
                    beam.end.location,
                    beam.end.anchor_x,
                    beam.end.anchor_y,
                    beam.direction,
                    beam.beam_class,
                    floor,
                    beam_width_mm,
                    beam_depth_mm,
                    wall_thickness_value,
                    wall_coverage_value,
                    beam.beam_start_x,
                    beam.beam_start_y,
                    beam.beam_end_x,
                    beam.beam_end_y,
                    beam.span_length_m,
                ]
            )

    autosize(ws_cols)
    autosize(ws_beams)
    format_primary_beams_sheet(ws_beams)

    if plinth_secondaries:
        ws_plinth = wb.create_sheet(RAW_PLINTH_SHEET)
        ws_plinth.append(
            [
                "No.",
                "Type",
                "Coordinate X1 (m)",
                "Coordinate Y1 (m)",
                "Coordinate X2 (m)",
                "Coordinate Y2 (m)",
                "Beam location",
                "Floor",
                "Present",
                "Beam width (mm)",
                "Beam depth (mm)",
                "Wall thickness (mm)",
            ]
        )
        for beam in plinth_secondaries:
            ws_plinth.append(
                secondary_row_values(
                    beam,
                    "Plinth",
                    walls_by_floor,
                    wall_alignment_tolerance_m,
                    edge_wall_coverage_threshold_pct,
                    interior_wall_coverage_threshold_pct,
                )
            )
        autosize(ws_plinth)
        format_secondary_sheet(ws_plinth)

    if nonplinth_secondaries:
        ws_nonplinth = wb.create_sheet(RAW_NONPLINTH_SHEET)
        ws_nonplinth.append(
            [
                "No.",
                "Type",
                "Coordinate X1 (m)",
                "Coordinate Y1 (m)",
                "Coordinate X2 (m)",
                "Coordinate Y2 (m)",
                "Beam location",
                "Floor",
                "Present",
                "Beam width (mm)",
                "Beam depth (mm)",
                "Wall thickness (mm)",
            ]
        )
        for beam in nonplinth_secondaries:
            for floor in NONPLINTH_FLOORS:
                ws_nonplinth.append(
                    secondary_row_values(
                        beam,
                        floor,
                        walls_by_floor,
                        wall_alignment_tolerance_m,
                        edge_wall_coverage_threshold_pct,
                        interior_wall_coverage_threshold_pct,
                    )
                )
        autosize(ws_nonplinth)
        format_secondary_sheet(ws_nonplinth)

    if secondary_audit_rows:
        ws_audit = wb.create_sheet("Secondary beam audit")
        headers = ["Floor group", "Rule", "Axis", "X1", "Y1", "X2", "Y2", "Beam class", "Detail", "Pass"]
        ws_audit.append(headers)
        for row in secondary_audit_rows:
            ws_audit.append([row.get(header, "") for header in headers])
        autosize(ws_audit)

    for sheet_name, rows in rectangle_sheet_rows.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(list(row))
        autosize(ws)

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a two-sheet workbook with normalized columns and pair-based primary beam data from the column and wall pipeline outputs."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT_PATH, help="Optional input column workbook path. If omitted, the script searches the current project folders.")
    parser.add_argument("--walls", type=Path, default=DEFAULT_WALL_INPUT_PATH, help="Optional wall workbook path. If omitted, the script infers the companion walls workbook beside the column workbook.")
    parser.add_argument("--rectangles-workbook", type=Path, default=None, help="Optional workbook containing the legacy Rectangle coordinates sheet used for lift/stair/shaft support edges.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH, help="Optional output workbook path. Defaults to <dxf_stem>_column_beam_pairs.xlsx beside the input workbook.")
    parser.add_argument("--group-tolerance-m", type=float, default=DEFAULT_GROUP_TOLERANCE_M, help="Tolerance in meters used to group columns onto the same X/Y beam family.")
    parser.add_argument("--boundary-tolerance-m", type=float, default=DEFAULT_BOUNDARY_TOLERANCE_M, help="Tolerance in meters used to derive edge/corner location tags from the chosen anchor coordinates.")
    parser.add_argument("--wall-alignment-tolerance-m", type=float, default=DEFAULT_WALL_ALIGNMENT_TOLERANCE_M, help="Maximum offset in meters between a beam line and a parallel wall centerline when assigning wall thickness.")
    parser.add_argument("--edge-wall-coverage-threshold-pct", type=float, default=DEFAULT_EDGE_WALL_COVERAGE_THRESHOLD_PCT, help="Minimum raw wall coverage percentage required to assign 115/230 wall thickness to edge beam pairs.")
    parser.add_argument("--interior-wall-coverage-threshold-pct", type=float, default=DEFAULT_INTERIOR_WALL_COVERAGE_THRESHOLD_PCT, help="Minimum raw wall coverage percentage required to assign 115/230 wall thickness to interior beam pairs.")
    parser.add_argument("--thickness-tolerance-mm", type=float, default=DEFAULT_THICKNESS_TOLERANCE_MM, help="Tolerance in millimeters used to snap detected wall thickness to 115 or 230.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = resolve_input_workbook(args.input)
    wall_path = resolve_wall_workbook(args.walls, input_path)
    rectangle_workbook = args.rectangles_workbook.resolve() if args.rectangles_workbook else resolve_rectangle_workbook(input_path)
    output_path = args.output.resolve() if args.output else input_path.with_name(
        f"{infer_dxf_stem_from_column_workbook(input_path)}_column_beam_pairs.xlsx"
    )

    source_columns = read_source_columns(input_path)
    derived_columns = derive_columns(source_columns, args.boundary_tolerance_m)
    walls = read_walls(wall_path, args.thickness_tolerance_mm)
    rectangles, rectangle_sheet_rows = read_zone_rectangles(rectangle_workbook)
    beams = generate_primary_beams(derived_columns, args.group_tolerance_m)
    walls_by_floor: dict[str, list[WallSegment]] = {}
    for wall in walls:
        if wall.dxf_source:
            walls_by_floor.setdefault(wall.dxf_source, []).append(wall)
    node_map = build_provisional_node_map(
        derived_columns,
        beams,
        walls_by_floor,
        wall_alignment_tolerance_m=args.wall_alignment_tolerance_m,
        edge_wall_coverage_threshold_pct=args.edge_wall_coverage_threshold_pct,
        interior_wall_coverage_threshold_pct=args.interior_wall_coverage_threshold_pct,
        node_level_tolerance_m=0.60,
    )
    primary_supports = build_node_primary_segments(beams, node_map)
    plinth_secondaries, plinth_audit = generate_secondary_group(
        derived_columns,
        primary_supports,
        node_map,
        [wall for wall in walls if wall.dxf_source == "plinth"],
        [],
        floor_group="plinth",
        floors=PLINTH_FLOORS,
        boundary_tolerance_m=args.boundary_tolerance_m,
        enabled_rules={"R4"},
    )
    nonplinth_secondaries, nonplinth_audit = generate_secondary_group(
        derived_columns,
        primary_supports,
        node_map,
        [wall for wall in walls if wall.dxf_source in {"typical", "terrace"}],
        rectangles,
        floor_group="nonplinth",
        floors=NONPLINTH_FLOORS,
        boundary_tolerance_m=args.boundary_tolerance_m,
        enabled_rules={"R4"},
    )
    write_output(
        output_path,
        columns=derived_columns,
        beams=beams,
        walls=walls,
        plinth_secondaries=plinth_secondaries,
        nonplinth_secondaries=nonplinth_secondaries,
        secondary_audit_rows=plinth_audit + nonplinth_audit,
        rectangle_sheet_rows=rectangle_sheet_rows,
        wall_alignment_tolerance_m=args.wall_alignment_tolerance_m,
        edge_wall_coverage_threshold_pct=args.edge_wall_coverage_threshold_pct,
        interior_wall_coverage_threshold_pct=args.interior_wall_coverage_threshold_pct,
        rectangles=rectangles,
    )

    print(f"Column workbook : {input_path}")
    print(f"Wall workbook   : {wall_path}")
    if rectangle_workbook is not None and rectangle_workbook.exists():
        print(f"Rect workbook   : {rectangle_workbook}")
    print(f"Columns found   : {len(derived_columns)}")
    print(f"Primary beams   : {len(beams)}")
    print(f"Plinth 2nds     : {len(plinth_secondaries)}")
    print(f"Nonplinth 2nds  : {len(nonplinth_secondaries)}")
    print(f"Rows written    : {len(beams) * len(FLOOR_SEQUENCE)}")
    print(f"Excel saved     : {output_path}")


if __name__ == "__main__":
    main()
