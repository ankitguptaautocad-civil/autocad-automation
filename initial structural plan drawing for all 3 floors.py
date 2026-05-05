from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT_PATH = None
DEFAULT_OUTPUT_PATH = None
PAGE_WIDTH_PT = 842.0
PAGE_HEIGHT_PT = 1190.0
PAGE_MARGIN_PT = 36.0
DRAWING_LEFT_OFFSET_PT = 82.0
DRAWING_RIGHT_RESERVED_PT = 148.0
FLOOR_PAGES = (
    ("Plinth", "Plinth"),
    ("Stilt roof", "Stilt Roof"),
    ("Typical floor roof", "Typical Floor"),
    ("Terrace", "Terrace"),
)
FINAL_SECONDARY_SHEET = "Secondary beam coordinates"
SHEAR_WALL_SOURCE_SHEETS = ("Final_Recommendation_Walls", "Shear wall landscape")
EPS = 1e-9


@dataclass(frozen=True)
class ColumnRect:
    type_name: str
    location: str
    node_x: float = 0.0
    node_y: float = 0.0
    draw_w: float = 0.0
    draw_h: float = 0.0


@dataclass(frozen=True)
class BeamStrip:
    beam_no: str
    floor: str
    direction: str
    beam_class: str
    beam_width_m: float
    beam_depth_m: float
    wall_thickness_m: float
    start_x: float
    start_y: float
    end_x: float
    end_y: float
    source_kind: str = "primary"


@dataclass(frozen=True)
class FeatureRect:
    label: str
    floor: str
    x1: float
    y1: float
    x2: float
    y2: float
    kind: str


@dataclass(frozen=True)
class SignedShearWall:
    wall_id: str
    x1: float
    y1: float
    x2: float
    y2: float
    thickness_m: float


def normalize_header(value: object) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def safe_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def safe_float(value: object) -> float | None:
    text = safe_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def synthesize_wall_label(node_x: float, node_y: float, direction: str) -> str:
    signed_direction = normalize_signed_direction(direction)
    direction_token = signed_direction.replace("+", "P").replace("-", "N")
    return f"SW_{round(float(node_x), 3):g}_{round(float(node_y), 3):g}_{direction_token}"


def candidate_search_dirs() -> list[Path]:
    dirs = []
    for candidate in ((Path.cwd() / "STD ANL model").resolve(), Path.cwd().resolve()):
        if candidate.exists() and candidate not in dirs:
            dirs.append(candidate)
    return dirs


def discover_single_workbook(patterns: tuple[str, ...], label: str) -> Path:
    matches: list[Path] = []
    for folder in candidate_search_dirs():
        for pattern in patterns:
            matches.extend(path.resolve() for path in folder.glob(pattern) if not path.name.startswith("~$"))
    unique_matches = list(set(matches))
    if not unique_matches:
        raise SystemExit(f"Could not find the {label} workbook. Pass it explicitly.")
    if len(unique_matches) > 1:
        chosen = max(unique_matches, key=lambda p: p.stat().st_mtime)
        others = ", ".join(path.name for path in unique_matches if path != chosen)
        print(f"[INFO] Multiple {label} workbooks found; using latest: {chosen.name} (skipped: {others})")
        return chosen
    return unique_matches[0]


def resolve_input_workbook(explicit_input: Path | None) -> Path:
    if explicit_input is not None:
        path = explicit_input.resolve()
        if not path.exists():
            raise SystemExit(f"Input workbook not found: {path}")
        return path
    return discover_single_workbook(
        (
            "node_coordinates_*.xlsx",
            "*secondary_nodes*.xlsx",
            "*with_nodes_v3.xlsx",
            "*with_nodes_v2.xlsx",
            "*with_nodes.xlsx",
        ),
        "node-appended beam pair",
    )


def resolve_other_geometry_workbook(input_path: Path) -> Path | None:
    candidates: list[Path] = []
    if "node_coordinates" in input_path.name:
        candidates.append(input_path.parent / input_path.name.replace("node_coordinates", "other_coordinates"))
    candidates.append(input_path.with_name(f"{input_path.stem}_other_coordinates.xlsx"))
    seen: set[Path] = set()
    for companion in candidates:
        resolved = companion.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists():
            return resolved
    return None


def load_columns(ws) -> list[ColumnRect]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(v): idx for idx, v in enumerate(headers)}
    required = {
        "type": "Type",
        "location": "Location",
        "nodecoordinatexm": "Node Coordinate X (m)",
        "nodecoordinateym": "Node Coordinate Y (m)",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Columns sheet missing required headers: {', '.join(missing)}")

    yd_key = next((k for k in header_map if k.startswith("yd") and k != "yd"), None)
    zd_key = next((k for k in header_map if k.startswith("zd") and k != "zd"), None)

    records: list[ColumnRect] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        column_no = ""
        if "columnno" in header_map and row[header_map["columnno"]] not in (None, ""):
            raw_column_no = str(row[header_map["columnno"]]).strip()
            column_no = raw_column_no if raw_column_no.upper().startswith("C") else f"C{raw_column_no}"
        node_x = float(row[header_map["nodecoordinatexm"]])
        node_y = float(row[header_map["nodecoordinateym"]])
        yd_raw = row[header_map[yd_key]] if yd_key is not None else None
        zd_raw = row[header_map[zd_key]] if zd_key is not None else None
        draw_w = float(yd_raw) / 1000.0 if yd_raw not in (None, "") else 0.3
        draw_h = float(zd_raw) / 1000.0 if zd_raw not in (None, "") else 0.3
        records.append(
            ColumnRect(
                type_name=column_no or str(row[header_map["type"]]).strip(),
                location=str(row[header_map["location"]] or "").strip(),
                node_x=node_x,
                node_y=node_y,
                draw_w=draw_w,
                draw_h=draw_h,
            )
        )
    return records


def load_beams(ws) -> list[BeamStrip]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(v): idx for idx, v in enumerate(headers)}
    required = {
        "beamno": "Beam No.",
        "floor": "Floor",
        "direction": "Direction",
        "beamclass": "Beam class",
        "beamwidthmm": "Beam width (mm)",
        "beamdepthmm": "Beam depth (mm)",
        "wallthicknessmm": "Wall thickness (mm)",
        "startnodexm": "Start Node X (m)",
        "startnodeym": "Start Node Y (m)",
        "endnodexm": "End Node X (m)",
        "endnodeym": "End Node Y (m)",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Primary Beams sheet missing required headers: {', '.join(missing)}")

    beams: list[BeamStrip] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        beams.append(
            BeamStrip(
                beam_no=str(row[header_map["beamno"]]).strip(),
                floor=str(row[header_map["floor"]]).strip(),
                direction=str(row[header_map["direction"]]).strip().upper(),
                beam_class=str(row[header_map["beamclass"]]).strip(),
                beam_width_m=float(row[header_map["beamwidthmm"]]) / 1000.0,
                beam_depth_m=float(row[header_map["beamdepthmm"]]) / 1000.0,
                wall_thickness_m=float(row[header_map["wallthicknessmm"]]) / 1000.0,
                start_x=float(row[header_map["startnodexm"]]),
                start_y=float(row[header_map["startnodeym"]]),
                end_x=float(row[header_map["endnodexm"]]),
                end_y=float(row[header_map["endnodeym"]]),
                source_kind="primary",
            )
        )
    return beams


def load_secondary_beams(ws) -> list[BeamStrip]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(v): idx for idx, v in enumerate(headers)}
    required = {
        "type": "Type",
        "floor": "Floor",
        "beamwidthmm": "Beam width (mm)",
        "beamdepthmm": "Beam depth (mm)",
        "wallthicknessmm": "Wall thickness (mm)",
        "snappedx1m": "Snapped X1 (m)",
        "snappedy1m": "Snapped Y1 (m)",
        "snappedx2m": "Snapped X2 (m)",
        "snappedy2m": "Snapped Y2 (m)",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Secondary beam sheet missing required headers: {', '.join(missing)}")

    beams: list[BeamStrip] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        start_x = float(row[header_map["snappedx1m"]])
        start_y = float(row[header_map["snappedy1m"]])
        end_x = float(row[header_map["snappedx2m"]])
        end_y = float(row[header_map["snappedy2m"]])
        direction = "X" if abs(start_y - end_y) <= 1e-6 else "Y"
        beams.append(
            BeamStrip(
                beam_no=str(row[header_map["type"]]).strip(),
                floor=str(row[header_map["floor"]]).strip(),
                direction=direction,
                beam_class="Secondary",
                beam_width_m=float(row[header_map["beamwidthmm"]]) / 1000.0,
                beam_depth_m=float(row[header_map["beamdepthmm"]]) / 1000.0,
                wall_thickness_m=float(row[header_map["wallthicknessmm"]]) / 1000.0,
                start_x=start_x,
                start_y=start_y,
                end_x=end_x,
                end_y=end_y,
                source_kind="secondary",
            )
        )
    return beams


def load_feature_rects(ws, floor: str, kind: str) -> list[FeatureRect]:
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(v): idx for idx, v in enumerate(headers)}
    has_snapped = all(k in header_map for k in ("snappedx1m", "snappedy1m", "snappedx2m", "snappedy2m"))
    has_raw = all(k in header_map for k in ("coordinatex1m", "coordinatey1m", "coordinatex2m", "coordinatey2m"))
    if not has_snapped and not has_raw:
        raise SystemExit(f"{ws.title} missing coordinate headers (need Snapped X1/Y1/X2/Y2 or Coordinate X1/Y1/X2/Y2)")

    label_key = None
    for candidate in ("location", "detailname", "type", "no"):
        if candidate in header_map:
            label_key = candidate
            break

    records: list[FeatureRect] = []
    x1_key = "snappedx1m" if has_snapped else "coordinatex1m"
    y1_key = "snappedy1m" if has_snapped else "coordinatey1m"
    x2_key = "snappedx2m" if has_snapped else "coordinatex2m"
    y2_key = "snappedy2m" if has_snapped else "coordinatey2m"
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        if any(row[header_map[key]] in (None, "") for key in (x1_key, y1_key, x2_key, y2_key)):
            continue
        label = str(row[header_map[label_key]]).strip() if label_key is not None and row[header_map[label_key]] not in (None, "") else kind.title()
        records.append(
            FeatureRect(
                label=label,
                floor=floor,
                x1=float(row[header_map[x1_key]]),
                y1=float(row[header_map[y1_key]]),
                x2=float(row[header_map[x2_key]]),
                y2=float(row[header_map[y2_key]]),
                kind=kind,
            )
        )
    return records


def normalize_signed_direction(value: object) -> str:
    text = safe_text(value).upper()
    mapping = {
        "X": "+X",
        "+X": "+X",
        "-X": "-X",
        "Y": "+Z",
        "+Y": "+Z",
        "-Y": "-Z",
        "Z": "+Z",
        "+Z": "+Z",
        "-Z": "-Z",
    }
    if text in mapping:
        return mapping[text]
    raise ValueError(f"Unsupported wall direction: {value!r}")


def signed_direction_sign(direction: str) -> float:
    return 1.0 if normalize_signed_direction(direction).startswith("+") else -1.0


def geometry_from_yd_zd(yd_m: float, zd_m: float) -> tuple[str, float, float]:
    abs_yd = abs(float(yd_m))
    abs_zd = abs(float(zd_m))
    if abs_yd <= EPS and abs_zd <= EPS:
        raise ValueError("Wall geometry cannot have both YD_m and ZD_m equal to zero.")
    if abs_yd >= abs_zd:
        return ("+X" if float(yd_m) >= 0.0 else "-X"), abs_zd, abs_yd
    return ("+Z" if float(zd_m) >= 0.0 else "-Z"), abs_yd, abs_zd


def build_signed_shear_wall(
    wall_id: str,
    node_x_m: float,
    node_y_m: float,
    direction: str,
    thickness_m: float,
    length_m: float,
) -> SignedShearWall:
    signed_direction = normalize_signed_direction(direction)
    thickness = abs(float(thickness_m))
    length = abs(float(length_m))
    if thickness <= EPS or length <= EPS:
        raise ValueError(f"{wall_id}: non-positive wall thickness/length.")
    sign = signed_direction_sign(signed_direction)
    dx = sign * length if signed_direction.endswith("X") else 0.0
    dy = sign * length if signed_direction.endswith("Z") else 0.0
    return SignedShearWall(
        wall_id=wall_id,
        x1=round(float(node_x_m), 3),
        y1=round(float(node_y_m), 3),
        x2=round(float(node_x_m) + dx, 3),
        y2=round(float(node_y_m) + dy, 3),
        thickness_m=round(thickness, 3),
    )


def load_signed_shear_wall_sheet(ws) -> list[SignedShearWall]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(value): idx for idx, value in enumerate(headers)}
    current_schema = {"nodexm", "nodezm", "ydm", "zdm"}
    intermediate_schema = {"nodexm", "nodezm", "direction", "thicknessmm", "maxlengthm"}
    legacy_schema = {"startxm", "startym", "endxm", "endym", "direction", "thicknessmm"}

    walls: list[SignedShearWall] = []
    warned_rows: set[str] = set()

    def warn_once(message: str) -> None:
        if message in warned_rows:
            return
        warned_rows.add(message)
        print(f"[WARN] {message}")

    if current_schema.issubset(header_map):
        for row in ws.iter_rows(min_row=2, values_only=True):
            node_x = safe_float(row[header_map["nodexm"]])
            node_y = safe_float(row[header_map["nodezm"]])
            yd_m = safe_float(row[header_map["ydm"]])
            zd_m = safe_float(row[header_map["zdm"]])
            wall_id = safe_text(row[header_map["wallid"]]) if "wallid" in header_map else ""
            if None in (node_x, node_y, yd_m, zd_m):
                warn_once(f"Skipping a wall in sheet '{ws.title}' due to blank signed geometry.")
                continue
            try:
                direction, thickness_m, length_m = geometry_from_yd_zd(yd_m, zd_m)
                if not wall_id:
                    wall_id = synthesize_wall_label(node_x, node_y, direction)
                walls.append(build_signed_shear_wall(wall_id, node_x, node_y, direction, thickness_m, length_m))
            except ValueError as exc:
                label = wall_id or f"({node_x}, {node_y})"
                warn_once(f"Skipping wall '{label}' in sheet '{ws.title}': {exc}")
        return walls

    if intermediate_schema.issubset(header_map):
        for row in ws.iter_rows(min_row=2, values_only=True):
            node_x = safe_float(row[header_map["nodexm"]])
            node_y = safe_float(row[header_map["nodezm"]])
            thickness_mm = safe_float(row[header_map["thicknessmm"]])
            length_m = safe_float(row[header_map["maxlengthm"]])
            direction = safe_text(row[header_map["direction"]])
            wall_id = safe_text(row[header_map["wallid"]]) if "wallid" in header_map else ""
            if None in (node_x, node_y, thickness_mm, length_m) or not direction:
                warn_once(f"Skipping a wall in sheet '{ws.title}' due to blank direction geometry.")
                continue
            try:
                if not wall_id:
                    wall_id = synthesize_wall_label(node_x, node_y, direction)
                walls.append(
                    build_signed_shear_wall(
                        wall_id,
                        node_x,
                        node_y,
                        direction,
                        float(thickness_mm) / 1000.0,
                        length_m,
                    )
                )
            except ValueError as exc:
                label = wall_id or f"({node_x}, {node_y})"
                warn_once(f"Skipping wall '{label}' in sheet '{ws.title}': {exc}")
        return walls

    if legacy_schema.issubset(header_map):
        for row in ws.iter_rows(min_row=2, values_only=True):
            start_x = safe_float(row[header_map["startxm"]])
            start_y = safe_float(row[header_map["startym"]])
            end_x = safe_float(row[header_map["endxm"]])
            end_y = safe_float(row[header_map["endym"]])
            thickness_mm = safe_float(row[header_map["thicknessmm"]])
            if None in (start_x, start_y, end_x, end_y, thickness_mm):
                warn_once(f"Skipping a wall in sheet '{ws.title}' due to blank endpoint geometry.")
                continue
            if "anchorside" in header_map and safe_text(row[header_map["anchorside"]]).lower() == "end":
                start_x, end_x = end_x, start_x
                start_y, end_y = end_y, start_y
            wall_id = safe_text(row[header_map["wallid"]]) if "wallid" in header_map else ""
            if not wall_id:
                wall_id = synthesize_wall_label(start_x, start_y, safe_text(row[header_map["direction"]]))
            walls.append(
                SignedShearWall(
                    wall_id=wall_id,
                    x1=round(float(start_x), 3),
                    y1=round(float(start_y), 3),
                    x2=round(float(end_x), 3),
                    y2=round(float(end_y), 3),
                    thickness_m=round(abs(float(thickness_mm)) / 1000.0, 3),
                )
            )
        return walls

    return []


def load_signed_shear_walls(wb) -> tuple[list[SignedShearWall], str | None]:
    for sheet_name in SHEAR_WALL_SOURCE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        walls = load_signed_shear_wall_sheet(wb[sheet_name])
        if walls:
            return walls, sheet_name
    return [], None


def load_visual_grid_lines(wb) -> tuple[list[float], list[float]]:
    xs: set[float] = set()
    ys: set[float] = set()

    if "Columns" in wb.sheetnames:
        ws = wb["Columns"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {normalize_header(v): idx for idx, v in enumerate(headers)}
        node_x_idx = header_map.get("nodecoordinatexm")
        node_y_idx = header_map.get("nodecoordinateym")
        if node_x_idx is not None and node_y_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(value in (None, "") for value in row):
                    continue
                if row[node_x_idx] not in (None, ""):
                    xs.add(round(float(row[node_x_idx]), 3))
                if row[node_y_idx] not in (None, ""):
                    ys.add(round(float(row[node_y_idx]), 3))

    if "Primary Beams" in wb.sheetnames:
        ws = wb["Primary Beams"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {normalize_header(v): idx for idx, v in enumerate(headers)}
        required = {"startnodexm", "startnodeym", "endnodexm", "endnodeym"}
        if required.issubset(header_map):
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(value in (None, "") for value in row):
                    continue
                xs.update([round(float(row[header_map["startnodexm"]]), 3), round(float(row[header_map["endnodexm"]]), 3)])
                ys.update([round(float(row[header_map["startnodeym"]]), 3), round(float(row[header_map["endnodeym"]]), 3)])

    if FINAL_SECONDARY_SHEET in wb.sheetnames:
        ws = wb[FINAL_SECONDARY_SHEET]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {normalize_header(v): idx for idx, v in enumerate(headers)}
        required = {"snappedx1m", "snappedy1m", "snappedx2m", "snappedy2m"}
        if required.issubset(header_map):
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(value in (None, "") for value in row):
                    continue
                xs.update([round(float(row[header_map["snappedx1m"]]), 3), round(float(row[header_map["snappedx2m"]]), 3)])
                ys.update([round(float(row[header_map["snappedy1m"]]), 3), round(float(row[header_map["snappedy2m"]]), 3)])

    return sorted(xs), sorted(ys)


def beam_rect(beam: BeamStrip, thickness_m: float) -> tuple[float, float, float, float]:
    if beam.direction == "X":
        x1 = min(beam.start_x, beam.end_x)
        x2 = max(beam.start_x, beam.end_x)
        y = (beam.start_y + beam.end_y) / 2.0
        half = thickness_m / 2.0
        return x1, y - half, x2 - x1, thickness_m
    x = (beam.start_x + beam.end_x) / 2.0
    y1 = min(beam.start_y, beam.end_y)
    y2 = max(beam.start_y, beam.end_y)
    half = thickness_m / 2.0
    return x - half, y1, thickness_m, y2 - y1


def shear_wall_rect(wall: SignedShearWall) -> tuple[float, float, float, float]:
    if abs(wall.y1 - wall.y2) <= EPS:
        x1 = min(wall.x1, wall.x2)
        x2 = max(wall.x1, wall.x2)
        y = (wall.y1 + wall.y2) / 2.0
        half = wall.thickness_m / 2.0
        return x1, y - half, x2 - x1, wall.thickness_m
    x = (wall.x1 + wall.x2) / 2.0
    y1 = min(wall.y1, wall.y2)
    y2 = max(wall.y1, wall.y2)
    half = wall.thickness_m / 2.0
    return x - half, y1, wall.thickness_m, y2 - y1


def vertical_grid_label(index: int) -> str:
    return str(index + 1)


def horizontal_grid_label(index: int) -> str:
    value = index
    result = ""
    while True:
        value, remainder = divmod(value, 26)
        result = chr(ord("A") + remainder) + result
        if value == 0:
            return result
        value -= 1


def compute_bounds(
    columns: list[ColumnRect],
    beams: list[BeamStrip],
    shear_walls: list[SignedShearWall] | None = None,
    features: list[FeatureRect] | None = None,
) -> tuple[float, float, float, float]:
    min_x = min(column.node_x - column.draw_w * 0.5 for column in columns)
    max_x = max(column.node_x + column.draw_w * 0.5 for column in columns)
    min_y = min(column.node_y - column.draw_h * 0.5 for column in columns)
    max_y = max(column.node_y + column.draw_h * 0.5 for column in columns)
    for beam in beams:
        bx, by, bw, bh = beam_rect(beam, beam.beam_width_m)
        min_x = min(min_x, bx)
        max_x = max(max_x, bx + bw)
        min_y = min(min_y, by)
        max_y = max(max_y, by + bh)
        if beam.wall_thickness_m > 0:
            wx, wy, ww, wh = beam_rect(beam, beam.wall_thickness_m)
            min_x = min(min_x, wx)
            max_x = max(max_x, wx + ww)
            min_y = min(min_y, wy)
            max_y = max(max_y, wy + wh)
    for wall in shear_walls or []:
        wx, wy, ww, wh = shear_wall_rect(wall)
        min_x = min(min_x, wx)
        max_x = max(max_x, wx + ww)
        min_y = min(min_y, wy)
        max_y = max(max_y, wy + wh)
    for feature in features or []:
        fx1, fx2 = sorted((feature.x1, feature.x2))
        fy1, fy2 = sorted((feature.y1, feature.y2))
        min_x = min(min_x, fx1)
        max_x = max(max_x, fx2)
        min_y = min(min_y, fy1)
        max_y = max(max_y, fy2)
    return min_x, min_y, max_x, max_y


class Transform:
    def __init__(self, min_x: float, min_y: float, max_x: float, max_y: float) -> None:
        span_x = max(max_x - min_x, 1.0)
        span_y = max(max_y - min_y, 1.0)
        scale_x = (PAGE_WIDTH_PT - DRAWING_LEFT_OFFSET_PT - DRAWING_RIGHT_RESERVED_PT) / span_x
        scale_y = (PAGE_HEIGHT_PT - 2 * PAGE_MARGIN_PT) / span_y
        self.scale = min(scale_x, scale_y)
        self.min_x = min_x
        self.min_y = min_y
        self.max_x = max_x
        self.max_y = max_y
        self.offset_x = DRAWING_LEFT_OFFSET_PT
        self.offset_y = PAGE_MARGIN_PT

    def point(self, x: float, y: float) -> tuple[float, float]:
        px = self.offset_x + (x - self.min_x) * self.scale
        py = self.offset_y + (y - self.min_y) * self.scale
        return round(px, 3), round(py, 3)

    def rect(self, x: float, y: float, width: float, height: float) -> tuple[float, float, float, float]:
        px, py = self.point(x, y)
        return round(px, 3), round(py, 3), round(width * self.scale, 3), round(height * self.scale, 3)


def pdf_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


class PdfCanvas:
    def __init__(self) -> None:
        self.commands: list[str] = []

    def stroke_color(self, r: float, g: float, b: float) -> None:
        self.commands.append(f"{r:.3f} {g:.3f} {b:.3f} RG")

    def fill_color(self, r: float, g: float, b: float) -> None:
        self.commands.append(f"{r:.3f} {g:.3f} {b:.3f} rg")

    def line_width(self, width: float) -> None:
        self.commands.append(f"{width:.3f} w")

    def dash(self, on: float | None = None, off: float | None = None) -> None:
        if on is None or off is None:
            self.commands.append("[] 0 d")
        else:
            self.commands.append(f"[{on:.3f} {off:.3f}] 0 d")

    def rect(self, x: float, y: float, width: float, height: float, mode: str) -> None:
        self.commands.append(f"{x:.3f} {y:.3f} {width:.3f} {height:.3f} re {mode}")

    def line(self, x1: float, y1: float, x2: float, y2: float) -> None:
        self.commands.append(f"{x1:.3f} {y1:.3f} m {x2:.3f} {y2:.3f} l S")

    def text(self, x: float, y: float, text: str, size: float = 10.0, font: str = "F1") -> None:
        self.commands.append(
            f"BT /{font} {size:.3f} Tf {x:.3f} {y:.3f} Td ({pdf_escape(text)}) Tj ET"
        )

    def content(self) -> str:
        return "\n".join(self.commands) + "\n"


class SimplePdfWriter:
    def __init__(self, page_width: float, page_height: float) -> None:
        self.page_width = page_width
        self.page_height = page_height
        self.pages: list[str] = []

    def add_page(self, content: str) -> None:
        self.pages.append(content)

    def save(self, path: Path) -> None:
        objects: list[bytes] = []

        def add_object(payload: str | bytes) -> int:
            data = payload if isinstance(payload, bytes) else payload.encode("latin-1")
            objects.append(data)
            return len(objects)

        catalog_id = add_object("placeholder")
        pages_id = add_object("placeholder")
        font_regular_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
        font_bold_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

        page_ids: list[int] = []
        for content in self.pages:
            content_bytes = content.encode("latin-1")
            content_id = add_object(
                b"<< /Length " + str(len(content_bytes)).encode("ascii") + b" >>\nstream\n" + content_bytes + b"endstream"
            )
            page_id = add_object(
                f"<< /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 {self.page_width:.3f} {self.page_height:.3f}] "
                f"/Resources << /Font << /F1 {font_regular_id} 0 R /F2 {font_bold_id} 0 R >> >> "
                f"/Contents {content_id} 0 R >>"
            )
            page_ids.append(page_id)

        objects[catalog_id - 1] = f"<< /Type /Catalog /Pages {pages_id} 0 R >>".encode("latin-1")
        kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
        objects[pages_id - 1] = f"<< /Type /Pages /Count {len(page_ids)} /Kids [{kids}] >>".encode("latin-1")

        offsets: list[int] = [0]
        pdf = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
        for obj_id, payload in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf.extend(f"{obj_id} 0 obj\n".encode("latin-1"))
            pdf.extend(payload)
            pdf.extend(b"\nendobj\n")

        xref_start = len(pdf)
        pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
        pdf.extend(b"0000000000 65535 f \n")
        for offset in offsets[1:]:
            pdf.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
        pdf.extend(
            f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode(
                "latin-1"
            )
        )
        path.write_bytes(pdf)


def boxes_overlap(a: tuple[float, float, float, float], b: tuple[float, float, float, float]) -> bool:
    return not (a[2] <= b[0] or b[2] <= a[0] or a[3] <= b[1] or b[3] <= a[1])


def _try_place_label(
    placed_boxes: list[tuple[float, float, float, float]],
    px: float,
    py: float,
    pw: float,
    ph: float,
    lbl: str,
    font_size: float,
) -> tuple[float, float, tuple[float, float, float, float]]:
    text_w = len(lbl) * font_size * 0.58
    text_h = font_size
    fracs = (0.50, 0.25, 0.75, 0.15, 0.85)
    if pw >= ph:
        for frac in fracs:
            cx = px + pw * frac
            text_x = cx - text_w * 0.5
            text_y = py + (ph * 0.5) - (font_size * 0.30)
            box = (text_x, text_y, text_x + text_w, text_y + text_h)
            if not any(boxes_overlap(box, placed) for placed in placed_boxes):
                return text_x, text_y, box
    else:
        for frac in fracs:
            cy = py + ph * frac
            text_x = px + (pw * 0.5) - text_w * 0.5
            text_y = cy - font_size * 0.30
            box = (text_x, text_y, text_x + text_w, text_y + text_h)
            if not any(boxes_overlap(box, placed) for placed in placed_boxes):
                return text_x, text_y, box
    text_x = px + (pw * 0.5) - (len(lbl) * font_size * 0.29)
    text_y = py + (ph * 0.5) - (font_size * 0.30)
    box = (text_x, text_y, text_x + text_w, text_y + text_h)
    return text_x, text_y, box


def draw_page(
    page_title: str,
    columns: list[ColumnRect],
    primary_beams: list[BeamStrip],
    secondary_beams: list[BeamStrip],
    shear_walls: list[SignedShearWall],
    rectangles: list[FeatureRect],
    balconies: list[FeatureRect],
    staircases: list[FeatureRect],
    grid_x: list[float],
    grid_y: list[float],
    transform: Transform,
) -> str:
    c = PdfCanvas()
    x_grid_items = [(vertical_grid_label(i), value) for i, value in enumerate(grid_x)]
    y_grid_items = [(horizontal_grid_label(i), value) for i, value in enumerate(grid_y)]

    # Title.
    c.fill_color(0, 0, 0)
    c.text(42, PAGE_HEIGHT_PT - 20, f"Initial Structural Plan - {page_title}", size=14, font="F2")

    # Compact right panel: grid tables then vertical legend.
    panel_x = PAGE_WIDTH_PT - 140
    _panel_y = PAGE_HEIGHT_PT - 36
    _row_h = 13

    c.fill_color(0, 0, 0)
    c.text(panel_x, _panel_y, "Vertical Grids", size=8, font="F2")
    _panel_y -= 11
    c.text(panel_x, _panel_y, "No.", size=7, font="F2")
    c.text(panel_x + 22, _panel_y, "X (m)", size=7, font="F2")
    for label, value in x_grid_items:
        _panel_y -= _row_h
        c.text(panel_x, _panel_y, label, size=7)
        c.text(panel_x + 22, _panel_y, f"{value:.3f}", size=7)

    _panel_y -= 16
    c.text(panel_x, _panel_y, "Horizontal Grids", size=8, font="F2")
    _panel_y -= 11
    c.text(panel_x, _panel_y, "No.", size=7, font="F2")
    c.text(panel_x + 22, _panel_y, "Y (m)", size=7, font="F2")
    for label, value in y_grid_items:
        _panel_y -= _row_h
        c.text(panel_x, _panel_y, label, size=7)
        c.text(panel_x + 22, _panel_y, f"{value:.3f}", size=7)

    _panel_y -= 18
    c.fill_color(0, 0, 0)
    c.text(panel_x, _panel_y, "Legend", size=8, font="F2")
    _panel_y -= 8
    c.text(panel_x, _panel_y, "C=corner  L=lift", size=6)
    _panel_y -= 13

    def _leg(lbl, r, g, b, filled, dashed=False):
        nonlocal _panel_y
        c.line_width(0.7)
        if dashed:
            c.stroke_color(r, g, b)
            c.dash(3.0, 2.0)
            c.rect(panel_x, _panel_y, 10, 6, "S")
            c.dash()
        elif filled:
            c.fill_color(r, g, b)
            c.stroke_color(r, g, b)
            c.rect(panel_x, _panel_y, 10, 6, "B")
        else:
            c.stroke_color(r, g, b)
            c.rect(panel_x, _panel_y, 10, 6, "S")
        c.fill_color(0, 0, 0)
        c.text(panel_x + 14, _panel_y, lbl, size=6.5)
        _panel_y -= 11

    _leg("Column", 0.78, 0.78, 0.78, True)
    _leg("Primary beam", 0, 0, 0, False)
    _leg("Secondary beam", 0.10, 0.32, 0.82, False)
    _leg("Primary wall", 0.88, 0.15, 0.15, True)
    _leg("Secondary wall", 0.55, 0.33, 0.16, True)
    _leg("Shear wall", 0.35, 0.10, 0.10, True)
    _leg("Balcony", 0.16, 0.50, 0.20, False, dashed=True)
    _leg("Staircase", 0.82, 0.46, 0.10, False, dashed=True)

    # Origin marker.
    ox, oy = transform.point(0.0, 0.0)
    c.stroke_color(0.15, 0.45, 0.15)
    c.line_width(0.9)
    c.line(ox - 8, oy, ox + 8, oy)
    c.line(ox, oy - 8, ox, oy + 8)
    c.fill_color(0.15, 0.45, 0.15)
    c.text(ox + 10, oy + 4, "(0,0)", size=8)

    # Grid lines behind the framing.
    c.stroke_color(0.72, 0.72, 0.72)
    c.line_width(0.55)
    c.dash(2.0, 4.0)
    for x in grid_x:
        gx1, gy1 = transform.point(x, transform.min_y)
        gx2, gy2 = transform.point(x, transform.max_y)
        c.line(gx1, gy1, gx2, gy2)
    for y in grid_y:
        gx1, gy1 = transform.point(transform.min_x, y)
        gx2, gy2 = transform.point(transform.max_x, y)
        c.line(gx1, gy1, gx2, gy2)
    c.dash()

    # Grid labels.
    c.fill_color(0.35, 0.35, 0.35)
    for label, x in x_grid_items:
        px, py = transform.point(x, transform.min_y)
        c.text(px - 6, max(14, py - 18), label, size=12, font="F2")
    for label, y in y_grid_items:
        px, py = transform.point(transform.min_x, y)
        c.text(max(4, px - 26), py - 5, label, size=12, font="F2")

    # Primary beams first.
    c.stroke_color(0, 0, 0)
    c.line_width(0.9)
    for beam in primary_beams:
        bx, by, bw, bh = beam_rect(beam, beam.beam_width_m)
        px, py, pw, ph = transform.rect(bx, by, bw, bh)
        c.rect(px, py, pw, ph, "S")

    # Secondary beams on top of primaries.
    c.stroke_color(0.10, 0.32, 0.82)
    c.line_width(0.9)
    for beam in secondary_beams:
        bx, by, bw, bh = beam_rect(beam, beam.beam_width_m)
        px, py, pw, ph = transform.rect(bx, by, bw, bh)
        c.rect(px, py, pw, ph, "S")

    # Primary walls.
    for beam in primary_beams:
        if beam.wall_thickness_m <= 0:
            continue
        wx, wy, ww, wh = beam_rect(beam, beam.wall_thickness_m)
        px, py, pw, ph = transform.rect(wx, wy, ww, wh)
        c.fill_color(0.88, 0.15, 0.15)
        c.stroke_color(0.88, 0.15, 0.15)
        c.rect(px, py, pw, ph, "B")

    # Secondary walls.
    for beam in secondary_beams:
        if beam.wall_thickness_m <= 0:
            continue
        wx, wy, ww, wh = beam_rect(beam, beam.wall_thickness_m)
        px, py, pw, ph = transform.rect(wx, wy, ww, wh)
        c.fill_color(0.55, 0.33, 0.16)
        c.stroke_color(0.55, 0.33, 0.16)
        c.rect(px, py, pw, ph, "B")

    # Signed shear walls from the wall recommendation/landscape sheet.
    for wall in shear_walls:
        wx, wy, ww, wh = shear_wall_rect(wall)
        px, py, pw, ph = transform.rect(wx, wy, ww, wh)
        c.fill_color(0.35, 0.10, 0.10)
        c.stroke_color(0.35, 0.10, 0.10)
        c.rect(px, py, pw, ph, "B")

        font_size = max(7.0, min(10.0, min(pw, ph) * 0.60))
        text_x = px + max(2.0, (pw * 0.5) - (len(wall.wall_id) * font_size * 0.17))
        text_y = py + max(1.5, (ph * 0.5) - (font_size * 0.25))
        c.fill_color(1.0, 1.0, 1.0)
        c.text(text_x, text_y, wall.wall_id, size=font_size, font="F2")

    # Primary and secondary beam labels with anti-overlap placement.
    placed_label_boxes: list[tuple[float, float, float, float]] = []

    for beam in primary_beams:
        bx, by, bw, bh = beam_rect(beam, beam.beam_width_m)
        px, py, pw, ph = transform.rect(bx, by, bw, bh)
        lbl = f"{beam.beam_no} ({int(round(beam.beam_width_m * 1000))}; {int(round(beam.beam_depth_m * 1000))}; {int(round(beam.wall_thickness_m * 1000))})"
        font_size = 13.0
        text_x, text_y, box = _try_place_label(placed_label_boxes, px, py, pw, ph, lbl, font_size)
        placed_label_boxes.append(box)
        c.fill_color(0, 0, 0)
        c.text(text_x, text_y, lbl, size=font_size, font="F2")

    for beam in secondary_beams:
        bx, by, bw, bh = beam_rect(beam, beam.beam_width_m)
        px, py, pw, ph = transform.rect(bx, by, bw, bh)
        lbl = f"{beam.beam_no} ({int(round(beam.beam_width_m * 1000))}; {int(round(beam.beam_depth_m * 1000))}; {int(round(beam.wall_thickness_m * 1000))})"
        font_size = 13.0
        text_x, text_y, box = _try_place_label(placed_label_boxes, px, py, pw, ph, lbl, font_size)
        placed_label_boxes.append(box)
        c.fill_color(0.05, 0.18, 0.55)
        c.text(text_x, text_y, lbl, size=font_size, font="F2")

    # Balcony outlines.
    c.stroke_color(0.16, 0.50, 0.20)
    c.line_width(1.0)
    c.dash(4.0, 3.0)
    for feature in balconies:
        fx1, fx2 = sorted((feature.x1, feature.x2))
        fy1, fy2 = sorted((feature.y1, feature.y2))
        px, py, pw, ph = transform.rect(fx1, fy1, fx2 - fx1, fy2 - fy1)
        c.rect(px, py, pw, ph, "S")
        c.fill_color(0.16, 0.50, 0.20)
        c.text(px + 3, py + max(3, ph * 0.5), feature.label, size=8)

    # Staircase outlines.
    c.stroke_color(0.82, 0.46, 0.10)
    c.line_width(1.0)
    c.dash(4.0, 3.0)
    for feature in staircases:
        fx1, fx2 = sorted((feature.x1, feature.x2))
        fy1, fy2 = sorted((feature.y1, feature.y2))
        px, py, pw, ph = transform.rect(fx1, fy1, fx2 - fx1, fy2 - fy1)
        c.rect(px, py, pw, ph, "S")
        c.fill_color(0.82, 0.46, 0.10)
        c.text(px + 3, py + max(3, ph * 0.5), feature.label, size=8)
    c.dash()

    # Rectangle outlines and labels.
    c.stroke_color(0.28, 0.28, 0.34)
    c.line_width(0.9)
    c.dash(6.0, 3.0)
    for feature in rectangles:
        fx1, fx2 = sorted((feature.x1, feature.x2))
        fy1, fy2 = sorted((feature.y1, feature.y2))
        px, py, pw, ph = transform.rect(fx1, fy1, fx2 - fx1, fy2 - fy1)
        c.rect(px, py, pw, ph, "S")
        font_size = max(9.0, min(12.0, min(pw, ph) * 0.35))
        text_x = px + max(4.0, (pw * 0.5) - (len(feature.label) * font_size * 0.18))
        text_y = py + max(4.0, (ph * 0.5) - (font_size * 0.20))
        c.fill_color(0.28, 0.28, 0.34)
        c.text(text_x, text_y, feature.label, size=font_size, font="F2")
    c.dash()

    # Columns on top.
    for column in columns:
        px, py, pw, ph = transform.rect(column.node_x - column.draw_w * 0.5, column.node_y - column.draw_h * 0.5, column.draw_w, column.draw_h)
        c.fill_color(0.78, 0.78, 0.78)
        c.stroke_color(0, 0, 0)
        c.line_width(0.8)
        c.rect(px, py, pw, ph, "B")

        mark = None
        if column.location == "Corner":
            mark = "C"
        elif column.location == "Lift":
            mark = "L"
        if mark:
            font_size = max(8.0, min(12.0, min(pw, ph) * 0.55))
            text_x = px + max(2.0, pw * 0.25)
            text_y = py + max(2.0, ph * 0.28)
            c.fill_color(0, 0, 0)
            c.text(text_x, text_y, mark, size=font_size, font="F2")

        id_font = max(6.0, min(8.5, min(pw, ph) * 0.42))
        id_x = px + 1.5
        id_y = py + max(1.2, ph - (id_font + 1.8))
        c.fill_color(0, 0, 0)
        c.text(id_x, id_y, column.type_name, size=id_font, font="F2")

    return c.content()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a simple 3-page PDF structural review drawing (Plinth, Typical Floor, Terrace) from the node-appended workbook."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help="Input workbook with appended node coordinates. Defaults to the single discovered workbook in the project folders.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="Output PDF path. Defaults to <input_stem>_initial_structural_plan.pdf beside the input workbook.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = resolve_input_workbook(args.input)
    output_path = args.output.resolve() if args.output else input_path.with_name(f"{input_path.stem}_initial_structural_plan.pdf")
    other_geometry_path = resolve_other_geometry_workbook(input_path)

    wb = load_workbook(input_path, data_only=True)
    if "Columns" not in wb.sheetnames or "Primary Beams" not in wb.sheetnames:
        raise SystemExit("Input workbook must contain 'Columns' and 'Primary Beams' sheets.")
    columns = load_columns(wb["Columns"])
    all_primary_beams = load_beams(wb["Primary Beams"])
    all_secondary_beams = load_secondary_beams(wb[FINAL_SECONDARY_SHEET]) if FINAL_SECONDARY_SHEET in wb.sheetnames else []
    signed_shear_walls, signed_wall_source = load_signed_shear_walls(wb)
    grid_x, grid_y = load_visual_grid_lines(wb)
    if signed_shear_walls:
        grid_x_vals = {round(value, 3) for value in grid_x}
        grid_y_vals = {round(value, 3) for value in grid_y}
        for wall in signed_shear_walls:
            grid_x_vals.update((wall.x1, wall.x2))
            grid_y_vals.update((wall.y1, wall.y2))
        grid_x = sorted(grid_x_vals)
        grid_y = sorted(grid_y_vals)
    floor_primary_beams = {floor: [beam for beam in all_primary_beams if beam.floor == floor] for floor, _ in FLOOR_PAGES}
    floor_secondary_beams = {floor: [beam for beam in all_secondary_beams if beam.floor == floor] for floor, _ in FLOOR_PAGES}
    balconies: list[FeatureRect] = []
    rectangles: list[FeatureRect] = []
    staircases: list[FeatureRect] = []
    if other_geometry_path is not None:
        other_wb = load_workbook(other_geometry_path, data_only=True)
        if "Rectangle coordinates" in other_wb.sheetnames:
            rectangles = load_feature_rects(other_wb["Rectangle coordinates"], "nonplinth", "rectangle")
        if "Balcony coordinates" in other_wb.sheetnames:
            balconies = load_feature_rects(other_wb["Balcony coordinates"], "nonplinth", "balcony")
        if "Staircase details" in other_wb.sheetnames:
            staircases = load_feature_rects(other_wb["Staircase details"], "nonplinth", "staircase")
        other_wb.close()

    all_features = rectangles + balconies + staircases
    bounds = compute_bounds(
        columns,
        [beam for floor in floor_primary_beams.values() for beam in floor] + [beam for floor in floor_secondary_beams.values() for beam in floor],
        signed_shear_walls,
        all_features,
    )
    transform = Transform(*bounds)

    pdf = SimplePdfWriter(PAGE_WIDTH_PT, PAGE_HEIGHT_PT)
    for floor_key, page_title in FLOOR_PAGES:
        floor_rectangles: list[FeatureRect] = []
        for feature in rectangles:
            label = feature.label.strip().lower()
            if label == "lift":
                floor_rectangles.append(feature)
            elif label in {"staircase", "shaft"} and floor_key != "Plinth":
                floor_rectangles.append(feature)
            elif label == "mumty" and floor_key == "Terrace":
                floor_rectangles.append(feature)
        floor_balconies = balconies if floor_key != "Plinth" else []
        floor_staircases = staircases if floor_key != "Plinth" else []
        pdf.add_page(
            draw_page(
                page_title,
                columns,
                floor_primary_beams[floor_key],
                floor_secondary_beams[floor_key],
                signed_shear_walls,
                floor_rectangles,
                floor_balconies,
                floor_staircases,
                grid_x,
                grid_y,
                transform,
            )
        )
    pdf.save(output_path)

    print(f"Input workbook : {input_path}")
    print(f"Output PDF     : {output_path}")
    print(f"Columns drawn  : {len(columns)}")
    if signed_shear_walls:
        print(f"Signed walls   : {len(signed_shear_walls)} from {signed_wall_source}")
    for floor_key, page_title in FLOOR_PAGES:
        print(f"{page_title}: {len(floor_primary_beams[floor_key])} primary, {len(floor_secondary_beams[floor_key])} secondary")


if __name__ == "__main__":
    main()
