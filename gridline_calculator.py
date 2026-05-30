"""
Gridline Calculator
Calculates beam/column rectangular footprint (X1, X2, Y1, Y2) at each grid intersection.
Uses raw anchor coordinates from node_coordinates Excel and YD(-2)/ZD(-2) from Filleddata.

Usage: python gridline_calculator.py
No arguments needed - uses hardcoded paths.
"""

import os
import glob
import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────
STD_ANL_FOLDER = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model"
FILLEDDATA_PATH = os.path.join(STD_ANL_FOLDER, "Datadata.xlsx")


# # ── Old ordinal words (kept for reference) ──
# ORDINAL_WORDS = [
#     "First", "Second", "Third", "Fourth", "Fifth",
#     "Sixth", "Seventh", "Eighth", "Ninth", "Tenth",
# ]
#
# def ordinal_word(n):
#     """Convert 1-based index to an ordinal word for floor labels."""
#     if 1 <= n <= len(ORDINAL_WORDS):
#         return ORDINAL_WORDS[n - 1]
#     return f"{n}th"


def ordinal_suffix(n):
    """Convert number to ordinal string: 1→'1st', 2→'2nd', 3→'3rd', etc."""
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def find_scalar_by_label(ws, label):
    """Search Datadata2 for a label and return the adjacent value."""
    target = str(label).strip()
    for row in ws.iter_rows(values_only=True):
        row_vals = list(row)
        for idx, val in enumerate(row_vals[:-1]):
            if isinstance(val, str) and val.strip() == target:
                return row_vals[idx + 1]
    return None


def build_floor_levels(y_values, mumty_height):
    """
    Build floor levels with tags:
      Index 0 (Y<0)      → "Foundation"
      Index 1 (Y=0)      → "Plinth"
      Index 2 (first +Y) → "Stilt roof"
      Index 3+            → "1st floor roof", "2nd floor roof", etc.
      Synthetic (last+mumty) → "Mumty"
    No separate "Terrace" — last Y value becomes last numbered floor roof.
    """
    # # ── Old floor tag logic (kept for reference) ──
    # floors = []
    # for idx, y_val in enumerate(y_values):
    #     if idx == 0:
    #         floor_type = "Foundation"
    #     elif idx == 1:
    #         floor_type = "Stilt floor"
    #     elif idx == 2:
    #         floor_type = "Stilt floor roof"
    #     elif idx == len(y_values) - 1:
    #         floor_type = "Terrace"
    #     else:
    #         floor_num = idx - 2
    #         floor_type = f"{ordinal_word(floor_num)} floor roof"
    #     floors.append((y_val, floor_type))

    floors = []
    for idx, y_val in enumerate(y_values):
        if idx == 0:
            floor_type = "Foundation"
        elif idx == 1:
            floor_type = "Plinth"
        elif idx == 2:
            floor_type = "Stilt roof"
        else:
            floor_num = idx - 2
            floor_type = f"{ordinal_suffix(floor_num)} floor roof"
        floors.append((y_val, floor_type))

    synthetic_y = None
    if y_values and mumty_height is not None:
        synthetic_y = round(float(y_values[-1]) + float(mumty_height), 3)
        if not any(abs(y - synthetic_y) <= 1e-6 for y in y_values):
            floors.append((synthetic_y, "Mumty"))

    return floors, synthetic_y


def find_latest_node_coordinates():
    """Find the latest node_coordinates_*.xlsx file in STD ANL folder."""
    pattern = os.path.join(STD_ANL_FOLDER, "node_coordinates_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"No node_coordinates_*.xlsx found in {STD_ANL_FOLDER}")
    latest = max(files, key=os.path.getmtime)
    print(f"Using node coordinates: {os.path.basename(latest)}")
    return latest


def read_node_coordinates(filepath):
    """
    Read node_coordinates Excel. Returns list of dicts, one per column (first row only,
    since anchor X/Y and locations are the same across all 3 floor rows).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Node coordinates"]

    columns = []
    seen_ids = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        col_no = row[0]
        if col_no in seen_ids:
            continue  # skip duplicate floor rows, only need first
        seen_ids.add(col_no)

        columns.append({
            "no": col_no,
            "name": row[1],
            "anchor_x": row[2],
            "anchor_y": row[3],
            "location": row[4],
            "beam_x_loc": row[6],       # Left / Right / Centre
            "opposite_x": row[7],       # for Centre X calc
            "beam_y_loc": row[13],      # Front / Back / Centre
            "opposite_y_x": row[14],    # opposite coord X for Y direction (unused)
            "opposite_y_y": row[15],    # opposite coord Y for Y direction Centre calc
            "orientation": row[20] if len(row) > 20 else "Vertical",
        })

    wb.close()
    print(f"Read {len(columns)} columns from node coordinates")
    return columns


def read_filleddata_all_yd_zd():
    """
    Read ALL YD/ZD pairs from Filleddata Datadata2 sheet (all floors).
    Returns:
      - floors: list of (y_value, floor_type) tuples, e.g. [(-2, "Foundation"), (0, "Stilt"), ...]
      - yd_zd_all: dict: node_number -> list of (yd, zd) per floor
    """
    wb = openpyxl.load_workbook(FILLEDDATA_PATH, read_only=True)
    ws = wb["Datadata2"]

    # Read headers to find YD/ZD pairs and extract Y values
    headers = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)

    # Find first YD( header dynamically
    yd_start = None
    for i, h in enumerate(headers):
        if h and str(h).startswith("YD("):
            yd_start = i
            break
    if yd_start is None:
        raise ValueError("No YD( header found in Datadata2 row 1")

    # Find all YD columns (every 2 columns from first YD)
    y_values = []
    yd_zd_col_pairs = []  # list of (yd_col_idx, zd_col_idx)
    for i in range(yd_start, len(headers), 2):
        h = headers[i]
        if h is None or not str(h).startswith("YD("):
            break
        # Extract Y value from header like "YD(-2)" or "YD(3.15)"
        y_str = h.replace("YD(", "").replace(")", "")
        y_val = float(y_str)
        y_values.append(y_val)
        yd_zd_col_pairs.append((i, i + 1))

    mumty_height = find_scalar_by_label(ws, "Mumty height")
    floors, synthetic_y = build_floor_levels(y_values, mumty_height)

    print(f"Found {len(floors)} floor levels:")
    for y_val, ft in floors:
        print(f"  Y={y_val} -> {ft}")

    # Read per-node YD/ZD for all floors
    yd_zd_all = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        node_no = row[0]
        if node_no is None:
            break
        pairs = []
        for yd_col, zd_col in yd_zd_col_pairs:
            yd = row[yd_col] if yd_col < len(row) else None
            zd = row[zd_col] if zd_col < len(row) else None
            pairs.append((yd, zd))
        if synthetic_y is not None and pairs:
            # Mumty roof repeats terrace footprint data; only the Y level changes.
            pairs.append(pairs[-1])
        yd_zd_all[node_no] = pairs

    wb.close()
    print(f"Read YD/ZD for {len(yd_zd_all)} nodes x {len(floors)} floors from Filleddata")
    return floors, yd_zd_all


def calculate_footprints(columns, floors, yd_zd_all):
    """
    Calculate rectangular footprint (X1, X2, Y1, Y2) per column PER FLOOR.
    Each column gets one row per floor level.
    Returns list of dicts with footprint data.
    """
    results = []

    for col in columns:
        node_no = col["no"]
        if node_no not in yd_zd_all:
            print(f"  WARNING: Node {node_no} ({col['name']}) not found in Filleddata, skipping")
            continue

        yd_zd_list = yd_zd_all[node_no]
        ax = col["anchor_x"]
        ay = col["anchor_y"]
        beam_x = col["beam_x_loc"]
        beam_y = col["beam_y_loc"]
        orientation = col.get("orientation", "Vertical")

        for floor_idx, (y_val, floor_type) in enumerate(floors):
            if floor_idx >= len(yd_zd_list):
                break
            yd, zd = yd_zd_list[floor_idx]
            if yd is None or zd is None:
                continue  # skip floors with no YD/ZD data

            # Swap YD/ZD based on column orientation
            if orientation == "Horizontal":
                yd, zd = max(yd, zd), min(yd, zd)
            else:
                yd, zd = min(yd, zd), max(yd, zd)

            # X direction
            if beam_x == "Left":
                x1 = ax
                x2 = ax + yd / 1000
            elif beam_x == "Right":
                x1 = ax - yd / 1000
                x2 = ax
            elif beam_x == "Centre":
                opp_x = col["opposite_x"]
                if opp_x is None:
                    x1 = ax
                    x2 = ax
                else:
                    mid = (ax + opp_x) / 2
                    x1 = mid - yd / 2000
                    x2 = mid + yd / 2000
            else:
                x1 = ax
                x2 = ax

            # Y direction
            if beam_y == "Front":
                y1 = ay
                y2 = ay + zd / 1000
            elif beam_y == "Back":
                y1 = ay - zd / 1000
                y2 = ay
            elif beam_y == "Centre":
                opp_y = col["opposite_y_y"]
                if opp_y is None:
                    y1 = ay
                    y2 = ay
                else:
                    mid = (ay + opp_y) / 2
                    y1 = mid - zd / 2000
                    y2 = mid + zd / 2000
            else:
                y1 = ay
                y2 = ay

            results.append({
                "no": node_no,
                "name": col["name"],
                "floor_type": floor_type,
                "y_value": y_val,
                "location": col["location"],
                "anchor_x": ax,
                "anchor_y": ay,
                "beam_x_loc": beam_x,
                "beam_y_loc": beam_y,
                "orientation": orientation,
                "yd": yd,
                "zd": zd,
                "x1": round(x1, 4),
                "x2": round(x2, 4),
                "y1": round(y1, 4),
                "y2": round(y2, 4),
            })

    return results


def write_output(results):
    """Write footprint results to Excel."""
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outpath = os.path.join(STD_ANL_FOLDER, f"gridline_coordinates_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gridline coordinates"

    # Headers (Floor Type and Y added before Location)
    headers = [
        "No.", "Name", "Floor Type", "Y",
        "Location",
        "Anchor X (m)", "Anchor Y (m)",
        "Beam X loc", "Beam Y loc",
        "Orientation",
        "YD (mm)", "ZD (mm)",
        "X1 (m)", "X2 (m)", "Y1 (m)", "Y2 (m)",
    ]
    ws.append(headers)

    # Data
    for r in results:
        ws.append([
            r["no"], r["name"], r["floor_type"], r["y_value"],
            r["location"],
            r["anchor_x"], r["anchor_y"],
            r["beam_x_loc"], r["beam_y_loc"],
            r["orientation"],
            r["yd"], r["zd"],
            r["x1"], r["x2"], r["y1"], r["y2"],
        ])

    wb.save(outpath)
    print(f"\nOutput saved: {os.path.basename(outpath)}")
    print(f"  {len(results)} columns processed")
    return outpath


def main():
    print("=" * 60)
    print("  GRIDLINE CALCULATOR")
    print("=" * 60)
    print()

    # 1. Find and read inputs
    node_path = find_latest_node_coordinates()
    columns = read_node_coordinates(node_path)
    floors, yd_zd_all = read_filleddata_all_yd_zd()

    # 2. Calculate footprints (per column per floor)
    print("\nCalculating footprints...")
    results = calculate_footprints(columns, floors, yd_zd_all)

    # 3. Print summary
    print(f"\n{'No.':<5} {'Name':<6} {'Floor':<14} {'Y':<8} {'Location':<12} "
          f"{'YD':<6} {'ZD':<6} {'X1':<8} {'X2':<8} {'Y1':<8} {'Y2':<8}")
    print("-" * 110)
    for r in results:
        print(f"{r['no']:<5} {r['name']:<6} {r['floor_type']:<14} {r['y_value']:<8} {r['location']:<12} "
              f"{r['yd']:<6} {r['zd']:<6} {r['x1']:<8.4f} {r['x2']:<8.4f} {r['y1']:<8.4f} {r['y2']:<8.4f}")

    # 4. Write output
    write_output(results)


if __name__ == "__main__":
    main()
