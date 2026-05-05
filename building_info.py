r"""
Fill building info into Filleddata_output.xlsx and Global_assumptions.csv.

Usage:
  python building_info.py

Reads building info from:
  D:\JARVIS back up 16092025\JARVIS backup\Code - Full set\building info.xlsx

Updates:
  1. Filleddata_output.xlsx (Datadata2 sheet, cols AE-AF) — A-B pairs only
  2. Global_assumptions.csv — E-F pairs (Seismic status, Fck, Fy, Slab thickness)

Run node_coordinate_calculator.py first to create Filleddata_output.xlsx.
"""

import argparse
import csv
from pathlib import Path
from openpyxl import load_workbook
try:
    import win32com.client as win32
except ImportError:  # pragma: no cover - optional dependency on Windows
    win32 = None

BUILDING_INFO_PATH = Path(r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\building info.xlsx")
FILLED_PATH = Path(r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\Datadata.xlsx")
GLOBAL_CSV_PATH = Path(r"D:\JARVIS back up 16092025\JARVIS backup\Code - Full set\Global_assumptions.csv")
_UNSET = object()


class _ExcelCellAdapter:
    def __init__(self, worksheet, row, column):
        self._worksheet = worksheet
        self._row = row
        self._column = column

    @property
    def value(self):
        return self._worksheet.Cells(self._row, self._column).Value

    @value.setter
    def value(self, new_value):
        self._worksheet.Cells(self._row, self._column).Value = new_value


class _ExcelSheetAdapter:
    def __init__(self, worksheet):
        self._worksheet = worksheet

    @property
    def max_row(self):
        used = self._worksheet.UsedRange
        return used.Row + used.Rows.Count - 1

    @property
    def max_column(self):
        used = self._worksheet.UsedRange
        return used.Column + used.Columns.Count - 1

    def cell(self, row, column, value=_UNSET):
        cell = _ExcelCellAdapter(self._worksheet, row, column)
        if value is not _UNSET:
            cell.value = value
        return cell


class _ExcelWorkbookAdapter:
    def __init__(self, excel_app, workbook, opened_here, created_app):
        self._excel_app = excel_app
        self._workbook = workbook
        self._opened_here = opened_here
        self._created_app = created_app

    def __getitem__(self, sheet_name):
        return _ExcelSheetAdapter(self._workbook.Worksheets(sheet_name))

    def save(self, _path=None):
        self._workbook.Save()

    def close(self):
        if self._opened_here:
            self._workbook.Close(SaveChanges=True)
        if self._created_app and self._excel_app.Workbooks.Count == 0:
            self._excel_app.Quit()


def _open_filled_workbook():
    return load_workbook(FILLED_PATH)


def read_building_info():
    """Read building info. Returns (ab_pairs, ef_pairs) separately."""
    wb = load_workbook(BUILDING_INFO_PATH, data_only=True)
    ws = wb["building info"]
    ab_pairs = []
    ef_pairs = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is not None:
            ab_pairs.append((str(row[0]).strip(), row[1]))
        if len(row) > 4 and row[4] is not None:
            ef_pairs.append((str(row[4]).strip(), row[5]))
    wb.close()
    return ab_pairs, ef_pairs


def update_global_csv(ef_pairs):
    """Update matching keys in Global_assumptions.csv with E-F values."""
    ef_dict = {k: v for k, v in ef_pairs}

    rows = []
    with open(GLOBAL_CSV_PATH, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if row and row[0] in ef_dict:
                row[1] = ef_dict[row[0]]
            rows.append(row)

    with open(GLOBAL_CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _canonical_key(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.casefold()


def _find_param_col(ws, first_key):
    """Locate the Datadata parameter key column from the first known key."""
    first_key_norm = _canonical_key(first_key)
    if first_key_norm:
        for c in range(1, ws.max_column + 1):
            cell_key = _canonical_key(ws.cell(2, c).value)
            if cell_key == first_key_norm:
                return c
    return 31


def _index_existing_params(ws, param_col):
    """Map existing parameter keys to their row and track the last used row."""
    existing_rows = {}
    last_row = 1

    for r in range(2, ws.max_row + 1):
        key_val = ws.cell(r, param_col).value
        data_val = ws.cell(r, param_col + 1).value
        if key_val is not None or data_val is not None:
            last_row = r

        key_norm = _canonical_key(key_val)
        if key_norm and key_norm not in existing_rows:
            existing_rows[key_norm] = r

    return existing_rows, last_row


def main():
    if not FILLED_PATH.exists():
        print(f"Error: {FILLED_PATH} not found. Run node_coordinate_calculator.py first.")
        return

    print(f"Reading building info from: {BUILDING_INFO_PATH}")
    ab_pairs, ef_pairs = read_building_info()
    skip_datadata_keys = {"Front Balcony Width", "Back Balcony Width"}
    datadata_ab_pairs = [(key, value) for key, value in ab_pairs if key not in skip_datadata_keys]

    # 1. Write A-B pairs to Filleddata — find column dynamically
    wb = _open_filled_workbook()
    try:
        ws = wb["Datadata2"]

        first_key = datadata_ab_pairs[0][0] if datadata_ab_pairs else None
        param_col = _find_param_col(ws, first_key)
        if param_col == 31 and _canonical_key(first_key) != _canonical_key(ws.cell(2, 31).value):
            print(f"  WARNING: Could not find '{first_key}' in row 2, using default col AE")

        existing_rows, last_param_row = _index_existing_params(ws, param_col)
        updated_count = 0
        appended_count = 0

        for key, value in datadata_ab_pairs:
            key_norm = _canonical_key(key)
            row = existing_rows.get(key_norm)

            if row is None:
                last_param_row += 1
                row = last_param_row
                ws.cell(row, param_col, key)
                existing_rows[key_norm] = row
                appended_count += 1
            else:
                updated_count += 1

            ws.cell(row, param_col + 1, value)

        wb.save(FILLED_PATH)
        print(
            f"Updated {updated_count} params and appended {appended_count} params "
            f"to col {param_col}-{param_col+1} in Datadata"
        )
    finally:
        wb.close()

    # 2. Write E-F pairs to Global_assumptions.csv
    update_global_csv(ef_pairs)
    print(f"Updated {len(ef_pairs)} params in Global_assumptions.csv:")
    for k, v in ef_pairs:
        print(f"  {k} = {v}")

    print(f"\nDone!")


def _parse_args() -> argparse.Namespace:
    """Optional CLI overrides for the 3 path constants.

    Used by the web pipeline (server/pipelines.py:run_building_info_pipeline)
    so the same script works for both local-run (no args, hardcoded paths)
    and web-run (each path supplied explicitly from the upload folder).
    """
    parser = argparse.ArgumentParser(
        description="Fill building info into Datadata.xlsx + Global_assumptions.csv."
    )
    parser.add_argument(
        "--building-info", dest="building_info", type=Path,
        default=BUILDING_INFO_PATH,
        help="Path to building info.xlsx (input only).",
    )
    parser.add_argument(
        "--datadata", type=Path, default=FILLED_PATH,
        help="Path to Datadata.xlsx (read existing AE-AF keys + write merged).",
    )
    parser.add_argument(
        "--globals", dest="globals_csv", type=Path, default=GLOBAL_CSV_PATH,
        help="Path to Global_assumptions.csv (patch matching keys).",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    BUILDING_INFO_PATH = args.building_info
    FILLED_PATH = args.datadata
    GLOBAL_CSV_PATH = args.globals_csv
    main()
