"""
AutoCAD Anchor Workflow — Multi-Sheet Data Entry with Menu System
-----------------------------------------------------------------
Usage:
  python autocad_anchor_workflow.py

Run this script while AutoCAD is open. On startup:
1) Connect to AutoCAD, detect units, set OSMODE
2) Pick an origin point (bottom-left corner of building) → becomes (0,0)
3) Show menu of available sheets
4) User picks a sheet → runs that workflow → marks sheet done
5) Loop until "Done" or all sheets complete → export single Excel

Supported sheets:
  1) Column coordinates      (21 cols, 3 rows/point)
  2) Rectangle coordinates   (13 cols, 1 row/rect)
  3) Secondary beam coords   (12 cols, 3 rows/beam)
  4) Balcony coordinates     (13 cols, 2 fixed rows)
  5) Plot boundary Y coords  (3 cols, 2 fixed rows)
"""

from __future__ import annotations

import io
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    from pyautocad import Autocad, APoint
    import pywintypes
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "pyautocad is required. Install with: pip install pyautocad pywin32"
    ) from exc

DEFAULT_LEVEL_TOLERANCE_M = 0.15

# INSUNITS values and their conversion to meters
INSUNITS_TO_METERS = {
    0: None,       # Unitless - ask user
    1: 0.0254,     # Inches
    2: 0.3048,     # Feet
    3: 1609.344,   # Miles
    4: 0.001,      # Millimeters
    5: 0.01,       # Centimeters
    6: 1.0,        # Meters
    7: 1000.0,     # Kilometers
    8: 0.0000254,  # Microinches
    9: 0.0000001,  # Mils (1/1000 inch)
    10: 0.9144,    # Yards
    11: 1.0e-10,   # Angstroms
    12: 1.0e-9,    # Nanometers
    13: 1.0e-6,    # Microns
    14: 10.0,      # Decimeters
    15: 10.0,      # Decameters
    16: 100.0,     # Hectometers
    17: 1.0e15,    # Gigameters
    18: 1.496e11,  # Astronomical units
    19: 9.461e15,  # Light years
    20: 3.086e16,  # Parsecs
}

INSUNITS_NAMES = {
    0: "Unitless", 1: "Inches", 2: "Feet", 3: "Miles",
    4: "Millimeters", 5: "Centimeters", 6: "Meters", 7: "Kilometers",
    8: "Microinches", 9: "Mils", 10: "Yards", 11: "Angstroms",
    12: "Nanometers", 13: "Microns", 14: "Decimeters", 15: "Decameters",
    16: "Hectometers", 17: "Gigameters", 18: "Astronomical units",
    19: "Light years", 20: "Parsecs",
}

# ─── Keyword maps ───────────────────────────────────────────────────

# Column workflow
LOCATION_KEYWORDS = [
    "Corner", "Lift", "Staircase", "Entry",
    "BackEdge", "FrontEdge", "LeftEdge", "RightEdge",
    "Interior", "Shaft", "Mumty", "LiftStaircase",
]
LOCATION_MAP = {
    "Corner": "Corner", "Lift": "Lift", "Staircase": "Staircase",
    "Entry": "Entry", "BackEdge": "Back edge", "FrontEdge": "Front edge",
    "LeftEdge": "Left edge", "RightEdge": "Right edge",
    "Interior": "Interior", "Shaft": "Shaft", "Mumty": "Mumty",
    "LiftStaircase": "Lift+staircase",
}
ANCHOR_KEYWORDS = ["BackLeft", "BackRight", "FrontLeft", "FrontRight"]
ANCHOR_MAP = {
    "BackLeft": "Back left", "BackRight": "Back right",
    "FrontLeft": "Front left", "FrontRight": "Front right",
}
BEAM_X_KEYWORDS = ["Left", "Right", "Centre"]
BEAM_Y_KEYWORDS = ["Back", "Front", "Centre"]
ORIENTATION_KEYWORDS = ["Horizontal", "Vertical"]
# # ── Old floor names (kept for reference) ──
# FLOOR_NAMES = ["Stilt floor", "Typical floor", "Terrace"]
FLOOR_NAMES = ["Plinth", "Stilt roof", "Typical floor roof", "Terrace"]

# Rectangle workflow
RECT_LOCATION_KEYWORDS = ["Lift", "STaircase", "Entry", "SHaft", "Mumty"]
RECT_LOCATION_MAP = {
    "Lift": "Lift", "STaircase": "Staircase", "Entry": "Entry",
    "SHaft": "Shaft", "Mumty": "Mumty",
}

# Secondary beam workflow
BEAM_LOC_KEYWORDS = ["Left", "Right", "Back", "Front"]
PRESENT_KEYWORDS = ["Yes", "No"]


# ─── Utility helpers ────────────────────────────────────────────────

def get_drawing_scale(doc):
    """Detect drawing units from INSUNITS and return meters per CAD unit."""
    try:
        insunits = int(doc.GetVariable("INSUNITS"))
    except Exception:
        insunits = 0

    unit_name = INSUNITS_NAMES.get(insunits, f"Unknown({insunits})")
    scale = INSUNITS_TO_METERS.get(insunits)

    if scale is not None:
        print(f"Detected drawing units: {unit_name} (INSUNITS={insunits})")
        print(f"Scale: {scale} meters per CAD unit")
        return scale, unit_name

    print("\nDrawing units are set to 'Unitless'. Please select:")
    print("  1 = Inches")
    print("  2 = Feet")
    print("  4 = Millimeters")
    print("  5 = Centimeters")
    print("  6 = Meters")

    try:
        choice = int(doc.Utility.GetInteger(
            "\nEnter unit code (1=in, 2=ft, 4=mm, 5=cm, 6=m): "
        ))
    except Exception:
        choice = 2

    if choice not in INSUNITS_TO_METERS or INSUNITS_TO_METERS[choice] is None:
        choice = 2

    scale = INSUNITS_TO_METERS[choice]
    unit_name = INSUNITS_NAMES[choice]
    print(f"Using: {unit_name} ({scale} meters per CAD unit)")
    return scale, unit_name


def get_bounding_box_origin(points):
    """Find the bottom-left-most actual point using normalized L2 distance."""
    min_x = min(p[0] for p in points)
    min_y = min(p[1] for p in points)
    x_range = max(p[0] for p in points) - min_x or 1
    y_range = max(p[1] for p in points) - min_y or 1

    def corner_dist(p):
        nx = (p[0] - min_x) / x_range
        ny = (p[1] - min_y) / y_range
        return nx ** 2 + ny ** 2

    best = min(points, key=corner_dist)
    return (best[0], best[1])


def to_cartesian(points, origin):
    """Convert CAD world coords to local Cartesian using origin."""
    ox, oy = origin
    return [(px - ox, py - oy) for px, py in points]


def group_into_levels(cart_points, tolerance):
    """Group points into horizontal levels. Returns BFS sequence entries."""
    if not cart_points:
        return []

    indexed = sorted(enumerate(cart_points), key=lambda t: t[1][1])

    levels = []
    current_level = [indexed[0]]
    current_y_avg = indexed[0][1][1]

    for idx, pt in indexed[1:]:
        if abs(pt[1] - current_y_avg) <= tolerance:
            current_level.append((idx, pt))
            current_y_avg = sum(p[1] for _, p in current_level) / len(current_level)
        else:
            levels.append(current_level)
            current_level = [(idx, pt)]
            current_y_avg = pt[1]
    levels.append(current_level)

    result = []
    for lv_num, level in enumerate(levels):
        level.sort(key=lambda t: t[1][0])
        for orig_idx, (cx, cy) in level:
            result.append((orig_idx, cx, cy, lv_num))

    return result


def to_xy(point3d):
    return (float(point3d[0]), float(point3d[1]))


def _toggle_snap(doc, enable):
    """Toggle AutoCAD object snap. OSMODE 45 = on, 0 = off."""
    try:
        doc.SetVariable("OSMODE", 45 if enable else 0)
    except Exception:
        pass


def _ask_snap_and_set(doc):
    """Ask user if they want snapping, toggle accordingly. Returns True if snap on."""
    use_snap = prompt_keyword(doc, "Use snapping?", PRESENT_KEYWORDS)
    snap_on = use_snap == "Yes"
    _toggle_snap(doc, snap_on)
    return snap_on


def prompt_point(doc, prompt_text, base_point=None, max_retries=3, ask_snap=True):
    """Prompt for point with retry logic for COM errors."""
    if ask_snap:
        _ask_snap_and_set(doc)

    bp = APoint(base_point[0], base_point[1], 0) if base_point else APoint(0, 0, 0)
    for attempt in range(max_retries):
        try:
            time.sleep(0.1)
            p = doc.Utility.GetPoint(bp, f"\n{prompt_text}: ")
            # Restore snapping after pick
            _toggle_snap(doc, True)
            return to_xy(p)
        except Exception as e:
            error_msg = str(e).lower()
            if "keyword" in error_msg:
                _toggle_snap(doc, True)
                raise
            if attempt < max_retries - 1:
                print(f"Retrying... (attempt {attempt + 2}/{max_retries})")
                time.sleep(0.5)
            else:
                _toggle_snap(doc, True)
                raise RuntimeError(
                    f"Failed to get point after {max_retries} attempts. "
                    "Make sure AutoCAD is in focus and no dialogs are open."
                ) from e


def safe_prompt(doc, message, max_retries=3):
    """Safely send prompt to AutoCAD with retry logic."""
    for attempt in range(max_retries):
        try:
            time.sleep(0.05)
            doc.Utility.Prompt(f"\n{message}")
            return
        except Exception:
            if attempt < max_retries - 1:
                time.sleep(0.3)
            else:
                print(f"[AutoCAD] {message}")


def prompt_real(doc, prompt_text, default_value):
    """Prompt for real number with default value support."""
    try:
        value = float(doc.Utility.GetReal(
            f"\n{prompt_text} (or press Enter for {default_value}): "
        ))
        return value if value > 0 else default_value
    except Exception:
        safe_prompt(doc, f"Using default: {default_value}")
        return default_value


def prompt_integer(doc, prompt_text, default_value=None):
    """Prompt for integer with error handling."""
    try:
        value = int(doc.Utility.GetInteger(f"\n{prompt_text}: "))
        return value
    except Exception as e:
        if default_value is not None:
            safe_prompt(doc, f"Using default: {default_value}")
            return default_value
        else:
            raise RuntimeError(
                f"Invalid input. {prompt_text} requires a number."
            ) from e


def prompt_keyword(doc, prompt_text, keywords, display_map=None, max_retries=3):
    """Prompt user to select from keyword list in AutoCAD command line."""
    kw_string = " ".join(keywords)
    display_str = "/".join(keywords)
    for attempt in range(max_retries):
        try:
            time.sleep(0.1)
            doc.Utility.InitializeUserInput(0, kw_string)
            kw = doc.Utility.GetKeyword(f"\n{prompt_text} [{display_str}]: ")
            if display_map:
                return display_map.get(kw, kw)
            return kw
        except Exception as e:
            error_msg = str(e).lower()
            if "rejected" in error_msg or "busy" in error_msg:
                if attempt < max_retries - 1:
                    time.sleep(0.5)
                    continue
            raise


def _add_entity(func, *args, max_retries=3):
    """Call a model.Add* function with retry logic for COM busy errors."""
    for attempt in range(max_retries):
        try:
            time.sleep(0.05)
            return func(*args)
        except Exception:
            if attempt < max_retries - 1:
                time.sleep(0.5)
            else:
                raise


MARKER_LAYER = "ANCHOR_WORKFLOW_TEMP"


def _ensure_marker_layer(doc):
    """Create the temporary marker layer if it doesn't exist."""
    try:
        doc.Layers.Add(MARKER_LAYER)
    except Exception:
        pass  # already exists


def draw_marker(model, x, y, label, marker_size):
    """Draw an X marker with label at a point. Returns list of entities."""
    entities = []
    half = marker_size / 2

    line1 = _add_entity(
        model.AddLine,
        APoint(x - half, y - half, 0),
        APoint(x + half, y + half, 0)
    )
    line1.Layer = MARKER_LAYER
    entities.append(line1)

    line2 = _add_entity(
        model.AddLine,
        APoint(x - half, y + half, 0),
        APoint(x + half, y - half, 0)
    )
    line2.Layer = MARKER_LAYER
    entities.append(line2)

    text_h = marker_size * 0.8
    offset = marker_size * 0.6
    txt = _add_entity(
        model.AddText,
        str(label), APoint(x + offset, y + offset, 0), text_h
    )
    txt.Layer = MARKER_LAYER
    entities.append(txt)

    return entities


def _emergency_cleanup_markers(model):
    """Delete ALL entities on the marker layer — used on crash/error."""
    try:
        to_delete = []
        for ent in model:
            try:
                if ent.Layer == MARKER_LAYER:
                    to_delete.append(ent)
            except Exception:
                continue
        for ent in to_delete:
            try:
                ent.Delete()
            except Exception:
                pass
        print(f"Emergency cleanup: deleted {len(to_delete)} marker entities.")
    except Exception as e:
        print(f"Emergency cleanup failed: {e}")


def delete_markers(temp_entities):
    """Delete all temporary marker entities from the drawing."""
    for ent in temp_entities:
        for attempt in range(3):
            try:
                time.sleep(0.05)
                ent.Delete()
                break
            except Exception:
                if attempt < 2:
                    time.sleep(0.3)


def get_marker_size(doc):
    """Calculate appropriate marker size based on current view."""
    try:
        viewsize = float(doc.GetVariable("VIEWSIZE"))
        return viewsize * 0.015
    except Exception:
        return 1.0


def bfs_reorder(raw_coords, scale_m_per_unit, origin):
    """Apply BFS ordering to a list of raw (x, y) coords.

    Returns list of (bfs_number_1based, original_index).
    """
    cart_points = to_cartesian(raw_coords, origin)
    level_tol_units = DEFAULT_LEVEL_TOLERANCE_M / scale_m_per_unit
    ordered = group_into_levels(cart_points, tolerance=level_tol_units)
    # ordered: list of (orig_idx, cx, cy, lv_num)
    return [(bfs_num + 1, orig_idx) for bfs_num, (orig_idx, *_) in enumerate(ordered)]


# ─── Beam / Anchor verification ──────────────────────────────────────

# Anchor "Front left" → expected beam_x="Left", beam_y="Front"
# Map anchor string → (expected_x, expected_y)
_ANCHOR_EXPECTED = {
    "Front left":  ("Left",  "Front"),
    "Front right": ("Right", "Front"),
    "Back left":   ("Left",  "Back"),
    "Back right":  ("Right", "Back"),
}

# Flip helpers
_FLIP_X = {"Left": "Right", "Right": "Left"}
_FLIP_Y = {"Front": "Back", "Back": "Front"}
_FLIP_ANCHOR_X = {
    "Front left": "Front right", "Front right": "Front left",
    "Back left": "Back right",   "Back right": "Back left",
}
_FLIP_ANCHOR_Y = {
    "Front left": "Back left",   "Back left": "Front left",
    "Front right": "Back right", "Back right": "Front right",
}


def _verify_beam_anchor(doc, model, anchor, beam_x_loc, beam_y_loc,
                        location, point, pick_num, marker_size, markers):
    """Check beam_x/beam_y match anchor location. Prompt user to fix mismatches.

    Loops until both beam_x and beam_y match the anchor location.
    If the user chooses to change the anchor, they re-pick the point and
    re-select the location in AutoCAD.
    Returns (anchor, beam_x_loc, beam_y_loc, location, point, markers).
    """
    while True:
        expected_x, expected_y = _ANCHOR_EXPECTED[anchor]
        x_ok = beam_x_loc == "Centre" or beam_x_loc == expected_x
        y_ok = beam_y_loc == "Centre" or beam_y_loc == expected_y

        if x_ok and y_ok:
            break

        # Check beam_x first
        if not x_ok:
            safe_prompt(doc, f"MISMATCH: Anchor='{anchor}' expects Beam X='{expected_x}', "
                             f"but got '{beam_x_loc}'.")
            new_anchor = _FLIP_ANCHOR_X[anchor]
            safe_prompt(doc, f"  BeamX = Change Beam X to '{expected_x}'")
            safe_prompt(doc, f"  Anchor = Change Anchor to '{new_anchor}' (re-pick point)")
            choice = prompt_keyword(doc, "Fix mismatch", ["BeamX", "Anchor"])
            if choice == "BeamX":
                beam_x_loc = expected_x
                safe_prompt(doc, f"Beam X changed to '{beam_x_loc}'.")
            else:
                anchor = new_anchor
                safe_prompt(doc, f"Anchor changed to '{anchor}'. Re-pick the point.")
                delete_markers(markers)
                point = prompt_point(doc, f"Re-pick anchor point #{pick_num} ({anchor})")
                markers = draw_marker(model, point[0], point[1], pick_num, marker_size)
                safe_prompt(doc, f"Re-picked #{pick_num}: ({point[0]:.3f}, {point[1]:.3f})")
                location = prompt_keyword(doc, "Location", LOCATION_KEYWORDS, LOCATION_MAP)
            continue  # re-check everything from the top

        # Check beam_y
        if not y_ok:
            safe_prompt(doc, f"MISMATCH: Anchor='{anchor}' expects Beam Y='{expected_y}', "
                             f"but got '{beam_y_loc}'.")
            new_anchor = _FLIP_ANCHOR_Y[anchor]
            safe_prompt(doc, f"  BeamY = Change Beam Y to '{expected_y}'")
            safe_prompt(doc, f"  Anchor = Change Anchor to '{new_anchor}' (re-pick point)")
            choice = prompt_keyword(doc, "Fix mismatch", ["BeamY", "Anchor"])
            if choice == "BeamY":
                beam_y_loc = expected_y
                safe_prompt(doc, f"Beam Y changed to '{beam_y_loc}'.")
            else:
                anchor = new_anchor
                safe_prompt(doc, f"Anchor changed to '{anchor}'. Re-pick the point.")
                delete_markers(markers)
                point = prompt_point(doc, f"Re-pick anchor point #{pick_num} ({anchor})")
                markers = draw_marker(model, point[0], point[1], pick_num, marker_size)
                safe_prompt(doc, f"Re-picked #{pick_num}: ({point[0]:.3f}, {point[1]:.3f})")
                location = prompt_keyword(doc, "Location", LOCATION_KEYWORDS, LOCATION_MAP)
            continue  # re-check everything from the top

    return anchor, beam_x_loc, beam_y_loc, location, point, markers


# ─── Workflow 1: Column coordinates ─────────────────────────────────

def collect_column_point(doc, model, pick_num, marker_size):
    """Pick one column point and collect all its data interactively."""
    _ask_snap_and_set(doc)
    try:
        time.sleep(0.1)
        p = doc.Utility.GetPoint(
            APoint(0, 0, 0),
            f"\nPick anchor point #{pick_num} (or press ENTER to finish): ",
        )
        point = to_xy(p)
    except Exception as e:
        _toggle_snap(doc, True)
        error_msg = str(e).lower()
        if "keyword" in error_msg or "cancel" in error_msg or "escape" in error_msg:
            return None, []
        raise
    _toggle_snap(doc, True)

    markers = draw_marker(model, point[0], point[1], pick_num, marker_size)
    safe_prompt(doc, f"Picked #{pick_num}: ({point[0]:.3f}, {point[1]:.3f})")

    location = prompt_keyword(doc, "Location", LOCATION_KEYWORDS, LOCATION_MAP)
    anchor = prompt_keyword(doc, "Anchor location", ANCHOR_KEYWORDS, ANCHOR_MAP)

    # ── Auto-derive beam_x/beam_y from anchor for Corner, Lift, Edge ──
    # Anchor maps: FrontLeft→(Left,Front), FrontRight→(Right,Front), etc.
    _ANCHOR_TO_BEAM = {
        "Front left": ("Left", "Front"),
        "Front right": ("Right", "Front"),
        "Back left": ("Left", "Back"),
        "Back right": ("Right", "Back"),
    }

    beam_x_opp = None
    beam_y_opp = None
    wall_on_top = None  # None = not asked, True/False = user answered
    wall_on_top_x = None  # per-direction: None = not asked / not internal
    wall_on_top_y = None
    # Custom beam dims for internal beam when wall_on_top=No
    bwx_custom, bdx_custom = None, None
    bwy_custom, bdy_custom = None, None

    if location in ("Corner",):
        # ── Corner: auto-derive, no beam X/Y prompts, no wall question ──
        beam_x_loc, beam_y_loc = _ANCHOR_TO_BEAM[anchor]
        safe_prompt(doc, f"Auto-derived: Beam X={beam_x_loc}, Beam Y={beam_y_loc}")

    elif location in ("Lift", "Lift+staircase"):
        # ── Lift: auto-derive, all defaults, no wall question ──
        beam_x_loc, beam_y_loc = _ANCHOR_TO_BEAM[anchor]
        safe_prompt(doc, f"Auto-derived: Beam X={beam_x_loc}, Beam Y={beam_y_loc}")

    elif location in ("Left edge", "Right edge", "Front edge", "Back edge", "Edges"):
        # ── Edge: auto-derive edge direction, ask perpendicular direction ──
        # Edge direction is known from anchor; perpendicular is asked
        auto_beam_x, auto_beam_y = _ANCHOR_TO_BEAM[anchor]

        if location in ("Left edge", "Right edge"):
            # X direction auto-derived (Left or Right from anchor)
            beam_x_loc = auto_beam_x
            safe_prompt(doc, f"Auto-derived: Beam X={beam_x_loc}")
            # Y direction: ask user (perpendicular to edge)
            beam_y_loc = prompt_keyword(doc, "Beam Y width location", BEAM_Y_KEYWORDS)
            if beam_y_loc == "Centre":
                safe_prompt(doc, "Pick opposite point for Beam Y...")
                beam_y_opp = prompt_point(doc, "Pick opposite point (Beam Y)",
                                          base_point=point)
        else:
            # Front/Back edge: Y direction auto-derived
            beam_y_loc = auto_beam_y
            safe_prompt(doc, f"Auto-derived: Beam Y={beam_y_loc}")
            # X direction: ask user (perpendicular to edge)
            beam_x_loc = prompt_keyword(doc, "Beam X width location", BEAM_X_KEYWORDS)
            if beam_x_loc == "Centre":
                safe_prompt(doc, "Pick opposite point for Beam X...")
                beam_x_opp = prompt_point(doc, "Pick opposite point (Beam X)",
                                          base_point=point)

        # Ask "Wall on top?" for the internal beam only
        if location in ("Left edge", "Right edge"):
            # Y is internal beam
            wot = prompt_keyword(doc, "Internal beam (Y): wall on top?", ["Yes", "No"])
            wall_on_top = (wot == "Yes")
            if not wall_on_top:
                bwy_custom = int(prompt_real(doc, "Internal beam width Y (mm)", 230))
                bdy_custom = int(prompt_real(doc, "Internal beam depth Y (mm)", 225))
        else:
            # X is internal beam
            wot = prompt_keyword(doc, "Internal beam (X): wall on top?", ["Yes", "No"])
            wall_on_top = (wot == "Yes")
            if not wall_on_top:
                bwx_custom = int(prompt_real(doc, "Internal beam width X (mm)", 230))
                bdx_custom = int(prompt_real(doc, "Internal beam depth X (mm)", 225))

    else:
        # ── Interior / Staircase / Other: ask beam X/Y ──
        beam_x_loc = prompt_keyword(doc, "Beam X width location", BEAM_X_KEYWORDS)
        if beam_x_loc == "Centre":
            safe_prompt(doc, "Pick opposite point for Beam X...")
            beam_x_opp = prompt_point(doc, "Pick opposite point (Beam X)", base_point=point)

        beam_y_loc = prompt_keyword(doc, "Beam Y width location", BEAM_Y_KEYWORDS)
        if beam_y_loc == "Centre":
            safe_prompt(doc, "Pick opposite point for Beam Y...")
            beam_y_opp = prompt_point(doc, "Pick opposite point (Beam Y)", base_point=point)

        # ── Verify beam directions match anchor location ──
        anchor, beam_x_loc, beam_y_loc, location, point, markers = _verify_beam_anchor(
            doc, model, anchor, beam_x_loc, beam_y_loc,
            location, point, pick_num, marker_size, markers
        )

        # Ask "Wall on top?" per direction (X and Y separately)
        wot_x = prompt_keyword(doc, "Beam X: wall on top?", ["Yes", "No"])
        wall_on_top_x = (wot_x == "Yes")
        if not wall_on_top_x:
            bwx_custom = int(prompt_real(doc, "Beam width X (mm)", 230))
            bdx_custom = int(prompt_real(doc, "Beam depth X (mm)", 225))

        wot_y = prompt_keyword(doc, "Beam Y: wall on top?", ["Yes", "No"])
        wall_on_top_y = (wot_y == "Yes")
        if not wall_on_top_y:
            bwy_custom = int(prompt_real(doc, "Beam width Y (mm)", 230))
            bdy_custom = int(prompt_real(doc, "Beam depth Y (mm)", 225))

    # Column orientation
    orientation = prompt_keyword(doc, "Column orientation", ORIENTATION_KEYWORDS)

    # ── Beam depth/width/wt defaults based on location + wall_on_top ──

    def _get_column_wt(location, floor_name, wall_on_top, is_internal_dir):
        """Get wall thickness for a column beam based on location, floor, and wall status.
        is_internal_dir: True if this is the internal-pointing beam direction.
        For edge columns: along-edge direction (is_internal=False) always gets boundary wt,
        internal direction follows wall_on_top rules."""
        is_lift = location in ("Lift", "Lift+staircase")
        is_boundary_loc = location in ("Corner", "Left edge", "Right edge",
                                       "Front edge", "Back edge", "Edges")
        is_staircase = location in ("Staircase",)

        # For boundary/edge columns: along-edge beam = boundary wt,
        # internal beam = wall_on_top rules
        is_boundary_beam = is_boundary_loc and not is_internal_dir

        if floor_name not in ("Plinth", "Terrace"):
            # ── Stilt roof / Typical floor roof (same rules) ──
            if is_lift:
                return 230
            if is_internal_dir and wall_on_top is False:
                return 0
            return 115

        # ── Plinth / Terrace ──
        if is_lift:
            return 230
        elif is_boundary_beam or is_staircase:
            # Along-edge beam on boundary, or staircase → always 115
            return 115
        elif is_internal_dir and wall_on_top is True:
            return 0  # interior with wall on Plinth/Terrace = 0
        else:
            return 0

    # Determine which direction is internal for edges
    # Left/Right edge: edge runs vertically (Y along edge), X is NOT internal, Y IS internal
    # Front/Back edge: edge runs horizontally (X along edge), Y is NOT internal, X IS internal
    x_is_internal = location in ("Front edge", "Back edge")
    y_is_internal = location in ("Left edge", "Right edge")
    # For non-edge: both are internal
    if location not in ("Corner", "Lift", "Lift+staircase",
                        "Left edge", "Right edge", "Front edge", "Back edge", "Edges"):
        x_is_internal = True
        y_is_internal = True

    # Beam depth defaults
    if location in ("Corner", "Lift", "Lift+staircase"):
        def_bdx, def_bdy = 300, 300
    elif location in ("Left edge", "Right edge"):
        def_bdx, def_bdy = 300, 225  # X along edge=300, Y internal=225
    elif location in ("Front edge", "Back edge"):
        def_bdx, def_bdy = 225, 300  # X internal=225, Y along edge=300
    else:
        def_bdx, def_bdy = 225, 225

    # Resolve per-direction wall_on_top flags
    # For corner/lift: wall_on_top_x/y stay None (not asked)
    # For edges: wall_on_top is single value, assigned to internal direction
    # For interior/staircase: wall_on_top_x/y already set per-direction above
    if wall_on_top is not None and wall_on_top_x is None and wall_on_top_y is None:
        # Edge case: single wall_on_top, assign to internal direction
        if x_is_internal:
            wall_on_top_x = wall_on_top
        if y_is_internal:
            wall_on_top_y = wall_on_top

    # Override beam dims if no wall and custom values entered
    def_bwx, def_bwy = 230, 230
    if bwx_custom is not None:
        def_bwx = bwx_custom
        def_bdx = bdx_custom
    if bwy_custom is not None:
        def_bwy = bwy_custom
        def_bdy = bdy_custom

    floors = {}
    for floor_name in FLOOR_NAMES:
        bwx = def_bwx
        bdx = int(def_bdx)
        bwy = def_bwy
        bdy = int(def_bdy)
        wtx = _get_column_wt(location, floor_name, wall_on_top_x, x_is_internal)
        wty = _get_column_wt(location, floor_name, wall_on_top_y, y_is_internal)
        # # ── Original prompts (uncomment to restore interactive mode) ──
        # safe_prompt(doc, f"--- {floor_name} ---")
        # bwx = prompt_real(doc, f"[{floor_name}] Beam width X (mm)", 230)
        # bdx = prompt_real(doc, f"[{floor_name}] Beam depth X (mm)", def_bdx)
        # wtx = prompt_real(doc, f"[{floor_name}] Wall thickness X (mm)", 115)
        # bwy = prompt_real(doc, f"[{floor_name}] Beam width Y (mm)", 230)
        # bdy = prompt_real(doc, f"[{floor_name}] Beam depth Y (mm)", def_bdy)
        # wty = prompt_real(doc, f"[{floor_name}] Wall thickness Y (mm)", 115)
        floors[floor_name] = {
            "bwx": int(bwx), "bdx": int(bdx), "wtx": int(wtx),
            "bwy": int(bwy), "bdy": int(bdy), "wty": int(wty),
        }

    point_data = {
        "raw_xy": point,
        "location": location,
        "anchor": anchor,
        "beam_x_loc": beam_x_loc,
        "beam_x_opp": beam_x_opp,
        "beam_y_loc": beam_y_loc,
        "beam_y_opp": beam_y_opp,
        "orientation": orientation,
        "floors": floors,
    }
    return point_data, markers


# ─── Coordinate leveling ─────────────────────────────────────────────

LEVEL_TOLERANCE_M = 0.025       # 25 mm
VALUE_MATCH_TOLERANCE_M = 0.001  # 1 mm — sub-tolerance for "same value" counting


def _find_leveled_value(values_m, origin_val_m):
    """Given a list of coordinate values (in meters, origin-relative),
    return the leveled value using the rules:
      1. If any value is 0 (origin) → return 0
      2. Majority value wins
      3. Tie → average
    """
    if not values_m:
        return None

    # Rule 1: origin takes priority
    for v in values_m:
        if abs(v) < VALUE_MATCH_TOLERANCE_M:
            return 0.0

    # Count occurrences of each unique value (within 1mm sub-tolerance)
    unique_vals = []
    counts = []
    for v in values_m:
        matched = False
        for i, uv in enumerate(unique_vals):
            if abs(v - uv) < VALUE_MATCH_TOLERANCE_M:
                counts[i] += 1
                matched = True
                break
        if not matched:
            unique_vals.append(v)
            counts.append(1)

    # Rule 2: majority wins
    max_count = max(counts)
    winners = [(uv, c) for uv, c in zip(unique_vals, counts) if c == max_count]

    if len(winners) == 1:
        return winners[0][0]

    # Rule 3: tie → average of ALL values
    return sum(values_m) / len(values_m)


def _level_column_coordinates(doc, all_points, origin, scale_m_per_unit):
    """Level X and Y coordinates for columns on the same grid lines.

    Modifies all_points[i]["raw_xy"] in place.
    Skips Centre beam columns for the corresponding axis.
    """
    ox, oy = origin
    n = len(all_points)
    if n < 2:
        return

    # Convert raw coords to meters (origin-relative) for comparison
    coords_m = []
    for pd in all_points:
        rx, ry = pd["raw_xy"]
        coords_m.append(((rx - ox) * scale_m_per_unit,
                         (ry - oy) * scale_m_per_unit))

    any_change = False

    # ── Y leveling (horizontal beam lines) ──
    # Exclude columns with beam_y == Centre
    y_eligible = [i for i in range(n) if all_points[i]["beam_y_loc"] != "Centre"]
    y_used = set()

    for i in y_eligible:
        if i in y_used:
            continue
        # Find all eligible columns within tolerance of this Y
        group = [i]
        y_used.add(i)
        for j in y_eligible:
            if j in y_used:
                continue
            if abs(coords_m[j][1] - coords_m[i][1]) <= LEVEL_TOLERANCE_M:
                group.append(j)
                y_used.add(j)

        if len(group) < 2:
            continue

        y_vals = [coords_m[idx][1] for idx in group]
        leveled_y_m = _find_leveled_value(y_vals, 0.0)

        # Check if any actual change
        if all(abs(v - leveled_y_m) < VALUE_MATCH_TOLERANCE_M for v in y_vals):
            continue

        # Apply: convert leveled meter value back to raw CAD units
        leveled_y_raw = leveled_y_m / scale_m_per_unit + oy
        col_names = [f"C{idx + 1}" for idx in group]
        old_ys = [f"{v:.3f}" for v in y_vals]
        safe_prompt(doc, f"Y-leveled {', '.join(col_names)}: "
                         f"Y = {leveled_y_m:.3f}m (was {', '.join(old_ys)})")
        for idx in group:
            rx, ry = all_points[idx]["raw_xy"]
            all_points[idx]["raw_xy"] = (rx, leveled_y_raw)
            coords_m[idx] = (coords_m[idx][0], leveled_y_m)
        any_change = True

    # ── X leveling (vertical grid lines) ──
    # Exclude columns with beam_x == Centre
    x_eligible = [i for i in range(n) if all_points[i]["beam_x_loc"] != "Centre"]
    x_used = set()

    for i in x_eligible:
        if i in x_used:
            continue
        group = [i]
        x_used.add(i)
        for j in x_eligible:
            if j in x_used:
                continue
            if abs(coords_m[j][0] - coords_m[i][0]) <= LEVEL_TOLERANCE_M:
                group.append(j)
                x_used.add(j)

        if len(group) < 2:
            continue

        x_vals = [coords_m[idx][0] for idx in group]
        leveled_x_m = _find_leveled_value(x_vals, 0.0)

        if all(abs(v - leveled_x_m) < VALUE_MATCH_TOLERANCE_M for v in x_vals):
            continue

        leveled_x_raw = leveled_x_m / scale_m_per_unit + ox
        col_names = [f"C{idx + 1}" for idx in group]
        old_xs = [f"{v:.3f}" for v in x_vals]
        safe_prompt(doc, f"X-leveled {', '.join(col_names)}: "
                         f"X = {leveled_x_m:.3f}m (was {', '.join(old_xs)})")
        for idx in group:
            rx, ry = all_points[idx]["raw_xy"]
            all_points[idx]["raw_xy"] = (leveled_x_raw, ry)
            coords_m[idx] = (leveled_x_m, coords_m[idx][1])
        any_change = True

    if not any_change:
        safe_prompt(doc, "Coordinate leveling: no adjustments needed.")


def workflow_columns(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect column coordinate data. Returns list of rows for Excel."""
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, "COLUMN COORDINATES")
    safe_prompt(doc, "Pick column points. ENTER to finish.")
    safe_prompt(doc, "=" * 50)

    temp_entities = []
    all_points = []
    pick_num = 1

    while True:
        pd, markers = collect_column_point(doc, model, pick_num, marker_size)
        if pd is None:
            safe_prompt(doc, f"Finished picking. Total columns: {len(all_points)}")
            # Validate lift column count (must be 4)
            lift_count = sum(1 for p in all_points if p["location"] == "Lift")
            while lift_count != 4 and all_points:
                safe_prompt(doc, f"WARNING: Lift has {lift_count} columns (expected 4).")
                choice = prompt_keyword(doc, "Add more columns or Continue?",
                                        ["Add", "Continue"])
                if choice == "Continue":
                    break
                # Add more columns
                while True:
                    pd2, markers2 = collect_column_point(doc, model, pick_num, marker_size)
                    if pd2 is None:
                        break
                    all_points.append(pd2)
                    temp_entities.extend(markers2)
                    safe_prompt(doc, f"Column #{pick_num} saved. Pick next or ENTER to finish.")
                    pick_num += 1
                lift_count = sum(1 for p in all_points if p["location"] == "Lift")
            break
        all_points.append(pd)
        temp_entities.extend(markers)
        safe_prompt(doc, f"Column #{pick_num} saved. Pick next or ENTER to finish.")
        pick_num += 1

    if not all_points:
        delete_markers(temp_entities)
        safe_prompt(doc, "No columns picked. Skipping.")
        return None

    # Level coordinates (snap columns on same grid line)
    _level_column_coordinates(doc, all_points, origin, scale_m_per_unit)

    # BFS ordering
    raw_coords = [pd["raw_xy"] for pd in all_points]
    bfs_order = bfs_reorder(raw_coords, scale_m_per_unit, origin)

    # Delete pick-order markers, redraw with BFS numbers
    delete_markers(temp_entities)
    time.sleep(0.5)
    bfs_entities = []
    bfs_ordered = []
    for bfs_num, orig_idx in bfs_order:
        pd = all_points[orig_idx]
        bfs_ordered.append((bfs_num, pd))
        rx, ry = pd["raw_xy"]
        time.sleep(0.1)
        ents = draw_marker(model, rx, ry, bfs_num, marker_size)
        bfs_entities.extend(ents)

    safe_prompt(doc, f"BFS reordered {len(bfs_ordered)} columns.")
    time.sleep(1)
    delete_markers(bfs_entities)

    return {"type": "columns", "bfs_ordered": bfs_ordered}


# ─── Workflow 2: Rectangle coordinates ──────────────────────────────

def workflow_rectangles(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect rectangle data (Lift/Staircase/etc). Returns data for Excel."""
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, "RECTANGLE COORDINATES")
    safe_prompt(doc, "Pick 2 corners per rectangle. ENTER to finish.")
    safe_prompt(doc, "=" * 50)

    temp_entities = []
    all_rects = []
    pick_num = 1

    while True:
        # Pick corner 1
        _ask_snap_and_set(doc)
        try:
            time.sleep(0.1)
            p1 = doc.Utility.GetPoint(
                APoint(0, 0, 0),
                f"\nRectangle #{pick_num} — Pick corner 1 (or ENTER to finish): ",
            )
            corner1 = to_xy(p1)
        except Exception as e:
            _toggle_snap(doc, True)
            error_msg = str(e).lower()
            if "keyword" in error_msg or "cancel" in error_msg or "escape" in error_msg:
                safe_prompt(doc, f"Finished picking. Total rectangles: {len(all_rects)}")
                break
            raise
        _toggle_snap(doc, True)

        # Corner 1: Anchor → auto-derive beam X/Y
        anchor1 = prompt_keyword(doc, "Corner 1 — Anchor location", ANCHOR_KEYWORDS, ANCHOR_MAP)
        _ANCHOR_TO_BEAM_RECT = {
            "Front left": ("Left", "Front"),
            "Front right": ("Right", "Front"),
            "Back left": ("Left", "Back"),
            "Back right": ("Right", "Back"),
        }
        beam_x_loc1, beam_y_loc1 = _ANCHOR_TO_BEAM_RECT[anchor1]
        safe_prompt(doc, f"Corner 1: auto-derived Beam X={beam_x_loc1}, Beam Y={beam_y_loc1}")
        # # ── Old prompts (kept for reference) ──
        # beam_x_loc1 = prompt_keyword(doc, "Corner 1 — Beam X width location", BEAM_X_KEYWORDS)
        # beam_y_loc1 = prompt_keyword(doc, "Corner 1 — Beam Y width location", BEAM_Y_KEYWORDS)

        # Pick corner 2 with rubber-band from corner 1
        corner2 = prompt_point(
            doc, f"Rectangle #{pick_num} — Pick corner 2", base_point=corner1
        )

        # Corner 2: Anchor → auto-derive beam X/Y
        anchor2 = prompt_keyword(doc, "Corner 2 — Anchor location", ANCHOR_KEYWORDS, ANCHOR_MAP)
        beam_x_loc2, beam_y_loc2 = _ANCHOR_TO_BEAM_RECT[anchor2]
        safe_prompt(doc, f"Corner 2: auto-derived Beam X={beam_x_loc2}, Beam Y={beam_y_loc2}")
        # # ── Old prompts (kept for reference) ──
        # beam_x_loc2 = prompt_keyword(doc, "Corner 2 — Beam X width location", BEAM_X_KEYWORDS)
        # beam_y_loc2 = prompt_keyword(doc, "Corner 2 — Beam Y width location", BEAM_Y_KEYWORDS)

        # Location keyword
        location = prompt_keyword(doc, "Location", RECT_LOCATION_KEYWORDS, RECT_LOCATION_MAP)

        # Midpoint for marker
        mx = (corner1[0] + corner2[0]) / 2
        my = (corner1[1] + corner2[1]) / 2
        markers = draw_marker(model, mx, my, f"R{pick_num}", marker_size)
        temp_entities.extend(markers)

        all_rects.append({
            "corner1": corner1,
            "corner2": corner2,
            "location": location,
            "anchor1": anchor1,
            "beam_x_loc1": beam_x_loc1,
            "beam_y_loc1": beam_y_loc1,
            "anchor2": anchor2,
            "beam_x_loc2": beam_x_loc2,
            "beam_y_loc2": beam_y_loc2,
        })
        safe_prompt(doc, f"Rectangle #{pick_num} saved. Pick next or ENTER to finish.")
        pick_num += 1

    if not all_rects:
        delete_markers(temp_entities)
        safe_prompt(doc, "No rectangles picked. Skipping.")
        return None

    # BFS ordering on midpoints
    midpoints = [
        ((r["corner1"][0] + r["corner2"][0]) / 2,
         (r["corner1"][1] + r["corner2"][1]) / 2)
        for r in all_rects
    ]
    bfs_order = bfs_reorder(midpoints, scale_m_per_unit, origin)

    # Redraw markers with BFS numbers
    delete_markers(temp_entities)
    time.sleep(0.5)
    bfs_entities = []
    bfs_ordered = []
    for bfs_num, orig_idx in bfs_order:
        rect = all_rects[orig_idx]
        bfs_ordered.append((bfs_num, rect))
        mx = (rect["corner1"][0] + rect["corner2"][0]) / 2
        my = (rect["corner1"][1] + rect["corner2"][1]) / 2
        time.sleep(0.1)
        ents = draw_marker(model, mx, my, f"R{bfs_num}", marker_size)
        bfs_entities.extend(ents)

    safe_prompt(doc, f"BFS reordered {len(bfs_ordered)} rectangles.")
    time.sleep(1)
    delete_markers(bfs_entities)

    return {"type": "rectangles", "bfs_ordered": bfs_ordered}


# ─── Workflow 3: Secondary beam coordinates ──────────────────────────

PLINTH_FLOOR_NAMES = ["Plinth"]
NONPLINTH_FLOOR_NAMES = ["Stilt roof", "Typical floor roof", "Terrace"]


def _workflow_secondary_beams_core(doc, model, origin, scale_m_per_unit, marker_size,
                                   floor_list, beam_type, label):
    """Core secondary beam collection. floor_list controls which floors are asked."""
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, f"SECONDARY BEAM COORDINATES ({label})")
    safe_prompt(doc, "Pick 2 endpoints per beam. ENTER to finish.")
    safe_prompt(doc, "=" * 50)

    temp_entities = []
    all_beams = []
    pick_num = 1

    while True:
        # Pick endpoint 1
        _ask_snap_and_set(doc)
        try:
            time.sleep(0.1)
            p1 = doc.Utility.GetPoint(
                APoint(0, 0, 0),
                f"\nBeam #{pick_num} — Pick endpoint 1 (or ENTER to finish): ",
            )
            ep1 = to_xy(p1)
        except Exception as e:
            _toggle_snap(doc, True)
            error_msg = str(e).lower()
            if "keyword" in error_msg or "cancel" in error_msg or "escape" in error_msg:
                safe_prompt(doc, f"Finished picking. Total beams: {len(all_beams)}")
                break
            raise
        _toggle_snap(doc, True)

        # Pick endpoint 2 with rubber-band
        ep2 = prompt_point(
            doc, f"Beam #{pick_num} — Pick endpoint 2", base_point=ep1
        )

        # Beam location
        beam_loc = prompt_keyword(doc, "Beam location", BEAM_LOC_KEYWORDS)

        # # ── EP face prompts (commented out — removed per user request) ──
        # # Endpoint wall face — for perpendicular axis offset (wt/2 toward centerline)
        # if beam_loc in ("Left", "Right"):
        #     ep1_face = prompt_keyword(doc, "EP1 wall face (Y axis)",
        #                               ["Front", "Back", "None"])
        #     ep2_face = prompt_keyword(doc, "EP2 wall face (Y axis)",
        #                               ["Front", "Back", "None"])
        # else:
        #     ep1_face = prompt_keyword(doc, "EP1 wall face (X axis)",
        #                               ["Left", "Right", "None"])
        #     ep2_face = prompt_keyword(doc, "EP2 wall face (X axis)",
        #                               ["Left", "Right", "None"])
        # _EP_OFFSET_M = 115.0 / 1000.0 / 2.0
        # ep1 = list(ep1); ep2 = list(ep2)
        # ... offset logic ...
        # ep1 = tuple(ep1); ep2 = tuple(ep2)
        ep1_face = "None"
        ep2_face = "None"

        # Wall on top? — determines wt and beam dims
        wot = prompt_keyword(doc, "Wall on top?", ["Yes", "No"])
        sec_wall_on_top = (wot == "Yes")
        sec_bw = 230
        sec_bd = 225
        if not sec_wall_on_top:
            # No wall → wt=0, ask beam width and depth
            sec_bw = int(prompt_real(doc, "Beam width (mm)", 230))
            sec_bd = int(prompt_real(doc, "Beam depth (mm)", 225))

        # Per-floor: assign defaults based on wall_on_top
        floors = {}
        for floor_name in floor_list:
            if len(floor_list) == 1:
                # Single floor (e.g., Plinth) — always present, no need to ask
                present = "Yes"
            else:
                safe_prompt(doc, f"--- {floor_name} ---")
                present = prompt_keyword(doc, f"[{floor_name}] Present?", PRESENT_KEYWORDS)
            if present == "Yes":
                bw = sec_bw
                bd = sec_bd
                if floor_name == "Terrace":
                    wt = 0  # Terrace always wt=0 for secondary beams
                elif sec_wall_on_top:
                    wt = 115
                else:
                    wt = 0
                # # ── Original prompts (uncomment to restore interactive mode) ──
                # bw = prompt_real(doc, f"[{floor_name}] Beam width (mm)", 230)
                # bd = prompt_real(doc, f"[{floor_name}] Beam depth (mm)", 225)
                # wt = prompt_real(doc, f"[{floor_name}] Wall thickness (mm)", 115)
                floors[floor_name] = {
                    "present": "YES",
                    "beam_width": int(bw),
                    "beam_depth": int(bd),
                    "wall_thickness": int(wt),
                }
            else:
                floors[floor_name] = {
                    "present": "NO",
                    "beam_width": "",
                    "beam_depth": "",
                    "wall_thickness": "",
                }

        # Midpoint for marker
        mx = (ep1[0] + ep2[0]) / 2
        my = (ep1[1] + ep2[1]) / 2
        markers = draw_marker(model, mx, my, f"{beam_type}{pick_num}", marker_size)
        temp_entities.extend(markers)

        all_beams.append({
            "ep1": ep1,
            "ep2": ep2,
            "beam_loc": beam_loc,
            "ep1_face": ep1_face,
            "ep2_face": ep2_face,
            "floors": floors,
        })
        safe_prompt(doc, f"Beam #{pick_num} saved. Pick next or ENTER to finish.")
        pick_num += 1

    if not all_beams:
        delete_markers(temp_entities)
        safe_prompt(doc, "No beams picked. Skipping.")
        return None

    # BFS ordering on midpoints
    midpoints = [
        ((b["ep1"][0] + b["ep2"][0]) / 2,
         (b["ep1"][1] + b["ep2"][1]) / 2)
        for b in all_beams
    ]
    bfs_order = bfs_reorder(midpoints, scale_m_per_unit, origin)

    # Redraw markers with BFS numbers
    delete_markers(temp_entities)
    time.sleep(0.5)
    bfs_entities = []
    bfs_ordered = []
    for bfs_num, orig_idx in bfs_order:
        beam = all_beams[orig_idx]
        bfs_ordered.append((bfs_num, beam))
        mx = (beam["ep1"][0] + beam["ep2"][0]) / 2
        my = (beam["ep1"][1] + beam["ep2"][1]) / 2
        time.sleep(0.1)
        ents = draw_marker(model, mx, my, f"{beam_type}{bfs_num}", marker_size)
        bfs_entities.extend(ents)

    safe_prompt(doc, f"BFS reordered {len(bfs_ordered)} secondary beams ({label}).")
    time.sleep(1)
    delete_markers(bfs_entities)

    return {"type": beam_type, "bfs_ordered": bfs_ordered, "floor_list": floor_list}


def workflow_secondary_beams_plinth(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect secondary beam data for Plinth floor only."""
    return _workflow_secondary_beams_core(
        doc, model, origin, scale_m_per_unit, marker_size,
        PLINTH_FLOOR_NAMES, "SP", "Plinth")


def workflow_secondary_beams_nonplinth(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect secondary beam data for non-Plinth floors (Stilt roof, Typical, Terrace)."""
    return _workflow_secondary_beams_core(
        doc, model, origin, scale_m_per_unit, marker_size,
        NONPLINTH_FLOOR_NAMES, "SO", "Non-Plinth")


# # ── Old single secondary beam workflow (kept for reference) ──
# def workflow_secondary_beams(doc, model, origin, scale_m_per_unit, marker_size,
#                              all_results=None):
#     """Collect secondary beam data. Returns data for Excel."""
#     ...


# ─── Workflow 4: Extra wall coordinates ──────────────────────────────

EXTRAWALL_FLOOR_NAMES = ["Stilt roof", "Typical floor roof", "Terrace"]


def workflow_extra_walls(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect extra wall data. Same logic as secondary beams (non-plinth only)."""
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, "EXTRA WALL COORDINATES")
    safe_prompt(doc, "Pick 2 endpoints per wall. ENTER to finish.")
    safe_prompt(doc, "=" * 50)

    temp_entities = []
    all_walls = []
    pick_num = 1

    while True:
        # Pick endpoint 1
        _ask_snap_and_set(doc)
        try:
            time.sleep(0.1)
            p1 = doc.Utility.GetPoint(
                APoint(0, 0, 0),
                f"\nExtra wall #{pick_num} — Pick endpoint 1 (or ENTER to finish): ",
            )
            ep1 = to_xy(p1)
        except Exception as e:
            _toggle_snap(doc, True)
            error_msg = str(e).lower()
            if "keyword" in error_msg or "cancel" in error_msg or "escape" in error_msg:
                safe_prompt(doc, f"Finished picking. Total extra walls: {len(all_walls)}")
                break
            raise
        _toggle_snap(doc, True)

        # Pick endpoint 2 with rubber-band
        ep2 = prompt_point(
            doc, f"Extra wall #{pick_num} — Pick endpoint 2", base_point=ep1
        )

        # Wall location
        wall_loc = prompt_keyword(doc, "Wall location", BEAM_LOC_KEYWORDS)

        # Wall on top? — same logic as secondary beams
        wot = prompt_keyword(doc, "Wall on top?", ["Yes", "No"])
        ew_wall_on_top = (wot == "Yes")

        # Per-floor: Present? wt based on wall_on_top, Terrace always 0
        floors = {}
        for floor_name in EXTRAWALL_FLOOR_NAMES:
            safe_prompt(doc, f"--- {floor_name} ---")
            present = prompt_keyword(doc, f"[{floor_name}] Present?", PRESENT_KEYWORDS)
            if present == "Yes":
                if floor_name == "Terrace":
                    wt = 0  # Terrace always wt=0
                elif ew_wall_on_top:
                    wt = 115
                else:
                    wt = 0
                floors[floor_name] = {
                    "present": "YES",
                    "wall_thickness": int(wt),
                }
            else:
                floors[floor_name] = {
                    "present": "NO",
                    "wall_thickness": "",
                }

        # Midpoint for marker
        mx = (ep1[0] + ep2[0]) / 2
        my = (ep1[1] + ep2[1]) / 2
        markers = draw_marker(model, mx, my, f"EW{pick_num}", marker_size)
        temp_entities.extend(markers)

        all_walls.append({
            "ep1": ep1,
            "ep2": ep2,
            "wall_loc": wall_loc,
            "floors": floors,
        })
        safe_prompt(doc, f"Extra wall #{pick_num} saved. Pick next or ENTER to finish.")
        pick_num += 1

    if not all_walls:
        delete_markers(temp_entities)
        safe_prompt(doc, "No extra walls picked. Skipping.")
        return None

    # BFS ordering on midpoints
    midpoints = [
        ((w["ep1"][0] + w["ep2"][0]) / 2,
         (w["ep1"][1] + w["ep2"][1]) / 2)
        for w in all_walls
    ]
    bfs_order = bfs_reorder(midpoints, scale_m_per_unit, origin)

    # Redraw markers with BFS numbers
    delete_markers(temp_entities)
    time.sleep(0.5)
    bfs_entities = []
    bfs_ordered = []
    for bfs_num, orig_idx in bfs_order:
        wall = all_walls[orig_idx]
        bfs_ordered.append((bfs_num, wall))
        mx = (wall["ep1"][0] + wall["ep2"][0]) / 2
        my = (wall["ep1"][1] + wall["ep2"][1]) / 2
        time.sleep(0.1)
        ents = draw_marker(model, mx, my, f"EW{bfs_num}", marker_size)
        bfs_entities.extend(ents)

    safe_prompt(doc, f"BFS reordered {len(bfs_ordered)} extra walls.")
    time.sleep(1)
    delete_markers(bfs_entities)

    return {"type": "extra_walls", "bfs_ordered": bfs_ordered}


# ─── Workflow 5: Balcony coordinates ────────────────────────────────

def workflow_balconies(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect balcony data (front + back). Returns data for Excel."""
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, "BALCONY COORDINATES")
    safe_prompt(doc, "Pick 2 corners each for Front and Back balcony.")
    safe_prompt(doc, "=" * 50)

    balconies = []
    temp_entities = []

    for label in ["Front balcony", "Back balcony"]:
        safe_prompt(doc, f"--- {label} ---")
        corner1 = prompt_point(doc, f"{label} -- Pick corner 1")

        # Corner 1 properties
        safe_prompt(doc, f"--- {label} Corner 1 properties ---")
        anchor1 = prompt_keyword(doc, "Corner 1 -- Anchor location", ANCHOR_KEYWORDS, ANCHOR_MAP)
        # ── Phase 1: Use default (prompt commented out) ──
        wt1 = 115
        # # ── Original prompt (uncomment to restore interactive mode) ──
        # wt1 = prompt_real(doc, "Corner 1 -- Wall thickness (mm)", 115)
        beam_x1 = prompt_keyword(doc, "Corner 1 -- Beam X width location", BEAM_X_KEYWORDS)
        beam_y1 = prompt_keyword(doc, "Corner 1 -- Beam Y width location", BEAM_Y_KEYWORDS)

        corner2 = prompt_point(doc, f"{label} -- Pick corner 2", base_point=corner1)

        # Corner 2 properties
        safe_prompt(doc, f"--- {label} Corner 2 properties ---")
        anchor2 = prompt_keyword(doc, "Corner 2 -- Anchor location", ANCHOR_KEYWORDS, ANCHOR_MAP)
        # ── Phase 1: Use default (prompt commented out) ──
        wt2 = 115
        # # ── Original prompt (uncomment to restore interactive mode) ──
        # wt2 = prompt_real(doc, "Corner 2 -- Wall thickness (mm)", 115)
        beam_x2 = prompt_keyword(doc, "Corner 2 -- Beam X width location", BEAM_X_KEYWORDS)
        beam_y2 = prompt_keyword(doc, "Corner 2 -- Beam Y width location", BEAM_Y_KEYWORDS)

        mx = (corner1[0] + corner2[0]) / 2
        my = (corner1[1] + corner2[1]) / 2
        markers = draw_marker(model, mx, my, label[0], marker_size)
        temp_entities.extend(markers)

        balconies.append({
            "label": label,
            "corner1": corner1,
            "corner2": corner2,
            "anchor1": anchor1,
            "wt1": int(wt1),
            "beam_x1": beam_x1,
            "beam_y1": beam_y1,
            "anchor2": anchor2,
            "wt2": int(wt2),
            "beam_x2": beam_x2,
            "beam_y2": beam_y2,
        })

    time.sleep(1)
    delete_markers(temp_entities)
    safe_prompt(doc, "Balcony coordinates collected.")

    return {"type": "balconies", "data": balconies}


# ─── Workflow 5: Plot boundary Y coordinates ────────────────────────

def workflow_plot_boundary(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect plot boundary Y data (+Y and -Y). Returns data for Excel."""
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, "PLOT BOUNDARY Y COORDINATES")
    safe_prompt(doc, "Pick 1 point each for +Y and -Y boundaries.")
    safe_prompt(doc, "=" * 50)

    boundaries = []
    temp_entities = []

    for label in ["Plot boundary +Y coordinates", "Plot boundary -Y coordinates"]:
        safe_prompt(doc, f"--- {label} ---")
        pt = prompt_point(doc, f"{label} — Pick point")

        markers = draw_marker(model, pt[0], pt[1], label[-15:-13], marker_size)
        temp_entities.extend(markers)

        boundaries.append({
            "label": label,
            "pt": pt,
        })

    time.sleep(1)
    delete_markers(temp_entities)
    safe_prompt(doc, "Plot boundary coordinates collected.")

    return {"type": "plot_boundary", "data": boundaries}


# ─── Workflow 6: Staircase details ───────────────────────────────────

def workflow_staircase_details(doc, model, origin, scale_m_per_unit, marker_size):
    """Collect staircase details: Entry Landing, Mid Landing, Number of Staircases."""
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, "STAIRCASE DETAILS")
    safe_prompt(doc, "Pick 2 points each for Entry Landing and Mid Landing.")
    safe_prompt(doc, "=" * 50)

    temp_entities = []
    details = []

    for detail_name in ["Entry Landing", "Mid Landing"]:
        safe_prompt(doc, f"--- {detail_name} ---")

        # Pick point 1
        pt1 = prompt_point(doc, f"{detail_name} — Pick point 1")
        markers1 = draw_marker(model, pt1[0], pt1[1], f"{detail_name[:3]}1", marker_size)
        temp_entities.extend(markers1)

        # Pick point 2
        pt2 = prompt_point(doc, f"{detail_name} — Pick point 2", base_point=pt1)
        markers2 = draw_marker(model, pt2[0], pt2[1], f"{detail_name[:3]}2", marker_size)
        temp_entities.extend(markers2)

        details.append({
            "name": detail_name,
            "pt1": pt1,
            "pt2": pt2,
        })
        safe_prompt(doc, f"{detail_name} saved.")

    # Number of staircases (must enter a value — no default)
    safe_prompt(doc, "--- Number of Staircases ---")
    while True:
        try:
            time.sleep(0.1)
            doc.Utility.InitializeUserInput(1)  # 1 = disallow empty input
            num_stairs = int(doc.Utility.GetInteger(
                "\nEnter number of staircases: "))
            break
        except Exception:
            safe_prompt(doc, "Invalid input. Please enter a number.")
    safe_prompt(doc, f"Number of staircases: {num_stairs}")

    # Staircase direction
    stair_direction = prompt_keyword(doc, "Staircase direction",
                                     ["Xdirection", "Ydirection"])
    stair_direction = "X-direction" if stair_direction == "Xdirection" else "Y-direction"
    safe_prompt(doc, f"Staircase direction: {stair_direction}")

    time.sleep(1)
    delete_markers(temp_entities)
    safe_prompt(doc, "Staircase details collected.")

    return {"type": "staircase_details", "data": details,
            "num_staircases": num_stairs, "stair_direction": stair_direction}


# ─── Excel export ────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_header(ws, headers, col_widths):
    """Write header row with formatting."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        col_letter = chr(64 + col_idx) if col_idx <= 26 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = width


def _write_row(ws, row_num, values):
    """Write a data row with formatting."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER


def build_column_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Column coordinates' sheet."""
    headers = [
        "No.", "Type", "Coordinate X (m)", "Coordinate Y (m)",
        "Location", "Anchor location",
        "For column: Beam X width location",
        "If centre, then opposite coordinate X (m)",
        "If centre, then opposite coordinate Y (m)",
        "Floor", "Beam width X (mm)", "Beam depth X (mm)", "Wall thickness (mm)",
        "For column: Beam Y width location",
        "If centre, then opposite coordinate X (m)",
        "If centre, then opposite coordinate Y (m)",
        "Floor", "Beam width Y (mm)", "Beam depth Y (mm)", "Wall thickness (mm)",
        "Column orientation",
    ]
    widths = [6, 8, 20, 20, 16, 16, 28, 28, 28, 16, 18, 18, 18, 28, 28, 28, 16, 18, 18, 18, 20]
    _write_header(ws, headers, widths)

    # Color palette for alternating columns (base hues, 4 shades each dark→light)
    _COL_COLORS = [
        ["4472C4", "6A93D4", "91B4E4", "B8D5F4"],  # Blue
        ["70AD47", "8FC46A", "AEDB8D", "CDF2B0"],  # Green
        ["ED7D31", "F19B5E", "F5B98B", "F9D7B8"],  # Orange
        ["9B59B6", "B07CC8", "C59FDA", "DAC2EC"],  # Purple
        ["E74C3C", "ED7669", "F3A096", "F9CAC3"],  # Red
        ["1ABC9C", "4DCDB3", "80DECA", "B3EFE1"],  # Teal
        ["F1C40F", "F4D03F", "F7DC6F", "FAEA9F"],  # Yellow
        ["E67E22", "EC9B50", "F2B87E", "F8D5AC"],  # Dark Orange
    ]
    num_floors = len(FLOOR_NAMES)

    ox, oy = origin
    row = 2

    for bfs_num, pd in data["bfs_ordered"]:
        rx, ry = pd["raw_xy"]
        cx_m = round((rx - ox) * scale_m_per_unit, 3)
        cy_m = round((ry - oy) * scale_m_per_unit, 3)
        type_label = f"C{bfs_num}"

        opp_x_xm, opp_x_ym = "", ""
        if pd["beam_x_opp"] is not None:
            opp_x_xm = round((pd["beam_x_opp"][0] - ox) * scale_m_per_unit, 3)
            opp_x_ym = round((pd["beam_x_opp"][1] - oy) * scale_m_per_unit, 3)

        opp_y_xm, opp_y_ym = "", ""
        if pd["beam_y_opp"] is not None:
            opp_y_xm = round((pd["beam_y_opp"][0] - ox) * scale_m_per_unit, 3)
            opp_y_ym = round((pd["beam_y_opp"][1] - oy) * scale_m_per_unit, 3)

        # Pick color palette for this column (cycles through available colors)
        col_colors = _COL_COLORS[(bfs_num - 1) % len(_COL_COLORS)]

        for floor_idx, floor_name in enumerate(FLOOR_NAMES):
            fd = pd["floors"][floor_name]
            _write_row(ws, row, [
                bfs_num, type_label, cx_m, cy_m,
                pd["location"], pd["anchor"], pd["beam_x_loc"],
                opp_x_xm, opp_x_ym, floor_name,
                fd["bwx"], fd["bdx"], fd["wtx"],
                pd["beam_y_loc"], opp_y_xm, opp_y_ym,
                floor_name, fd["bwy"], fd["bdy"], fd["wty"],
                pd["orientation"],
            ])
            # Apply color shading: dark (floor 0) → light (floor 3)
            shade_idx = min(floor_idx, len(col_colors) - 1)
            fill = PatternFill("solid", fgColor=col_colors[shade_idx])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = fill
            row += 1


def build_rectangle_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Rectangle coordinates' sheet."""
    headers = [
        "No.", "Type",
        "Coordinate X1 (m)", "Coordinate Y1 (m)",
        "Coordinate X2 (m)", "Coordinate Y2 (m)",
        "Location",
        "Anchor location 1", "Beam X width location 1", "Beam Y width location 1",
        "Anchor location 2", "Beam X width location 2", "Beam Y width location 2",
    ]
    widths = [6, 8, 18, 18, 18, 18, 16, 16, 20, 20, 16, 20, 20]
    _write_header(ws, headers, widths)

    ox, oy = origin
    row = 2

    for bfs_num, rect in data["bfs_ordered"]:
        c1 = rect["corner1"]
        c2 = rect["corner2"]
        _write_row(ws, row, [
            bfs_num,
            f"R{bfs_num}",
            round((c1[0] - ox) * scale_m_per_unit, 3),
            round((c1[1] - oy) * scale_m_per_unit, 3),
            round((c2[0] - ox) * scale_m_per_unit, 3),
            round((c2[1] - oy) * scale_m_per_unit, 3),
            rect["location"],
            rect["anchor1"],
            rect["beam_x_loc1"],
            rect["beam_y_loc1"],
            rect["anchor2"],
            rect["beam_x_loc2"],
            rect["beam_y_loc2"],
        ])
        row += 1


def _build_secondary_beam_sheet_generic(ws, data, scale_m_per_unit, origin, prefix):
    """Populate a secondary beam coordinates sheet (generic for plinth/nonplinth)."""
    headers = [
        "No.", "Type",
        "Coordinate X1 (m)", "Coordinate Y1 (m)",
        "Coordinate X2 (m)", "Coordinate Y2 (m)",
        "Beam location", "Floor", "Present",
        "Beam width (mm)", "Beam depth (mm)", "Wall thickness (mm)",
    ]
    widths = [6, 8, 18, 18, 18, 18, 16, 16, 10, 18, 18, 18]
    _write_header(ws, headers, widths)

    ox, oy = origin
    row = 2
    floor_list = data.get("floor_list", FLOOR_NAMES)

    # Color palette for alternating beams (3 shades for nonplinth, 1 for plinth)
    _SEC_COLORS = [
        ["4472C4", "6A93D4", "91B4E4"],  # Blue
        ["70AD47", "8FC46A", "AEDB8D"],  # Green
        ["ED7D31", "F19B5E", "F5B98B"],  # Orange
        ["9B59B6", "B07CC8", "C59FDA"],  # Purple
        ["E74C3C", "ED7669", "F3A096"],  # Red
        ["1ABC9C", "4DCDB3", "80DECA"],  # Teal
        ["F1C40F", "F4D03F", "F7DC6F"],  # Yellow
        ["E67E22", "EC9B50", "F2B87E"],  # Dark Orange
    ]
    num_floors = len(floor_list)

    for bfs_num, beam in data["bfs_ordered"]:
        e1 = beam["ep1"]
        e2 = beam["ep2"]
        x1_m = round((e1[0] - ox) * scale_m_per_unit, 3)
        y1_m = round((e1[1] - oy) * scale_m_per_unit, 3)
        x2_m = round((e2[0] - ox) * scale_m_per_unit, 3)
        y2_m = round((e2[1] - oy) * scale_m_per_unit, 3)

        beam_colors = _SEC_COLORS[(bfs_num - 1) % len(_SEC_COLORS)]

        for floor_idx, floor_name in enumerate(floor_list):
            fd = beam["floors"][floor_name]
            _write_row(ws, row, [
                bfs_num,
                f"{prefix}{bfs_num}",
                x1_m, y1_m, x2_m, y2_m,
                beam["beam_loc"],
                floor_name,
                fd["present"],
                fd["beam_width"],
                fd["beam_depth"],
                fd["wall_thickness"],
            ])
            shade_idx = min(floor_idx, len(beam_colors) - 1)
            fill = PatternFill("solid", fgColor=beam_colors[shade_idx])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = fill
            row += 1


def build_secondary_beam_plinth_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Secondary beam coordinates_plinth' sheet."""
    _build_secondary_beam_sheet_generic(ws, data, scale_m_per_unit, origin, "SP")


def build_secondary_beam_nonplinth_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Secondary beam coordinates_nonplinth' sheet."""
    _build_secondary_beam_sheet_generic(ws, data, scale_m_per_unit, origin, "SO")


def build_extra_wall_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Extra wall coordinates' sheet."""
    headers = [
        "No.", "Type",
        "Coordinate X1 (m)", "Coordinate Y1 (m)",
        "Coordinate X2 (m)", "Coordinate Y2 (m)",
        "wall location", "Floor", "Present", "Wall thickness (mm)",
    ]
    widths = [6, 8, 18, 18, 18, 18, 16, 16, 10, 18]
    _write_header(ws, headers, widths)

    ox, oy = origin
    row = 2

    # Same color palette as column sheet
    _EW_COLORS = [
        ["4472C4", "6A93D4", "91B4E4"],  # Blue (3 floors)
        ["70AD47", "8FC46A", "AEDB8D"],  # Green
        ["ED7D31", "F19B5E", "F5B98B"],  # Orange
        ["9B59B6", "B07CC8", "C59FDA"],  # Purple
        ["E74C3C", "ED7669", "F3A096"],  # Red
        ["1ABC9C", "4DCDB3", "80DECA"],  # Teal
        ["F1C40F", "F4D03F", "F7DC6F"],  # Yellow
        ["E67E22", "EC9B50", "F2B87E"],  # Dark Orange
    ]

    for bfs_num, wall in data["bfs_ordered"]:
        e1 = wall["ep1"]
        e2 = wall["ep2"]
        x1_m = round((e1[0] - ox) * scale_m_per_unit, 3)
        y1_m = round((e1[1] - oy) * scale_m_per_unit, 3)
        x2_m = round((e2[0] - ox) * scale_m_per_unit, 3)
        y2_m = round((e2[1] - oy) * scale_m_per_unit, 3)

        ew_colors = _EW_COLORS[(bfs_num - 1) % len(_EW_COLORS)]

        for floor_idx, floor_name in enumerate(EXTRAWALL_FLOOR_NAMES):
            fd = wall["floors"][floor_name]
            _write_row(ws, row, [
                bfs_num,
                f"EW{bfs_num}",
                x1_m, y1_m, x2_m, y2_m,
                wall["wall_loc"],
                floor_name,
                fd["present"],
                fd["wall_thickness"],
            ])
            shade_idx = min(floor_idx, len(ew_colors) - 1)
            fill = PatternFill("solid", fgColor=ew_colors[shade_idx])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = fill
            row += 1


def build_balcony_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Balcony coordinates' sheet."""
    headers = [
        "Location",
        "Coordinate X1 (m)", "Coordinate Y1 (m)",
        "Anchor location 1", "Wall thickness 1 (mm)",
        "Beam X width location 1", "Beam Y width location 1",
        "Coordinate X2 (m)", "Coordinate Y2 (m)",
        "Anchor location 2", "Wall thickness 2 (mm)",
        "Beam X width location 2", "Beam Y width location 2",
    ]
    widths = [20, 18, 18, 16, 18, 20, 20, 18, 18, 16, 18, 20, 20]
    _write_header(ws, headers, widths)

    ox, oy = origin
    row = 2

    for bal in data["data"]:
        c1 = bal["corner1"]
        c2 = bal["corner2"]
        _write_row(ws, row, [
            bal["label"],
            round((c1[0] - ox) * scale_m_per_unit, 3),
            round((c1[1] - oy) * scale_m_per_unit, 3),
            bal["anchor1"],
            bal["wt1"],
            bal["beam_x1"],
            bal["beam_y1"],
            round((c2[0] - ox) * scale_m_per_unit, 3),
            round((c2[1] - oy) * scale_m_per_unit, 3),
            bal["anchor2"],
            bal["wt2"],
            bal["beam_x2"],
            bal["beam_y2"],
        ])
        row += 1


def build_plot_boundary_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Plot boundary Y coordinates' sheet."""
    headers = [
        "",
        "Coordinate X1 (m)", "Coordinate Y1 (m)",
    ]
    widths = [30, 18, 18]
    _write_header(ws, headers, widths)

    ox, oy = origin
    row = 2

    for bd in data["data"]:
        pt = bd["pt"]
        _write_row(ws, row, [
            bd["label"],
            round((pt[0] - ox) * scale_m_per_unit, 3),
            round((pt[1] - oy) * scale_m_per_unit, 3),
        ])
        row += 1


def build_staircase_details_sheet(ws, data, scale_m_per_unit, origin):
    """Populate 'Staircase details' sheet."""
    headers = [
        "Detail Name",
        "Coordinate X1 (m)", "Coordinate Y1 (m)",
        "Coordinate X2 (m)", "Coordinate Y2 (m)",
    ]
    widths = [22, 18, 18, 18, 18]
    _write_header(ws, headers, widths)

    ox, oy = origin
    row = 2

    for detail in data["data"]:
        pt1 = detail["pt1"]
        pt2 = detail["pt2"]
        _write_row(ws, row, [
            detail["name"],
            round((pt1[0] - ox) * scale_m_per_unit, 3),
            round((pt1[1] - oy) * scale_m_per_unit, 3),
            round((pt2[0] - ox) * scale_m_per_unit, 3),
            round((pt2[1] - oy) * scale_m_per_unit, 3),
        ])
        row += 1

    # 3 rows gap, then number of staircases and direction
    row += 3
    _write_row(ws, row, ["Number of Staircases", data["num_staircases"]])
    row += 1
    _write_row(ws, row, ["Staircase Direction", data.get("stair_direction", "")])


# Sheet name → builder function
SHEET_BUILDERS = {
    "columns": ("Column coordinates", build_column_sheet),
    "rectangles": ("Rectangle coordinates", build_rectangle_sheet),
    # # ── Old single secondary beam builder (kept for reference) ──
    # "secondary_beams": ("Secondary beam coordinates", build_secondary_beam_sheet),
    "SP": ("Secondary beam coordinates_plinth", build_secondary_beam_plinth_sheet),
    "SO": ("Secondary beam coordinates_nonplinth", build_secondary_beam_nonplinth_sheet),
    # Note: "SP"/"SO" are the type keys returned by the workflow functions,
    # while "SecPlinth"/"SecOther" are the menu keys for user selection.
    "extra_walls": ("Extra wall coordinates", build_extra_wall_sheet),
    "balconies": ("Balcony coordinates", build_balcony_sheet),
    "plot_boundary": ("Plot boundary Y coordinates", build_plot_boundary_sheet),
    "staircase_details": ("Staircase details", build_staircase_details_sheet),
}


def build_multi_sheet_excel(all_results, scale_m_per_unit, origin):
    """Build a single Excel workbook with one sheet per completed workflow."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for result in all_results:
        data_type = result["type"]
        sheet_name, builder_fn = SHEET_BUILDERS[data_type]
        ws = wb.create_sheet(title=sheet_name)
        builder_fn(ws, result, scale_m_per_unit, origin)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Menu & main ─────────────────────────────────────────────────────

SHEET_MENU = {
    "C": ("Column coordinates", workflow_columns),
    "R": ("Rectangle coordinates", workflow_rectangles),
    # # ── Old single secondary beam menu (kept for reference) ──
    # "S": ("Secondary beam coordinates", workflow_secondary_beams),
    "SecPlinth": ("Secondary beam coordinates (Plinth)", workflow_secondary_beams_plinth),
    "SecOther": ("Secondary beam coordinates (Non-Plinth)", workflow_secondary_beams_nonplinth),
    "E": ("Extra wall coordinates", workflow_extra_walls),
    "B": ("Balcony coordinates", workflow_balconies),
    "P": ("Plot boundary Y coordinates", workflow_plot_boundary),
    "StaircaseDet": ("Staircase details", workflow_staircase_details),
}

# Display names for keys shown in the prompt
_MENU_DISPLAY = {
    "C": "Column", "R": "Rectangle",
    "SecPlinth": "SecPlinth", "SecOther": "SecOther",
    "E": "Extra", "B": "Balcony", "P": "Plot",
    "StaircaseDet": "StaircaseDet",
}


def show_menu(doc, remaining_keys):
    """Display menu and get user choice via keyword prompt."""
    safe_prompt(doc, "")
    safe_prompt(doc, "=" * 50)
    safe_prompt(doc, "SELECT A SHEET TO FILL:")
    for key in sorted(remaining_keys):
        name = SHEET_MENU[key][0]
        safe_prompt(doc, f"  {key} = {name}")
    safe_prompt(doc, "  Done = Finish and export")
    safe_prompt(doc, "=" * 50)

    # Actual keywords for input: single letters + Done
    keywords = sorted(remaining_keys) + ["Done"]
    kw_string = " ".join(keywords)

    # Display string uses full words
    display_parts = [_MENU_DISPLAY[k] for k in sorted(remaining_keys)] + ["Done"]
    display_str = "/".join(display_parts)

    time.sleep(0.1)
    doc.Utility.InitializeUserInput(0, kw_string)
    try:
        kw = doc.Utility.GetKeyword(f"\nSelect sheet [{display_str}]: ")
    except Exception:
        return "Done"

    return kw


def main():
    print("Connecting to AutoCAD...")
    print("Make sure AutoCAD window is active and no dialogs are open.")

    try:
        acad = Autocad(create_if_not_exists=False)
    except Exception as e:
        raise SystemExit(f"Cannot connect to AutoCAD. Is it running? Error: {e}")

    doc = acad.doc
    time.sleep(0.3)

    safe_prompt(doc, "=" * 60)
    safe_prompt(doc, "Multi-Sheet AutoCAD Data Collection")
    safe_prompt(doc, "=" * 60)

    # Auto-detect drawing units
    scale_m_per_unit, unit_name = get_drawing_scale(doc)
    safe_prompt(doc, f"Drawing units: {unit_name} ({scale_m_per_unit:.6f} m per unit)")

    # Save and set OSMODE
    try:
        original_osmode = int(doc.GetVariable("OSMODE"))
    except Exception:
        original_osmode = 0
    try:
        doc.SetVariable("OSMODE", 45)
        safe_prompt(doc, "Object Snap: Endpoint, Center, Node, Intersection")
    except Exception:
        pass

    marker_size = get_marker_size(doc)
    model = acad.doc.ModelSpace
    _ensure_marker_layer(doc)

    # Pick origin point ONCE for all sheets
    safe_prompt(doc, "-" * 60)
    safe_prompt(doc, "STEP 1: Pick the ORIGIN point (bottom-left corner of building).")
    safe_prompt(doc, "This becomes (0, 0) for ALL coordinate sheets.")
    safe_prompt(doc, "-" * 60)

    origin = prompt_point(doc, "Pick ORIGIN point (bottom-left corner)", ask_snap=False)
    safe_prompt(doc, f"Origin set: ({origin[0]:.3f}, {origin[1]:.3f})")

    # Draw a small origin marker
    origin_markers = draw_marker(model, origin[0], origin[1], "O", marker_size)

    # Menu loop
    remaining = set(SHEET_MENU.keys())
    all_results = []

    try:
        while remaining:
            choice = show_menu(doc, remaining)

            if choice == "Done":
                break

            if choice not in remaining:
                safe_prompt(doc, f"Invalid choice: {choice}. Try again.")
                continue

            sheet_name, workflow_fn = SHEET_MENU[choice]
            safe_prompt(doc, f"Starting: {sheet_name}")

            result = workflow_fn(doc, model, origin, scale_m_per_unit, marker_size)
            if result is not None:
                all_results.append(result)
                remaining.discard(choice)
                safe_prompt(doc, f"Completed: {sheet_name}")
            else:
                safe_prompt(doc, f"Skipped: {sheet_name} (no data entered)")
                # Keep in menu so user can retry

    except Exception as e:
        # Emergency cleanup: delete ALL marker entities left in drawing
        safe_prompt(doc, f"Error occurred: {e}")
        safe_prompt(doc, "Cleaning up markers...")
        print(f"Error: {e}")
        _emergency_cleanup_markers(model)

    finally:
        # Always clean up origin marker and restore OSMODE
        delete_markers(origin_markers)
        try:
            doc.SetVariable("OSMODE", original_osmode)
        except Exception:
            pass

    if not all_results:
        safe_prompt(doc, "No data collected. Nothing to export.")
        print("No data collected. Nothing to export.")
        return

    # ── Export 3 separate Excel files ──
    # File 1: Column coordinates only
    # File 2: Secondary beam plinth only
    # File 3: Secondary beam non-plinth + all other sheets (Rectangle, Balcony, etc.)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model")

    # Split results into 3 groups
    column_results = [r for r in all_results if r["type"] == "columns"]
    plinth_results = [r for r in all_results if r["type"] == "SP"]
    nonplinth_results = [r for r in all_results
                         if r["type"] not in ("columns", "SP")]

    exported_files = []

    # File 1: Column coordinates
    if column_results:
        wb = Workbook()
        wb.remove(wb.active)
        for result in column_results:
            sheet_name, builder_fn = SHEET_BUILDERS[result["type"]]
            ws = wb.create_sheet(title=sheet_name)
            builder_fn(ws, result, scale_m_per_unit, origin)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        f1 = out_dir / f"floor_coordinates_column_coordinates_{ts}.xlsx"
        f1.write_bytes(buf.getvalue())
        exported_files.append(f1)
        print(f"Exported: {f1.name}")

    # File 2: Secondary beam plinth
    if plinth_results:
        wb = Workbook()
        wb.remove(wb.active)
        for result in plinth_results:
            sheet_name, builder_fn = SHEET_BUILDERS[result["type"]]
            ws = wb.create_sheet(title=sheet_name)
            builder_fn(ws, result, scale_m_per_unit, origin)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        f2 = out_dir / f"floor_coordinates_secondary_coordinates_plinth{ts}.xlsx"
        f2.write_bytes(buf.getvalue())
        exported_files.append(f2)
        print(f"Exported: {f2.name}")

    # File 3: Secondary beam non-plinth + Rectangle + Balcony + Extra walls + Plot + Staircase
    if nonplinth_results:
        wb = Workbook()
        wb.remove(wb.active)
        for result in nonplinth_results:
            sheet_name, builder_fn = SHEET_BUILDERS[result["type"]]
            ws = wb.create_sheet(title=sheet_name)
            builder_fn(ws, result, scale_m_per_unit, origin)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        f3 = out_dir / f"floor_coordinates_secondary_coordinates_nonplinth_{ts}.xlsx"
        f3.write_bytes(buf.getvalue())
        exported_files.append(f3)
        print(f"Exported: {f3.name}")

    # # ── Old single-file export (kept for reference) ──
    # out_bytes = build_multi_sheet_excel(all_results, scale_m_per_unit, origin)
    # out_file = out_dir / f"floor_coordinates_{ts}.xlsx"
    # out_file.write_bytes(out_bytes)

    try:
        doc.Regen(0)
    except Exception:
        pass

    safe_prompt(doc, f"Done. Exported {len(exported_files)} file(s).")
    for f in exported_files:
        safe_prompt(doc, f"  {f.name}")
    print(f"Done. Exported {len(exported_files)} file(s):")
    for f in exported_files:
        print(f"  {f}")


if __name__ == "__main__":
    main()
