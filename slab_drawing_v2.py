"""
Slab Drawing Generator V2
--------------------------
Generates ALL FLOOR LEVEL SLAB PLAN DXF with rebar details.

Inputs:
  Databeametabsupdated.xlsx         — beam segments
  Rectangle_geometry_filled.xlsx    — slab panel coordinates
  node_coordinates_*.xlsx           — grid X/Y values
  gridline_coordinates_*.xlsx       — per-floor column footprints
  other_coordinates_*.xlsx          — lift/staircase/shaft rectangles

Usage: python slab_drawing_v2.py
"""

import os
import sys
import glob
import openpyxl
import ezdxf
from ezdxf.enums import TextEntityAlignment
from datetime import datetime

# ── Paths ──────────────────────────────────────────────────────────────
DEFAULT_INPUT = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\beam_geometry.csv"
RECT_GEOM_PATH = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\Rectangle_geometry_filled.xlsx"
REBAR_INPUT_PATH = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\Slab drawing input details.xlsx"
STD_ANL_FOLDER = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model"
OUTPUT_FOLDER = os.path.join(STD_ANL_FOLDER, "final")

# Notes panel
NOTES_WIDTH = 5500

# ── Drawing constants (all in mm) ─────────────────────────────────────
SCALE = 1000  # metres to mm

# Grid
GRID_EXTEND_TOP = 5500   # vertical extension (X labels at top) — keep large
GRID_EXTEND_LEFT = 1500  # horizontal extension (Y labels at left) — compact
GRID_EXTEND = 5500       # for border/detail calculations (use max)
CIRCLE_RADIUS = 350
GRID_LABEL_HEIGHT = 250

# Dimensions
DIM_OFFSET_X1 = 2500     # X dims at top (individual)
DIM_OFFSET_X2 = 3500     # X dims at top (overall)
DIM_OFFSET_Y1 = 600      # Y dims at left (individual)
DIM_OFFSET_Y2 = 1100     # Y dims at left (overall)
DIM_TEXT_HEIGHT = 120
DIM_TICK_SIZE = 80

# Layout
FLOOR_GAP = 5000
MUMTY_GAP = 4000

# Title
TITLE_HEIGHT = 300
TITLE_GAP = 800

# Colors (AutoCAD Color Index)
CLR_WHITE = 7
CLR_RED = 1
CLR_YELLOW = 2
CLR_GREEN = 3
CLR_CYAN = 4
CLR_BLUE = 5
CLR_MAGENTA = 6
CLR_GREY = 253


# ══════════════════════════════════════════════════════════════════════
#  DATA READING
# ══════════════════════════════════════════════════════════════════════

def read_beam_data(filepath):
    """Read beam data from CSV (beam_geometry.csv). Returns list of beam dicts."""
    import csv
    beams = []
    with open(filepath, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row.get("member_no"):
                continue
            loc = row.get("Location", "")
            if loc in ("Front Balcony", "Back Balcony"):
                beam_type = "BALCONY_BEAM"
            elif row.get("span_role", "") == "secondary":
                beam_type = "SECONDARY_BEAM"
            else:
                beam_type = "PRIMARY_BEAM"
            beams.append({
                "member_no": int(row["member_no"]),
                "start_node": row["start_node_id"],
                "end_node": row["end_node_id"],
                "start_x": float(row["start_x"]),
                "start_z": float(row["start_z"]),
                "end_x": float(row["end_x"]),
                "end_z": float(row["end_z"]),
                "length": float(row["length_m"]),
                "beam_type": beam_type,
                "floor": row["Floor"],
                "yd": float(row["YD"] or 300),
                "zd": float(row["ZD"] or 230),
                "wt": float(row["Wall Thickness"] or 115),
                "plan_span_tag": row.get("plan_span_tag", ""),
            })
    return beams


def classify_beams(beams):
    """Classify beams as horizontal or vertical based on coordinates."""
    for b in beams:
        dz = abs(b["end_z"] - b["start_z"])
        dx = abs(b["end_x"] - b["start_x"])
        if dz < 0.001:
            b["direction"] = "horizontal"
        elif dx < 0.001:
            b["direction"] = "vertical"
        else:
            b["direction"] = "diagonal"
    return beams


def get_floor_beams(beams, floor_key):
    """Filter beams for a specific floor."""
    return [b for b in beams if b["floor"] == floor_key]


# ══════════════════════════════════════════════════════════════════════
#  OLD DATA READERS (for lift, staircase, mumty from other_coordinates)
# ══════════════════════════════════════════════════════════════════════

def find_latest_file(pattern):
    """Find the latest file matching a glob pattern in STD ANL folder."""
    files = glob.glob(os.path.join(STD_ANL_FOLDER, pattern))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def read_rectangles(filepath):
    """Read rectangle coordinates from other_coordinates Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    rects = []
    if "Rectangle coordinates" in wb.sheetnames:
        ws = wb["Rectangle coordinates"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rects.append({
                "no": row[0], "type": row[1],
                "x1": row[2], "y1": row[3], "x2": row[4], "y2": row[5],
                "location": row[6],
                "snapped_x1": row[13] if len(row) > 13 and row[13] else row[2],
                "snapped_y1": row[14] if len(row) > 14 and row[14] else row[3],
                "snapped_x2": row[15] if len(row) > 15 and row[15] else row[4],
                "snapped_y2": row[16] if len(row) > 16 and row[16] else row[5],
            })
    wb.close()
    return rects


GRID_SNAP_TOL = 300


def read_slab_panels(filepath):
    """Read slab panel geometry. Returns dict: floor_type -> list of panels."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Sheet1"]
    panels = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        floor_type = row[12]
        zone = row[11]
        if zone and str(zone).strip() in ("None", ""):
            zone = None
        panel = {
            "rect_no": row[1],
            "lx": row[2],
            "clear_lx": row[3],
            "lz": row[4],
            "clear_lz": row[5],
            "thickness": row[6],
            "x_start": row[7],
            "x_end": row[8],
            "z_start": row[9],
            "z_end": row[10],
            "zone": zone,
        }
        panels.setdefault(floor_type, []).append(panel)
    wb.close()
    # Deduplicate: keep only first occurrence of each rect_no per floor type
    deduped = {}
    for ft, pnls in panels.items():
        seen = set()
        unique = []
        for p in pnls:
            if p["rect_no"] not in seen:
                seen.add(p["rect_no"])
                unique.append(p)
        deduped[ft] = unique
    return deduped


def read_slab_panels_by_y(filepath):
    """Read slab panel geometry grouped by Y value. Returns dict: Y -> list of panels."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Sheet1"]
    panels = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        y_val = row[0]
        zone = row[11]
        if zone and str(zone).strip() in ("None", ""):
            zone = None
        panel = {
            "rect_no": row[1],
            "lx": row[2], "clear_lx": row[3],
            "lz": row[4], "clear_lz": row[5],
            "thickness": row[6],
            "x_start": row[7], "x_end": row[8],
            "z_start": row[9], "z_end": row[10],
            "zone": zone,
            "floor_type": row[12],
        }
        panels.setdefault(y_val, []).append(panel)
    wb.close()
    return panels


def read_rebar_input(filepath):
    """Read rebar specs. Returns dict: (Y_value, rect_no) -> rebar dict."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Reinforcement table"]
    rebar = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        key = (row[0], row[1])
        rebar[key] = {
            "hollow_zone": row[2],
            "slab_type": row[3],
            "tag_top": row[4],
            "tag_left": row[5],
            "tag_bot": row[6],
            "tag_right": row[7],
            "main_dir": row[9],
            "main_dia": row[10],
            "main_spacing": row[11],
            "cross_dir": row[13],
            "cross_dia": row[14],
            "cross_spacing": row[15],
        }
    wb.close()
    return rebar


def read_unique_grid_values(filepath):
    """Read node_coordinates and extract unique grid X/Y values in mm."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Node coordinates"]
    xs, ys = set(), set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        xs.add(round(row[21] * SCALE, 1))
        ys.add(round(row[22] * SCALE, 1))
    wb.close()
    return sorted(xs), sorted(ys)


def read_gridline_coordinates(filepath):
    """Read gridline_coordinates Excel (new format with Floor Type and Y).
    Returns dict: floor_type -> list of column dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    all_columns = {}  # floor_type -> [columns]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # New format: No, Name, FloorType, Y, Location, AnchorX, AnchorY,
        #             BeamXloc, BeamYloc, Orientation, YD, ZD, X1, X2, Y1, Y2
        floor_type = row[2]
        raw_x1 = row[12] * SCALE
        raw_x2 = row[13] * SCALE
        raw_y1 = row[14] * SCALE
        raw_y2 = row[15] * SCALE
        col = {
            "no": row[0], "name": row[1],
            "floor_type": floor_type, "y_value": row[3],
            "location": row[4],
            "anchor_x": row[5] * SCALE, "anchor_y": row[6] * SCALE,
            "beam_x_loc": row[7], "beam_y_loc": row[8],
            "orientation": row[9], "yd": row[10], "zd": row[11],
            "x1": min(raw_x1, raw_x2), "x2": max(raw_x1, raw_x2),
            "y1": min(raw_y1, raw_y2), "y2": max(raw_y1, raw_y2),
        }
        all_columns.setdefault(floor_type, []).append(col)
    wb.close()
    return all_columns


def get_columns_for_floor(all_columns, beam_floor):
    """Get columns for a specific beam floor.
    Maps beam floor names to gridline floor types."""
    floor_map = {
        "Stilt": "Stilt",
        "Typical": "1 Floor",
        "Terrace": "Terrace",
        "Mumty": "Mumty Roof",
    }
    floor_type = floor_map.get(beam_floor, beam_floor)
    cols = all_columns.get(floor_type, [])
    if not cols:
        # Fallback: try Stilt if specific floor not found
        cols = all_columns.get("Stilt", [])
    return cols


# ══════════════════════════════════════════════════════════════════════
#  DXF DOCUMENT SETUP
# ══════════════════════════════════════════════════════════════════════

def setup_document():
    """Create DXF document with layers and linetypes."""
    doc = ezdxf.new("R2010")

    style = doc.styles.get("Standard")
    style.dxf.font = "romans.shx"

    doc.linetypes.add("CENTER", pattern="A,1250,-250,250,-250",
                       description="Center ____ _ ____ _ ____")
    doc.linetypes.add("DASHED", pattern="A,200,-100",
                       description="Dashed ---- ----")

    doc.layers.add("S-GRID", color=CLR_RED, linetype="CENTER")
    doc.layers.add("S-COL.", color=CLR_GREEN)
    doc.layers.add("S-COL-HATCH", color=CLR_GREY)
    doc.layers.add("S-BEAM", color=CLR_CYAN, dxfattribs={"lineweight": 600})
    doc.layers.add("S-LINE", color=CLR_RED)
    doc.layers.add("S-DIM", color=CLR_BLUE, dxfattribs={"lineweight": 500})
    doc.layers.add("S-DIM-TEXT", color=CLR_CYAN)
    doc.layers.add("TEXT", color=CLR_GREEN)
    doc.layers.add("S-BEAM-NO", color=CLR_YELLOW, dxfattribs={"lineweight": 500})
    doc.layers.add("NOTES", color=CLR_WHITE)
    doc.layers.add("TITLE_BLOCK", color=CLR_WHITE)
    doc.layers.add("BORDER", color=CLR_MAGENTA)
    doc.layers.add("REBAR_BOT", color=CLR_YELLOW)
    doc.layers.add("REBAR_TOP", color=CLR_YELLOW, linetype="DASHED")
    doc.layers.add("REBAR_CRANK", color=CLR_YELLOW)
    doc.layers.add("REBAR_DIM", color=CLR_CYAN)

    return doc


# ══════════════════════════════════════════════════════════════════════
#  DIMENSION TEXT
# ══════════════════════════════════════════════════════════════════════

def _dim_text(value_mm):
    """Format dimension in mm as feet-inches with half-inch rounding."""
    total_inches = value_mm / 25.4
    half_inches = round(total_inches * 2)
    total_half = half_inches

    feet = int(total_half // 24)
    remaining_half = int(total_half % 24)
    whole_inches = remaining_half // 2
    has_half = remaining_half % 2 == 1

    if feet == 0:
        if has_half:
            if whole_inches == 0:
                return '%%189"'
            return f'{whole_inches}%%189"'
        else:
            return f'{whole_inches}"'
    else:
        if whole_inches == 0 and not has_half:
            return f"{feet}'-0\""
        elif has_half:
            if whole_inches == 0:
                return f"{feet}'-%%189\""
            return f"{feet}'-{whole_inches}%%189\""
        else:
            return f"{feet}'-{whole_inches}\""


# ══════════════════════════════════════════════════════════════════════
#  DRAWING FUNCTIONS
# ══════════════════════════════════════════════════════════════════════

def draw_grid_stubs(msp, grid_x, grid_y, ox, oy):
    """Draw grid line stubs from circles to building edge."""
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)
    extension = 400
    attribs = {"layer": "S-GRID", "ltscale": 50.0}

    # Vertical stubs (top, X labels) — use GRID_EXTEND_TOP
    top_offset = GRID_EXTEND_TOP + CIRCLE_RADIUS + 200
    for x in grid_x:
        msp.add_line((ox + x, oy + max_y - extension),
                     (ox + x, oy + max_y + top_offset - CIRCLE_RADIUS),
                     dxfattribs=attribs)

    # Horizontal stubs (left, Y labels) — use GRID_EXTEND_LEFT
    left_offset = GRID_EXTEND_LEFT + CIRCLE_RADIUS + 200
    for y in grid_y:
        msp.add_line((ox + min_x + extension, oy + y),
                     (ox + min_x - left_offset + CIRCLE_RADIUS, oy + y),
                     dxfattribs=attribs)


def draw_grid_labels(msp, grid_x, grid_y, ox, oy):
    """Draw grid labels: numbers at left (Y), letters at top (X)."""
    min_x = min(grid_x)
    max_y = max(grid_y)
    left_label_offset = GRID_EXTEND_LEFT + CIRCLE_RADIUS + 200
    top_label_offset = GRID_EXTEND_TOP + CIRCLE_RADIUS + 200

    # Y labels (1, 2, 3...) at left
    min_gap = CIRCLE_RADIUS * 2
    stagger_offset = 800
    y_staggered = set()
    for i in range(len(grid_y) - 1):
        if grid_y[i + 1] - grid_y[i] < min_gap:
            y_staggered.add(i + 1)

    for i, y in enumerate(grid_y):
        label = str(i + 1)
        base_lx = ox + min_x - left_label_offset
        ly = oy + y

        if i in y_staggered:
            lx = base_lx - stagger_offset
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_line((lx + CIRCLE_RADIUS, ly),
                         (base_lx - CIRCLE_RADIUS, ly),
                         dxfattribs={"layer": "S-GRID", "ltscale": 50.0})
        else:
            lx = base_lx
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)

    # X labels (A, B, C...) at top
    x_staggered = set()
    for i in range(len(grid_x) - 1):
        if grid_x[i + 1] - grid_x[i] < min_gap:
            x_staggered.add(i + 1)

    for i, x in enumerate(grid_x):
        # Support beyond Z (AA, AB, etc.)
        if i < 26:
            label = chr(65 + i)
        else:
            label = chr(65 + (i // 26) - 1) + chr(65 + (i % 26))

        lx = ox + x
        base_ly = oy + max_y + top_label_offset

        if i in x_staggered:
            ly = base_ly + stagger_offset
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)
            msp.add_line((lx, ly - CIRCLE_RADIUS),
                         (lx, base_ly + CIRCLE_RADIUS),
                         dxfattribs={"layer": "S-GRID", "ltscale": 50.0})
        else:
            ly = base_ly
            msp.add_circle((lx, ly), CIRCLE_RADIUS,
                           dxfattribs={"layer": "S-LINE"})
            msp.add_text(label, height=GRID_LABEL_HEIGHT,
                         dxfattribs={"layer": "TEXT"}).set_placement(
                (lx, ly), align=TextEntityAlignment.MIDDLE_CENTER)


def draw_beams(msp, floor_beams, columns, ox, oy):
    """Draw each beam as a pair of parallel lines showing beam width (ZD).
    At terminus endpoints (where no other beam of the same direction continues),
    retract inward by half_zd so edge lines meet perpendicular beam edges."""

    # Count how many H and V beams meet at each node
    node_counts = {}
    for b in floor_beams:
        sx = round(b["start_x"] * SCALE, 1)
        sy = round(b["start_z"] * SCALE, 1)
        ex = round(b["end_x"] * SCALE, 1)
        ey = round(b["end_z"] * SCALE, 1)
        for nx, ny in [(sx, sy), (ex, ey)]:
            if (nx, ny) not in node_counts:
                node_counts[(nx, ny)] = {"h": 0, "v": 0}
            if b["direction"] == "horizontal":
                node_counts[(nx, ny)]["h"] += 1
            elif b["direction"] == "vertical":
                node_counts[(nx, ny)]["v"] += 1

    retracted = 0
    for b in floor_beams:
        sx = b["start_x"] * SCALE
        sy = b["start_z"] * SCALE
        ex = b["end_x"] * SCALE
        ey = b["end_z"] * SCALE
        half_zd = b["zd"] / 2

        if b["direction"] == "horizontal":
            left_x = min(sx, ex)
            right_x = max(sx, ex)
            left_node = (round(left_x, 1), round(sy, 1))
            right_node = (round(right_x, 1), round(sy, 1))

            # Retract left end only if this is a terminus (no other H beam continues left)
            if node_counts.get(left_node, {}).get("h", 0) <= 1:
                left_x += half_zd
                retracted += 1
            # Retract right end only if this is a terminus
            if node_counts.get(right_node, {}).get("h", 0) <= 1:
                right_x -= half_zd
                retracted += 1

            msp.add_line((ox + left_x, oy + sy - half_zd),
                         (ox + right_x, oy + sy - half_zd),
                         dxfattribs={"layer": "S-BEAM"})
            msp.add_line((ox + left_x, oy + sy + half_zd),
                         (ox + right_x, oy + sy + half_zd),
                         dxfattribs={"layer": "S-BEAM"})
            # Close balcony beam ends
            if b["beam_type"] == "BALCONY_BEAM":
                msp.add_line((ox + left_x, oy + sy - half_zd),
                             (ox + left_x, oy + sy + half_zd),
                             dxfattribs={"layer": "S-BEAM"})
                msp.add_line((ox + right_x, oy + sy - half_zd),
                             (ox + right_x, oy + sy + half_zd),
                             dxfattribs={"layer": "S-BEAM"})

        elif b["direction"] == "vertical":
            bot_y = min(sy, ey)
            top_y = max(sy, ey)
            bot_node = (round(sx, 1), round(bot_y, 1))
            top_node = (round(sx, 1), round(top_y, 1))

            # Retract bottom only if terminus
            if node_counts.get(bot_node, {}).get("v", 0) <= 1:
                bot_y += half_zd
                retracted += 1
            # Retract top only if terminus
            if node_counts.get(top_node, {}).get("v", 0) <= 1:
                top_y -= half_zd
                retracted += 1

            msp.add_line((ox + sx - half_zd, oy + bot_y),
                         (ox + ex - half_zd, oy + top_y),
                         dxfattribs={"layer": "S-BEAM"})
            msp.add_line((ox + sx + half_zd, oy + bot_y),
                         (ox + ex + half_zd, oy + top_y),
                         dxfattribs={"layer": "S-BEAM"})
            # Close balcony beam ends
            if b["beam_type"] == "BALCONY_BEAM":
                msp.add_line((ox + sx - half_zd, oy + bot_y),
                             (ox + sx + half_zd, oy + bot_y),
                             dxfattribs={"layer": "S-BEAM"})
                msp.add_line((ox + ex - half_zd, oy + top_y),
                             (ox + ex + half_zd, oy + top_y),
                             dxfattribs={"layer": "S-BEAM"})

        elif b["direction"] == "diagonal":
            msp.add_line((ox + sx, oy + sy),
                         (ox + ex, oy + ey),
                         dxfattribs={"layer": "S-BEAM"})

    print(f"      ({retracted} terminus endpoints retracted)")

    # ── Balcony closing lines (outer edge connecting all balcony beams) ──
    # Balcony beams are vertical — they run perpendicular to the building edge.
    # The closing line is horizontal, connecting the outer tips of all front/back balcony beams.
    col_y_min = min(c["y1"] for c in columns)
    col_y_max = max(c["y2"] for c in columns)
    front_bal_beams = [b for b in floor_beams if b["beam_type"] == "BALCONY_BEAM"
                       and min(b["start_z"], b["end_z"]) * SCALE < col_y_min - 100]
    back_bal_beams = [b for b in floor_beams if b["beam_type"] == "BALCONY_BEAM"
                      and max(b["start_z"], b["end_z"]) * SCALE > col_y_max + 100]

    for bal_group, is_front in [(front_bal_beams, True), (back_bal_beams, False)]:
        if len(bal_group) < 2:
            continue
        half_zd = bal_group[0]["zd"] / 2
        if is_front:
            # Outer Y = min of all Z values, retracted inward by half_zd (terminus retraction)
            outer_y = min(min(b["start_z"], b["end_z"]) * SCALE for b in bal_group) + half_zd
        else:
            # Outer Y = max of all Z values, retracted inward by half_zd (terminus retraction)
            outer_y = max(max(b["start_z"], b["end_z"]) * SCALE for b in bal_group) - half_zd
        # X range: leftmost to rightmost beam (using X ± half_zd for beam width)
        all_left = []
        all_right = []
        for b in bal_group:
            sx = b["start_x"] * SCALE
            all_left.append(sx - half_zd)
            all_right.append(sx + half_zd)
        left_x = min(all_left)
        right_x = max(all_right)
        msp.add_line((ox + left_x, oy + outer_y),
                     (ox + right_x, oy + outer_y),
                     dxfattribs={"layer": "S-BEAM"})
        side = "Front" if is_front else "Back"
        print(f"      {side} balcony closing line: X={left_x:.0f}-{right_x:.0f}, Y={outer_y:.0f}")


def draw_columns(msp, columns, ox, oy):
    """Draw column rectangles with opaque fill + green outline.
    Drawn LAST so opaque fill hides beam overlaps at junctions."""
    for col in columns:
        x1, y1 = ox + col["x1"], oy + col["y1"]
        x2, y2 = ox + col["x2"], oy + col["y2"]

        # Opaque 2D SOLID fill
        msp.add_solid(
            [(x1, y1), (x2, y1), (x1, y2), (x2, y2)],
            dxfattribs={"layer": "S-COL-HATCH", "color": CLR_GREY})

        # Green outline
        msp.add_lwpolyline(
            [(x1, y1), (x2, y1), (x2, y2), (x1, y2)],
            dxfattribs={"layer": "S-COL."}, close=True)


def draw_dimensions(msp, grid_x, grid_y, ox, oy):
    """Draw grid spacing dimensions at top (X) and left (Y)."""
    min_x, max_x = min(grid_x), max(grid_x)
    min_y, max_y = min(grid_y), max(grid_y)
    tick = DIM_TICK_SIZE
    th = DIM_TEXT_HEIGHT

    # X dims at top
    for i in range(len(grid_x) - 1):
        x1, x2 = grid_x[i], grid_x[i + 1]
        y_dim = max_y + DIM_OFFSET_X1
        msp.add_line((ox + x1, oy + y_dim), (ox + x2, oy + y_dim),
                     dxfattribs={"layer": "S-DIM"})
        for tx in [x1, x2]:
            msp.add_line((ox + tx - tick * 0.5, oy + y_dim - tick * 0.5),
                         (ox + tx + tick * 0.5, oy + y_dim + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + tx, oy + max_y + 200),
                         (ox + tx, oy + y_dim - tick),
                         dxfattribs={"layer": "S-LINE"})
        dist = x2 - x1
        msp.add_text(_dim_text(dist), height=th,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + (x1 + x2) / 2, oy + y_dim + th * 0.5),
            align=TextEntityAlignment.BOTTOM_CENTER)

    # X overall
    if len(grid_x) > 2:
        y_dim = max_y + DIM_OFFSET_X2
        msp.add_line((ox + grid_x[0], oy + y_dim),
                     (ox + grid_x[-1], oy + y_dim),
                     dxfattribs={"layer": "S-DIM"})
        for tx in [grid_x[0], grid_x[-1]]:
            msp.add_line((ox + tx - tick * 0.5, oy + y_dim - tick * 0.5),
                         (ox + tx + tick * 0.5, oy + y_dim + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + tx, oy + max_y + 200),
                         (ox + tx, oy + y_dim - tick),
                         dxfattribs={"layer": "S-LINE"})
        dist = grid_x[-1] - grid_x[0]
        msp.add_text(_dim_text(dist), height=th,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + (grid_x[0] + grid_x[-1]) / 2, oy + y_dim + th * 0.5),
            align=TextEntityAlignment.BOTTOM_CENTER)

    # Y dims at left
    for i in range(len(grid_y) - 1):
        y1, y2 = grid_y[i], grid_y[i + 1]
        x_dim = min_x - DIM_OFFSET_Y1
        msp.add_line((ox + x_dim, oy + y1), (ox + x_dim, oy + y2),
                     dxfattribs={"layer": "S-DIM"})
        for ty in [y1, y2]:
            msp.add_line((ox + x_dim - tick * 0.5, oy + ty - tick * 0.5),
                         (ox + x_dim + tick * 0.5, oy + ty + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + min_x - 200, oy + ty),
                         (ox + x_dim - tick, oy + ty),
                         dxfattribs={"layer": "S-LINE"})
        dist = y2 - y1
        msp.add_text(_dim_text(dist), height=th, rotation=90,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + x_dim - th * 0.5, oy + (y1 + y2) / 2),
            align=TextEntityAlignment.BOTTOM_CENTER)

    # Y overall
    if len(grid_y) > 2:
        x_dim = min_x - DIM_OFFSET_Y2
        msp.add_line((ox + x_dim, oy + grid_y[0]),
                     (ox + x_dim, oy + grid_y[-1]),
                     dxfattribs={"layer": "S-DIM"})
        for ty in [grid_y[0], grid_y[-1]]:
            msp.add_line((ox + x_dim - tick * 0.5, oy + ty - tick * 0.5),
                         (ox + x_dim + tick * 0.5, oy + ty + tick * 0.5),
                         dxfattribs={"layer": "S-DIM"})
            msp.add_line((ox + min_x - 200, oy + ty),
                         (ox + x_dim - tick, oy + ty),
                         dxfattribs={"layer": "S-LINE"})
        dist = grid_y[-1] - grid_y[0]
        msp.add_text(_dim_text(dist), height=th, rotation=90,
                     dxfattribs={"layer": "S-DIM-TEXT"}).set_placement(
            (ox + x_dim - th * 0.5, oy + (grid_y[0] + grid_y[-1]) / 2),
            align=TextEntityAlignment.BOTTOM_CENTER)


def select_rebar_panels(panels):
    """Select 4 corner + 2 interior + all balcony panels for rebar drawing."""
    if not panels:
        return []
    regular = [p for p in panels if p.get("zone") is None]
    balcony = [p for p in panels if p.get("zone") in ("BB1", "BF1")]

    if not regular:
        return balcony

    # Find edges
    max_xe = max(p["x_end"] for p in regular)
    max_ze = max(p["z_end"] for p in regular)
    min_xs = min(p["x_start"] for p in regular)
    min_zs = min(p["z_start"] for p in regular)

    def panel_area(p):
        return abs(p["x_end"] - p["x_start"]) * abs(p["z_end"] - p["z_start"])

    corners, interior = [], []
    for p in regular:
        at_left = abs(p["x_start"] - min_xs) < 0.01
        at_right = abs(p["x_end"] - max_xe) < 0.01
        at_bot = abs(p["z_start"] - min_zs) < 0.01
        at_top = abs(p["z_end"] - max_ze) < 0.01
        if (at_left or at_right) and (at_bot or at_top):
            corners.append(p)
        elif not (at_left or at_right) and not (at_bot or at_top):
            interior.append(p)

    # Sort by area — largest first, so rebar goes on big panels not small ones
    corners.sort(key=panel_area, reverse=True)
    interior.sort(key=panel_area, reverse=True)

    return corners[:4] + interior[:2] + balcony


CRANK_OFFSET = 50  # mm visual offset for crank


def draw_rebar_for_panel(msp, panel, rebar_spec, ox, oy):
    """Draw exactly 4 rebar bars for one panel."""
    if not rebar_spec:
        return

    x1 = min(panel["x_start"], panel["x_end"]) * SCALE
    x2 = max(panel["x_start"], panel["x_end"]) * SCALE
    y1 = min(panel["z_start"], panel["z_end"]) * SCALE
    y2 = max(panel["z_start"], panel["z_end"]) * SCALE

    clear_lx = panel["clear_lx"] * SCALE
    clear_lz = panel["clear_lz"] * SCALE
    panel_w = x2 - x1
    panel_h = y2 - y1

    # Half beam width = ZD/2 = 115mm (panel edge is at beam face, extend to beam outer face)
    beam_w = 115
    beam_h = 115

    lx_q = clear_lx / 4
    lz_q = clear_lz / 4
    neigh_lx_q = clear_lx / 4  # approx neighbour span
    neigh_lz_q = clear_lz / 4

    tag_l = rebar_spec["tag_left"]
    tag_r = rebar_spec["tag_right"]
    tag_t = rebar_spec["tag_top"]
    tag_b = rebar_spec["tag_bot"]
    co = CRANK_OFFSET

    dim_h = 100
    dim_off = 150
    ar = 60  # arrow size

    def _h_dim(mx1, mx2, my, label="L/4"):
        """Draw horizontal dimension line with arrows and label."""
        lyr = {"layer":"REBAR_DIM"}
        msp.add_line((ox+mx1,oy+my),(ox+mx2,oy+my), dxfattribs=lyr)
        # Left arrow →
        msp.add_line((ox+mx1,oy+my),(ox+mx1+ar,oy+my+ar*0.4), dxfattribs=lyr)
        msp.add_line((ox+mx1,oy+my),(ox+mx1+ar,oy+my-ar*0.4), dxfattribs=lyr)
        # Right arrow ←
        msp.add_line((ox+mx2,oy+my),(ox+mx2-ar,oy+my+ar*0.4), dxfattribs=lyr)
        msp.add_line((ox+mx2,oy+my),(ox+mx2-ar,oy+my-ar*0.4), dxfattribs=lyr)
        msp.add_text(label, height=dim_h, dxfattribs=lyr).set_placement(
            (ox+(mx1+mx2)/2, oy+my+dim_h*0.3), align=TextEntityAlignment.BOTTOM_CENTER)

    def _v_dim(mx, my1, my2, label="L/4"):
        """Draw vertical dimension line with arrows and label."""
        lyr = {"layer":"REBAR_DIM"}
        msp.add_line((ox+mx,oy+my1),(ox+mx,oy+my2), dxfattribs=lyr)
        # Bottom arrow ↑
        msp.add_line((ox+mx,oy+my1),(ox+mx+ar*0.4,oy+my1+ar), dxfattribs=lyr)
        msp.add_line((ox+mx,oy+my1),(ox+mx-ar*0.4,oy+my1+ar), dxfattribs=lyr)
        # Top arrow ↓
        msp.add_line((ox+mx,oy+my2),(ox+mx+ar*0.4,oy+my2-ar), dxfattribs=lyr)
        msp.add_line((ox+mx,oy+my2),(ox+mx-ar*0.4,oy+my2-ar), dxfattribs=lyr)
        msp.add_text(label, height=dim_h, rotation=90, dxfattribs=lyr).set_placement(
            (ox+mx-dim_h*0.3, oy+(my1+my2)/2), align=TextEntityAlignment.BOTTOM_CENTER)

    # Small gap so the two horizontal/vertical bars don't overlap at center
    gap = co * 1.5  # 75mm offset from center line
    is_two_way = rebar_spec.get("slab_type", "TWO-WAY").upper() != "ONE-WAY"

    # ── Bar 1: Horizontal LEFT→RIGHT at center (slightly above) ──
    bar_y = y1 + panel_h / 2 + gap
    sx = (x1 - neigh_lx_q) if tag_l == "CONT" else (x1 - beam_w)
    crank_x = x1 + lx_q
    ex = x2 + beam_w
    if is_two_way:
        msp.add_line((ox+sx, oy+bar_y), (ox+crank_x, oy+bar_y),
                     dxfattribs={"layer":"REBAR_BOT"})
        msp.add_line((ox+crank_x, oy+bar_y), (ox+crank_x+co, oy+bar_y-co),
                     dxfattribs={"layer":"REBAR_CRANK"})
        msp.add_line((ox+crank_x+co, oy+bar_y-co), (ox+ex, oy+bar_y-co),
                     dxfattribs={"layer":"REBAR_TOP", "ltscale": 30.0})
    else:
        # ONE-WAY: short line matching L/4 dim extent only
        msp.add_line((ox+x1+beam_w, oy+bar_y), (ox+crank_x, oy+bar_y),
                     dxfattribs={"layer":"REBAR_BOT"})
        if tag_l == "CONT":
            msp.add_line((ox+sx, oy+bar_y), (ox+x1-beam_w, oy+bar_y),
                         dxfattribs={"layer":"REBAR_BOT"})
    # L/4 crank dim (from beam inner face to crank) — ABOVE bar (always drawn)
    _h_dim(x1 + beam_w, crank_x, bar_y + dim_off)
    # L/4 neighbour dim (CONT only — from bar start to beam inner face)
    if tag_l == "CONT":
        _h_dim(sx, x1 - beam_w, bar_y + dim_off * 2)

    # ── Bar 2: Horizontal RIGHT→LEFT at center (slightly below) ──
    bar_y = y1 + panel_h / 2 - gap
    sx = (x2 + neigh_lx_q) if tag_r == "CONT" else (x2 + beam_w)
    crank_x = x2 - lx_q
    ex = x1 - beam_w
    if is_two_way:
        msp.add_line((ox+sx, oy+bar_y), (ox+crank_x, oy+bar_y),
                     dxfattribs={"layer":"REBAR_BOT"})
        msp.add_line((ox+crank_x, oy+bar_y), (ox+crank_x-co, oy+bar_y-co),
                     dxfattribs={"layer":"REBAR_CRANK"})
        msp.add_line((ox+crank_x-co, oy+bar_y-co), (ox+ex, oy+bar_y-co),
                     dxfattribs={"layer":"REBAR_TOP", "ltscale": 30.0})
    else:
        # ONE-WAY: short line matching L/4 dim extent only
        msp.add_line((ox+crank_x, oy+bar_y), (ox+x2-beam_w, oy+bar_y),
                     dxfattribs={"layer":"REBAR_BOT"})
        if tag_r == "CONT":
            msp.add_line((ox+x2+beam_w, oy+bar_y), (ox+sx, oy+bar_y),
                         dxfattribs={"layer":"REBAR_BOT"})
    # L/4 crank dim (from crank to beam inner face) — ABOVE bar (always drawn)
    _h_dim(crank_x, x2 - beam_w, bar_y + dim_off)
    # L/4 neighbour dim (CONT only)
    if tag_r == "CONT":
        _h_dim(x2 + beam_w, sx, bar_y + dim_off * 2)

    # ── Bar 3: Vertical BOTTOM→TOP at center (slightly right) ──
    bar_x = x1 + panel_w / 2 + gap
    sy = (y1 - neigh_lz_q) if tag_b == "CONT" else (y1 - beam_h)
    crank_y = y1 + lz_q
    ey = y2 + beam_h
    msp.add_line((ox+bar_x, oy+sy), (ox+bar_x, oy+crank_y),
                 dxfattribs={"layer":"REBAR_BOT"})
    msp.add_line((ox+bar_x, oy+crank_y), (ox+bar_x+co, oy+crank_y+co),
                 dxfattribs={"layer":"REBAR_CRANK"})
    msp.add_line((ox+bar_x+co, oy+crank_y+co), (ox+bar_x+co, oy+ey),
                 dxfattribs={"layer":"REBAR_TOP", "ltscale": 30.0})
    # L/4 crank dim (from beam inner face to crank) — RIGHT of bar
    _v_dim(bar_x + dim_off, y1 + beam_h, crank_y)
    # L/4 neighbour dim (CONT only)
    if tag_b == "CONT":
        _v_dim(bar_x + dim_off * 2, sy, y1 - beam_h)

    # ── Bar 4: Vertical TOP→BOTTOM at center (slightly left) ──
    bar_x = x1 + panel_w / 2 - gap
    sy = (y2 + neigh_lz_q) if tag_t == "CONT" else (y2 + beam_h)
    crank_y = y2 - lz_q
    ey = y1 - beam_h
    msp.add_line((ox+bar_x, oy+sy), (ox+bar_x, oy+crank_y),
                 dxfattribs={"layer":"REBAR_BOT"})
    msp.add_line((ox+bar_x, oy+crank_y), (ox+bar_x+co, oy+crank_y-co),
                 dxfattribs={"layer":"REBAR_CRANK"})
    msp.add_line((ox+bar_x+co, oy+crank_y-co), (ox+bar_x+co, oy+ey),
                 dxfattribs={"layer":"REBAR_TOP", "ltscale": 30.0})
    # L/4 crank dim (from crank to beam inner face) — RIGHT of bar
    _v_dim(bar_x + dim_off, crank_y, y2 - beam_h)
    # L/4 neighbour dim (CONT only)
    if tag_t == "CONT":
        _v_dim(bar_x + dim_off * 2, y2 + beam_h, sy)


def draw_slab_labels(msp, panels, ox, oy, rebar_rect_nos=None):
    """Draw S1, S2, S3... labels on slab panels. Skip Lift/Stair/Shaft.
    For panels with rebars, offset label to clear zone (avoid 1/3 & 2/3 bar lines)."""
    if not panels:
        return
    if rebar_rect_nos is None:
        rebar_rect_nos = set()
    counter = 1
    for p in panels:
        zone = p.get("zone")
        if zone and str(zone).strip().lower() in ("lift", "stair", "shaft"):
            continue
        x1 = min(p["x_start"], p["x_end"]) * SCALE
        x2 = max(p["x_start"], p["x_end"]) * SCALE
        y1 = min(p["z_start"], p["z_end"]) * SCALE
        y2 = max(p["z_start"], p["z_end"]) * SCALE
        pw = x2 - x1
        ph = y2 - y1
        if p["rect_no"] in rebar_rect_nos:
            # Offset to bottom-right clear zone (between 2/3 and edge)
            cx = x1 + pw * 5 / 6
            cy = y1 + ph / 6
        else:
            cx = (x1 + x2) / 2
            cy = (y1 + y2) / 2
        label = f"S{counter}"
        r = 150 + len(label) * 30  # tight circle based on text length
        msp.add_circle((ox + cx, oy + cy), r,
                      dxfattribs={"layer": "TEXT", "lineweight": 50})
        msp.add_text(label, height=150,
                     dxfattribs={"layer": "TEXT", "lineweight": 50}).set_placement(
            (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        counter += 1


def draw_beam_names(msp, floor_beams, grid_x, grid_y, ox, oy):
    """Draw beam names (B1, B2... BF1, BB1...) grouped by grid line."""
    if not floor_beams:
        return

    SNAP = 300  # mm snap tolerance for grouping to grid lines

    # Separate balcony beams
    front_bal = [b for b in floor_beams if b["beam_type"] == "BALCONY_BEAM"
                 and min(b["start_z"], b["end_z"]) * SCALE < min(grid_y) - 100]
    back_bal = [b for b in floor_beams if b["beam_type"] == "BALCONY_BEAM"
                and max(b["start_z"], b["end_z"]) * SCALE > max(grid_y) + 100]
    regular = [b for b in floor_beams if b["beam_type"] != "BALCONY_BEAM"]

    # Group regular beams by grid line
    h_groups = {}  # grid_y_val -> list of beams
    v_groups = {}  # grid_x_val -> list of beams

    for b in regular:
        sy = round(b["start_z"] * SCALE, 1)
        sx = round(b["start_x"] * SCALE, 1)

        if b["direction"] == "horizontal":
            # Find nearest grid Y
            for gy in grid_y:
                if abs(sy - gy) < SNAP:
                    h_groups.setdefault(gy, []).append(b)
                    break
        elif b["direction"] == "vertical":
            # Find nearest grid X
            for gx in grid_x:
                if abs(sx - gx) < SNAP:
                    v_groups.setdefault(gx, []).append(b)
                    break

    b_counter = 1
    text_h = 200

    # Label horizontal beam groups
    for gy in sorted(h_groups.keys()):
        beams = h_groups[gy]
        # Find full span (min start to max end)
        all_x = []
        for b in beams:
            all_x.extend([b["start_x"] * SCALE, b["end_x"] * SCALE])
        min_x, max_x = min(all_x), max(all_x)
        cx = (min_x + max_x) / 2
        label = f"B{b_counter}"
        msp.add_text(label, height=text_h,
                     dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
            (ox + cx, oy + gy), align=TextEntityAlignment.MIDDLE_CENTER)
        b_counter += 1

    # Label vertical beam groups
    for gx in sorted(v_groups.keys()):
        beams = v_groups[gx]
        all_y = []
        for b in beams:
            all_y.extend([b["start_z"] * SCALE, b["end_z"] * SCALE])
        min_y, max_y_val = min(all_y), max(all_y)
        cy = (min_y + max_y_val) / 2
        label = f"B{b_counter}"
        msp.add_text(label, height=text_h, rotation=90,
                     dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
            (ox + gx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        b_counter += 1

    # Label front balcony beams
    bf_counter = 1
    for b in front_bal:
        sx = b["start_x"] * SCALE
        sy = b["start_z"] * SCALE
        ex = b["end_x"] * SCALE
        ey = b["end_z"] * SCALE
        cx, cy = (sx + ex) / 2, (sy + ey) / 2
        if b["direction"] == "horizontal":
            msp.add_text(f"BF{bf_counter}", height=text_h,
                         dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
                (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        else:
            msp.add_text(f"BF{bf_counter}", height=text_h, rotation=90,
                         dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
                (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        bf_counter += 1

    # Label back balcony beams
    bb_counter = 1
    for b in back_bal:
        sx = b["start_x"] * SCALE
        sy = b["start_z"] * SCALE
        ex = b["end_x"] * SCALE
        ey = b["end_z"] * SCALE
        cx, cy = (sx + ex) / 2, (sy + ey) / 2
        if b["direction"] == "horizontal":
            msp.add_text(f"BB{bb_counter}", height=text_h,
                         dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
                (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        else:
            msp.add_text(f"BB{bb_counter}", height=text_h, rotation=90,
                         dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
                (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        bb_counter += 1

    total = (b_counter - 1) + (bf_counter - 1) + (bb_counter - 1)
    print(f"    Beam names: {b_counter-1} regular + {bf_counter-1} BF + {bb_counter-1} BB = {total}")


def draw_beam_tags(msp, floor_beams, ox, oy):
    """Draw beam names using plan_span_tag from beam_geometry.csv.
    Groups segments by tag, places label at midpoint of overall beam line."""
    if not floor_beams:
        return

    # Group beams by plan_span_tag
    tag_groups = {}
    for b in floor_beams:
        tag = b.get("plan_span_tag", "")
        if not tag:
            continue
        tag_groups.setdefault(tag, []).append(b)

    text_h = 140
    placed = 0

    for tag, beams in sorted(tag_groups.items()):
        direction = beams[0]["direction"]
        if direction == "horizontal":
            # Horizontal beam: find full X extent, use common Y
            all_x = []
            for b in beams:
                all_x.extend([b["start_x"] * SCALE, b["end_x"] * SCALE])
            cx = (min(all_x) + max(all_x)) / 2
            cy = beams[0]["start_z"] * SCALE
            msp.add_text(tag, height=text_h,
                         dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
                (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        elif direction == "vertical":
            # Vertical beam: find full Y extent, use common X
            all_y = []
            for b in beams:
                all_y.extend([b["start_z"] * SCALE, b["end_z"] * SCALE])
            cx = beams[0]["start_x"] * SCALE
            cy = (min(all_y) + max(all_y)) / 2
            msp.add_text(tag, height=text_h, rotation=90,
                         dxfattribs={"layer": "S-BEAM-NO"}).set_placement(
                (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)
        placed += 1

    print(f"    Beam tags: {placed}")


def draw_typ_detail(msp, grid_x, grid_y, ox, oy):
    """Draw TYP DETAIL OF TOP EXTRA BAR — exact copy of reference template."""
    cx = (min(grid_x) + max(grid_x)) / 2
    top_y = max(grid_y) + GRID_EXTEND + 4000
    lyr_c = {"layer": "REBAR_DIM"}  # cyan
    lyr_g = {"layer": "S-BEAM"}     # green (beam layer)
    lyr_y = {"layer": "REBAR_BOT"}  # yellow (rebar)
    th = 120
    th_sm = 100

    y = top_y + 2000

    # ── 1. "Ld = Development length" (cyan, centered) ──
    msp.add_text("Ld = Development length", height=th,
                 dxfattribs=lyr_c).set_placement(
        (ox+cx, oy+y), align=TextEntityAlignment.MIDDLE_CENTER)
    y -= th * 3

    # ── 2. Box with note (cyan, LEFT-aligned like reference) ──
    box_x = cx - 3000
    box_w = 4800
    box_h = 550
    msp.add_lwpolyline(
        [(ox+box_x, oy+y), (ox+box_x+box_w, oy+y),
         (ox+box_x+box_w, oy+y-box_h), (ox+box_x, oy+y-box_h)],
        dxfattribs=lyr_c, close=True)
    msp.add_text("TOP EXTRA BAR AT DOUBLE SPACING", height=th_sm,
                 dxfattribs=lyr_c).set_placement(
        (ox+box_x+120, oy+y-box_h*0.3), align=TextEntityAlignment.MIDDLE_LEFT)
    msp.add_text("DIA OF BAR SAME AS PANEL BAR", height=th_sm,
                 dxfattribs=lyr_c).set_placement(
        (ox+box_x+250, oy+y-box_h*0.7), align=TextEntityAlignment.MIDDLE_LEFT)
    y -= box_h + th * 2

    # ── 3. "DISTRIBUTION STEEL 8#@10"C/C" (cyan, left-aligned) ──
    msp.add_text('DISTRIBUTION STEEL 8#@10"C/C', height=th,
                 dxfattribs=lyr_c).set_placement(
        (ox+box_x, oy+y), align=TextEntityAlignment.MIDDLE_LEFT)
    y -= th * 3

    # ── 4. Cross-section diagram ──
    sec_w = 5500
    sec_h = 400
    beam_w = 600
    beam_d = 1400

    sx1 = cx - sec_w / 2     # slab left
    sx2 = cx + sec_w / 2     # slab right
    slab_top = y
    slab_bot = y - sec_h

    # Slab rectangle (green)
    msp.add_lwpolyline(
        [(ox+sx1, oy+slab_top), (ox+sx2, oy+slab_top),
         (ox+sx2, oy+slab_bot), (ox+sx1, oy+slab_bot)],
        dxfattribs=lyr_g, close=True)

    # Beam/column on RIGHT side going down (green)
    bx1 = sx2 - beam_w
    bx2 = sx2
    by1 = slab_bot
    by2 = slab_bot - beam_d
    msp.add_lwpolyline(
        [(ox+bx1, oy+by1), (ox+bx2, oy+by1),
         (ox+bx2, oy+by2), (ox+bx1, oy+by2)],
        dxfattribs=lyr_g, close=True)

    # Hatching in beam (diagonal lines)
    num_hatch = 12
    for i in range(num_hatch):
        hy = by1 - i * beam_d / num_hatch
        hx_end = min(bx1 + beam_w * 0.8, bx2)
        msp.add_line((ox+bx1, oy+hy),
                     (ox+hx_end, oy+hy - beam_d/num_hatch),
                     dxfattribs=lyr_g)

    # Dashed centerline through slab (center linetype)
    mid_y = (slab_top + slab_bot) / 2
    msp.add_line((ox+sx1-300, oy+mid_y), (ox+sx2+300, oy+mid_y),
                 dxfattribs={"layer": "S-GRID", "ltscale": 30.0})

    # Bottom rebar line (yellow, solid, full width)
    bot_bar_y = slab_bot + sec_h * 0.25
    msp.add_line((ox+sx1, oy+bot_bar_y), (ox+sx2, oy+bot_bar_y),
                 dxfattribs=lyr_y)

    # Yellow dots on bottom bar (rebar cross-sections)
    dot_r = 40
    num_dots = 8
    for i in range(num_dots):
        dx = sx1 + beam_w/2 + (i + 0.5) * (sec_w - beam_w) / num_dots
        msp.add_circle((ox+dx, oy+bot_bar_y), dot_r,
                       dxfattribs={"layer": "REBAR_BOT", "color": CLR_YELLOW})
        # Fill dot (small solid)
        msp.add_solid(
            [(ox+dx-dot_r*0.7, oy+bot_bar_y), (ox+dx+dot_r*0.7, oy+bot_bar_y),
             (ox+dx, oy+bot_bar_y+dot_r*0.7), (ox+dx, oy+bot_bar_y-dot_r*0.7)],
            dxfattribs={"layer": "REBAR_BOT", "color": CLR_YELLOW})

    # Top rebar with crank (zigzag going up from slab)
    top_bar_y = slab_top - sec_h * 0.25
    crank_x = sx1 + sec_w * 0.15
    crank_peak = slab_top + sec_h * 1.5  # peak of crank above slab

    # Top bar: dashed line from left to beam
    msp.add_line((ox+sx1-200, oy+top_bar_y), (ox+crank_x, oy+top_bar_y),
                 dxfattribs={"layer": "REBAR_TOP", "ltscale": 20.0})

    # Crank zigzag going UP
    msp.add_line((ox+crank_x, oy+top_bar_y), (ox+crank_x+100, oy+crank_peak),
                 dxfattribs=lyr_g)
    msp.add_line((ox+crank_x+100, oy+crank_peak), (ox+crank_x+200, oy+top_bar_y+100),
                 dxfattribs=lyr_g)

    # Continuation line from crank going right (dashed, into slab)
    msp.add_line((ox+crank_x+200, oy+top_bar_y+100), (ox+bx1, oy+top_bar_y+100),
                 dxfattribs={"layer": "REBAR_TOP", "ltscale": 20.0})

    # Ld line going down at beam face
    msp.add_line((ox+bx1, oy+top_bar_y+100), (ox+bx1, oy+by2+200),
                 dxfattribs=lyr_g)

    # (L/4 dimension removed from typ detail)

    # (8" TYP removed)

    y = by2 - th * 4

    # ── 5. Title (underlined, cyan, centered) ──
    lyr_title = {"layer": "TEXT"}  # green
    msp.add_text("%%UTYP  DETAIL  OF  TOP  EXTRA  BAR  IN  SLAB", height=th * 1.3,
                 dxfattribs=lyr_title).set_placement(
        (ox+cx, oy+y), align=TextEntityAlignment.MIDDLE_CENTER)
    y -= th * 2.5
    msp.add_text("%%UAT  DISCONTINUOUS  EDGE  FOR  ALL  PANEL", height=th * 1.3,
                 dxfattribs=lyr_title).set_placement(
        (ox+cx, oy+y), align=TextEntityAlignment.MIDDLE_CENTER)
    y -= th * 2.5
    msp.add_text("%%USECTION  X-X", height=th * 1.3,
                 dxfattribs=lyr_title).set_placement(
        (ox+cx, oy+y), align=TextEntityAlignment.MIDDLE_CENTER)


def draw_steel_chart(msp, grid_x, grid_y, slab_panels, rebar_data, y_val, ox, oy):
    """Draw STEEL CHART FOR SLAB table to the right of typ detail."""
    if not slab_panels or not rebar_data:
        return

    # Align table to top-right corner of inner border
    border_right = max(grid_x) + GRID_EXTEND + DIM_OFFSET_X2 + 100  # inner border right
    border_top = max(grid_y) + GRID_EXTEND + 6100  # inner border top
    table_x = border_right - sum([900, 1000, 800, 1000, 1000, 1200, 1200])
    table_y = border_top

    lyr = {"layer": "REBAR_DIM"}
    th = 100
    row_h = 250
    col_w = [900, 1000, 800, 1000, 1000, 1200, 1200]  # MARK, TYPE, THK, MAIN LEN, CROSS LEN, MAIN BAR, CROSS BAR
    total_w_tbl = sum(col_w)

    # Collect slab data (skip Lift/Stair/Shaft)
    rows = []
    counter = 1
    for p in slab_panels:
        zone = p.get("zone")
        if zone and str(zone).strip().lower() in ("lift", "stair", "shaft"):
            continue
        key = (y_val, p["rect_no"])
        spec = rebar_data.get(key)
        thk = p.get("thickness", 150)
        thk_in = int(round(thk / 25.4))

        # Main length = min(lx, lz), Cross length = max(lx, lz)
        lx = p.get("lx", 0) or 0
        lz = p.get("lz", 0) or 0
        main_len = min(lx, lz) * SCALE  # mm
        cross_len = max(lx, lz) * SCALE
        main_len_str = _dim_text(main_len) if main_len > 0 else "-"
        cross_len_str = _dim_text(cross_len) if cross_len > 0 else "-"

        main_str = ""
        cross_str = ""
        if spec:
            md = int(spec["main_dia"]) if spec["main_dia"] else 8
            ms = int(spec["main_spacing"]) if spec["main_spacing"] else 200
            cd = int(spec["cross_dia"]) if spec["cross_dia"] else 8
            cs = int(spec["cross_spacing"]) if spec["cross_spacing"] else 200
            ms_in = int(round(ms / 25.4))
            cs_in = int(round(cs / 25.4))
            main_str = f'{md}@{ms_in}"C/C'
            cross_str = f'{cd}@{cs_in}"C/C'

        slab_type = spec.get("slab_type", "") if spec else ""
        rows.append((f"S{counter}", slab_type, f'{thk_in}" THK', main_len_str, cross_len_str, main_str, cross_str))
        counter += 1

    if not rows:
        return

    # ── Draw table ──
    y = table_y

    # Title row
    msp.add_lwpolyline(
        [(ox+table_x, oy+y), (ox+table_x+total_w_tbl, oy+y),
         (ox+table_x+total_w_tbl, oy+y-row_h), (ox+table_x, oy+y-row_h)],
        dxfattribs=lyr, close=True)
    msp.add_text("STEEL CHART FOR SLAB", height=th*1.1,
                 dxfattribs=lyr).set_placement(
        (ox+table_x+total_w_tbl/2, oy+y-row_h/2), align=TextEntityAlignment.MIDDLE_CENTER)
    y -= row_h

    # Header row
    headers = ["SLAB MARK", "TYPE", "SLAB THK", "MAIN LENGTH", "CROSS LENGTH", "MAIN BAR", "CROSS BAR"]
    hx = table_x
    for i, h in enumerate(headers):
        msp.add_lwpolyline(
            [(ox+hx, oy+y), (ox+hx+col_w[i], oy+y),
             (ox+hx+col_w[i], oy+y-row_h), (ox+hx, oy+y-row_h)],
            dxfattribs=lyr, close=True)
        msp.add_text(h, height=th*0.9,
                     dxfattribs=lyr).set_placement(
            (ox+hx+col_w[i]/2, oy+y-row_h/2), align=TextEntityAlignment.MIDDLE_CENTER)
        hx += col_w[i]
    y -= row_h

    # Data rows (skin/peach color = ACI 40)
    for row_data in rows:
        hx = table_x
        for i, val in enumerate(row_data):
            msp.add_lwpolyline(
                [(ox+hx, oy+y), (ox+hx+col_w[i], oy+y),
                 (ox+hx+col_w[i], oy+y-row_h), (ox+hx, oy+y-row_h)],
                dxfattribs=lyr, close=True)
            msp.add_text(val, height=th*0.85,
                         dxfattribs={"layer": "REBAR_DIM", "color": 40}).set_placement(
                (ox+hx+col_w[i]/2, oy+y-row_h/2), align=TextEntityAlignment.MIDDLE_CENTER)
            hx += col_w[i]
        y -= row_h

    print(f"    Steel chart: {len(rows)} rows")


def draw_title_block_br(msp, floor_title, grid_x, grid_y, ox, oy):
    """Draw full right-side panel: title block + sections + tables. Stacks upward."""
    b_right = max(grid_x) + GRID_EXTEND_TOP + DIM_OFFSET_X2 + 500
    b_bottom = min(grid_y) - DIM_OFFSET_Y2 - TITLE_GAP - TITLE_HEIGHT - 1000

    w = 3500
    th, th_big, th_label, pad = 80, 130, 55, 80
    lyr = {"layer": "S-LINE", "color": CLR_WHITE}
    x = b_right - 400 - w  # right-aligned with inner border
    y = b_bottom + 400      # start at inner border bottom

    def row(cy, h):
        """Draw a box at cy going UP by h. Returns new cy (top of box)."""
        msp.add_lwpolyline(
            [(ox+x, oy+cy), (ox+x+w, oy+cy),
             (ox+x+w, oy+cy+h), (ox+x, oy+cy+h)],
            dxfattribs=lyr, close=True)
        return cy + h

    def tbl_row(cy, h, cols, headers=None, values=None):
        """Draw a table row with vertical dividers."""
        msp.add_lwpolyline(
            [(ox+x, oy+cy), (ox+x+w, oy+cy),
             (ox+x+w, oy+cy+h), (ox+x, oy+cy+h)],
            dxfattribs=lyr, close=True)
        cx = x
        for i, cw in enumerate(cols):
            if i > 0:
                msp.add_line((ox+cx, oy+cy), (ox+cx, oy+cy+h), dxfattribs=lyr)
            if headers and i < len(headers):
                msp.add_text(headers[i], height=th*0.7, dxfattribs=lyr).set_placement(
                    (ox+cx+cw/2, oy+cy+h*0.65), align=TextEntityAlignment.MIDDLE_CENTER)
            if values and i < len(values) and values[i]:
                msp.add_text(values[i], height=th*0.9, dxfattribs=lyr).set_placement(
                    (ox+cx+cw/2, oy+cy+h*0.3), align=TextEntityAlignment.MIDDLE_CENTER)
            cx += cw
        return cy + h

    # ══ Stack upward from bottom ══

    # 1. STRUCTURAL DRAWING | DRAWING NO.
    c1w = w * 0.35
    y = tbl_row(y, 350, [c1w, w-c1w],
                ["STRUCTURAL", None], ["DRAWING", "DRAWING NO.:-ST/01"])

    # 2. SHEET NO. | REV. NO. | DATE
    date_str = datetime.now().strftime("%d-%m-%Y")
    y = tbl_row(y, 350, [w*0.33, w*0.22, w*0.45],
                ["SHEET NO.", "REV. NO.", "DATE"],
                ["01", "R00", date_str])

    # 3. SCALE N.T.S.
    y = tbl_row(y, 300, [w*0.3, w*0.35, w*0.35],
                ["SCALE", None, None], ["N.T.S.", None, None])

    # 4. TITLE + Floor name
    h = 700
    y = row(y, h)
    msp.add_text("TITLE-", height=th_label,
                 dxfattribs={"layer": "S-LINE", "color": CLR_RED}).set_placement(
        (ox+x+pad, oy+y-pad), align=TextEntityAlignment.TOP_LEFT)
    msp.add_text(floor_title, height=th_big,
                 dxfattribs={"layer": "S-LINE", "color": CLR_RED}).set_placement(
        (ox+x+pad, oy+y-h*0.55), align=TextEntityAlignment.MIDDLE_LEFT)

    # 5. PROJECT section
    h = 800
    y = row(y, h)
    msp.add_text("PROJECT:-", height=th_label,
                 dxfattribs={"layer": "S-LINE", "color": CLR_RED}).set_placement(
        (ox+x+pad, oy+y-pad), align=TextEntityAlignment.TOP_LEFT)

    # 6. ARCHITECTS section
    h = 1200
    y = row(y, h)
    msp.add_text("ARCHITECTS", height=th*0.9, dxfattribs=lyr).set_placement(
        (ox+x+pad, oy+y-pad), align=TextEntityAlignment.TOP_LEFT)

    # 7. Empty section
    y = row(y, 600)

    # 8. Revision table (empty rows first, header on top)
    tc = [w*0.15, w*0.1, w*0.5, w*0.25]
    for _ in range(5):
        y = tbl_row(y, 200, tc)
    y = tbl_row(y, 200, tc, ["DATE", "NO", "REMARK", "BY"])

    # Small gap
    y += 50

    # 9. Print table (empty rows first, header on top)
    for _ in range(4):
        y = tbl_row(y, 200, tc)
    y = tbl_row(y, 200, tc, ["DATE", "NO", "PRINT ISSUED TO", "BY"])

    # 10. Notes sections (build upward, but content reads top-to-bottom)
    # So draw in reverse order: Reinforcing Steel first (bottom), Concrete middle, General Notes last (top)
    y += 100
    notes_th = 70
    notes_sp = notes_th * 1.8
    indent = 150

    lyr_y = {"layer": "S-BEAM-NO"}  # yellow text

    def note_block(cy, title, lines):
        """Draw a notes section. Lines are in reading order (top to bottom).
        We reverse them since we build upward."""
        block_h = (len(lines) + 2) * notes_sp  # +2 for title + padding
        start_y = cy
        # Draw box first (white)
        msp.add_lwpolyline(
            [(ox+x, oy+cy), (ox+x+w, oy+cy),
             (ox+x+w, oy+cy+block_h), (ox+x, oy+cy+block_h)],
            dxfattribs=lyr, close=True)
        # Title at top of box (yellow)
        top_y = cy + block_h - notes_sp * 0.5
        msp.add_text(title, height=notes_th*1.2, dxfattribs=lyr_y).set_placement(
            (ox+x+indent, oy+top_y), align=TextEntityAlignment.TOP_LEFT)
        # Lines below title (yellow)
        line_y = top_y - notes_sp * 1.2
        for line in lines:
            msp.add_text(line, height=notes_th, dxfattribs=lyr_y).set_placement(
                (ox+x+indent, oy+line_y), align=TextEntityAlignment.TOP_LEFT)
            line_y -= notes_sp
        return cy + block_h

    # REINFORCING STEEL (drawn first = bottom position)
    y = note_block(y, "%%UREINFORCING  STEEL:-", [
        "1. ALL REINFORCEING STEEL WILL BE OF TESTED QUALITY",
        "   CONFORMING TO IS:1786 LATEST.",
        "2. PREFER TO HIGH YEILD STRENGTH DEFORMED BARS",
        "   WITH CHARACTERISTIC STRENGTH OF 500 N/sq. mm.",
        "3. CLEAR COVER TO MAIN REINFORCEMENT SHALL BE",
        "   * FOUNDATION     50 mm. ALL AROUND",
        "   * COLUMNS        40 mm. ALL AROUND",
        "   * BEAMS          30 mm. ALL AROUND",
        "   * SLABS          20 mm. TOP & BOTTOM",
        "   * WALLS          25 mm. EARTH FACE",
        "   * WALLS          20 mm. INNER FACE",
        "4. LAP LENGTH TO BE 50xDIA OF BAR MINIMUM.",
        "5. SLAB BARS IN SHORTER DIRECTION, SHALL BE",
        "   BELOW BARS FOR THE LONGER DIRECTION",
        "6. IN BEAMS, FIRST STIRRUP SHALL BE AT NO MORE THAN",
        "   40 mm FROM FACE OF THE SUPPORTING MEMBER.",
        "7. IN BEAMS TOP BARS ARE NOT TO BE SPLICED IN THE",
        "   END QUARTERS OF THE SPAN, AND THE BOTTOM BARS",
        "   ARE NOT TO BE SPLICED AT MIDDLE HALF OF THE SPAN.",
        "8. (--------) SIGN SHOWS TOP BARS",
        "9. (________) SIGN SHOWS BOTTOM BARS",
        "10.     O     SIGN SHOWS SIMILAR SLAB PANALS",
        "11.           SIGN SHOWS SUNKEN SLAB PANALS",
        "12. DO NOT SCALE FOLLOW WRITTEN DIMENSION ONLY.",
    ])

    # CONCRETE (middle)
    y += 50
    y = note_block(y, "%%UCONCRETE:-", [
        "   ALL R.C.C WORK SHALL BE IN MIX M25",
    ])

    # GENERAL NOTES (drawn last = top position)
    y += 50
    y = note_block(y, "%%UGENERAL  NOTES:-", [
        "G1. DO NOT SCALE THE DRAWING FOLLOW ONLY FIGURED",
        "     DIMENSIONS.",
        "G2. ALL STRUCTURAL DRAWINGS SHOULD  BE  READ IN",
        "     CONJUCTION WITH RELEVANT ARCHITECTURAL DRAWINGS.",
        "     ANY DISCREPANCY OR AMBIGUITY IN EITHER SHOULD",
        "     BE BROUGHT TO THE NOTICE OF THE ARCHITECT",
        "G3. ALL DIMENSIONS ARE IN FEET & INCH",
    ])


def draw_floor_title(msp, title, grid_x, grid_y, ox, oy, thickness_mm=150):
    """Draw underlined floor title + distribution steel note below the plan."""
    cx = (min(grid_x) + max(grid_x)) / 2
    ty = min(grid_y) - DIM_OFFSET_Y2 - TITLE_GAP

    # Line 1: Floor name + slab thickness (underlined)
    thk_inches = int(round(thickness_mm / 25.4))
    line1 = f"%%U{title}"
    msp.add_text(line1, height=TITLE_HEIGHT,
                 dxfattribs={"layer": "TEXT"}).set_placement(
        (ox + cx, oy + ty), align=TextEntityAlignment.MIDDLE_CENTER)

    # Line 2: Distribution steel note
    line2 = 'DISTRIBUTION STEEL IN SLAB WHERE NOT SHOWN SHALL BE 8#@10"C/C'
    msp.add_text(line2, height=TITLE_HEIGHT * 0.6,
                 dxfattribs={"layer": "TEXT"}).set_placement(
        (ox + cx, oy + ty - TITLE_HEIGHT * 1.5), align=TextEntityAlignment.MIDDLE_CENTER)


# ══════════════════════════════════════════════════════════════════════
#  STAIRCASE, LIFT, NOTES, TITLE BLOCK (from old code)
# ══════════════════════════════════════════════════════════════════════

def draw_staircase(msp, stair_rects, ox, oy):
    """Draw unified U-shaped staircase."""
    if not stair_rects:
        return
    all_x, all_y = [], []
    for rect in stair_rects:
        all_x.extend([rect["x1"] * SCALE, rect["x2"] * SCALE])
        all_y.extend([rect["y1"] * SCALE, rect["y2"] * SCALE])
    x1, x2 = min(all_x), max(all_x)
    y1, y2 = min(all_y), max(all_y)
    w, h = x2 - x1, y2 - y1
    if min(w, h) < 1200:
        return
    msp.add_lwpolyline(
        [(ox+x1,oy+y1),(ox+x2,oy+y1),(ox+x2,oy+y2),(ox+x1,oy+y2)],
        dxfattribs={"layer":"S-LINE"}, close=True)
    if h >= w:
        fw, ld = w*0.42, h*0.25
        vx1, vx2 = x1+fw, x2-fw
        vy1, vy2 = y1+ld, y2-ld
        msp.add_lwpolyline(
            [(ox+vx1,oy+vy1),(ox+vx2,oy+vy1),(ox+vx2,oy+vy2),(ox+vx1,oy+vy2)],
            dxfattribs={"layer":"S-LINE"}, close=True)
        flight_h = vy2 - vy1
        nt = max(4, round(flight_h/250))
        sp = flight_h/nt
        for t in range(1, nt):
            ty = vy1 + t*sp
            msp.add_line((ox+x1,oy+ty),(ox+vx1,oy+ty), dxfattribs={"layer":"S-LINE"})
            msp.add_line((ox+vx2,oy+ty),(ox+x2,oy+ty), dxfattribs={"layer":"S-LINE"})
        ax = vx2 + (x2-vx2)/2
        msp.add_line((ox+ax,oy+vy1+sp),(ox+ax,oy+vy2-sp), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+ax-75,oy+vy2-sp-150),(ox+ax,oy+vy2-sp), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+ax+75,oy+vy2-sp-150),(ox+ax,oy+vy2-sp), dxfattribs={"layer":"S-LINE"})
        zy = vy2 - 2*sp
        zx = x1 + fw/2
        zw = fw*0.3
        msp.add_lwpolyline(
            [(ox+x1-50,oy+zy),(ox+zx-zw/2,oy+zy),(ox+zx-zw/4,oy+zy+150),
             (ox+zx+zw/4,oy+zy-150),(ox+zx+zw/2,oy+zy),(ox+vx1+50,oy+zy)],
            dxfattribs={"layer":"S-LINE"})
    else:
        fw, ld = h*0.42, w*0.25
        vx1, vx2 = x1+ld, x2-ld
        vy1, vy2 = y1+fw, y2-fw
        msp.add_lwpolyline(
            [(ox+vx1,oy+vy1),(ox+vx2,oy+vy1),(ox+vx2,oy+vy2),(ox+vx1,oy+vy2)],
            dxfattribs={"layer":"S-LINE"}, close=True)
        flight_w = vx2 - vx1
        nt = max(4, round(flight_w/250))
        sp = flight_w/nt
        for t in range(1, nt):
            tx = vx1 + t*sp
            msp.add_line((ox+tx,oy+y1),(ox+tx,oy+vy1), dxfattribs={"layer":"S-LINE"})
            msp.add_line((ox+tx,oy+vy2),(ox+tx,oy+y2), dxfattribs={"layer":"S-LINE"})
        ay = y1 + fw/2
        msp.add_line((ox+vx1+sp,oy+ay),(ox+vx2-sp,oy+ay), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+vx2-sp-150,oy+ay-75),(ox+vx2-sp,oy+ay), dxfattribs={"layer":"S-LINE"})
        msp.add_line((ox+vx2-sp-150,oy+ay+75),(ox+vx2-sp,oy+ay), dxfattribs={"layer":"S-LINE"})
        zx = vx2 - 2*sp
        zy = y2 - fw/2
        zw = fw*0.3
        msp.add_lwpolyline(
            [(ox+zx,oy+y2+50),(ox+zx,oy+zy+zw/2),(ox+zx+150,oy+zy+zw/4),
             (ox+zx-150,oy+zy-zw/4),(ox+zx,oy+zy-zw/2),(ox+zx,oy+vy2-50)],
            dxfattribs={"layer":"S-LINE"})

    # "Refer staircase detail" text in center
    cx = (x1 + x2) / 2
    cy = (y1 + y2) / 2
    msp.add_text("Refer staircase detail", height=120,
                 dxfattribs={"layer": "TEXT", "color": 7}).set_placement(
        (ox + cx, oy + cy), align=TextEntityAlignment.MIDDLE_CENTER)


def draw_lift(msp, rect, ox, oy):
    """Draw lift X-cross + LIFT text."""
    x1 = rect.get("snapped_x1", rect["x1"]) * SCALE
    y1 = rect.get("snapped_y1", rect["y1"]) * SCALE
    x2 = rect.get("snapped_x2", rect["x2"]) * SCALE
    y2 = rect.get("snapped_y2", rect["y2"]) * SCALE
    if x1 > x2: x1, x2 = x2, x1
    if y1 > y2: y1, y2 = y2, y1
    msp.add_line((ox+x1,oy+y1),(ox+x2,oy+y2), dxfattribs={"layer":"S-LINE"})
    msp.add_line((ox+x2,oy+y1),(ox+x1,oy+y2), dxfattribs={"layer":"S-LINE"})
    cx, cy = (x1+x2)/2, (y1+y2)/2
    msp.add_text("LIFT", height=200,
                 dxfattribs={"layer":"S-BEAM-NO"}).set_placement(
        (ox+cx, oy+cy), align=TextEntityAlignment.MIDDLE_CENTER)


def draw_shaft(msp, rect, ox, oy):
    """Draw shaft markings: X-cross + SHAFT text (same style as lift)."""
    x1 = rect.get("snapped_x1", rect["x1"]) * SCALE
    y1 = rect.get("snapped_y1", rect["y1"]) * SCALE
    x2 = rect.get("snapped_x2", rect["x2"]) * SCALE
    y2 = rect.get("snapped_y2", rect["y2"]) * SCALE
    if x1 > x2: x1, x2 = x2, x1
    if y1 > y2: y1, y2 = y2, y1
    msp.add_line((ox+x1,oy+y1),(ox+x2,oy+y2), dxfattribs={"layer":"S-LINE"})
    msp.add_line((ox+x2,oy+y1),(ox+x1,oy+y2), dxfattribs={"layer":"S-LINE"})
    cx, cy = (x1+x2)/2, (y1+y2)/2
    msp.add_text("SHAFT", height=200,
                 dxfattribs={"layer":"S-BEAM-NO"}).set_placement(
        (ox+cx, oy+cy), align=TextEntityAlignment.MIDDLE_CENTER)


def _group_nearby_rects(rects, gap_threshold=1.0):
    """Group rectangles by spatial proximity using union-find."""
    if not rects:
        return []
    boxes = []
    for r in rects:
        boxes.append((min(r["x1"],r["x2"]), min(r["y1"],r["y2"]),
                       max(r["x1"],r["x2"]), max(r["y1"],r["y2"])))
    parent = list(range(len(rects)))
    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i
    def union(i, j):
        pi, pj = find(i), find(j)
        if pi != pj: parent[pi] = pj
    for i in range(len(rects)):
        for j in range(i+1, len(rects)):
            x_gap = max(0, max(boxes[i][0]-boxes[j][2], boxes[j][0]-boxes[i][2]))
            y_gap = max(0, max(boxes[i][1]-boxes[j][3], boxes[j][1]-boxes[i][3]))
            if x_gap < gap_threshold and y_gap < gap_threshold:
                union(i, j)
    groups = {}
    for i in range(len(rects)):
        groups.setdefault(find(i), []).append(rects[i])
    return list(groups.values())


def filter_mumty_columns(columns, rectangles):
    """Filter columns to only those within the staircase/lift zone for Mumty."""
    if not rectangles:
        return columns
    all_x, all_y = [], []
    for r in rectangles:
        all_x.extend([r["snapped_x1"] * SCALE, r["snapped_x2"] * SCALE])
        all_y.extend([r["snapped_y1"] * SCALE, r["snapped_y2"] * SCALE])
    zone_x1, zone_x2 = min(all_x) - 500, max(all_x) + 500
    zone_y1, zone_y2 = min(all_y) - 500, max(all_y) + 500
    return [col for col in columns
            if zone_x1 <= (col["x1"]+col["x2"])/2 <= zone_x2
            and zone_y1 <= (col["y1"]+col["y2"])/2 <= zone_y2]


def filter_mumty_grids(grid_x, grid_y, mumty_columns):
    """Filter grid values to only those that have mumty columns on them."""
    if not mumty_columns:
        return grid_x, grid_y
    col_xs, col_ys = set(), set()
    for col in mumty_columns:
        cx = (col["x1"] + col["x2"]) / 2
        cy = (col["y1"] + col["y2"]) / 2
        for gx in grid_x:
            if abs(gx - cx) < GRID_SNAP_TOL: col_xs.add(gx)
        for gy in grid_y:
            if abs(gy - cy) < GRID_SNAP_TOL: col_ys.add(gy)
    return sorted(col_xs), sorted(col_ys)


def _fill_mumty_corners(mumty_cols, mumty_gx, mumty_gy):
    """Add 300x600mm columns at any empty corners of mumty bounding box."""
    corners = [
        (min(mumty_gx), min(mumty_gy)),  # bottom-left
        (max(mumty_gx), min(mumty_gy)),  # bottom-right
        (min(mumty_gx), max(mumty_gy)),  # top-left
        (max(mumty_gx), max(mumty_gy)),  # top-right
    ]
    for cx, cy in corners:
        # Check if a column already exists near this corner
        has_col = False
        for col in mumty_cols:
            col_cx = (col["x1"] + col["x2"]) / 2
            col_cy = (col["y1"] + col["y2"]) / 2
            if abs(col_cx - cx) < GRID_SNAP_TOL and abs(col_cy - cy) < GRID_SNAP_TOL:
                has_col = True
                break
        if not has_col:
            # Add a 300x600mm column centered on this corner
            w, h = 300, 600
            mumty_cols.append({
                "no": f"M_{cx}_{cy}",
                "name": f"MC",
                "x1": cx - w / 2,
                "x2": cx + w / 2,
                "y1": cy - h / 2,
                "y2": cy + h / 2,
            })
            print(f"    Added corner column at ({cx}, {cy})")
    return mumty_cols


def filter_columns_by_beams(columns, beams):
    """Keep only columns that have at least one beam endpoint near them."""
    if not beams:
        return columns
    beam_pts = set()
    for b in beams:
        beam_pts.add((round(b["start_x"] * SCALE, 1), round(b["start_z"] * SCALE, 1)))
        beam_pts.add((round(b["end_x"] * SCALE, 1), round(b["end_z"] * SCALE, 1)))
    result = []
    for col in columns:
        cx = (col["x1"] + col["x2"]) / 2
        cy = (col["y1"] + col["y2"]) / 2
        for bx, by in beam_pts:
            if abs(bx - cx) < GRID_SNAP_TOL and abs(by - cy) < GRID_SNAP_TOL:
                result.append(col)
                break
    return result


def filter_rects_to_mumty_zone(rectangles):
    """Filter rectangles to only those within the mumty zone."""
    # Find mumty bounding box from the Mumty rectangle
    mumty_rects = [r for r in rectangles
                   if "mumty" in str(r.get("location", "")).lower()]
    if not mumty_rects:
        return []
    all_x = [v * SCALE for r in mumty_rects for v in [r["x1"], r["x2"]]]
    all_y = [v * SCALE for r in mumty_rects for v in [r["y1"], r["y2"]]]
    zx1, zx2 = min(all_x) - 500, max(all_x) + 500
    zy1, zy2 = min(all_y) - 500, max(all_y) + 500
    # Keep rectangles whose center falls within the mumty zone
    result = []
    for r in rectangles:
        cx = (r["x1"] + r["x2"]) / 2 * SCALE
        cy = (r["y1"] + r["y2"]) / 2 * SCALE
        if zx1 <= cx <= zx2 and zy1 <= cy <= zy2:
            result.append(r)
    return result


def draw_rectangles_on_floor(msp, rectangles, floor_key, ox, oy):
    """Draw lift and staircase markings from rectangle data."""
    if not rectangles:
        return
    lift_rect = None
    for rect in rectangles:
        if "lift" in str(rect.get("location", "")).lower():
            lift_rect = rect
            break
    lift_y_min = min(lift_rect["y1"], lift_rect["y2"]) if lift_rect else None
    lift_y_max = max(lift_rect["y1"], lift_rect["y2"]) if lift_rect else None

    stair_rects = []
    for rect in rectangles:
        rloc = str(rect.get("location", "")).lower()
        if "lift" in rloc:
            print(f"    Lift: {rect['no']} ({rect['type']})")
            draw_lift(msp, rect, ox, oy)
        elif "shaft" in rloc:
            print(f"    Shaft: {rect['no']} ({rect['type']})")
            draw_shaft(msp, rect, ox, oy)
        elif "stair" in rloc:
            if lift_rect is not None:
                ry_min = min(rect["y1"], rect["y2"])
                ry_max = max(rect["y1"], rect["y2"])
                overlap = max(0, min(ry_max, lift_y_max) - max(ry_min, lift_y_min))
                ry_range = ry_max - ry_min
                rx_min = min(rect["x1"], rect["x2"])
                rx_max = max(rect["x1"], rect["x2"])
                lx_min = min(lift_rect["x1"], lift_rect["x2"])
                lx_max = max(lift_rect["x1"], lift_rect["x2"])
                x_gap = max(0, max(rx_min - lx_max, lx_min - rx_max))
                if ry_range > 0 and overlap / ry_range > 0.5 and x_gap < 0.5:
                    print(f"    Skip shaft-adjacent: {rect['no']} ({rect['type']})")
                    continue
            print(f"    Staircase rect: {rect['no']} ({rect['type']})")
            stair_rects.append(rect)

    stair_groups = _group_nearby_rects(stair_rects, gap_threshold=1.0)
    for group in stair_groups:
        names = ", ".join(f"{r['no']}({r['type']})" for r in group)
        print(f"    Drawing staircase group: [{names}]")
        draw_staircase(msp, group, ox, oy)


def draw_notes_panel(msp, notes_x, notes_y, notes_h):
    """Draw General Notes and Legend."""
    x, y = notes_x, notes_y + notes_h
    th, hh = 80, 110
    line_sp, indent, wrap_indent = th*2.0, 250, 400
    cy = y - 100
    msp.add_text("%%UGENERAL  NOTES:-", height=hh,
                 dxfattribs={"layer":"NOTES"}).set_placement(
        (x, cy), align=TextEntityAlignment.TOP_LEFT)
    cy -= hh + line_sp*0.5
    notes = [
        ("1.","STRUCTURAL DRAWINGS SHALL BE READ IN CONJUNCTION WITH",
              "RELEVANT ARCHITECTURAL AND SERVICES DRAWINGS."),
        ("2.","DO NOT SCALE.  FOLLOW WRITTEN DIMENSIONS ONLY.",None),
        ("3.","ALL DIMENSIONS ARE IN MM. UNLESS OTHERWISE SPECIFIED.",None),
        ("4.","UNLESS SPECIFIED OTHERWISE, ALL LEVELS SHOWN IN STRUCTURAL",
              "DRAWINGS ARE STRUCTURAL LEVELS ONLY."),
        ("5.","COVER",None), (None,"a) BEAM                : 1\"",None),
        ("6.","CONCRETE MIX",None), (None,"a) BEAM                : M25",None),
        ("7.","REINFORCEMENT   : REINF. STEEL SHALL BE TMT BARS OF",
              "GRADE Fe500D CONFORMING TO IS 1786-2008."),
        ("8.","ALL LEVELS ARE TO BE TAKEN FROM ARCHITECTURAL DRAWINGS.",None),
        ("9.","CONFIRM LOCATION OF BEAM FOR WALLS WITH RELEVANT","ARCH. DRGS."),
        ("10.","ALL R.C.C. TO BE MACHINE MIXED, VIBRATED AND CURED",
               "THOROUGHLY AS PER I.S 456-LATEST."),
        ("11.","ALL DIMENSIONS MUST BE CHECKED WITH ARCHITECT'S DRGS.& IN",
               "CASE OF ANY DISCREPANCY ARCHITECTS DRGS. SHALL PREVAIL."),
        ("12.","ALL CONSTRUCTION JOINTS SHALL BE APPROVED BY CONSULTANT",
               "ON THE BASIS OF SCHEME PREPARED BY CONTRACTOR."),
        ("13.","TOP AND BOTTOM EXTRA BARS IN BEAMS TO EXTEND BEYOND THE",
               "FACE OF SUPPORT AS SHOWN IN DRG UNLESS OTHERWISE SHOWN."),
        ("14.","THE FIRST STIRRUPS IN BEAMS SHALL BE AT A DISTANCE OF",
               "50MM FROM THE JOINT FACE."),
        ("15.","All ANGLES ARE RIGHT ANGLES UNLESS OTHERWISE SPECIFIED.",None),
    ]
    for item in notes:
        num = item[0]
        lines = [l for l in item[1:] if l is not None]
        if num:
            msp.add_text(f"{num}  {lines[0]}", height=th,
                         dxfattribs={"layer":"NOTES"}).set_placement(
                (x+indent, cy), align=TextEntityAlignment.TOP_LEFT)
            cy -= line_sp
            for cont in lines[1:]:
                msp.add_text(cont, height=th,
                             dxfattribs={"layer":"NOTES"}).set_placement(
                    (x+wrap_indent, cy), align=TextEntityAlignment.TOP_LEFT)
                cy -= line_sp
        else:
            msp.add_text(f"     {lines[0]}", height=th,
                         dxfattribs={"layer":"NOTES"}).set_placement(
                (x+wrap_indent, cy), align=TextEntityAlignment.TOP_LEFT)
            cy -= line_sp
    cy -= line_sp*0.8
    msp.add_text("%%ULEGEND:-", height=hh,
                 dxfattribs={"layer":"NOTES"}).set_placement(
        (x, cy), align=TextEntityAlignment.TOP_LEFT)
    cy -= hh + line_sp*0.5
    for line in ["TOB     = TOP OF BEAM","T.O      = TOP OF",
                 "NGL     = NATURAL GROUND LEVEL","FFL      = FLOOR FINISH LEVEL",
                 "TOS     = TOP OF SLAB"]:
        msp.add_text(line, height=th,
                     dxfattribs={"layer":"NOTES"}).set_placement(
            (x+indent, cy), align=TextEntityAlignment.TOP_LEFT)
        cy -= line_sp


def draw_title_block(msp, x, y, w):
    """Draw title block."""
    layer = "NOTES"
    th, th_med, th_big, th_label, pad = 70, 90, 130, 55, 100
    sp = 170
    def box(cy, h):
        msp.add_lwpolyline([(x,cy),(x+w,cy),(x+w,cy-h),(x,cy-h)],
                           dxfattribs={"layer":layer}, close=True)
        return cy - h
    def label(text, cy):
        msp.add_text(text, height=th_label,
                     dxfattribs={"layer":layer,"color":1}).set_placement(
            (x+pad, cy-40), align=TextEntityAlignment.TOP_LEFT)
    def center_text(text, cy, h):
        msp.add_text(text, height=h,
                     dxfattribs={"layer":layer}).set_placement(
            (x+w/2, cy), align=TextEntityAlignment.MIDDLE_CENTER)
    cy = y
    cy = box(cy, 350)  # FOR REVIEW
    msp.add_text("FOR REVIEW", height=th_big,
                 dxfattribs={"layer":layer}).set_placement(
        (x+pad, cy+175), align=TextEntityAlignment.MIDDLE_LEFT)
    for _ in range(4): cy = box(cy, 160)  # revision rows
    rh = 180; cy_top = cy; cy = box(cy, rh)  # rev header
    c1,c2,c3,c4 = x+w*0.08, x+w*0.18, x+w*0.65, x+w*0.82
    for cx_div in [c1,c2,c3,c4]:
        msp.add_line((cx_div,cy_top),(cx_div,cy), dxfattribs={"layer":layer})
    for lbl,lx,rx in [("Rev",x,c1),("Date",c1,c2),("Description",c2,c3),
                       ("Drn",c3,c4),("Chk",c4,x+w)]:
        msp.add_text(lbl, height=th_label,
                     dxfattribs={"layer":layer,"color":1}).set_placement(
            ((lx+rx)/2, cy+rh/2), align=TextEntityAlignment.MIDDLE_CENTER)
    h=600; cy_top=cy; cy=box(cy,h); label("CLIENT:",cy_top)
    center_text("-------------------", cy+h/2-50, th_med)
    h=1400; cy_top=cy; cy=box(cy,h); label("ARCHITECTS:",cy_top)
    ty=cy_top-250; center_text("PLACEHOLDER & ASSOCIATES",ty,th_big)
    h=1200; cy_top=cy; cy=box(cy,h); label("STRUCTURAL CONSULTANT:",cy_top)
    ty=cy_top-250; center_text("Placeholder Consultants",ty,th_big)
    h=900; cy_top=cy; cy=box(cy,h); label("PROJECT:",cy_top)
    ty=cy_top-280; center_text("PROPOSED FLOOR PLAN FOR",ty,th_big)
    h=800; cy_top=cy; cy=box(cy,h); label("Drawing Title:",cy_top)
    ty=cy_top-300; center_text("ALL FLOOR LEVEL",ty,th_big*1.2)
    ty-=sp*1.5; center_text("FRAMING PLAN",ty,th_big*1.2)
    h=280; cy_top=cy; cy=box(cy,h)
    c1,c2,c3 = x+w*0.22, x+w*0.68, x+w*0.85
    for cx_div in [c1,c2,c3]:
        msp.add_line((cx_div,cy_top),(cx_div,cy), dxfattribs={"layer":layer})


def draw_north_arrow(msp, cx, cy):
    """Draw north arrow symbol."""
    r = 400
    msp.add_circle((cx,cy), r, dxfattribs={"layer":"S-LINE"})
    msp.add_line((cx,cy+r*0.3),(cx,cy+r*1.3), dxfattribs={"layer":"S-LINE"})
    msp.add_line((cx,cy+r*1.3),(cx-r*0.2,cy+r*1.0), dxfattribs={"layer":"S-LINE"})
    msp.add_line((cx,cy+r*1.3),(cx+r*0.2,cy+r*1.0), dxfattribs={"layer":"S-LINE"})
    msp.add_text("N", height=250,
                 dxfattribs={"layer":"TEXT"}).set_placement(
        (cx, cy+r*1.6), align=TextEntityAlignment.BOTTOM_CENTER)


def draw_sheet_border(msp, bounds):
    """Draw 3-line border (violet-cyan-violet) with cyan L-marks at corners."""
    bx1, by1, bx2, by2 = bounds
    m1 = 400   # inner violet margin
    m2 = 500   # middle cyan margin
    m3 = 600   # outer violet margin
    lm = 800   # L-mark length beyond outer border

    # Inner border (violet/magenta)
    msp.add_lwpolyline(
        [(bx1-m1,by1-m1),(bx2+m1,by1-m1),(bx2+m1,by2+m1),(bx1-m1,by2+m1)],
        dxfattribs={"layer":"BORDER"}, close=True)

    # Middle border (cyan)
    msp.add_lwpolyline(
        [(bx1-m2,by1-m2),(bx2+m2,by1-m2),(bx2+m2,by2+m2),(bx1-m2,by2+m2)],
        dxfattribs={"layer":"BORDER", "color": CLR_CYAN}, close=True)

    # Outer border (violet/magenta)
    msp.add_lwpolyline(
        [(bx1-m3,by1-m3),(bx2+m3,by1-m3),(bx2+m3,by2+m3),(bx1-m3,by2+m3)],
        dxfattribs={"layer":"BORDER"}, close=True)

    # Cyan L-marks at 4 corners (outside outer border with gap)
    cl = {"layer": "BORDER", "color": CLR_CYAN}
    gap = 300  # gap between outer border and L-mark
    corners = [
        (bx1-m3-gap, by1-m3-gap),  # bottom-left
        (bx2+m3+gap, by1-m3-gap),  # bottom-right
        (bx1-m3-gap, by2+m3+gap),  # top-left
        (bx2+m3+gap, by2+m3+gap),  # top-right
    ]
    dirs = [
        ((lm, 0), (0, lm)),     # BL: right + up (inward)
        ((-lm, 0), (0, lm)),    # BR: left + up (inward)
        ((lm, 0), (0, -lm)),    # TL: right + down (inward)
        ((-lm, 0), (0, -lm)),   # TR: left + down (inward)
    ]
    for (cx, cy), ((dx1, dy1), (dx2, dy2)) in zip(corners, dirs):
        msp.add_line((cx, cy), (cx+dx1, cy+dy1), dxfattribs=cl)
        msp.add_line((cx, cy), (cx+dx2, cy+dy2), dxfattribs=cl)


def draw_right_panel(msp, grid_x, grid_y, typ_ox, draw_w,
                     border_top, border_bottom):
    """Draw right-side panel with KEY PLAN, NOTES, TITLE BLOCK."""
    tb_width = NOTES_WIDTH
    panel_x = typ_ox + draw_w + 2000
    panel_right = panel_x + tb_width
    panel_top, panel_bottom = border_top, border_bottom

    msp.add_lwpolyline(
        [(panel_x,panel_bottom),(panel_right,panel_bottom),
         (panel_right,panel_top),(panel_x,panel_top)],
        dxfattribs={"layer":"NOTES"}, close=True)

    # KEY PLAN
    keyplan_h = 3500
    keyplan_bottom = panel_top - keyplan_h
    msp.add_line((panel_x,keyplan_bottom),(panel_right,keyplan_bottom),
                 dxfattribs={"layer":"NOTES"})
    msp.add_text("%%UKEY PLAN", height=120,
                 dxfattribs={"layer":"NOTES"}).set_placement(
        (panel_x+tb_width/2, keyplan_bottom+100),
        align=TextEntityAlignment.BOTTOM_CENTER)
    draw_north_arrow(msp, panel_right-600, keyplan_bottom+keyplan_h/2+200)

    # Title block
    tb_total_h = 350+4*160+180+600+1400+1200+900+800+280
    tb_top_y = panel_bottom + tb_total_h
    draw_title_block(msp, panel_x, tb_top_y, tb_width)
    msp.add_line((panel_x,tb_top_y),(panel_right,tb_top_y),
                 dxfattribs={"layer":"NOTES"})

    # Notes
    notes_bottom = tb_top_y + 100
    notes_top = keyplan_bottom - 100
    notes_h = max(notes_top - notes_bottom, 8000)
    draw_notes_panel(msp, panel_x, notes_bottom, notes_h)

    return panel_right + 500  # border_right


# ══════════════════════════════════════════════════════════════════════
#  FLOOR PLAN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════

def draw_floor_plan(msp, title, floor_beams, grid_x, grid_y, columns,
                    rectangles, floor_key, ox, oy,
                    slab_panels=None, rebar_data=None, y_val=None):
    """Draw a complete floor plan at offset (ox, oy)."""
    print(f"    Grid stubs")
    draw_grid_stubs(msp, grid_x, grid_y, ox, oy)

    print(f"    Grid labels")
    draw_grid_labels(msp, grid_x, grid_y, ox, oy)

    print(f"    Beams ({len(floor_beams)})")
    draw_beams(msp, floor_beams, columns, ox, oy)

    # Staircase and lift from old rectangle data
    if rectangles:
        draw_rectangles_on_floor(msp, rectangles, floor_key, ox, oy)

    # Compute rebar panel set (needed for label offset)
    rebar_rect_nos = set()
    if slab_panels and rebar_data and y_val is not None:
        for p in select_rebar_panels(slab_panels):
            key = (y_val, p["rect_no"])
            if rebar_data.get(key):
                rebar_rect_nos.add(p["rect_no"])

    # Slab panel labels
    if slab_panels:
        print(f"    Slab labels ({len(slab_panels)})")
        draw_slab_labels(msp, slab_panels, ox, oy, rebar_rect_nos)

    # Beam names from plan_span_tag
    draw_beam_tags(msp, floor_beams, ox, oy)

    # Typ detail block above plan
    draw_typ_detail(msp, grid_x, grid_y, ox, oy)

    # Title block at bottom-right
    draw_title_block_br(msp, title, grid_x, grid_y, ox, oy)

    # Steel chart table
    if slab_panels and rebar_data and y_val is not None:
        draw_steel_chart(msp, grid_x, grid_y, slab_panels, rebar_data, y_val, ox, oy)

    # Rebar on selected panels
    if slab_panels and rebar_data and y_val is not None:
        selected = select_rebar_panels(slab_panels)
        drawn = 0
        for p in selected:
            key = (y_val, p["rect_no"])
            spec = rebar_data.get(key)
            if spec:
                draw_rebar_for_panel(msp, p, spec, ox, oy)
                drawn += 1
        print(f"    Rebar on {drawn} panels")

    # Columns drawn LAST (opaque fill hides beam overlaps)
    print(f"    Columns ({len(columns)})")
    draw_columns(msp, columns, ox, oy)

    print(f"    Dimensions")
    draw_dimensions(msp, grid_x, grid_y, ox, oy)

    print(f"    Title: {title}")
    draw_floor_title(msp, title, grid_x, grid_y, ox, oy)


# ══════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  SLAB DRAWING GENERATOR V2")
    print("=" * 60)
    print()

    # Input file
    input_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}")
        return
    print(f"Input: {input_path}")

    # Read and classify beams
    all_beams = read_beam_data(input_path)
    all_beams = classify_beams(all_beams)

    # Summary
    h_count = sum(1 for b in all_beams if b["direction"] == "horizontal")
    v_count = sum(1 for b in all_beams if b["direction"] == "vertical")
    d_count = sum(1 for b in all_beams if b["direction"] == "diagonal")
    floors = sorted(set(b["floor"] for b in all_beams))
    types = sorted(set(b["beam_type"] for b in all_beams))
    print(f"Total beams: {len(all_beams)} ({h_count} H, {v_count} V, {d_count} diagonal)")
    print(f"Floors: {floors}")
    print(f"Types: {types}")

    # ── Read OLD data (grids, columns, rectangles) ──
    node_path = find_latest_file("node_coordinates_*.xlsx")
    gridline_path = find_latest_file("gridline_coordinates_*.xlsx")
    other_path = find_latest_file("other_coordinates_*.xlsx")

    if not node_path or not gridline_path:
        print("ERROR: node_coordinates or gridline_coordinates not found in STD ANL folder")
        return

    print(f"Node coordinates: {os.path.basename(node_path)}")
    print(f"Gridline coordinates: {os.path.basename(gridline_path)}")

    # Grids from old node_coordinates (5 Y values, 6 X values)
    grid_x, grid_y = read_unique_grid_values(node_path)
    print(f"Grid X ({len(grid_x)}): {grid_x}")
    print(f"Grid Y ({len(grid_y)}): {grid_y}")

    # Columns from gridline_coordinates (per-floor)
    all_columns = read_gridline_coordinates(gridline_path)
    for ft, cols in all_columns.items():
        print(f"  {ft}: {len(cols)} columns")

    # Rectangles for lift/staircase
    rectangles = []
    if other_path:
        print(f"Other coordinates: {os.path.basename(other_path)}")
        rectangles = read_rectangles(other_path)
        print(f"{len(rectangles)} rectangles (lift/staircase)")

    # Slab panels for labels
    all_slab_panels = read_slab_panels(RECT_GEOM_PATH)
    for ft, pnls in all_slab_panels.items():
        print(f"  Slab panels {ft}: {len(pnls)}")

    # Rebar specs
    all_rebar = read_rebar_input(REBAR_INPUT_PATH)
    print(f"  Rebar specs: {len(all_rebar)} entries")

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # ── PER-FLOOR SLAB DXFs ──
    print("\n" + "=" * 60)
    print("  GENERATING PER-FLOOR SLAB DRAWINGS")
    print("=" * 60)

    panels_by_y = read_slab_panels_by_y(RECT_GEOM_PATH)
    all_y_sorted = sorted(panels_by_y.keys())

    # Floor naming using Y value
    def y_to_floor_name(yv):
        return f"Y{yv}"

    # Floor type to gridline column mapping
    def y_to_col_type(yv):
        if yv == 0: return "Stilt"
        if yv == all_y_sorted[-1]: return "Mumty Roof"
        if yv == all_y_sorted[-2]: return "Terrace"
        stilt_idx = next((i for i, v in enumerate(all_y_sorted) if v == 0), 0)
        fn = all_y_sorted.index(yv) - stilt_idx
        return f"{fn} Floor"

    generated = 0
    for y_val in all_y_sorted:
        if y_val <= 0:
            continue

        floor_name = y_to_floor_name(y_val)
        col_type = y_to_col_type(y_val)
        floor_panels = panels_by_y[y_val]
        floor_cols = all_columns.get(col_type, all_columns.get("Stilt", []))

        # Get beams for this floor
        beam_floor_map = {
            "Stilt": "Stilt", "Mumty_Roof": "Mumty", "Terrace": "Terrace"
        }
        beam_floor = beam_floor_map.get(floor_name, "Typical")
        floor_beams = get_floor_beams(all_beams, beam_floor)

        print(f"\n[{floor_name.upper()} - Y={y_val}]")
        print(f"  Panels: {len(floor_panels)}, Columns: {len(floor_cols)}, Beams: {len(floor_beams)}")

        doc_f = setup_document()
        msp_f = doc_f.modelspace()

        # Dynamic floor title
        y_above_0 = [yv for yv in all_y_sorted if yv > 0]
        y_idx = y_above_0.index(y_val) if y_val in y_above_0 else -1
        if y_idx == 0:
            ftitle = "STILT FLOOR ROOF SLAB"
        elif y_val == y_above_0[-1]:
            ftitle = "MUMTY ROOF SLAB"
        elif y_val == y_above_0[-2]:
            ftitle = "TERRACE FLOOR ROOF SLAB"
        else:
            n = y_idx  # 1-based floor number (idx 1 = 1st floor)
            suffix = "TH"
            if n == 1: suffix = "ST"
            elif n == 2: suffix = "ND"
            elif n == 3: suffix = "RD"
            ftitle = f"{n}{suffix} FLOOR ROOF SLAB"
        draw_floor_plan(msp_f, ftitle,
                        floor_beams, grid_x, grid_y, floor_cols,
                        rectangles, "Stilt floor",
                        0, 0,
                        slab_panels=floor_panels,
                        rebar_data=all_rebar, y_val=y_val)

        # Sheet border
        b_left = min(grid_x) - GRID_EXTEND_LEFT - DIM_OFFSET_Y2 - 500
        b_right = max(grid_x) + GRID_EXTEND_TOP + DIM_OFFSET_X2 + 500
        b_bottom = min(grid_y) - DIM_OFFSET_Y2 - TITLE_GAP - TITLE_HEIGHT - 1000
        b_top = max(grid_y) + GRID_EXTEND + 6500  # space for typ detail
        draw_sheet_border(msp_f, (b_left, b_bottom, b_right, b_top))

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fpath = os.path.join(OUTPUT_FOLDER, f"slab_plan_{floor_name}_{ts}.dxf")
        doc_f.saveas(fpath)
        print(f"  Saved: {os.path.basename(fpath)}")
        generated += 1

    print(f"\nGenerated {generated} per-floor slab drawings.")


if __name__ == "__main__":
    main()
