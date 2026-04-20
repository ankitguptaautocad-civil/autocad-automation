r"""
Node Coordinate Calculator
--------------------------
Usage:
  # Input: floor_coordinates Excel from autocad_anchor_workflow.py
  # Both input and output are in: D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\

  # Single file (all sheets in one Excel):
  python node_coordinate_calculator.py "D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\floor_coordinates_YYYYMMDD_HHMMSS.xlsx"

  # Two files (columns separate from rectangle/balcony):
  python node_coordinate_calculator.py columns.xlsx rect_balcony.xlsx

Input:
  1. floor_coordinates with "Column coordinates" sheet
  2. floor_coordinates with "Rectangle coordinates" and "Balcony coordinates" sheets
  (Can be the same file or two separate files)

Output:
  1. node_coordinates_*.xlsx (same folder as input) with sheets:
     - Node coordinates (calculated from columns)
     - Secondary beam coordinates (snapped to node coordinates)
  2. other_coordinates_*.xlsx with sheets:
     - Primary beam (generated from node grid lines)
     - Rectangle coordinates (snapped to node coordinates)
     - Balcony coordinates (snapped to node coordinates)
     - Staircase details (pass-through)
     - debug_x, debug_y

Node calculation rules:
  X: Left -> +bwx/2,  Right -> -bwx/2,  Centre -> midpoint(anchor_x, opposite_x)
  Y: Front -> +bwy/2, Back  -> -bwy/2,  Centre -> midpoint(anchor_y, opposite_y)

Snapping: Each coordinate value in rectangle/balcony/secondary beam is matched
against node coordinates within +/-0.5m. If a match is found, the value is
replaced with the node coordinate. X columns match node X, Y columns match node Y.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SNAP_TOLERANCE = 0.5  # meters
# # ── Old threshold (kept for reference) ──
DEBUG_DIFF_THRESHOLD = 0.50  # 500mm - matches snap tolerance; white = should have merged

# ─── Styles ──────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(horizontal="center")
RED_FILL = PatternFill("solid", fgColor="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_header(ws, headers, widths):
    for ci, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        col_letter = chr(64 + ci)
        ws.column_dimensions[col_letter].width = width


def _write_row(ws, row_num, values):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER


# ─── Node coordinate calculation ─────────────────────────────────────

def read_columns(input_path: str | Path):
    """Read Column coordinates sheet. Returns one dict per unique column.
    Also stores all raw rows (3 per column) for output."""
    wb = load_workbook(input_path, data_only=True)
    ws = wb["Column coordinates"]

    columns = {}
    raw_rows_by_col = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        col_no = row[0]
        if col_no is None:
            continue

        # Store all raw rows
        if col_no not in raw_rows_by_col:
            raw_rows_by_col[col_no] = []
        raw_rows_by_col[col_no].append(list(row))

        if col_no in columns:
            continue

        columns[col_no] = {
            "no": col_no,
            "type": row[1],           # B
            "anchor_x": row[2],       # C
            "anchor_y": row[3],       # D
            "location": row[4],       # E
            "anchor_loc": row[5],     # F
            "beam_x_loc": row[6],     # G
            "opp_x_x": row[7],       # H
            "opp_x_y": row[8],       # I
            "beam_width_x_mm": row[10],  # K
            "beam_depth_x_mm": row[11],  # L
            "beam_y_loc": row[13],    # N
            "opp_y_x": row[14],      # O
            "opp_y_y": row[15],      # P
            "beam_width_y_mm": row[17],  # R
            "beam_depth_y_mm": row[18],  # S
        }

    wb.close()
    sorted_keys = sorted(columns.keys())
    col_list = [columns[k] for k in sorted_keys]
    raw_rows = [raw_rows_by_col[k] for k in sorted_keys]
    return col_list, raw_rows


def calc_node_coordinate(col: dict) -> tuple[float, float]:
    """Calculate node X, Y from anchor + beam width offsets."""
    ax = float(col["anchor_x"])
    ay = float(col["anchor_y"])
    bwx_m = float(col["beam_width_x_mm"]) / 1000.0
    bwy_m = float(col["beam_width_y_mm"]) / 1000.0

    beam_x = col["beam_x_loc"]
    beam_y = col["beam_y_loc"]

    if beam_x == "Left":
        nx = ax + bwx_m / 2
    elif beam_x == "Right":
        nx = ax - bwx_m / 2
    elif beam_x == "Centre":
        opp_x = float(col["opp_x_x"])
        nx = (ax + opp_x) / 2
    else:
        raise ValueError(f"Unknown beam X location: {beam_x}")

    if beam_y == "Front":
        ny = ay + bwy_m / 2
    elif beam_y == "Back":
        ny = ay - bwy_m / 2
    elif beam_y == "Centre":
        opp_y = float(col["opp_y_y"])
        ny = (ay + opp_y) / 2
    else:
        raise ValueError(f"Unknown beam Y location: {beam_y}")

    return round(nx, 3), round(ny, 3)


# ─── Node leveling (post-calculation alignment) ─────────────────────

LEVEL_TOLERANCE = 0.50  # 500mm - merge node values closer than this (covers beam width diffs up to 530mm)


def level_node_coords(node_coords):
    """Level node coordinates: if X or Y values are within 50mm, average them.
    This fixes small differences from picking precision in AutoCAD.
    Returns new list of (nx, ny) tuples."""
    xs = [nx for nx, ny in node_coords]
    ys = [ny for nx, ny in node_coords]

    leveled_x = _level_values(xs)
    leveled_y = _level_values(ys)

    return list(zip(leveled_x, leveled_y))


def _level_values(values):
    """Group values within LEVEL_TOLERANCE, replace each group with average."""
    n = len(values)
    used = [False] * n
    result = list(values)

    for i in range(n):
        if used[i]:
            continue
        # Find all values within tolerance of values[i]
        group = [i]
        used[i] = True
        for j in range(i + 1, n):
            if used[j]:
                continue
            if abs(values[j] - values[i]) <= LEVEL_TOLERANCE:
                group.append(j)
                used[j] = True

        if len(group) < 2:
            continue

        # Average the group
        avg = round(sum(values[idx] for idx in group) / len(group), 3)
        for idx in group:
            result[idx] = avg

    return result


# ─── Snapping logic ─────────────────────────────────────────────────

def snap_value(value, node_values, tolerance=SNAP_TOLERANCE):
    """Snap a value to the closest node coordinate within tolerance.
    Returns the snapped value, or original if no match."""
    if value is None:
        return value

    best_match = None
    best_dist = tolerance + 1

    for nv in node_values:
        dist = abs(value - nv)
        if dist <= tolerance and dist < best_dist:
            best_match = nv
            best_dist = dist

    return best_match if best_match is not None else value


RECT_DEFAULT_WT_MM = 115  # default wall thickness for rectangle offset


def _apply_rect_offset_single(x, y, beam_x_loc, beam_y_loc):
    """Apply wall thickness offset to a single rectangle corner.
    Beam X location controls X offset, Beam Y location controls Y offset.
    Default wall thickness = 115mm.
    """
    wt_m = RECT_DEFAULT_WT_MM / 1000.0 / 2.0  # half wall thickness in metres
    if beam_x_loc == "Left":
        x = round(x - wt_m, 3)
    elif beam_x_loc == "Right":
        x = round(x + wt_m, 3)
    if beam_y_loc == "Front":
        y = round(y - wt_m, 3)
    elif beam_y_loc == "Back":
        y = round(y + wt_m, 3)
    return x, y


def _group_and_average(values_with_indices, tolerance):
    """Group values within tolerance, return mapping of index → averaged value.
    values_with_indices: list of (index, value) tuples.
    Returns dict: index → new_value (only for values that were grouped)."""
    if not values_with_indices:
        return {}

    # Sort by value
    sorted_vals = sorted(values_with_indices, key=lambda x: x[1])
    used = set()
    mapping = {}

    for i, (idx_i, val_i) in enumerate(sorted_vals):
        if idx_i in used:
            continue
        group = [(idx_i, val_i)]
        used.add(idx_i)
        for j in range(i + 1, len(sorted_vals)):
            idx_j, val_j = sorted_vals[j]
            if idx_j in used:
                continue
            if val_j - val_i <= tolerance:
                group.append((idx_j, val_j))
                used.add(idx_j)
            else:
                break  # sorted, so no more matches

        if len(group) >= 2:
            avg = round(sum(v for _, v in group) / len(group), 3)
            for idx, _ in group:
                mapping[idx] = avg

    return mapping


def read_and_snap_rectangles(input_path, node_x_vals, node_y_vals,
                             sec_beam_rows=None):
    """Read Rectangle coordinates sheet. 3-step snapping logic:
    Step 1: Offset all → group nearby rect values within 150mm → average
    Step 2: Unmatched values → swap with nearest node within 150mm
    Step 3: Still unmatched → swap with nearest sec beam value within 150mm
    Step 4: Fallback → keep offset value as-is."""
    wb = load_workbook(input_path, data_only=True)
    if "Rectangle coordinates" not in wb.sheetnames:
        wb.close()
        return None

    ws = wb["Rectangle coordinates"]
    raw_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        no = row[0]
        rtype = row[1]
        orig_x1 = row[2]
        orig_y1 = row[3]
        orig_x2 = row[4]
        orig_y2 = row[5]
        location = row[6]
        anchor1 = row[7] if len(row) > 7 else None
        beam_x_loc1 = row[8] if len(row) > 8 else None
        beam_y_loc1 = row[9] if len(row) > 9 else None
        anchor2 = row[10] if len(row) > 10 else None
        beam_x_loc2 = row[11] if len(row) > 11 else None
        beam_y_loc2 = row[12] if len(row) > 12 else None

        # Apply wall thickness offset per corner
        off_x1, off_y1 = _apply_rect_offset_single(orig_x1, orig_y1, beam_x_loc1, beam_y_loc1)
        off_x2, off_y2 = _apply_rect_offset_single(orig_x2, orig_y2, beam_x_loc2, beam_y_loc2)

        raw_data.append({
            "no": no, "rtype": rtype,
            "orig": (orig_x1, orig_y1, orig_x2, orig_y2),
            "location": location,
            "anchor1": anchor1, "bx1": beam_x_loc1, "by1": beam_y_loc1,
            "anchor2": anchor2, "bx2": beam_x_loc2, "by2": beam_y_loc2,
            "off": [off_x1, off_y1, off_x2, off_y2],  # mutable list
        })

    wb.close()
    if not raw_data:
        return None

    RECT_SNAP_TOL = SNAP_TOLERANCE  # 150mm

    # ── Step 1: Group nearby rectangle offset values → average ──
    # Collect all X and Y offset values with indices
    all_x = []  # (rect_idx * 2 + corner, value)
    all_y = []
    for i, rd in enumerate(raw_data):
        all_x.append((i * 2, rd["off"][0]))      # X1
        all_x.append((i * 2 + 1, rd["off"][2]))  # X2
        all_y.append((i * 2, rd["off"][1]))       # Y1
        all_y.append((i * 2 + 1, rd["off"][3]))   # Y2

    x_map = _group_and_average(all_x, RECT_SNAP_TOL)
    y_map = _group_and_average(all_y, RECT_SNAP_TOL)

    # Apply averages
    grouped_x = set()
    grouped_y = set()
    for i, rd in enumerate(raw_data):
        if i * 2 in x_map:
            rd["off"][0] = x_map[i * 2]
            grouped_x.add(i * 2)
        if i * 2 + 1 in x_map:
            rd["off"][2] = x_map[i * 2 + 1]
            grouped_x.add(i * 2 + 1)
        if i * 2 in y_map:
            rd["off"][1] = y_map[i * 2]
            grouped_y.add(i * 2)
        if i * 2 + 1 in y_map:
            rd["off"][3] = y_map[i * 2 + 1]
            grouped_y.add(i * 2 + 1)

    if x_map or y_map:
        print(f"  Rect step 1: grouped {len(x_map)} X + {len(y_map)} Y values (averaged)")

    # ── Step 2: ALL values (including averaged) → swap with nearest node within 150mm ──
    node_swaps = 0
    for i, rd in enumerate(raw_data):
        for ci, axis_vals in [(0, node_x_vals), (2, node_x_vals),
                              (1, node_y_vals), (3, node_y_vals)]:
            val = rd["off"][ci]
            best = None
            best_d = RECT_SNAP_TOL + 1
            for nv in axis_vals:
                d = abs(val - nv)
                if d <= RECT_SNAP_TOL and d < best_d:
                    best = nv
                    best_d = d
            if best is not None and best != val:
                rd["off"][ci] = best
                node_swaps += 1

    if node_swaps:
        print(f"  Rect step 2: swapped {node_swaps} values with nodes")

    # ── Step 3: Values not on a node → swap with nearest sec beam value within 150mm ──
    sec_swaps = 0
    if sec_beam_rows:
        sec_x_vals = sorted(set(r[12] for r in sec_beam_rows if r[12] is not None) |
                            set(r[14] for r in sec_beam_rows if r[14] is not None))
        sec_y_vals = sorted(set(r[13] for r in sec_beam_rows if r[13] is not None) |
                            set(r[15] for r in sec_beam_rows if r[15] is not None))

        for i, rd in enumerate(raw_data):
            for ci, sec_vals in [(0, sec_x_vals), (2, sec_x_vals),
                                 (1, sec_y_vals), (3, sec_y_vals)]:
                val = rd["off"][ci]
                # Skip if already on a node (step 2 handled it)
                is_x = ci in (0, 2)
                axis_nodes = node_x_vals if is_x else node_y_vals
                if val in axis_nodes:
                    continue
                # Find nearest sec beam value within tolerance
                best = None
                best_d = RECT_SNAP_TOL + 1
                for sv in sec_vals:
                    d = abs(val - sv)
                    if d <= RECT_SNAP_TOL and d < best_d:
                        best = sv
                        best_d = d
                if best is not None and best != val:
                    rd["off"][ci] = best
                    sec_swaps += 1

    if sec_swaps:
        print(f"  Rect step 3: swapped {sec_swaps} values with sec beams")

    # Build output rows
    rows = []
    for rd in raw_data:
        rows.append((rd["no"], rd["rtype"],
                     rd["orig"][0], rd["orig"][1], rd["orig"][2], rd["orig"][3],
                     rd["location"], rd["anchor1"], rd["bx1"], rd["by1"],
                     rd["anchor2"], rd["bx2"], rd["by2"],
                     rd["off"][0], rd["off"][1], rd["off"][2], rd["off"][3]))

    return rows if rows else None


def read_and_snap_balconies(input_path, node_x_vals, node_y_vals,
                            sec_beam_rows=None):
    """Read Balcony coordinates sheet. 3-step snapping:
    Step 0: Apply wt/2 offset per corner (beam_x/beam_y location)
    Step 1: Swap with nearest node within 150mm
    Step 2: If not on node → swap with nearest sec beam value within 150mm
    Step 3: Fallback → keep offset value as-is."""
    wb = load_workbook(input_path, data_only=True)
    if "Balcony coordinates" not in wb.sheetnames:
        wb.close()
        return None

    BAL_SNAP_TOL = SNAP_TOLERANCE  # 150mm

    # Collect sec beam values for step 2
    sec_x_vals = []
    sec_y_vals = []
    if sec_beam_rows:
        sec_x_vals = sorted(set(r[12] for r in sec_beam_rows if r[12] is not None) |
                            set(r[14] for r in sec_beam_rows if r[14] is not None))
        sec_y_vals = sorted(set(r[13] for r in sec_beam_rows if r[13] is not None) |
                            set(r[15] for r in sec_beam_rows if r[15] is not None))

    ws = wb["Balcony coordinates"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        label = row[0]

        # Support both old 5-col format and new 13-col format
        if len(row) >= 13 and isinstance(row[3], str):
            # New format with beam locations and wall thickness
            raw_x1, raw_y1 = row[1], row[2]
            bx1, by1 = row[5], row[6]
            raw_x2, raw_y2 = row[7], row[8]
            bx2, by2 = row[11], row[12]

            # Step 0: Apply wt/2 offset per corner
            x1, y1 = _apply_rect_offset_single(raw_x1, raw_y1, bx1, by1)
            x2, y2 = _apply_rect_offset_single(raw_x2, raw_y2, bx2, by2)
        else:
            # Old format: no beam locations, no offset
            x1, y1 = row[1], row[2]
            x2, y2 = row[3], row[4]

        # Step 1: Swap with nearest node within 150mm
        coords = [x1, y1, x2, y2]
        for ci, axis_vals in [(0, node_x_vals), (2, node_x_vals),
                              (1, node_y_vals), (3, node_y_vals)]:
            val = coords[ci]
            best = None
            best_d = BAL_SNAP_TOL + 1
            for nv in axis_vals:
                d = abs(val - nv)
                if d <= BAL_SNAP_TOL and d < best_d:
                    best = nv
                    best_d = d
            if best is not None:
                coords[ci] = best

        # Step 2: If not on node → swap with nearest sec beam value within 150mm
        for ci, sec_vals in [(0, sec_x_vals), (2, sec_x_vals),
                             (1, sec_y_vals), (3, sec_y_vals)]:
            val = coords[ci]
            is_x = ci in (0, 2)
            axis_nodes = node_x_vals if is_x else node_y_vals
            if val in axis_nodes:
                continue  # already on a node, skip
            best = None
            best_d = BAL_SNAP_TOL + 1
            for sv in sec_vals:
                d = abs(val - sv)
                if d <= BAL_SNAP_TOL and d < best_d:
                    best = sv
                    best_d = d
            if best is not None and best != val:
                coords[ci] = best

        rows.append((label, coords[0], coords[1], coords[2], coords[3]))

    wb.close()
    return rows if rows else None


def _apply_wall_offset(x1, y1, x2, y2, location, wall_thickness_mm):
    """Apply wall thickness offset to secondary beam coordinates.

    Own axis (beam_loc direction):
      Left  → x1, x2 minus wt/2
      Right → x1, x2 plus wt/2
      Back  → y1, y2 plus wt/2
      Front → y1, y2 minus wt/2

    Perpendicular axis (endpoints picked at inner wall face → offset outward to centerline):
      For Left/Right beams (X is own axis, Y is perpendicular):
        smaller Y → subtract wt/2,  larger Y → add wt/2
      For Front/Back beams (Y is own axis, X is perpendicular):
        smaller X → subtract wt/2,  larger X → add wt/2
    """
    if wall_thickness_mm is None or wall_thickness_mm == "":
        return x1, y1, x2, y2
    wt_m = float(wall_thickness_mm) / 1000.0 / 2.0

    # ── Own axis offset (beam_loc direction) ──
    if location == "Left":
        x1 = round(x1 - wt_m, 3)
        x2 = round(x2 - wt_m, 3)
    elif location == "Right":
        x1 = round(x1 + wt_m, 3)
        x2 = round(x2 + wt_m, 3)
    elif location == "Back":
        y1 = round(y1 + wt_m, 3)
        y2 = round(y2 + wt_m, 3)
    elif location == "Front":
        y1 = round(y1 - wt_m, 3)
        y2 = round(y2 - wt_m, 3)

    # ── Perpendicular axis offset (inner face → centerline) ──
    # User picks inner face of wall; centerline is behind (further from beam)
    # Smaller value → subtract wt/2, Larger value → add wt/2
    if location in ("Left", "Right"):
        # Y is perpendicular
        if y1 < y2:
            y1 = round(y1 - wt_m, 3)
            y2 = round(y2 + wt_m, 3)
        elif y2 < y1:
            y2 = round(y2 - wt_m, 3)
            y1 = round(y1 + wt_m, 3)
        # if y1 == y2: no perpendicular offset needed
    elif location in ("Front", "Back"):
        # X is perpendicular
        if x1 < x2:
            x1 = round(x1 - wt_m, 3)
            x2 = round(x2 + wt_m, 3)
        elif x2 < x1:
            x2 = round(x2 - wt_m, 3)
            x1 = round(x1 + wt_m, 3)
        # if x1 == x2: no perpendicular offset needed

    return x1, y1, x2, y2


def read_and_offset_secondary_beams(input_path):
    """Read Secondary beam coordinates sheet and apply wall offset only.
    Returns list of mutable lists (not tuples) for post-processing."""
    wb = load_workbook(input_path, data_only=True)
    # Find sheet by partial match: "Secondary beam coordinates" prefix
    sec_sheet = None
    for sn in wb.sheetnames:
        if sn.lower().startswith("secondary beam coordinates"):
            sec_sheet = sn
            break
    if sec_sheet is None:
        wb.close()
        return None

    ws = wb[sec_sheet]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        no = row[0]
        stype = row[1]
        orig_x1 = row[2]
        orig_y1 = row[3]
        orig_x2 = row[4]
        orig_y2 = row[5]
        location = row[6]
        floor = row[7]
        present = row[8]
        beam_width = row[9]
        beam_depth = row[10]
        wall_thickness = row[11]

        # Apply wall thickness offset (own axis + perpendicular)
        off_x1, off_y1, off_x2, off_y2 = _apply_wall_offset(
            orig_x1, orig_y1, orig_x2, orig_y2, location, wall_thickness)

        # Store as list (mutable) with offset values in snap positions (12-15)
        rows.append([no, stype, orig_x1, orig_y1, orig_x2, orig_y2,
                     location, floor, present, beam_width, beam_depth,
                     wall_thickness, off_x1, off_y1, off_x2, off_y2])

    wb.close()
    return rows if rows else None


def snap_secondary_beams(sec_beam_rows, node_x_vals, node_y_vals):
    """Apply 2-step snapping to merged secondary beam rows.
    Step 1: Group all offset values within 150mm → average
    Step 2: Swap averaged values with nearest node within 150mm
    Fallback: keep value as-is."""
    if not sec_beam_rows:
        return sec_beam_rows

    SEC_SNAP_TOL = SNAP_TOLERANCE  # 150mm

    # ── Step 1: Group nearby sec beam values → average ──
    # Collect unique beam values (first row per beam type)
    first_rows = {}
    for row in sec_beam_rows:
        if row[1] not in first_rows:
            first_rows[row[1]] = row

    # Collect X and Y from first rows only (avoid duplicating per-floor rows)
    all_x = []  # (beam_type, coord_index, value)
    all_y = []
    for btype, row in first_rows.items():
        all_x.append((btype, 12, row[12]))  # snap_x1
        all_x.append((btype, 14, row[14]))  # snap_x2
        all_y.append((btype, 13, row[13]))  # snap_y1
        all_y.append((btype, 15, row[15]))  # snap_y2

    # Group and average X values
    x_indexed = [(i, v) for i, (_, _, v) in enumerate(all_x)]
    y_indexed = [(i, v) for i, (_, _, v) in enumerate(all_y)]

    x_map = _group_and_average(x_indexed, SEC_SNAP_TOL)
    y_map = _group_and_average(y_indexed, SEC_SNAP_TOL)

    # Build replacement maps: (beam_type, coord_index) → new_value
    replace = {}
    for i, (btype, ci, _) in enumerate(all_x):
        if i in x_map:
            replace[(btype, ci)] = x_map[i]
    for i, (btype, ci, _) in enumerate(all_y):
        if i in y_map:
            replace[(btype, ci)] = y_map[i]

    # Apply averages to first rows
    if replace:
        grouped_count = len(replace)
        for row in sec_beam_rows:
            btype = row[1]
            for ci in (12, 13, 14, 15):
                key = (btype, ci)
                if key in replace:
                    row[ci] = replace[key]
        print(f"  Sec step 1: grouped {grouped_count} values (averaged)")

    # ── Step 2: ALL values → swap with nearest node within 150mm ──
    node_swaps = 0
    # Only process first row per beam, then copy to other floor rows
    processed = {}
    for row in sec_beam_rows:
        btype = row[1]
        if btype not in processed:
            for ci, axis_vals in [(12, node_x_vals), (14, node_x_vals),
                                  (13, node_y_vals), (15, node_y_vals)]:
                val = row[ci]
                best = None
                best_d = SEC_SNAP_TOL + 1
                for nv in axis_vals:
                    d = abs(val - nv)
                    if d <= SEC_SNAP_TOL and d < best_d:
                        best = nv
                        best_d = d
                if best is not None and best != val:
                    row[ci] = best
                    node_swaps += 1
            processed[btype] = (row[12], row[13], row[14], row[15])
        else:
            # Copy from first row of same beam type
            row[12], row[13], row[14], row[15] = processed[btype]

    if node_swaps:
        print(f"  Sec step 2: swapped {node_swaps} values with nodes")

    # Convert back to tuples
    return [tuple(row) for row in sec_beam_rows]


# ─── Cross-snap & final leveling (post-processing) ──────────────────

CROSS_SNAP_TOL = 0.060   # 60mm - snap face values to nearest centerline
# # ── Old tolerance (kept for reference) ──
# FINAL_LEVEL_TOL = 0.005  # 5mm
FINAL_LEVEL_TOL = 0.050  # 10mm - merge remaining picking precision splits


def cross_snap_secondary_beams(sec_beam_rows, rect_rows, node_x_vals, node_y_vals):
    """Cross-snap un-offset secondary beam coordinates to centerline values.

    Problem: Left/Right beams only get X offset, their Y stays at wall face.
             Front/Back beams only get Y offset, their X stays at wall face.
             At junctions, one beam is at centerline, the perpendicular is at face → 57mm gap.

    Fix: Snap Front/Back beams' X to trusted centerline X values (within 60mm).
         Snap Left/Right beams' Y to trusted centerline Y values (within 60mm).

    Trusted X = node X + rect snapped X + Left/Right sec beam snapped X (all had X offset).
    Trusted Y = node Y + rect snapped Y + Front/Back sec beam snapped Y (all had Y offset).
    """
    if not sec_beam_rows:
        return sec_beam_rows

    # Build trusted centerline values
    trusted_x = set(node_x_vals)
    trusted_y = set(node_y_vals)

    if rect_rows:
        for row in rect_rows:
            trusted_x.add(row[13])  # snapped X1
            trusted_x.add(row[15])  # snapped X2
            trusted_y.add(row[14])  # snapped Y1
            trusted_y.add(row[16])  # snapped Y2

    # First pass: collect offset values from sec beams themselves
    # (Left/Right contribute to trusted X, Front/Back contribute to trusted Y)
    # Use row[1] (type like SP1/SO1) as unique key, not row[0] (number)
    first_rows = {}
    for row in sec_beam_rows:
        if row[1] not in first_rows:
            first_rows[row[1]] = row
    for row in first_rows.values():
        loc = row[6]
        if loc in ("Left", "Right"):
            trusted_x.add(row[12])  # snapped X1
            trusted_x.add(row[14])  # snapped X2
        elif loc in ("Front", "Back"):
            trusted_y.add(row[13])  # snapped Y1
            trusted_y.add(row[15])  # snapped Y2

    trusted_x = sorted(trusted_x)
    trusted_y = sorted(trusted_y)

    # Cross-snap: modify sec beam rows
    new_rows = []
    changes = []
    processed_beams = set()

    for row in sec_beam_rows:
        row = list(row)  # make mutable
        beam_key = row[1]  # Use type (SP1/SO1) not number (1) as unique key
        loc = row[6]
        is_first = beam_key not in processed_beams

        if is_first:
            processed_beams.add(beam_key)

            if loc in ("Front", "Back"):
                # X values were NOT offset - cross-snap to trusted X
                for idx in (12, 14):  # snapped X1, X2
                    old_val = row[idx]
                    new_val = _snap_to_nearest(old_val, trusted_x, CROSS_SNAP_TOL)
                    if new_val != old_val:
                        changes.append(
                            f"  {row[1]} ({loc}) X: {old_val} -> {new_val} "
                            f"({abs(new_val - old_val)*1000:.0f}mm)")
                        row[idx] = new_val

            elif loc in ("Left", "Right"):
                # Y values were NOT offset - cross-snap to trusted Y
                for idx in (13, 15):  # snapped Y1, Y2
                    old_val = row[idx]
                    new_val = _snap_to_nearest(old_val, trusted_y, CROSS_SNAP_TOL)
                    if new_val != old_val:
                        changes.append(
                            f"  {row[1]} ({loc}) Y: {old_val} -> {new_val} "
                            f"({abs(new_val - old_val)*1000:.0f}mm)")
                        row[idx] = new_val

        else:
            # Non-first rows: copy coords from first row of same beam
            for r in new_rows:
                if r[1] == beam_key:
                    row[12], row[13], row[14], row[15] = r[12], r[13], r[14], r[15]
                    break

        new_rows.append(tuple(row))

    if changes:
        print(f"\nCross-snapped {len(changes)} coordinates (face -> centerline):")
        for c in changes:
            print(c)
    else:
        print("\nCross-snap: no adjustments needed.")

    return new_rows


def _snap_to_nearest(value, targets, tolerance):
    """Snap value to nearest target within tolerance. Returns original if no match."""
    best = value
    best_d = tolerance + 1
    for t in targets:
        d = abs(value - t)
        if 0.001 < d <= tolerance and d < best_d:
            best = t
            best_d = d
    return best


def final_level_all(rect_rows, sec_beam_rows, balcony_rows,
                    node_x_vals, node_y_vals, ew_rows=None):
    """Final leveling pass: merge values within 10mm across all elements.
    Fixes picking precision splits (e.g. 6.846 vs 6.848).
    Node values always take priority - non-node values snap to nearby node.
    """
    # Collect all final values
    all_x = set(node_x_vals)
    all_y = set(node_y_vals)
    if rect_rows:
        for row in rect_rows:
            all_x.add(row[13]); all_x.add(row[15])
            all_y.add(row[14]); all_y.add(row[16])
    if sec_beam_rows:
        for row in sec_beam_rows:
            all_x.add(row[12]); all_x.add(row[14])
            all_y.add(row[13]); all_y.add(row[15])
    if balcony_rows:
        for row in balcony_rows:
            all_x.add(row[1]); all_x.add(row[3])
            all_y.add(row[2]); all_y.add(row[4])
    if ew_rows:
        for row in ew_rows:
            all_x.add(row[10]); all_x.add(row[12])
            all_y.add(row[11]); all_y.add(row[13])

    x_map = _build_level_map(sorted(all_x), set(node_x_vals))
    y_map = _build_level_map(sorted(all_y), set(node_y_vals))

    if not x_map and not y_map:
        print("\nFinal leveling: no adjustments needed.")
        return rect_rows, sec_beam_rows, balcony_rows, ew_rows

    changes = []
    for axis, m in [("X", x_map), ("Y", y_map)]:
        for old, new in sorted(m.items()):
            changes.append(f"  {axis}: {old} -> {new}")

    if changes:
        print(f"\nFinal leveling ({FINAL_LEVEL_TOL*1000:.0f}mm tolerance):")
        for c in changes:
            print(c)

    # Apply mappings
    if rect_rows:
        rect_rows = [
            row[:13] + (
                x_map.get(row[13], row[13]),
                y_map.get(row[14], row[14]),
                x_map.get(row[15], row[15]),
                y_map.get(row[16], row[16]),
            )
            for row in rect_rows
        ]
    if sec_beam_rows:
        sec_beam_rows = [
            row[:12] + (
                x_map.get(row[12], row[12]),
                y_map.get(row[13], row[13]),
                x_map.get(row[14], row[14]),
                y_map.get(row[15], row[15]),
            )
            for row in sec_beam_rows
        ]
    if balcony_rows:
        balcony_rows = [
            (row[0],
             x_map.get(row[1], row[1]),
             y_map.get(row[2], row[2]),
             x_map.get(row[3], row[3]),
             y_map.get(row[4], row[4]))
            for row in balcony_rows
        ]
    if ew_rows:
        ew_rows = [
            row[:10] + (
                x_map.get(row[10], row[10]),
                y_map.get(row[11], row[11]),
                x_map.get(row[12], row[12]),
                y_map.get(row[13], row[13]),
            )
            for row in ew_rows
        ]

    return rect_rows, sec_beam_rows, balcony_rows, ew_rows


def _build_level_map(sorted_vals, node_vals):
    """Group values within FINAL_LEVEL_TOL. Node values take priority as representative.
    Returns dict mapping old_value -> new_value (only for changed values)."""
    mapping = {}
    used = set()
    for i, val in enumerate(sorted_vals):
        if val in used:
            continue
        group = [val]
        used.add(val)
        for j in range(i + 1, len(sorted_vals)):
            if sorted_vals[j] in used:
                continue
            if sorted_vals[j] - val <= FINAL_LEVEL_TOL:
                group.append(sorted_vals[j])
                used.add(sorted_vals[j])
        if len(group) < 2:
            continue
        # Pick representative: prefer node value, else average
        node_in_group = [v for v in group if v in node_vals]
        if node_in_group:
            rep = node_in_group[0]
        else:
            rep = round(sum(group) / len(group), 3)
        for v in group:
            if v != rep:
                mapping[v] = rep
    return mapping


# ─── Validation warnings ────────────────────────────────────────────

def print_validation_warnings(node_x_vals, node_y_vals, rect_rows, sec_beam_rows):
    """Print warnings for edge cases that need engineer review."""
    warnings = []

    if sec_beam_rows:
        # 1. Near-miss node snaps (150-300mm from a node)
        first_rows = {}
        for row in sec_beam_rows:
            if row[0] not in first_rows:
                first_rows[row[0]] = row

        for row in first_rows.values():
            for label, val, nodes, axis in [
                ("X1", row[12], node_x_vals, "X"),
                ("Y1", row[13], node_y_vals, "Y"),
                ("X2", row[14], node_x_vals, "X"),
                ("Y2", row[15], node_y_vals, "Y"),
            ]:
                nearest = min(nodes, key=lambda n: abs(n - val))
                gap = abs(val - nearest)
                if SNAP_TOLERANCE < gap < 0.300:
                    warnings.append(
                        f"  {row[1]} {label}={val:.3f} is {gap*1000:.0f}mm from "
                        f"node {axis}={nearest:.3f} (>{SNAP_TOLERANCE*1000:.0f}mm snap tolerance) "
                        f"- check picking in AutoCAD")

        # 2. Secondary beam passing through a rectangle zone
        if rect_rows:
            for sb in first_rows.values():
                sx1, sy1, sx2, sy2 = sb[12], sb[13], sb[14], sb[15]
                for rect in rect_rows:
                    rx1, ry1, rx2, ry2 = rect[13], rect[14], rect[15], rect[16]
                    # Normalize rect coords (min/max)
                    r_xmin, r_xmax = min(rx1, rx2), max(rx1, rx2)
                    r_ymin, r_ymax = min(ry1, ry2), max(ry1, ry2)

                    # Check if beam is horizontal (same Y) passing through rect Y range
                    if abs(sy1 - sy2) < 0.01:  # horizontal beam
                        beam_y = sy1
                        if r_ymin < beam_y < r_ymax:
                            # Check X overlap
                            b_xmin, b_xmax = min(sx1, sx2), max(sx1, sx2)
                            if b_xmin < r_xmax and b_xmax > r_xmin:
                                warnings.append(
                                    f"  {sb[1]} at Y={beam_y:.3f} passes through "
                                    f"{rect[1]} zone Y=({r_ymin:.3f}-{r_ymax:.3f}) "
                                    f"- verify in drawing")

                    # Check if beam is vertical (same X) passing through rect X range
                    if abs(sx1 - sx2) < 0.01:  # vertical beam
                        beam_x = sx1
                        if r_xmin < beam_x < r_xmax:
                            # Check Y overlap
                            b_ymin, b_ymax = min(sy1, sy2), max(sy1, sy2)
                            if b_ymin < r_ymax and b_ymax > r_ymin:
                                warnings.append(
                                    f"  {sb[1]} at X={beam_x:.3f} passes through "
                                    f"{rect[1]} zone X=({r_xmin:.3f}-{r_xmax:.3f}) "
                                    f"- verify in drawing")

    if warnings:
        print(f"\n{'='*60}")
        print(f"VALIDATION WARNINGS ({len(warnings)}) - review in AutoCAD:")
        print(f"{'='*60}")
        for w in warnings:
            print(w)
    else:
        print("\nValidation: no warnings.")


# ─── Primary beam generation ─────────────────────────────────────────

# # ─── Primary beam generation (COMMENTED OUT - will need later) ───────
# FLOOR_NAMES = ["Stilt floor", "Typical floor", "Terrace", "Mumty"]
# COORD_MATCH_TOL = 0.001  # 1mm tolerance for grouping same grid line
#
#
# def generate_primary_beams(node_coords):
#     """Generate primary beams by connecting adjacent nodes on same grid lines.
#
#     Groups nodes by same X (vertical beams) and same Y (horizontal beams).
#     Adjacent nodes on each grid line form one beam.
#     Returns list of beam dicts with (x1, y1, x2, y2).
#     """
#     beams = []
#
#     # Horizontal beams: group by same Y, sort by X, connect adjacent
#     y_groups = {}
#     for nx, ny in node_coords:
#         matched = False
#         for key_y in y_groups:
#             if abs(ny - key_y) < COORD_MATCH_TOL:
#                 y_groups[key_y].append((nx, ny))
#                 matched = True
#                 break
#         if not matched:
#             y_groups[ny] = [(nx, ny)]
#
#     for key_y in sorted(y_groups):
#         points = sorted(y_groups[key_y], key=lambda p: p[0])
#         for i in range(len(points) - 1):
#             beams.append({
#                 "x1": points[i][0],
#                 "y1": points[i][1],
#                 "x2": points[i + 1][0],
#                 "y2": points[i + 1][1],
#             })
#
#     # Vertical beams: group by same X, sort by Y, connect adjacent
#     x_groups = {}
#     for nx, ny in node_coords:
#         matched = False
#         for key_x in x_groups:
#             if abs(nx - key_x) < COORD_MATCH_TOL:
#                 x_groups[key_x].append((nx, ny))
#                 matched = True
#                 break
#         if not matched:
#             x_groups[nx] = [(nx, ny)]
#
#     for key_x in sorted(x_groups):
#         points = sorted(x_groups[key_x], key=lambda p: p[1])
#         for i in range(len(points) - 1):
#             beams.append({
#                 "x1": points[i][0],
#                 "y1": points[i][1],
#                 "x2": points[i + 1][0],
#                 "y2": points[i + 1][1],
#             })
#
#     return beams


# ─── Debug sheets (debug_x / debug_y) ─────────────────────────────────

def _collect_debug_values(node_coords, beam_rows, rect_rows, balcony_rows, sec_beam_rows,
                          ew_rows=None):
    """Collect all final X and Y values from all sheets with their source labels.
    Returns (x_entries, y_entries) where each entry is (value, source_label)."""
    x_entries = []
    y_entries = []

    # Node coordinates
    for nx, ny in node_coords:
        x_entries.append((nx, "Node"))
        y_entries.append((ny, "Node"))

    # Primary beams
    if beam_rows:
        for beam in beam_rows:
            x_entries.append((beam["x1"], "Primary beam"))
            x_entries.append((beam["x2"], "Primary beam"))
            y_entries.append((beam["y1"], "Primary beam"))
            y_entries.append((beam["y2"], "Primary beam"))

    # Rectangles (snapped values: indices 13-16)
    if rect_rows:
        for row in rect_rows:
            x_entries.append((row[13], "Rectangle"))
            x_entries.append((row[15], "Rectangle"))
            y_entries.append((row[14], "Rectangle"))
            y_entries.append((row[16], "Rectangle"))

    # Balconies (indices 1-4: x1, y1, x2, y2)
    if balcony_rows:
        for row in balcony_rows:
            x_entries.append((row[1], "Balcony"))
            x_entries.append((row[3], "Balcony"))
            y_entries.append((row[2], "Balcony"))
            y_entries.append((row[4], "Balcony"))

    # Secondary beams (snapped values: indices 12-15)
    if sec_beam_rows:
        for row in sec_beam_rows:
            x_entries.append((row[12], "Secondary beam"))
            x_entries.append((row[14], "Secondary beam"))
            y_entries.append((row[13], "Secondary beam"))
            y_entries.append((row[15], "Secondary beam"))

    # Extra walls (snapped values: indices 10-13)
    if ew_rows:
        for row in ew_rows:
            x_entries.append((row[10], "Extra wall"))
            x_entries.append((row[12], "Extra wall"))
            y_entries.append((row[11], "Extra wall"))
            y_entries.append((row[13], "Extra wall"))

    return x_entries, y_entries


def _build_debug_rows(entries):
    """Build sorted unique rows with sources and differences.
    Returns list of (source_str, value, diff_or_None, is_red)."""
    # Group sources by unique value
    value_sources = {}
    for val, src in entries:
        if val is None:
            continue
        val = round(val, 3)
        if val not in value_sources:
            value_sources[val] = set()
        value_sources[val].add(src)

    # Sort by value ascending
    sorted_vals = sorted(value_sources.keys())

    rows = []
    for i, val in enumerate(sorted_vals):
        sources = ", ".join(sorted(value_sources[val]))
        if i == 0:
            diff = None
            is_red = False
        else:
            diff = round(val - sorted_vals[i - 1], 3)
            is_red = diff > DEBUG_DIFF_THRESHOLD
        rows.append((sources, val, diff, is_red))

    return rows


def write_debug_sheets(wb, node_coords, beam_rows, rect_rows, balcony_rows, sec_beam_rows,
                       ew_rows=None):
    """Add debug_x and debug_y sheets to the workbook."""
    x_entries, y_entries = _collect_debug_values(
        node_coords, beam_rows, rect_rows, balcony_rows, sec_beam_rows, ew_rows=ew_rows)

    for axis, entries in [("x", x_entries), ("y", y_entries)]:
        debug_rows = _build_debug_rows(entries)
        if not debug_rows:
            continue

        ws = wb.create_sheet(f"debug_{axis}")
        headers = ["Source", f"{axis.upper()} Value (m)", "Difference (m)"]
        widths = [30, 18, 18]
        _write_header(ws, headers, widths)

        for ri, (sources, val, diff, is_red) in enumerate(debug_rows, 2):
            _write_row(ws, ri, [sources, val, diff])
            if is_red:
                ws.cell(ri, 3).fill = RED_FILL




# ─── Extra wall reading & snapping ───────────────────────────────────

def read_and_offset_extra_walls(input_path):
    """Read Extra wall coordinates sheet and apply wall offset only.
    Returns list of mutable lists for post-processing."""
    wb = load_workbook(input_path, data_only=True)
    if "Extra wall coordinates" not in wb.sheetnames:
        wb.close()
        return None

    ws = wb["Extra wall coordinates"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        no = row[0]
        etype = row[1]
        orig_x1 = row[2]
        orig_y1 = row[3]
        orig_x2 = row[4]
        orig_y2 = row[5]
        wall_loc = row[6]
        floor = row[7]
        present = row[8]
        wall_thickness = row[9]

        # Apply wall offset (own axis + perpendicular)
        off_x1, off_y1, off_x2, off_y2 = _apply_wall_offset(
            orig_x1, orig_y1, orig_x2, orig_y2, wall_loc, wall_thickness)

        rows.append([no, etype, orig_x1, orig_y1, orig_x2, orig_y2,
                     wall_loc, floor, present, wall_thickness,
                     off_x1, off_y1, off_x2, off_y2])

    wb.close()
    return rows if rows else None


def snap_extra_walls(ew_rows, node_x_vals, node_y_vals, sec_beam_rows=None):
    """Apply 3-step snapping to extra wall rows.
    Step 1: Group all offset values within 150mm → average
    Step 2: Swap with nearest node within 150mm
    Step 3: If not on node → swap with nearest sec beam value within 150mm."""
    if not ew_rows:
        return ew_rows

    EW_SNAP_TOL = SNAP_TOLERANCE  # 150mm

    # ── Step 1: Group nearby values → average ──
    first_rows = {}
    for row in ew_rows:
        if row[1] not in first_rows:
            first_rows[row[1]] = row

    all_x = []
    all_y = []
    for i, (etype, row) in enumerate(first_rows.items()):
        all_x.append((i * 2, row[10]))      # off_x1
        all_x.append((i * 2 + 1, row[12]))  # off_x2
        all_y.append((i * 2, row[11]))       # off_y1
        all_y.append((i * 2 + 1, row[13]))   # off_y2

    x_map = _group_and_average(all_x, EW_SNAP_TOL)
    y_map = _group_and_average(all_y, EW_SNAP_TOL)

    # Build replacement maps
    etypes = list(first_rows.keys())
    replace = {}
    for i, etype in enumerate(etypes):
        if i * 2 in x_map:
            replace[(etype, 10)] = x_map[i * 2]
        if i * 2 + 1 in x_map:
            replace[(etype, 12)] = x_map[i * 2 + 1]
        if i * 2 in y_map:
            replace[(etype, 11)] = y_map[i * 2]
        if i * 2 + 1 in y_map:
            replace[(etype, 13)] = y_map[i * 2 + 1]

    if replace:
        for row in ew_rows:
            for ci in (10, 11, 12, 13):
                key = (row[1], ci)
                if key in replace:
                    row[ci] = replace[key]
        print(f"  EW step 1: grouped {len(replace)} values (averaged)")

    # ── Step 2: Swap with nearest node within 150mm ──
    node_swaps = 0
    processed = {}
    for row in ew_rows:
        etype = row[1]
        if etype not in processed:
            for ci, axis_vals in [(10, node_x_vals), (12, node_x_vals),
                                  (11, node_y_vals), (13, node_y_vals)]:
                val = row[ci]
                best = None
                best_d = EW_SNAP_TOL + 1
                for nv in axis_vals:
                    d = abs(val - nv)
                    if d <= EW_SNAP_TOL and d < best_d:
                        best = nv
                        best_d = d
                if best is not None and best != val:
                    row[ci] = best
                    node_swaps += 1
            processed[etype] = (row[10], row[11], row[12], row[13])
        else:
            row[10], row[11], row[12], row[13] = processed[etype]

    if node_swaps:
        print(f"  EW step 2: swapped {node_swaps} values with nodes")

    # ── Step 3: If not on node → swap with nearest sec beam value ──
    sec_swaps = 0
    if sec_beam_rows:
        sec_x_vals = sorted(set(r[12] for r in sec_beam_rows if r[12] is not None) |
                            set(r[14] for r in sec_beam_rows if r[14] is not None))
        sec_y_vals = sorted(set(r[13] for r in sec_beam_rows if r[13] is not None) |
                            set(r[15] for r in sec_beam_rows if r[15] is not None))

        processed2 = {}
        for row in ew_rows:
            etype = row[1]
            if etype not in processed2:
                for ci, sec_vals in [(10, sec_x_vals), (12, sec_x_vals),
                                     (11, sec_y_vals), (13, sec_y_vals)]:
                    val = row[ci]
                    is_x = ci in (10, 12)
                    axis_nodes = node_x_vals if is_x else node_y_vals
                    if val in axis_nodes:
                        continue
                    best = None
                    best_d = EW_SNAP_TOL + 1
                    for sv in sec_vals:
                        d = abs(val - sv)
                        if d <= EW_SNAP_TOL and d < best_d:
                            best = sv
                            best_d = d
                    if best is not None and best != val:
                        row[ci] = best
                        sec_swaps += 1
                processed2[etype] = (row[10], row[11], row[12], row[13])
            else:
                row[10], row[11], row[12], row[13] = processed2[etype]

    if sec_swaps:
        print(f"  EW step 3: swapped {sec_swaps} values with sec beams")

    return [tuple(row) for row in ew_rows]


# ─── Excel output ────────────────────────────────────────────────────

def read_staircase_details(filepath):
    """Read Staircase details sheet (pass-through, no processing)."""
    wb = load_workbook(filepath, data_only=True)
    if "Staircase details" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["Staircase details"]
    rows = []
    for r in range(1, ws.max_row + 1):
        row_data = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        rows.append(row_data)
    wb.close()
    return rows


def write_output(columns, node_coords, raw_rows, beam_rows, rect_rows, balcony_rows,
                 sec_beam_rows, output_path, staircase_rows=None, ew_rows=None):
    """Write 2 output files:
    1. node_coordinates_*.xlsx - Node coordinates sheet only
    2. other_coordinates_*.xlsx - Primary beam, Rectangle, Balcony, Secondary beam, Staircase details
    """
    # ─── File 1: Node coordinates ───
    wb_node = Workbook()
    ws_node = wb_node.active
    ws_node.title = "Node coordinates"
    raw_headers = [
        "No.", "Type", "Coordinate X (m)", "Coordinate Y (m)", "Location",
        "Anchor location", "Beam X width location",
        "Opposite coordinate X (m)", "Opposite coordinate Y (m)",
        "Floor", "Beam width X (mm)", "Beam depth X (mm)", "Wall thickness X (mm)",
        "Beam Y width location",
        "Opposite coordinate X (m)", "Opposite coordinate Y (m)",
        "Floor", "Beam width Y (mm)", "Beam depth Y (mm)", "Wall thickness Y (mm)",
        "Column orientation",
        "Node X (m)", "Node Y (m)",
    ]
    raw_widths = [6, 8, 18, 18, 16, 16, 20, 18, 18, 16, 18, 18, 18, 20, 18, 18, 16, 18, 18, 18, 20, 18, 18]
    _write_header(ws_node, raw_headers, raw_widths)
    ri = 2
    for col_rows, (nx, ny) in zip(raw_rows, node_coords):
        for floor_row in col_rows:
            _write_row(ws_node, ri, floor_row + [nx, ny])
            ri += 1
    # Sheet 2: Secondary beam coordinates (in node_coordinates file)
    if sec_beam_rows:
        ws_sec = wb_node.create_sheet("Secondary beam coordinates")
        headers = ["No.", "Type", "Coordinate X1 (m)", "Coordinate Y1 (m)",
                   "Coordinate X2 (m)", "Coordinate Y2 (m)", "Beam location",
                   "Floor", "Present", "Beam width (mm)", "Beam depth (mm)",
                   "Wall thickness (mm)", "Snapped X1 (m)", "Snapped Y1 (m)",
                   "Snapped X2 (m)", "Snapped Y2 (m)"]
        widths = [6, 8, 18, 18, 18, 18, 16, 16, 10, 18, 18, 18, 18, 18, 18, 18]
        _write_header(ws_sec, headers, widths)
        for ri, row_data in enumerate(sec_beam_rows, 2):
            _write_row(ws_sec, ri, list(row_data))

    wb_node.save(output_path)
    print(f"Node coordinates saved: {output_path}")

    # ─── File 2: Other coordinates ───
    other_path = output_path.parent / output_path.name.replace("node_coordinates", "other_coordinates")
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Sheet 1: Primary beam
    if beam_rows:
        ws_beam = wb.create_sheet("Primary beam")
        headers = ["No.", "Type", "Coordinate X1 (m)", "Coordinate Y1 (m)",
                   "Coordinate X2 (m)", "Coordinate Y2 (m)", "Floor",
                   "Present", "Beam width X (mm)", "Wall thickness (mm)"]
        widths = [6, 8, 18, 18, 18, 18, 16, 10, 18, 18]
        _write_header(ws_beam, headers, widths)
        ri = 2
        for beam_num, beam in enumerate(beam_rows, 1):
            for floor in FLOOR_NAMES:
                _write_row(ws_beam, ri, [
                    beam_num, f"P{beam_num}",
                    beam["x1"], beam["y1"], beam["x2"], beam["y2"],
                    floor, "", "", "",
                ])
                ri += 1

    # Sheet 2: Rectangle coordinates
    if rect_rows:
        ws_rect = wb.create_sheet("Rectangle coordinates")
        headers = ["No.", "Type", "Coordinate X1 (m)", "Coordinate Y1 (m)",
                   "Coordinate X2 (m)", "Coordinate Y2 (m)", "Location",
                   "Anchor location 1", "Beam X width location 1", "Beam Y width location 1",
                   "Anchor location 2", "Beam X width location 2", "Beam Y width location 2",
                   "Snapped X1 (m)", "Snapped Y1 (m)", "Snapped X2 (m)", "Snapped Y2 (m)"]
        widths = [6, 8, 18, 18, 18, 18, 16, 16, 20, 20, 16, 20, 20, 18, 18, 18, 18]
        _write_header(ws_rect, headers, widths)
        for ri, row_data in enumerate(rect_rows, 2):
            _write_row(ws_rect, ri, list(row_data))

    # Sheet 3: Balcony coordinates
    if balcony_rows:
        ws_bal = wb.create_sheet("Balcony coordinates")
        headers = ["Location", "Coordinate X1 (m)", "Coordinate Y1 (m)",
                   "Coordinate X2 (m)", "Coordinate Y2 (m)"]
        widths = [20, 18, 18, 18, 18]
        _write_header(ws_bal, headers, widths)
        for ri, row_data in enumerate(balcony_rows, 2):
            _write_row(ws_bal, ri, list(row_data))

    # Sheet 4: Extra wall coordinates
    if ew_rows:
        ws_ew = wb.create_sheet("Extra wall coordinates")
        headers = ["No.", "Type", "Coordinate X1 (m)", "Coordinate Y1 (m)",
                   "Coordinate X2 (m)", "Coordinate Y2 (m)",
                   "Wall location", "Floor", "Present", "Wall thickness (mm)",
                   "Snapped X1 (m)", "Snapped Y1 (m)", "Snapped X2 (m)", "Snapped Y2 (m)"]
        widths = [6, 8, 18, 18, 18, 18, 16, 16, 10, 18, 18, 18, 18, 18]
        _write_header(ws_ew, headers, widths)
        for ri, row_data in enumerate(ew_rows, 2):
            _write_row(ws_ew, ri, list(row_data))

    # Sheet 5: Staircase details (pass-through from floor_coordinates)
    if staircase_rows:
        ws_stair = wb.create_sheet("Staircase details")
        for ri, row_data in enumerate(staircase_rows, 1):
            for ci, val in enumerate(row_data, 1):
                ws_stair.cell(ri, ci, val)

    # Sheet 6-7: Debug sheets (debug_x, debug_y)
    write_debug_sheets(wb, node_coords, beam_rows, rect_rows, balcony_rows, sec_beam_rows,
                       ew_rows=ew_rows)

    wb.save(other_path)
    print(f"Other coordinates saved: {other_path}")


# ─── Main ────────────────────────────────────────────────────────────

def _find_latest_by_pattern(folder, pattern):
    """Find the latest file matching a glob pattern in a folder."""
    import glob, os
    matches = glob.glob(str(Path(folder) / pattern))
    if not matches:
        return None
    return Path(max(matches, key=os.path.getmtime))


def main():
    STD_ANL = Path(r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model")

    if len(sys.argv) >= 2:
        # ── Manual mode: file paths from command line ──
        columns_path = Path(sys.argv[1])
        if not columns_path.exists():
            print(f"Error: File not found: {columns_path}")
            sys.exit(1)

        if (len(sys.argv) >= 3
                and not sys.argv[2].lower().endswith("output.xlsx")
                and Path(sys.argv[2]).exists()
                and Path(sys.argv[2]).suffix.lower() == ".xlsx"):
            rect_bal_path = Path(sys.argv[2])
            output_path = Path(sys.argv[3]) if len(sys.argv) >= 4 else None
        else:
            rect_bal_path = columns_path
            output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else None

        # Try to find plinth secondary beams file in same folder
        sec_plinth_path = _find_latest_by_pattern(
            columns_path.parent, "floor_coordinates_secondary_coordinates_plinth*.xlsx")

        if output_path is None:
            from datetime import datetime as _dt
            timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
            output_path = columns_path.parent / f"node_coordinates_{timestamp}.xlsx"

    else:
        # ── Auto-find mode: find latest 3 files from STD ANL model ──
        print("Auto-finding input files...")
        columns_path = _find_latest_by_pattern(
            STD_ANL, "floor_coordinates_column_coordinates_*.xlsx")
        rect_bal_path = _find_latest_by_pattern(
            STD_ANL, "floor_coordinates_secondary_coordinates_nonplinth_*.xlsx")
        sec_plinth_path = _find_latest_by_pattern(
            STD_ANL, "floor_coordinates_secondary_coordinates_plinth*.xlsx")

        if columns_path is None:
            # Fallback: try old single-file format
            columns_path = _find_latest_by_pattern(STD_ANL, "floor_coordinates_*.xlsx")
            rect_bal_path = columns_path
            sec_plinth_path = None

        if columns_path is None:
            print("Error: No floor_coordinates files found in STD ANL model.")
            sys.exit(1)

        if rect_bal_path is None:
            rect_bal_path = columns_path

        from datetime import datetime as _dt
        timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        output_path = columns_path.parent / f"node_coordinates_{timestamp}.xlsx"

    print(f"Columns file:     {columns_path}")
    print(f"Rect/Balcony file: {rect_bal_path}")
    if sec_plinth_path:
        print(f"Sec plinth file:  {sec_plinth_path}")
    else:
        print(f"Sec plinth file:  (not found)")

    # Step 1: Calculate node coordinates
    columns, raw_rows = read_columns(columns_path)
    print(f"\nRead {len(columns)} columns")

    node_coords = []
    for col in columns:
        nx, ny = calc_node_coordinate(col)
        node_coords.append((nx, ny))
        print(f"  {col['type']:>4s}  anchor=({col['anchor_x']}, {col['anchor_y']})  "
              f"->  node=({nx}, {ny})  [{col['location']}]")

    # Step 1.5: Level node coordinates (merge values within 50mm)
    leveled_coords = level_node_coords(node_coords)
    leveled_count = sum(1 for a, b in zip(node_coords, leveled_coords) if a != b)
    if leveled_count:
        print(f"\nLeveled {leveled_count} node values (50mm tolerance):")
        for col, old, new in zip(columns, node_coords, leveled_coords):
            if old != new:
                print(f"  {col['type']}: ({old[0]}, {old[1]}) -> ({new[0]}, {new[1]})")
    node_coords = leveled_coords

    # Step 2: Extract unique node X and Y values for snapping
    unique_x = sorted(set(nx for nx, ny in node_coords))
    unique_y = sorted(set(ny for nx, ny in node_coords))
    print(f"\nUnique node X values: {unique_x}")
    print(f"Unique node Y values: {unique_y}")

    # # Step 3: Generate primary beams from node grid lines (COMMENTED OUT - will need later)
    # beam_rows = generate_primary_beams(node_coords)
    # print(f"\nGenerated {len(beam_rows)} primary beams:")
    # for i, b in enumerate(beam_rows, 1):
    #     print(f"  P{i}: ({b['x1']}, {b['y1']}) -> ({b['x2']}, {b['y2']})")
    beam_rows = None

    # Step 4: Read secondary beams (offset only, no snap yet)
    sec_beam_rows_nonplinth = read_and_offset_secondary_beams(rect_bal_path)
    sec_beam_rows_plinth = None
    if sec_plinth_path:
        sec_beam_rows_plinth = read_and_offset_secondary_beams(sec_plinth_path)

    # Merge plinth + nonplinth secondary beams
    sec_beam_rows = None
    if sec_beam_rows_plinth and sec_beam_rows_nonplinth:
        sec_beam_rows = list(sec_beam_rows_plinth) + list(sec_beam_rows_nonplinth)
    elif sec_beam_rows_plinth:
        sec_beam_rows = list(sec_beam_rows_plinth)
    elif sec_beam_rows_nonplinth:
        sec_beam_rows = list(sec_beam_rows_nonplinth)

    # Step 4.5: Apply 2-step snap (group+average → node swap)
    sec_beam_rows = snap_secondary_beams(sec_beam_rows, unique_x, unique_y)

    if sec_beam_rows:
        seen = set()
        unique_beams = []
        for row in sec_beam_rows:
            key = (row[0], row[1])
            if key not in seen:
                seen.add(key)
                unique_beams.append(row)
        print(f"\nSnapped {len(unique_beams)} secondary beams:")
        for row in unique_beams:
            print(f"  {row[1]}: ({row[12]}, {row[13]}) -> ({row[14]}, {row[15]})  [{row[6]}]")
    else:
        print("\nNo secondary beam data found.")

    # Step 4.7: Read and snap extra wall coordinates (3-step: group → node → sec beam)
    ew_rows = read_and_offset_extra_walls(rect_bal_path)
    ew_rows = snap_extra_walls(ew_rows, unique_x, unique_y, sec_beam_rows=sec_beam_rows)
    if ew_rows:
        seen_ew = set()
        unique_ews = []
        for row in ew_rows:
            if row[1] not in seen_ew:
                seen_ew.add(row[1])
                unique_ews.append(row)
        print(f"\nSnapped {len(unique_ews)} extra walls:")
        for row in unique_ews:
            print(f"  {row[1]}: ({row[10]}, {row[11]}) -> ({row[12]}, {row[13]})  [{row[6]}]")
    else:
        print("\nNo extra wall data found.")

    # Step 5: Read and snap rectangle coordinates (3-step logic, uses sec beam values)
    rect_rows = read_and_snap_rectangles(rect_bal_path, unique_x, unique_y,
                                         sec_beam_rows=sec_beam_rows)
    if rect_rows:
        print(f"\nSnapped {len(rect_rows)} rectangles:")
        for row in rect_rows:
            print(f"  {row[1]}: ({row[13]}, {row[14]}) - ({row[15]}, {row[16]})  [{row[6]}]")
    else:
        print("\nNo rectangle data found.")

    # Step 5.5: Read and snap balcony coordinates (offset + node swap + sec beam swap)
    balcony_rows = read_and_snap_balconies(rect_bal_path, unique_x, unique_y,
                                           sec_beam_rows=sec_beam_rows)
    if balcony_rows:
        print(f"\nSnapped {len(balcony_rows)} balconies:")
        for row in balcony_rows:
            print(f"  {row[0]}: ({row[1]}, {row[2]}) - ({row[3]}, {row[4]})")
    else:
        print("No balcony data found.")

    # # ── Old cross-snap (replaced by 2-step snap in snap_secondary_beams) ──
    # sec_beam_rows = cross_snap_secondary_beams(
    #     sec_beam_rows, rect_rows, unique_x, unique_y)

    # Step 7: Final leveling pass (5mm) for picking precision
    rect_rows, sec_beam_rows, balcony_rows, ew_rows = final_level_all(
        rect_rows, sec_beam_rows, balcony_rows, unique_x, unique_y, ew_rows=ew_rows)

    # Step 8: Validation warnings
    print_validation_warnings(unique_x, unique_y, rect_rows, sec_beam_rows)

    # Step 8.5: Read staircase details (pass-through)
    staircase_rows = read_staircase_details(rect_bal_path)
    if staircase_rows:
        print(f"\nStaircase details: {len(staircase_rows) - 1} rows (pass-through)")
    else:
        print("\nNo staircase details found.")

    # Step 9: Write output
    write_output(columns, node_coords, raw_rows, beam_rows, rect_rows, balcony_rows,
                 sec_beam_rows, output_path, staircase_rows, ew_rows=ew_rows)

    # Step 10: Check for close value pairs and show popup if any
    _show_close_pairs_popup(node_coords, beam_rows, rect_rows, balcony_rows, sec_beam_rows,
                            ew_rows=ew_rows)


def _show_close_pairs_popup(node_coords, beam_rows, rect_rows, balcony_rows, sec_beam_rows,
                            ew_rows=None):
    """Check debug values for close pairs (<= threshold) and show Windows popup if found."""
    x_entries, y_entries = _collect_debug_values(
        node_coords, beam_rows, rect_rows, balcony_rows, sec_beam_rows, ew_rows=ew_rows)

    close_pairs = []
    for axis, entries in [("X", x_entries), ("Y", y_entries)]:
        debug_rows = _build_debug_rows(entries)
        for sources, val, diff, is_red in debug_rows:
            if diff is not None and not is_red and diff > 0.001:
                # White row with actual difference (> 1mm, <= 60mm)
                close_pairs.append(
                    f"{axis}={val:.3f}  diff={diff*1000:.0f}mm  ({sources})")

    if not close_pairs:
        return

    # Build message
    msg = "Close value pairs found (< 60mm difference):\n\n"
    for pair in close_pairs:
        msg += f"  {pair}\n"

    print(f"\nWARNING: {len(close_pairs)} close value pair(s) found:")
    for p in close_pairs:
        print(f"  {p}")

    # Show Windows popup
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()  # hide main window
        messagebox.showwarning("Close Value Pairs Found", msg)
        root.destroy()
    except Exception:
        pass  # if tkinter not available, console warning is enough


if __name__ == "__main__":
    main()
