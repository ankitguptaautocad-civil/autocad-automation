"""
Elevation Calculator
--------------------
Reads building info and writes YD/ZD elevation columns (F onwards) to Filleddata Datadata2 sheet.

Usage: python elevation_calculator.py
No arguments needed - uses hardcoded paths.

Pipeline step 4: Run AFTER node_coordinate_calculator.py, BEFORE gridline_calculator.py.
"""

from openpyxl import load_workbook

# ── Paths ──────────────────────────────────────────────────────────────
BUILDING_INFO_PATH = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\building info.xlsx"
FILLEDDATA_PATH = r"D:\JARVIS back up 16092025\JARVIS backup\STD ANL model\Datadata.xlsx"


def read_building_info():
    """Read building info from the fixed Excel path.
    Returns info_dict with keys like 'Floors', 'Story height', etc."""
    wb = load_workbook(BUILDING_INFO_PATH, data_only=True)
    ws = wb["building info"]
    info = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is not None:
            key = str(row[0]).strip()
            info[key] = row[1]
        if len(row) > 4 and row[4] is not None:
            key = str(row[4]).strip()
            info[key] = row[5]
    wb.close()
    return info


def generate_elevation_columns(info):
    """Generate YD/ZD header pairs and default values from building info.

    Returns:
        headers: list of header strings e.g. ["YD(-2)", "ZD(-2)", "YD(0)", "ZD(0)", ...]
        defaults: list of default values for each header
    """
    y0_shift = float(info["Y = 0 Shift Above Base"])
    floors = int(info["Floors"])
    story_height = float(info["Story height"])
    mumty_height = float(info["Mumty height"])

    # Build elevation list: -Y0_shift, 0, then (floors-1) regular floors, then mumty
    elevations = []
    elevations.append(-y0_shift)           # foundation (e.g. -2)
    elevations.append(0)                   # base always 0
    cumulative = 0
    for _ in range(floors - 1):            # 4 regular floors (0 is already added)
        cumulative = round(cumulative + story_height, 3)
        elevations.append(cumulative)
    elevations.append(round(cumulative + mumty_height, 3))  # mumty

    # Generate headers and defaults
    headers = []
    defaults = []
    for i, elev in enumerate(elevations):
        # Format elevation: remove trailing zeros for cleaner display
        elev_str = f"{elev:g}"
        headers.extend([f"YD({elev_str})", f"ZD({elev_str})"])

        # Default ZD: 600 for all floors
        yd = 300
        zd = 600
        defaults.extend([yd, zd])

    return headers, defaults


def write_yd_zd_to_filleddata(headers, defaults):
    """Write YD/ZD elevation columns to Datadata2 sheet.
    Finds existing YD/ZD range dynamically from row 1 headers, clears only that range."""
    wb = load_workbook(FILLEDDATA_PATH)
    ws = wb["Datadata2"]

    # Find existing YD/ZD range by scanning row 1
    first_yd_col = None
    last_zd_col = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val and str(val).startswith("YD(") and first_yd_col is None:
            first_yd_col = c
        if val and str(val).startswith("ZD("):
            last_zd_col = c

    # If no existing YD/ZD found (first run), find "Location" header and start after it
    if first_yd_col is None:
        for c in range(1, ws.max_column + 1):
            if ws.cell(1, c).value and str(ws.cell(1, c).value).strip() == "Location":
                first_yd_col = c + 1
                break
        if first_yd_col is None:
            first_yd_col = 6  # fallback to col F

    # Clear old YD/ZD columns (only the range where they exist)
    clear_end = last_zd_col if last_zd_col else first_yd_col
    print(f"Clearing old YD/ZD: col {first_yd_col} to {clear_end}")
    for r in range(1, ws.max_row + 1):
        for c in range(first_yd_col, clear_end + 1):
            ws.cell(r, c).value = None

    # Write new headers in row 1
    num_elev_cols = len(headers)
    for ci, header in enumerate(headers):
        ws.cell(1, first_yd_col + ci, header)

    # Count how many node rows exist (check col A from row 2)
    node_count = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        node_count += 1

    # Write defaults for each node row
    for i in range(2, node_count + 2):
        for ci, val in enumerate(defaults):
            ws.cell(i, first_yd_col + ci, val)

    wb.save(FILLEDDATA_PATH)
    print(f"YD/ZD columns written ({num_elev_cols} cols starting at col {first_yd_col}, {node_count} nodes)")
    print(f"Saved: {FILLEDDATA_PATH}")


def main():
    print("=" * 60)
    print("  ELEVATION CALCULATOR (YD/ZD)")
    print("=" * 60)
    print()

    # 1. Read building info
    print(f"Reading building info from: {BUILDING_INFO_PATH}")
    info = read_building_info()
    print(f"  Floors: {info.get('Floors')}")
    print(f"  Story height: {info.get('Story height')}")
    print(f"  Mumty height: {info.get('Mumty height')}")
    print(f"  Y=0 Shift: {info.get('Y = 0 Shift Above Base')}")

    # 2. Generate elevation columns
    headers, defaults = generate_elevation_columns(info)
    print(f"\nElevation levels: {headers[::2]}")
    print(f"YD defaults: {defaults[::2]}")
    print(f"ZD defaults: {defaults[1::2]}")

    # 3. Write to Filleddata
    print()
    write_yd_zd_to_filleddata(headers, defaults)


if __name__ == "__main__":
    main()
