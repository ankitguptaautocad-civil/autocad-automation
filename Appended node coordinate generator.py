from __future__ import annotations

import argparse
import importlib.util
import sys
from datetime import datetime as _dt
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_PATH = None
DEFAULT_OUTPUT_PATH = None
DEFAULT_WALL_INPUT_PATH = None
DEFAULT_GEOMETRY_INPUT_PATH = None
DEFAULT_DXF_COLUMNS_INPUT_PATH = None
DEFAULT_NODE_LEVEL_TOLERANCE_M = 0.60
BEAM_FACE_MAX_DISTANCE_M = 0.250  # max distance (in metres) between two columns' beam-attaching faces for them to share a gridline. Nominal beam width is 230mm plus 20mm drafting/extraction-noise tolerance. Uses each column's Left/Right (for X) and Front/Back (for Y) labels to pick the relevant face — the FACE the beam connects to, not arbitrary rectangle overlap.
NODE_SOURCE_FLOOR = "Typical floor roof"
RAW_SECONDARY_SHEETS = ("Secondary beam coordinates_plin", "Secondary beam coordinates_nonp")
FINAL_SECONDARY_SHEET = "Secondary beam coordinates"
SHEAR_WALL_TEMPLATE_SHEET = "Shear wall landscape"
COLUMN_LANDSCAPE_TEMPLATE_SHEET = "Column landscape"
SECONDARY_ENDPOINT_MATCH_TOL_M = 0.25
SECONDARY_SNAP_TOL_M = 0.10
EXTRA_WALL_ORTHO_TOL_M = 0.20
EXTRA_WALL_BEAM_EXCLUDE_TOL_M = 0.30
EXTRA_WALL_MIN_EXPORT_LENGTH_M = 0.60
EXTRA_WALL_MIN_FILTERED_LENGTH_M = 2.0
EXTRA_WALL_BALCONY_MIN_FILTERED_LENGTH_M = 1.0
EXTRA_WALL_PRESENT_DEFAULTS = {
    "Plinth": ("YES", "chain"),
    "Stilt roof": ("YES", "chain"),
    "Typical floor roof": ("YES", "chain"),
    "Terrace": ("NO", None),
}
PLINTH_FLOORS = ("Plinth",)
NONPLINTH_FLOORS = ("Stilt roof", "Typical floor roof", "Terrace")


@dataclass(frozen=True)
class ColumnRecord:
    row_idx: int
    column_no: int
    type_name: str
    xmin: float
    xmax: float
    ymin: float
    ymax: float
    left_right: str
    front_back: str
    anchor_x: float
    anchor_y: float
    location: str = ""
    anchor_location: str = ""
    orientation: str = ""


@dataclass(frozen=True)
class PrimaryBeamRecord:
    start_c: str
    end_c: str
    direction: str
    beam_class: str
    floor: str
    beam_width_mm: float
    beam_depth_mm: float
    wall_thickness_mm: float


def normalize_header(value: object) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def candidate_search_dirs() -> list[Path]:
    dirs = []
    for candidate in ((Path.cwd() / "STD ANL model").resolve(), Path.cwd().resolve()):
        if candidate.exists() and candidate not in dirs:
            dirs.append(candidate)
    return dirs


def discover_single_workbook(patterns: tuple[str, ...], label: str) -> Path:
    matches: list[Path] = []
    for folder in candidate_search_dirs():
        for pattern in patterns:
            matches.extend(path.resolve() for path in folder.glob(pattern))
    unique_matches = sorted(set(matches))
    if not unique_matches:
        raise SystemExit(f"Could not find the {label} workbook. Pass it explicitly.")
    if len(unique_matches) == 1:
        return unique_matches[0]
    return max(unique_matches, key=lambda path: (path.stat().st_mtime_ns, path.name.lower()))


def resolve_input_workbook(explicit_input: Path | None) -> Path:
    if explicit_input is not None:
        path = explicit_input.resolve()
        if not path.exists():
            raise SystemExit(f"Input workbook not found: {path}")
        return path
    return discover_single_workbook(
        (
            "*column_beam_pairs_v2.xlsx",
            "*column_beam_pairs.xlsx",
        ),
        "column beam pair",
    )


def resolve_wall_workbook(explicit_walls: Path | None) -> Path:
    if explicit_walls is not None:
        path = explicit_walls.resolve()
        if not path.exists():
            raise SystemExit(f"Wall workbook not found: {path}")
        return path
    return discover_single_workbook(
        (
            "*walls_m_v2.xlsx",
            "*walls*.xlsx",
        ),
        "wall",
    )


def resolve_geometry_workbook(explicit_geometry: Path | None) -> Path:
    if explicit_geometry is not None:
        path = explicit_geometry.resolve()
        if not path.exists():
            raise SystemExit(f"Geometry workbook not found: {path}")
        return path
    return discover_single_workbook(
        (
            "*floor_coordinates_secondary_coordinates_nonplinth*.xlsx",
            "*secondary_coordinates_nonplinth*.xlsx",
            "*floor*nonplinth*.xlsx",
        ),
        "nonplinth floor geometry",
    )


def resolve_dxf_columns_workbook(explicit_path: Path | None) -> Path:
    if explicit_path is not None:
        path = explicit_path.resolve()
        if not path.exists():
            raise SystemExit(f"DXF column rectangles workbook not found: {path}")
        return path
    return discover_single_workbook(
        ("*_col_rectangles_m_v2_wall_assisted.xlsx",),
        "DXF column rectangles",
    )


def load_dxf_column_geometry(path: Path) -> dict[int, dict[str, float]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(str(v or "")): idx for idx, v in enumerate(headers)}
    required = {
        "columnno": "Column No",
        "xminm": "Xmin (m)",
        "xmaxm": "Xmax (m)",
        "yminm": "Ymin (m)",
        "ymaxm": "Ymax (m)",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"DXF column rectangles workbook missing headers: {', '.join(missing)}")
    geometry: dict[int, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        col_no_raw = row[header_map["columnno"]]
        text = str(col_no_raw or "").strip()
        col_no = int(text[1:]) if text.upper().startswith("C") and text[1:].isdigit() else (int(text) if text.isdigit() else -1)
        if col_no < 0:
            continue
        geometry[col_no] = {
            "xmin": round(float(row[header_map["xminm"]]), 3),
            "xmax": round(float(row[header_map["xmaxm"]]), 3),
            "ymin": round(float(row[header_map["yminm"]]), 3),
            "ymax": round(float(row[header_map["ymaxm"]]), 3),
        }
    return geometry


def load_unfiltered_module():
    module_path = (Path(__file__).parent / "Unfiltered column coordinates generator.py").resolve()
    spec = importlib.util.spec_from_file_location("unfiltered_generator_runtime", module_path)
    if spec is None or spec.loader is None:
        raise SystemExit(f"Could not load helper module: {module_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_legacy_node_module():
    module_path = (Path(__file__).parent / "node_coordinate_calculator.py").resolve()
    if not module_path.exists():
        return None
    spec = importlib.util.spec_from_file_location("legacy_node_coordinate_runtime", module_path)
    if spec is None or spec.loader is None:
        raise SystemExit(f"Could not load helper module: {module_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def parse_column_number(raw_value: object, fallback_idx: int) -> int:
    text = str(raw_value or "").strip()
    if text.upper().startswith("C") and text[1:].isdigit():
        return int(text[1:])
    if text.isdigit():
        return int(text)
    return fallback_idx


def _select_anchor(min_val: float, max_val: float, tag: str) -> float:
    if tag in {"Left", "Front"}:
        return round(min_val, 3)
    if tag in {"Right", "Back"}:
        return round(max_val, 3)
    return round((min_val + max_val) / 2.0, 3)


def load_columns(ws, dxf_geometry: dict[int, dict[str, float]]) -> tuple[list[ColumnRecord], dict[str, int]]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(value): idx for idx, value in enumerate(headers)}
    required = {
        "columnno": "Column No.",
        "type": "Type",
        "leftright": "Left/Right",
        "frontback": "Front/Back",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Columns sheet is missing required headers: {', '.join(missing)}")

    rows: list[ColumnRecord] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue
        idx = parse_column_number(row[header_map["columnno"]], len(rows) + 1)
        left_right = str(row[header_map["leftright"]] or "").strip().title()
        front_back = str(row[header_map["frontback"]] or "").strip().title()
        geo = dxf_geometry.get(idx, {})
        xmin = geo.get("xmin", 0.0)
        xmax = geo.get("xmax", 0.0)
        ymin = geo.get("ymin", 0.0)
        ymax = geo.get("ymax", 0.0)
        anchor_x = _select_anchor(xmin, xmax, left_right)
        anchor_y = _select_anchor(ymin, ymax, front_back)
        rows.append(
            ColumnRecord(
                row_idx=row_idx,
                column_no=idx,
                type_name=str(row[header_map["type"]]).strip(),
                xmin=xmin,
                xmax=xmax,
                ymin=ymin,
                ymax=ymax,
                left_right=left_right,
                front_back=front_back,
                anchor_x=anchor_x,
                anchor_y=anchor_y,
                location=str(row[header_map["location"]]).strip() if "location" in header_map and row[header_map["location"]] not in (None, "") else "",
                anchor_location=str(row[header_map["anchorlocation"]]).strip() if "anchorlocation" in header_map and row[header_map["anchorlocation"]] not in (None, "") else "",
                orientation=str(row[header_map["orientation"]]).strip() if "orientation" in header_map and row[header_map["orientation"]] not in (None, "") else "",
            )
        )
    return rows, header_map


def collect_beam_widths(ws) -> tuple[dict[str, list[float]], dict[str, list[float]], dict[str, int]]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {normalize_header(value): idx for idx, value in enumerate(headers)}
    required = {
        "startc": "StartC",
        "endc": "EndC",
        "direction": "Direction",
        "floor": "Floor",
        "beamwidthmm": "Beam width (mm)",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Primary Beams sheet is missing required headers: {', '.join(missing)}")

    x_beam_widths_by_column: dict[str, list[float]] = {}
    y_beam_widths_by_column: dict[str, list[float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        if str(row[header_map["floor"]] or "").strip() != NODE_SOURCE_FLOOR:
            continue
        direction = str(row[header_map["direction"]] or "").strip().upper()
        beam_width_mm = float(row[header_map["beamwidthmm"]])
        for key in ("startc", "endc"):
            col = str(row[header_map[key]] or "").strip()
            if not col:
                continue
            if direction == "X":
                x_beam_widths_by_column.setdefault(col, []).append(beam_width_mm)
            elif direction == "Y":
                y_beam_widths_by_column.setdefault(col, []).append(beam_width_mm)
    return x_beam_widths_by_column, y_beam_widths_by_column, header_map


def load_primary_beam_rows(ws, header_map: dict[str, int]) -> list[PrimaryBeamRecord]:
    rows: list[PrimaryBeamRecord] = []
    required = {
        "startc": "StartC",
        "endc": "EndC",
        "direction": "Direction",
        "beamclass": "Beam class",
        "floor": "Floor",
        "beamwidthmm": "Beam width (mm)",
        "beamdepthmm": "Beam depth (mm)",
        "wallthicknessmm": "Wall thickness (mm)",
    }
    missing = [label for key, label in required.items() if key not in header_map]
    if missing:
        raise SystemExit(f"Primary Beams sheet is missing required headers for secondary generation: {', '.join(missing)}")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        if str(row[header_map["floor"]] or "").strip() != NODE_SOURCE_FLOOR:
            continue
        rows.append(
            PrimaryBeamRecord(
                start_c=str(row[header_map["startc"]] or "").strip(),
                end_c=str(row[header_map["endc"]] or "").strip(),
                direction=str(row[header_map["direction"]] or "").strip().upper(),
                beam_class=str(row[header_map["beamclass"]] or "").strip(),
                floor=str(row[header_map["floor"]] or "").strip(),
                beam_width_mm=float(row[header_map["beamwidthmm"]] or 0),
                beam_depth_mm=float(row[header_map["beamdepthmm"]] or 0),
                wall_thickness_mm=float(row[header_map["wallthicknessmm"]] or 0),
            )
        )
    return rows


def merge_width_maps(*maps: dict[str, list[float]]) -> dict[str, list[float]]:
    merged: dict[str, list[float]] = {}
    for current in maps:
        for key, values in current.items():
            merged.setdefault(key, []).extend(values)
    return merged


def match_column_key(columns: list[ColumnRecord], x: float, y: float, tolerance_m: float = SECONDARY_ENDPOINT_MATCH_TOL_M) -> str | None:
    best_key = None
    best_dist = tolerance_m + 1.0
    for column in columns:
        dx = abs(column.anchor_x - x)
        dy = abs(column.anchor_y - y)
        if dx <= tolerance_m and dy <= tolerance_m:
            dist = dx + dy
            if dist < best_dist:
                best_key = column.type_name
                best_dist = dist
    return best_key


def load_raw_secondary_rows(wb) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for sheet_name in RAW_SECONDARY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {normalize_header(value): idx for idx, value in enumerate(headers)}
        required = {
            "no": "No.",
            "type": "Type",
            "coordinatex1m": "Coordinate X1 (m)",
            "coordinatey1m": "Coordinate Y1 (m)",
            "coordinatex2m": "Coordinate X2 (m)",
            "coordinatey2m": "Coordinate Y2 (m)",
            "beamlocation": "Beam location",
            "floor": "Floor",
            "present": "Present",
            "beamwidthmm": "Beam width (mm)",
            "beamdepthmm": "Beam depth (mm)",
            "wallthicknessmm": "Wall thickness (mm)",
        }
        missing = [label for key, label in required.items() if key not in header_map]
        if missing:
            raise SystemExit(f"{sheet_name} is missing required headers: {', '.join(missing)}")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(value in (None, "") for value in row):
                continue
            rows.append(
                {
                    "sheet_name": sheet_name,
                    "row_idx": row_idx,
                    "no": row[header_map["no"]],
                    "type_name": str(row[header_map["type"]]).strip(),
                    "x1": float(row[header_map["coordinatex1m"]]),
                    "y1": float(row[header_map["coordinatey1m"]]),
                    "x2": float(row[header_map["coordinatex2m"]]),
                    "y2": float(row[header_map["coordinatey2m"]]),
                    "beam_location": str(row[header_map["beamlocation"]] or "").strip().title(),
                    "floor": str(row[header_map["floor"]] or "").strip(),
                    "present": row[header_map["present"]],
                    "beam_width_mm": float(row[header_map["beamwidthmm"]]),
                    "beam_depth_mm": float(row[header_map["beamdepthmm"]]),
                    "wall_thickness_mm": float(row[header_map["wallthicknessmm"]] or 0),
                }
            )
    return rows


def collect_secondary_beam_widths(columns: list[ColumnRecord], rows: list[dict[str, object]]) -> tuple[dict[str, list[float]], dict[str, list[float]]]:
    x_by_column: dict[str, list[float]] = {}
    y_by_column: dict[str, list[float]] = {}
    for row in rows:
        if row["floor"] != NODE_SOURCE_FLOOR:
            continue
        x1 = float(row["x1"])
        y1 = float(row["y1"])
        x2 = float(row["x2"])
        y2 = float(row["y2"])
        direction = "X" if abs(y1 - y2) <= 1e-6 else "Y"
        width_mm = float(row["beam_width_mm"])
        for point in ((x1, y1), (x2, y2)):
            column_key = match_column_key(columns, point[0], point[1])
            if column_key is None:
                continue
            if direction == "X":
                x_by_column.setdefault(column_key, []).append(width_mm)
            else:
                y_by_column.setdefault(column_key, []).append(width_mm)
    return x_by_column, y_by_column


def choose_width(primary: list[float] | None, fallback: list[float] | None) -> float:
    if primary:
        return max(primary)
    if fallback:
        return max(fallback)
    return 0.0


def calc_raw_node(column: ColumnRecord, x_axis_width_mm: float, y_axis_width_mm: float) -> tuple[float, float]:
    if column.left_right == "Left":
        node_x = column.anchor_x + (x_axis_width_mm / 1000.0) / 2.0
    elif column.left_right == "Right":
        node_x = column.anchor_x - (x_axis_width_mm / 1000.0) / 2.0
    else:
        node_x = (column.xmin + column.xmax) / 2.0

    if column.front_back == "Front":
        node_y = column.anchor_y + (y_axis_width_mm / 1000.0) / 2.0
    elif column.front_back == "Back":
        node_y = column.anchor_y - (y_axis_width_mm / 1000.0) / 2.0
    else:
        node_y = (column.ymin + column.ymax) / 2.0

    return round(node_x, 3), round(node_y, 3)


def level_values(values: list[float], tolerance_m: float) -> list[float]:
    n = len(values)
    used = [False] * n
    result = list(values)
    for i in range(n):
        if used[i]:
            continue
        group = [i]
        used[i] = True
        for j in range(i + 1, n):
            if used[j]:
                continue
            if abs(values[j] - values[i]) <= tolerance_m:
                group.append(j)
                used[j] = True
        if len(group) < 2:
            continue
        avg = round(sum(values[idx] for idx in group) / len(group), 3)
        for idx in group:
            result[idx] = avg
    return result


def beam_face(col: ColumnRecord, axis: str) -> float:
    """Return the coordinate of the column FACE where the beam attaches.

    For axis='X' the relevant face is determined by Left/Right:
      - "Left"  → Xmin (left face)
      - "Right" → Xmax (right face)
      - other   → column center (column without a clear side label)

    For axis='Y' the relevant face is determined by Front/Back:
      - "Front" → Ymin (front face)
      - "Back"  → Ymax (back face)
      - other   → column center
    """
    if axis == "X":
        if col.left_right == "Left":
            return col.xmin
        if col.left_right == "Right":
            return col.xmax
        return (col.xmin + col.xmax) / 2
    # axis == "Y"
    if col.front_back == "Front":
        return col.ymin
    if col.front_back == "Back":
        return col.ymax
    return (col.ymin + col.ymax) / 2


def level_nodes_by_face_distance(
    columns: list[ColumnRecord],
    node_values: list[float],
    axis: str,
    max_distance_m: float,
) -> list[float]:
    """Group columns whose beam-attaching faces are within `max_distance_m`
    of each other on the given axis, then average their node values within
    each group.

    For axis='X' (X-leveling: columns share a vertical gridline), the face is
    Xmin for "Left" columns and Xmax for "Right" columns. A beam connecting
    two columns must land on each column's beam face — so if the faces are
    farther apart than the beam can span, the columns cannot share a gridline.

    For axis='Y', the face is Ymin for "Front" and Ymax for "Back".

    Two columns at "Front Ymin=6.023" and "Front Ymin=6.366" have face distance
    343mm — beyond a 230mm beam — so they are NOT on the same gridline, even
    if their full rectangles overlap by 267mm. Rectangle overlap alone is
    geometrically misleading; only the FACE distance determines beam
    connectivity.

    Grouping is transitive (union-find): if A's face is within tolerance of
    B's face, and B's face is within tolerance of C's face, all three end up
    in the same group even if A and C are farther apart directly.

    Output preserves input order and length so the caller can zip it with the
    original `columns` list.
    """
    n = len(columns)
    if n == 0:
        return list(node_values)

    parent = list(range(n))

    def find(i: int) -> int:
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def union(i: int, j: int) -> None:
        ri, rj = find(i), find(j)
        if ri != rj:
            parent[ri] = rj

    faces = [beam_face(col, axis) for col in columns]
    for i in range(n):
        for j in range(i + 1, n):
            if abs(faces[i] - faces[j]) <= max_distance_m:
                union(i, j)

    groups: dict[int, list[int]] = {}
    for i in range(n):
        groups.setdefault(find(i), []).append(i)

    result = list(node_values)
    for members in groups.values():
        if len(members) < 2:
            continue
        avg = round(sum(node_values[idx] for idx in members) / len(members), 3)
        for idx in members:
            result[idx] = avg
    return result


def compute_leveled_nodes(
    columns: list[ColumnRecord],
    x_beam_widths_by_column: dict[str, list[float]],
    y_beam_widths_by_column: dict[str, list[float]],
    tolerance_m: float,
) -> tuple[dict[str, tuple[float, float]], dict[str, tuple[float, float]]]:
    raw_nodes: dict[str, tuple[float, float]] = {}
    x_values: list[float] = []
    y_values: list[float] = []

    for column in columns:
        column_key = column.type_name
        x_axis_width_mm = choose_width(
            primary=y_beam_widths_by_column.get(column_key),
            fallback=x_beam_widths_by_column.get(column_key),
        )
        y_axis_width_mm = choose_width(
            primary=x_beam_widths_by_column.get(column_key),
            fallback=y_beam_widths_by_column.get(column_key),
        )
        node_x, node_y = calc_raw_node(column, x_axis_width_mm, y_axis_width_mm)
        raw_nodes[column_key] = (node_x, node_y)
        x_values.append(node_x)
        y_values.append(node_y)

    leveled_x = level_nodes_by_face_distance(columns, x_values, "X", BEAM_FACE_MAX_DISTANCE_M)
    leveled_y = level_nodes_by_face_distance(columns, y_values, "Y", BEAM_FACE_MAX_DISTANCE_M)
    leveled_nodes: dict[str, tuple[float, float]] = {}
    for column, node_x, node_y in zip(columns, leveled_x, leveled_y):
        leveled_nodes[column.type_name] = (round(node_x, 3), round(node_y, 3))
    return raw_nodes, leveled_nodes


def set_column_width(ws, header_map: dict[str, int], header_name: str, width: float) -> None:
    idx = header_map.get(normalize_header(header_name))
    if idx is None:
        return
    ws.column_dimensions[get_column_letter(idx + 1)].width = width


def _find_building_info_xlsx() -> Path | None:
    for d in candidate_search_dirs():
        p = d / "building info.xlsx"
        if p.exists():
            return p
    return None


def _read_building_info(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["building info"]
    info: dict = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and row[0] is not None:
            info[str(row[0]).strip()] = row[1]
        if row and len(row) > 4 and row[4] is not None:
            info[str(row[4]).strip()] = row[5]
    wb.close()
    return info


def _generate_elevation_yd_zd(info: dict) -> tuple[list[str], list[int]]:
    foundation_height = float(info["Foundation height"])
    stilt_height = float(info["Stilt height"])
    floors = int(info["Floors"])
    story_height = float(info["Story height"])
    default_yd = int(round(float(info["DEFAULT_COLUMN_YD_MM"])))
    default_zd = int(round(float(info["DEFAULT_COLUMN_ZD_MM"])))
    elevations: list[float] = [-foundation_height, 0.0, round(stilt_height, 3)]
    cumulative = stilt_height
    for _ in range(floors - 1):
        cumulative = round(cumulative + story_height, 3)
        elevations.append(cumulative)
    headers: list[str] = []
    defaults: list[int] = []
    for elev in elevations:
        elev_str = f"{elev:g}"
        headers.extend([f"YD({elev_str})", f"ZD({elev_str})"])
        defaults.extend([default_yd, default_zd])
    return headers, defaults


def append_node_columns(ws_cols, columns: list[ColumnRecord], leveled_nodes: dict[str, tuple[float, float]]) -> None:
    info_path = _find_building_info_xlsx()
    if info_path is None:
        raise SystemExit(
            "building info.xlsx not found. Place it in cwd or cwd/STD ANL model — "
            "YD/ZD elevation columns are derived from it."
        )
    info = _read_building_info(info_path)
    yd_zd_headers, yd_zd_defaults = _generate_elevation_yd_zd(info)

    headers = ["Node coordinate X (m)", "Node coordinate Y (m)"] + yd_zd_headers
    start_col = ws_cols.max_column + 1
    header_positions: dict[str, int] = {}
    for offset, header in enumerate(headers):
        col_idx = start_col + offset
        ws_cols.cell(row=1, column=col_idx, value=header)
        header_positions[header] = col_idx
    for column in columns:
        node_x, node_y = leveled_nodes[column.type_name]
        ws_cols.cell(row=column.row_idx, column=header_positions["Node coordinate X (m)"], value=node_x)
        ws_cols.cell(row=column.row_idx, column=header_positions["Node coordinate Y (m)"], value=node_y)
        for yzh, yzv in zip(yd_zd_headers, yd_zd_defaults):
            ws_cols.cell(row=column.row_idx, column=header_positions[yzh], value=yzv)
    width_map: dict[str, float] = {"Node coordinate X (m)": 18, "Node coordinate Y (m)": 18}
    for yzh in yd_zd_headers:
        width_map[yzh] = 10
    for header, width in width_map.items():
        ws_cols.column_dimensions[get_column_letter(header_positions[header])].width = width


def write_shear_wall_template_sheet(wb) -> None:
    if SHEAR_WALL_TEMPLATE_SHEET in wb.sheetnames:
        del wb[SHEAR_WALL_TEMPLATE_SHEET]
    ws = wb.create_sheet(SHEAR_WALL_TEMPLATE_SHEET)
    headers = [
        "wall_id",
        "node_x_m",
        "node_z_m",
        "thickness_m",
        "direction",
        "priority_tag",
        "partial_length_allowed",
        "symmetry_pair_node_x_m",
        "symmetry_pair_node_z_m",
        "min_length_m",
        "max_length_m",
        "zone_class",
        "wall_function_class",
    ]
    ws.append(headers)
    widths = [14, 12, 12, 12, 12, 12, 22, 18, 18, 14, 14, 12, 9]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_map = {header: idx for idx, header in enumerate(headers, start=1)}
    dropdowns = {
        "direction": '"+X,-X,+Y,-Y"',
        "priority_tag": '"' + ",".join(f"P{i}" for i in range(1, 11)) + '"',
        "partial_length_allowed": '"yes,no"',
        "zone_class": '"boundary_common_area,boundary_utility_area,boundary_bedroom,internal"',
        "wall_function_class": '"lift,staircase,other"',
    }
    for header, formula in dropdowns.items():
        col_letter = get_column_letter(header_map[header])
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"{col_letter}2:{col_letter}5000")


def write_column_landscape_sheet(wb, ws_cols) -> None:
    if COLUMN_LANDSCAPE_TEMPLATE_SHEET in wb.sheetnames:
        del wb[COLUMN_LANDSCAPE_TEMPLATE_SHEET]
    ws = wb.create_sheet(COLUMN_LANDSCAPE_TEMPLATE_SHEET)
    headers = [
        "node_x_m",
        "node_z_m",
        "base_long_axis",
        "allowed_orientations",
        "priority_tag",
        "zone_class",
        "symmetry_pair_node_x_m",
        "symmetry_pair_node_z_m",
    ]
    ws.append(headers)
    widths = [12, 12, 16, 18, 12, 24, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_cells = list(ws_cols.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    source_header_map = {normalize_header(value): idx for idx, value in enumerate(header_cells)}
    node_x_idx = source_header_map.get(normalize_header("Node coordinate X (m)"))
    node_z_idx = source_header_map.get(normalize_header("Node coordinate Y (m)"))
    seen_coords: set[tuple[float, float]] = set()
    if node_x_idx is not None and node_z_idx is not None:
        for row in ws_cols.iter_rows(min_row=2, values_only=True):
            node_x = row[node_x_idx]
            node_z = row[node_z_idx]
            if node_x in (None, "") or node_z in (None, ""):
                continue
            coord_key = (round(float(node_x), 3), round(float(node_z), 3))
            if coord_key in seen_coords:
                continue
            seen_coords.add(coord_key)
            ws.append(
                [
                    coord_key[0],
                    coord_key[1],
                    "",
                    "as_is",
                    "must_keep",
                    "",
                    "",
                    "",
                ]
            )

    header_map = {header: idx for idx, header in enumerate(headers, start=1)}
    dropdowns = {
        "base_long_axis": {
            "formula": '"X,Y"',
            "title": "Base Long Axis",
            "message": "Fill only for new nodes that do not currently have a column. X means the seed longer side is along X; Y means along Y.",
        },
        "allowed_orientations": {
            "formula": '"as_is,both"',
            "title": "Allowed Orientations",
            "message": "Use as_is to lock the current or seed orientation, or both to allow 90-degree rotation.",
        },
        "priority_tag": {
            "formula": '"must_keep,optional"',
            "title": "Priority Tag",
            "message": "Use must_keep for columns that cannot be removed. Use optional where future optimizer deletion or non-use is architecturally acceptable.",
        },
        "zone_class": {
            "formula": '"boundary_common_area,boundary_utility_area,boundary_bedroom,internal"',
            "title": "Zone Class",
            "message": "Classify the node location for future optimizer and learner use.",
        },
    }
    for header, config in dropdowns.items():
        col_letter = get_column_letter(header_map[header])
        validation = DataValidation(type="list", formula1=config["formula"], allow_blank=True)
        validation.promptTitle = config["title"]
        validation.prompt = config["message"]
        ws.add_data_validation(validation)
        validation.add(f"{col_letter}2:{col_letter}5000")


def rewrite_primary_beams(ws_beams, leveled_nodes: dict[str, tuple[float, float]], header_map: dict[str, int]) -> None:
    header_updates = {
        "startanchorxm": "Start Node X (m)",
        "startanchorym": "Start Node Y (m)",
        "endanchorxm": "End Node X (m)",
        "endanchorym": "End Node Y (m)",
    }
    for key, header in header_updates.items():
        col_idx = header_map[key] + 1
        ws_beams.cell(row=1, column=col_idx, value=header)

    start_c_idx = header_map["startc"]
    end_c_idx = header_map["endc"]
    start_node_x_idx = header_map["startanchorxm"]
    start_node_y_idx = header_map["startanchorym"]
    end_node_x_idx = header_map["endanchorxm"]
    end_node_y_idx = header_map["endanchorym"]
    beam_start_x_idx = header_map["beamstartxm"]
    beam_start_y_idx = header_map["beamstartym"]
    beam_end_x_idx = header_map["beamendxm"]
    beam_end_y_idx = header_map["beamendym"]

    for row in ws_beams.iter_rows(min_row=2):
        start_c = str(row[start_c_idx].value or "").strip()
        end_c = str(row[end_c_idx].value or "").strip()
        if not start_c or not end_c:
            continue
        start_node = leveled_nodes[start_c]
        end_node = leveled_nodes[end_c]
        row[start_node_x_idx].value = start_node[0]
        row[start_node_y_idx].value = start_node[1]
        row[end_node_x_idx].value = end_node[0]
        row[end_node_y_idx].value = end_node[1]
        row[beam_start_x_idx].value = start_node[0]
        row[beam_start_y_idx].value = start_node[1]
        row[beam_end_x_idx].value = end_node[0]
        row[beam_end_y_idx].value = end_node[1]

    for key in ("startanchorxm", "startanchorym", "endanchorxm", "endanchorym"):
        ws_beams.column_dimensions[get_column_letter(header_map[key] + 1)].width = 11


def _group_and_average(values_with_indices: list[tuple[int, float]], tolerance_m: float) -> dict[int, float]:
    if not values_with_indices:
        return {}
    mapping: dict[int, float] = {}
    ordered = sorted(values_with_indices, key=lambda item: item[1])
    used: set[int] = set()
    for i, (idx_i, value_i) in enumerate(ordered):
        if idx_i in used:
            continue
        group = [(idx_i, value_i)]
        used.add(idx_i)
        for idx_j, value_j in ordered[i + 1 :]:
            if idx_j in used:
                continue
            if value_j - value_i <= tolerance_m:
                group.append((idx_j, value_j))
                used.add(idx_j)
            else:
                break
        if len(group) >= 2:
            avg = round(sum(value for _, value in group) / len(group), 3)
            for idx, _ in group:
                mapping[idx] = avg
    return mapping


def _snap_to_nearest(value: float, targets: list[float], tolerance_m: float) -> float:
    best = value
    best_delta = tolerance_m + 1.0
    for target in targets:
        delta = abs(value - target)
        if 0.001 < delta <= tolerance_m and delta < best_delta:
            best = target
            best_delta = delta
    return best


def _apply_wall_offset(x1: float, y1: float, x2: float, y2: float, location: str, wall_thickness_mm: float) -> tuple[float, float, float, float]:
    if wall_thickness_mm in (None, "", 0, 0.0) or location in {"", "Centre"}:
        return round(x1, 3), round(y1, 3), round(x2, 3), round(y2, 3)
    half = float(wall_thickness_mm) / 1000.0 / 2.0
    if location == "Left":
        x1 -= half
        x2 -= half
    elif location == "Right":
        x1 += half
        x2 += half
    elif location == "Back":
        y1 += half
        y2 += half
    elif location == "Front":
        y1 -= half
        y2 -= half

    if location in {"Left", "Right"}:
        if y1 < y2:
            y1 -= half
            y2 += half
        elif y2 < y1:
            y2 -= half
            y1 += half
    elif location in {"Front", "Back"}:
        if x1 < x2:
            x1 -= half
            x2 += half
        elif x2 < x1:
            x2 -= half
            x1 += half
    return round(x1, 3), round(y1, 3), round(x2, 3), round(y2, 3)


def build_other_output_path(main_output_path: Path) -> Path:
    if "node_coordinates" in main_output_path.name:
        return main_output_path.parent / main_output_path.name.replace("node_coordinates", "other_coordinates")
    return main_output_path.with_name(f"{main_output_path.stem}_other_coordinates.xlsx")


def rectangle_axes_from_rows(rect_rows) -> tuple[list[float], list[float]]:
    xs: list[float] = []
    ys: list[float] = []
    if not rect_rows:
        return xs, ys
    for row in rect_rows:
        xs.extend([round(float(row[13]), 3), round(float(row[15]), 3)])
        ys.extend([round(float(row[14]), 3), round(float(row[16]), 3)])
    return unique_sorted(xs), unique_sorted(ys)


def build_harmonized_other_rows(
    legacy_node,
    geometry_input_path: Path,
    node_x_vals: list[float],
    node_y_vals: list[float],
    sec_beam_rows: list[list[object]] | None = None,
) -> tuple[list[tuple[object, ...]] | None, list[tuple[object, ...]] | None, list[list[object]] | None]:
    rect_rows = legacy_node.read_and_snap_rectangles(geometry_input_path, node_x_vals, node_y_vals, sec_beam_rows=sec_beam_rows)
    balcony_rows = legacy_node.read_and_snap_balconies(geometry_input_path, node_x_vals, node_y_vals, sec_beam_rows=sec_beam_rows)
    staircase_rows = legacy_node.read_staircase_details(geometry_input_path)
    rect_rows, _, balcony_rows, _ = legacy_node.final_level_all(rect_rows, sec_beam_rows, balcony_rows, node_x_vals, node_y_vals)
    if staircase_rows:
        _snapped_stair: list[list[object]] = []
        for _sr in staircase_rows:
            if len(_sr) >= 5 and all(isinstance(_sr[_i], (int, float)) for _i in range(1, 5)):
                _r = list(_sr)
                _r[1] = round(_snap_to_nearest(float(_sr[1]), node_x_vals, SECONDARY_SNAP_TOL_M), 3)
                _r[2] = round(_snap_to_nearest(float(_sr[2]), node_y_vals, SECONDARY_SNAP_TOL_M), 3)
                _r[3] = round(_snap_to_nearest(float(_sr[3]), node_x_vals, SECONDARY_SNAP_TOL_M), 3)
                _r[4] = round(_snap_to_nearest(float(_sr[4]), node_y_vals, SECONDARY_SNAP_TOL_M), 3)
                _snapped_stair.append(_r)
            else:
                _snapped_stair.append(list(_sr))
        staircase_rows = _snapped_stair
    return rect_rows, balcony_rows, staircase_rows


def build_extra_wall_rows(
    unfiltered,
    legacy_node,
    wall_path: Path,
    node_x_vals: list[float],
    node_y_vals: list[float],
    primary_rows: list[PrimaryBeamRecord],
    leveled_nodes: dict[str, tuple[float, float]],
    final_secondary_rows: list[list[object]],
) -> list[tuple[object, ...]] | None:
    raw_walls = unfiltered.read_walls(wall_path, unfiltered.DEFAULT_THICKNESS_TOLERANCE_MM)
    corrected_walls = []
    for wall in raw_walls:
        dx = abs(float(wall.end_x) - float(wall.start_x))
        dy = abs(float(wall.end_y) - float(wall.start_y))
        if dy <= EXTRA_WALL_ORTHO_TOL_M and dx > 0.001:
            orientation = "Horizontal"
            fixed = round((float(wall.start_y) + float(wall.end_y)) / 2.0, 3)
            start = round(min(float(wall.start_x), float(wall.end_x)), 3)
            end = round(max(float(wall.start_x), float(wall.end_x)), 3)
            start_x, start_y, end_x, end_y = start, fixed, end, fixed
        elif dx <= EXTRA_WALL_ORTHO_TOL_M and dy > 0.001:
            orientation = "Vertical"
            fixed = round((float(wall.start_x) + float(wall.end_x)) / 2.0, 3)
            start = round(min(float(wall.start_y), float(wall.end_y)), 3)
            end = round(max(float(wall.start_y), float(wall.end_y)), 3)
            start_x, start_y, end_x, end_y = fixed, start, fixed, end
        else:
            continue
        corrected_walls.append(
            unfiltered.WallSegment(
                wall_no=wall.wall_no,
                start_x=start_x,
                start_y=start_y,
                end_x=end_x,
                end_y=end_y,
                thickness_m=wall.thickness_m,
                thickness_class_mm=wall.thickness_class_mm,
                orientation=orientation,
                source=wall.source,
                dxf_source=wall.dxf_source,
            )
        )

    primary_supports: list[object] = []
    for idx, row in enumerate(primary_rows, start=1):
        start_node = leveled_nodes[row.start_c]
        end_node = leveled_nodes[row.end_c]
        fixed, start, end = unfiltered.axis_major_values(row.direction, start_node[0], start_node[1], end_node[0], end_node[1])
        primary_supports.append(
            unfiltered.SupportSegment(
                axis=row.direction,
                fixed=fixed,
                start=start,
                end=end,
                source_kind="primary",
                source_name=f"B{idx}",
            )
        )

    secondary_supports: list[object] = []
    for row in final_secondary_rows:
        x1 = float(row[12]); y1 = float(row[13]); x2 = float(row[14]); y2 = float(row[15])
        axis = "X" if abs(y1 - y2) <= 1e-6 else "Y"
        fixed, start, end = unfiltered.axis_major_values(axis, x1, y1, x2, y2)
        secondary_supports.append(
            unfiltered.SupportSegment(
                axis=axis,
                fixed=fixed,
                start=start,
                end=end,
                source_kind="secondary",
                source_name=str(row[1]),
            )
        )
    support_segments = unfiltered.merge_support_segments(primary_supports + secondary_supports)

    def chain_supported_by_beam(chain) -> bool:
        for segment in support_segments:
            if segment.axis != chain.axis or abs(segment.fixed - chain.fixed) > EXTRA_WALL_BEAM_EXCLUDE_TOL_M:
                continue
            overlap = unfiltered.overlap_1d(chain.start, chain.end, segment.start, segment.end)
            if overlap is None:
                continue
            if (overlap[1] - overlap[0]) >= max(0.50, 0.60 * chain.length):
                return True
        return False

    floor_map = {
        "plinth": ("Plinth",),
        "typical": ("Stilt roof", "Typical floor roof"),
        "terrace": ("Terrace",),
    }
    raw_rows: list[list[object]] = []
    counter = 1
    for dxf_source, floors in floor_map.items():
        walls_for_source = [wall for wall in corrected_walls if wall.dxf_source == dxf_source]
        if not walls_for_source:
            continue
        chains = unfiltered.merge_wall_chains(walls_for_source)
        for chain in chains:
            if chain.length < EXTRA_WALL_MIN_EXPORT_LENGTH_M:
                continue
            if chain_supported_by_beam(chain):
                continue
            if chain.axis == "X":
                x1, y1, x2, y2 = round(chain.start, 3), round(chain.fixed, 3), round(chain.end, 3), round(chain.fixed, 3)
            else:
                x1, y1, x2, y2 = round(chain.fixed, 3), round(chain.start, 3), round(chain.fixed, 3), round(chain.end, 3)
            ew_type = f"EW{counter}"
            for floor in floors:
                present, thickness_mode = EXTRA_WALL_PRESENT_DEFAULTS.get(floor, ("YES", "chain"))
                export_thickness = int(chain.thickness_mm) if thickness_mode == "chain" else None
                raw_rows.append(
                    [
                        counter,
                        ew_type,
                        x1,
                        y1,
                        x2,
                        y2,
                        "Centre",
                        floor,
                        present,
                        export_thickness,
                        x1,
                        y1,
                        x2,
                        y2,
                    ]
                )
            counter += 1

    if not raw_rows:
        return None, []
    return legacy_node.snap_extra_walls(raw_rows, node_x_vals, node_y_vals, sec_beam_rows=final_secondary_rows), list(support_segments)


def write_other_coordinates_workbook(
    output_path: Path,
    rect_rows: list[tuple[object, ...]] | None,
    balcony_rows: list[tuple[object, ...]] | None,
    staircase_rows: list[list[object]] | None,
    ew_rows: list[tuple[object, ...]] | None = None,
    ew_support_segments: list | None = None,
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    if rect_rows:
        ws_rect = wb.create_sheet("Rectangle coordinates")
        headers = [
            "No.",
            "Type",
            "Location",
            "Anchor location 1",
            "Beam X width location 1",
            "Beam Y width location 1",
            "Anchor location 2",
            "Beam X width location 2",
            "Beam Y width location 2",
            "Snapped X1 (m)",
            "Snapped Y1 (m)",
            "Snapped X2 (m)",
            "Snapped Y2 (m)",
        ]
        widths = [6, 8, 16, 16, 20, 20, 16, 20, 20, 18, 18, 18, 18]
        _skip_raw = {2, 3, 4, 5}
        ws_rect.append(headers)
        for row in rect_rows:
            ws_rect.append([v for i, v in enumerate(row) if i not in _skip_raw])
        for idx, width in enumerate(widths, start=1):
            ws_rect.column_dimensions[get_column_letter(idx)].width = width if idx <= 2 else round(width * 0.8, 1)
        for row in ws_rect.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if balcony_rows:
        ws_bal = wb.create_sheet("Balcony coordinates")
        headers = ["Location", "Coordinate X1 (m)", "Coordinate Y1 (m)", "Coordinate X2 (m)", "Coordinate Y2 (m)", "Cantilever slab (YES/NO)"]
        widths = [20, 18, 18, 18, 18, 22]
        ws_bal.append(headers)
        for row in balcony_rows:
            ws_bal.append(list(row) + [None])
        for idx, width in enumerate(widths, start=1):
            ws_bal.column_dimensions[get_column_letter(idx)].width = width
        if len(balcony_rows) > 0:
            cantilever_dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
            cantilever_dv.error = "Please choose YES or NO from the dropdown."
            cantilever_dv.errorTitle = "Invalid entry"
            cantilever_dv.prompt = "Select YES if this balcony is a cantilever slab, otherwise NO."
            cantilever_dv.promptTitle = "Cantilever slab"
            ws_bal.add_data_validation(cantilever_dv)
            cantilever_dv.add(f"F2:F{1 + len(balcony_rows)}")
        for row in ws_bal.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if staircase_rows:
        ws_stair = wb.create_sheet("Staircase details")
        for row in staircase_rows:
            ws_stair.append(list(row))
        for row in ws_stair.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if ew_rows:
        ws_ew = wb.create_sheet("Extra wall coordinates")
        headers = [
            "No.",
            "Type",
            "Wall location",
            "Floor",
            "Present",
            "Wall thickness (mm)",
            "Snapped X1 (m)",
            "Snapped Y1 (m)",
            "Snapped X2 (m)",
            "Snapped Y2 (m)",
        ]
        widths = [6, 8, 16, 16, 10, 18, 18, 18, 18, 18]
        _skip_raw_ew = {2, 3, 4, 5}
        ws_ew.append(headers)
        for row in ew_rows:
            ws_ew.append([v for i, v in enumerate(row) if i not in _skip_raw_ew])
        for idx, width in enumerate(widths, start=1):
            ws_ew.column_dimensions[get_column_letter(idx)].width = width
        for row in ws_ew.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        _balcony_rects: list[tuple[float, float, float, float]] = []
        if rect_rows:
            for _rr in rect_rows:
                if str(_rr[6]).strip().lower() == "balcony":
                    _balcony_rects.append((
                        min(float(_rr[13]), float(_rr[15])),
                        max(float(_rr[13]), float(_rr[15])),
                        min(float(_rr[14]), float(_rr[16])),
                        max(float(_rr[14]), float(_rr[16])),
                    ))

        valid_ew_rows = []
        for row in ew_rows:
            present = str(row[8]).strip().upper() if row[8] is not None else ""
            thickness = row[9]
            if present != "YES":
                continue
            if thickness in (None, ""):
                continue
            sx1, sy1, sx2, sy2 = float(row[10]), float(row[11]), float(row[12]), float(row[13])
            snapped_length = max(abs(sx2 - sx1), abs(sy2 - sy1))
            _axis = "X" if abs(sy1 - sy2) < 0.001 else "Y"
            _fixed = sy1 if _axis == "X" else sx1
            _span_s = min(sx1, sx2) if _axis == "X" else min(sy1, sy2)
            _span_e = max(sx1, sx2) if _axis == "X" else max(sy1, sy2)
            _on_balcony = False
            for _rx1, _rx2, _ry1, _ry2 in _balcony_rects:
                _perp_lo, _perp_hi = (_ry1, _ry2) if _axis == "X" else (_rx1, _rx2)
                _par_lo, _par_hi = (_rx1, _rx2) if _axis == "X" else (_ry1, _ry2)
                if _perp_lo - 0.15 <= _fixed <= _perp_hi + 0.15 and min(_span_e, _par_hi) > max(_span_s, _par_lo):
                    _on_balcony = True
                    break
            _min_len = EXTRA_WALL_BALCONY_MIN_FILTERED_LENGTH_M if _on_balcony else EXTRA_WALL_MIN_FILTERED_LENGTH_M
            if snapped_length < _min_len:
                continue
            if ew_support_segments:
                on_beam = False
                for seg in ew_support_segments:
                    if seg.axis != _axis or abs(seg.fixed - _fixed) > EXTRA_WALL_BEAM_EXCLUDE_TOL_M:
                        continue
                    lo = max(_span_s, seg.start)
                    hi = min(_span_e, seg.end)
                    if hi - lo >= max(0.50, 0.60 * snapped_length):
                        on_beam = True
                        break
                if on_beam:
                    continue
            valid_ew_rows.append(row)

        if valid_ew_rows:
            ws_ewf = wb.create_sheet("Extra walls (filtered)")
            filtered_headers = [
                "Extra Wall",
                "start_x",
                "start_z",
                "end_x",
                "end_z",
                "Wall thickness (mm)",
                "Floor",
                "Present",
            ]
            filtered_widths = [12, 14, 14, 14, 14, 18, 14, 10]
            ws_ewf.append(filtered_headers)
            prev_no = None
            for row_data in valid_ew_rows:
                ew_no = row_data[0]
                is_first = (ew_no != prev_no)
                ws_ewf.append([
                    ew_no if is_first else None,   # Extra Wall number only on first row
                    row_data[10],                   # start_x (snapped X1)
                    row_data[11],                   # start_z (snapped Y1)
                    row_data[12],                   # end_x (snapped X2)
                    row_data[13],                   # end_z (snapped Y2)
                    row_data[9],                    # Wall thickness (mm)
                    row_data[7],                    # Floor
                    row_data[8],                    # Present
                ])
                prev_no = ew_no
            for idx, width in enumerate(filtered_widths, start=1):
                ws_ewf.column_dimensions[get_column_letter(idx)].width = width
            for row in ws_ewf.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(output_path)


def convert_columns_to_unfiltered(unfiltered, columns: list[ColumnRecord]):
    result = []
    for column in columns:
        result.append(
            unfiltered.ColumnRecord(
                idx=column.column_no,
                type_name=column.type_name,
                xmin=column.xmin,
                xmax=column.xmax,
                ymin=column.ymin,
                ymax=column.ymax,
                width=round(column.xmax - column.xmin, 3),
                height=round(column.ymax - column.ymin, 3),
                left_right=column.left_right,
                front_back=column.front_back,
                source_location=None,
                location=column.location,
                anchor_location=column.anchor_location or "",
                anchor_x=column.anchor_x,
                anchor_y=column.anchor_y,
                orientation=column.orientation or ("Vertical" if (column.ymax - column.ymin) >= (column.xmax - column.xmin) else "Horizontal"),
            )
        )
    return result


def convert_primary_rows_to_unfiltered(unfiltered, primary_rows: list[PrimaryBeamRecord], uf_columns, leveled_nodes: dict[str, tuple[float, float]]):
    column_map = {column.type_name: column for column in uf_columns}
    beams = []
    for idx, row in enumerate(primary_rows, start=1):
        start_node = leveled_nodes[row.start_c]
        end_node = leveled_nodes[row.end_c]
        beams.append(
            unfiltered.BeamPair(
                beam_no=f"B{idx}",
                start=column_map[row.start_c],
                end=column_map[row.end_c],
                direction=row.direction,
                beam_class=row.beam_class,
                beam_start_x=start_node[0],
                beam_start_y=start_node[1],
                beam_end_x=end_node[0],
                beam_end_y=end_node[1],
                span_length_m=round(abs(end_node[0] - start_node[0]) if row.direction == "X" else abs(end_node[1] - start_node[1]), 3),
                group_coordinate_m=round(start_node[1] if row.direction == "X" else start_node[0], 3),
            )
        )
    return beams


def build_preaccepted_r4(unfiltered, raw_rows: list[dict[str, object]], floor_group: str, floors: tuple[str, ...], uf_columns) -> list[object]:
    result = []
    seen: set[tuple[str, float, float, float]] = set()
    target_floors = set(floors)
    for row in raw_rows:
        if str(row["floor"]) not in target_floors:
            continue
        axis = "X" if abs(float(row["y1"]) - float(row["y2"])) <= 1e-6 else "Y"
        fixed, start, end = unfiltered.axis_major_values(axis, float(row["x1"]), float(row["y1"]), float(row["x2"]), float(row["y2"]))
        key = (axis, fixed, start, end)
        if key in seen:
            continue
        seen.add(key)
        result.append(
            unfiltered.SecondaryBeamCandidate(
                axis=axis,
                x1=round(float(row["x1"]), 3),
                y1=round(float(row["y1"]), 3),
                x2=round(float(row["x2"]), 3),
                y2=round(float(row["y2"]), 3),
                beam_location=str(row["beam_location"]),
                floor_group=floor_group,
                floors=floors,
                rule_code="R4",
                beam_class=unfiltered.build_secondary_beam_class(axis, fixed, uf_columns, unfiltered.DEFAULT_BOUNDARY_TOLERANCE_M),
                detail=f"raw_r4:{row['type_name']}",
                score=35.0,
            )
        )
    return result


def build_generated_secondary_raw_rows(
    unfiltered,
    other_geometry_path: Path,
    wall_path: Path,
    columns: list[ColumnRecord],
    leveled_nodes: dict[str, tuple[float, float]],
    primary_rows: list[PrimaryBeamRecord],
    raw_secondary_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    uf_columns = convert_columns_to_unfiltered(unfiltered, columns)
    uf_primary_beams = convert_primary_rows_to_unfiltered(unfiltered, primary_rows, uf_columns, leveled_nodes)
    primary_supports = unfiltered.build_node_primary_segments(uf_primary_beams, leveled_nodes)
    walls = unfiltered.read_walls(wall_path, unfiltered.DEFAULT_THICKNESS_TOLERANCE_MM)
    walls_by_floor: dict[str, list[object]] = {}
    for wall in walls:
        if wall.dxf_source:
            walls_by_floor.setdefault(wall.dxf_source, []).append(wall)
    rectangles, _ = unfiltered.read_zone_rectangles(other_geometry_path)
    plinth_rectangles = [rect for rect in rectangles if str(rect.location).strip().lower() == "lift"]
    lift_rectangles = [rect for rect in rectangles if str(rect.location).strip().lower() == "lift"]

    def beam_is_inside_lift(beam) -> bool:
        fixed, start, end = unfiltered.axis_major_values(beam.axis, beam.x1, beam.y1, beam.x2, beam.y2)
        for rect in lift_rectangles:
            rx1, rx2 = sorted((rect.x1, rect.x2))
            ry1, ry2 = sorted((rect.y1, rect.y2))
            if beam.axis == "X":
                if not (ry1 - SECONDARY_SNAP_TOL_M <= fixed <= ry2 + SECONDARY_SNAP_TOL_M):
                    continue
                overlap = unfiltered.overlap_1d(start, end, rx1, rx2)
            else:
                if not (rx1 - SECONDARY_SNAP_TOL_M <= fixed <= rx2 + SECONDARY_SNAP_TOL_M):
                    continue
                overlap = unfiltered.overlap_1d(start, end, ry1, ry2)
            if overlap is not None and (overlap[1] - overlap[0]) > 0.01:
                return True
        return False

    plinth_r4 = build_preaccepted_r4(unfiltered, raw_secondary_rows, "plinth", PLINTH_FLOORS, uf_columns)
    nonplinth_r4 = build_preaccepted_r4(unfiltered, raw_secondary_rows, "nonplinth", NONPLINTH_FLOORS, uf_columns)

    plinth_beams, _ = unfiltered.generate_secondary_group(
        uf_columns,
        primary_supports,
        leveled_nodes,
        walls_by_floor.get("plinth", []),
        plinth_rectangles,
        floor_group="plinth",
        floors=PLINTH_FLOORS,
        boundary_tolerance_m=unfiltered.DEFAULT_BOUNDARY_TOLERANCE_M,
        enabled_rules={"R1", "R2", "R5"},
        preaccepted=plinth_r4,
    )
    nonplinth_beams, _ = unfiltered.generate_secondary_group(
        uf_columns,
        primary_supports,
        leveled_nodes,
        walls_by_floor.get("typical", []) + walls_by_floor.get("terrace", []),
        rectangles,
        floor_group="nonplinth",
        floors=NONPLINTH_FLOORS,
        boundary_tolerance_m=unfiltered.DEFAULT_BOUNDARY_TOLERANCE_M,
        enabled_rules={"R1", "R2", "R5"},
        preaccepted=nonplinth_r4,
    )

    rows: list[dict[str, object]] = []
    sheet_by_floor = {"Plinth": RAW_SECONDARY_SHEETS[0], "Stilt roof": RAW_SECONDARY_SHEETS[1], "Typical floor roof": RAW_SECONDARY_SHEETS[1], "Terrace": RAW_SECONDARY_SHEETS[1]}
    for beam in list(plinth_beams) + list(nonplinth_beams):
        for floor in beam.floors:
            values = unfiltered.secondary_row_values(
                beam,
                floor,
                walls_by_floor,
                unfiltered.DEFAULT_WALL_ALIGNMENT_TOLERANCE_M,
                unfiltered.DEFAULT_EDGE_WALL_COVERAGE_THRESHOLD_PCT,
                unfiltered.DEFAULT_INTERIOR_WALL_COVERAGE_THRESHOLD_PCT,
            )
            if beam_is_inside_lift(beam):
                values[9] = 230
                values[10] = 300
                values[11] = 230
            rows.append(
                {
                    "sheet_name": sheet_by_floor[floor],
                    "row_idx": 0,
                    "no": values[0],
                    "type_name": values[1],
                    "x1": values[2],
                    "y1": values[3],
                    "x2": values[4],
                    "y2": values[5],
                    "beam_location": values[6],
                    "floor": values[7],
                    "present": values[8],
                    "beam_width_mm": values[9],
                    "beam_depth_mm": values[10],
                    "wall_thickness_mm": values[11],
                }
            )
    return rows


def build_final_secondary_rows(
    raw_rows: list[dict[str, object]],
    leveled_nodes: dict[str, tuple[float, float]],
    rectangle_x: list[float],
    rectangle_y: list[float],
) -> list[list[object]]:
    if not raw_rows:
        return []

    trusted_x = unique_sorted([coords[0] for coords in leveled_nodes.values()] + rectangle_x)
    trusted_y = unique_sorted([coords[1] for coords in leveled_nodes.values()] + rectangle_y)

    first_rows: dict[str, list[object]] = {}
    for row in raw_rows:
        beam_type = str(row["type_name"])
        if beam_type in first_rows:
            continue
        x1, y1, x2, y2 = _apply_wall_offset(
            float(row["x1"]),
            float(row["y1"]),
            float(row["x2"]),
            float(row["y2"]),
            str(row["beam_location"]),
            float(row["wall_thickness_mm"]),
        )
        first_rows[beam_type] = [
            row["no"],
            beam_type,
            row["x1"],
            row["y1"],
            row["x2"],
            row["y2"],
            row["beam_location"],
            row["floor"],
            row["present"],
            row["beam_width_mm"],
            row["beam_depth_mm"],
            row["wall_thickness_mm"],
            x1,
            y1,
            x2,
            y2,
        ]

    all_x = []
    all_y = []
    for idx, (_, row) in enumerate(first_rows.items()):
        all_x.extend([(idx * 2, float(row[12])), (idx * 2 + 1, float(row[14]))])
        all_y.extend([(idx * 2, float(row[13])), (idx * 2 + 1, float(row[15]))])
    x_map = _group_and_average(all_x, SECONDARY_SNAP_TOL_M)
    y_map = _group_and_average(all_y, SECONDARY_SNAP_TOL_M)

    ordered_types = list(first_rows.keys())
    for idx, beam_type in enumerate(ordered_types):
        row = first_rows[beam_type]
        if idx * 2 in x_map:
            row[12] = x_map[idx * 2]
        if idx * 2 + 1 in x_map:
            row[14] = x_map[idx * 2 + 1]
        if idx * 2 in y_map:
            row[13] = y_map[idx * 2]
        if idx * 2 + 1 in y_map:
            row[15] = y_map[idx * 2 + 1]

        row[12] = _snap_to_nearest(float(row[12]), trusted_x, SECONDARY_SNAP_TOL_M)
        row[14] = _snap_to_nearest(float(row[14]), trusted_x, SECONDARY_SNAP_TOL_M)
        row[13] = _snap_to_nearest(float(row[13]), trusted_y, SECONDARY_SNAP_TOL_M)
        row[15] = _snap_to_nearest(float(row[15]), trusted_y, SECONDARY_SNAP_TOL_M)

        if abs(float(row[13]) - float(row[15])) <= SECONDARY_SNAP_TOL_M:
            avg_y = round((float(row[13]) + float(row[15])) / 2.0, 3)
            row[13] = avg_y
            row[15] = avg_y
        if abs(float(row[12]) - float(row[14])) <= SECONDARY_SNAP_TOL_M:
            avg_x = round((float(row[12]) + float(row[14])) / 2.0, 3)
            row[12] = avg_x
            row[14] = avg_x

    final_rows: list[list[object]] = []
    for row in raw_rows:
        snapped = first_rows[str(row["type_name"])]
        final_rows.append(
            [
                row["no"],
                row["type_name"],
                row["x1"],
                row["y1"],
                row["x2"],
                row["y2"],
                row["beam_location"],
                row["floor"],
                row["present"],
                row["beam_width_mm"],
                row["beam_depth_mm"],
                row["wall_thickness_mm"],
                snapped[12],
                snapped[13],
                snapped[14],
                snapped[15],
            ]
        )
    return final_rows


def write_final_secondary_sheet(wb, rows: list[list[object]]) -> tuple[list[float], list[float]]:
    if FINAL_SECONDARY_SHEET in wb.sheetnames:
        del wb[FINAL_SECONDARY_SHEET]
    if not rows:
        return [], []
    ws = wb.create_sheet(FINAL_SECONDARY_SHEET)
    ws.append(
        [
            "No.",
            "Type",
            "Floor",
            "Beam width (mm)",
            "Beam depth (mm)",
            "Wall thickness (mm)",
            "Snapped X1 (m)",
            "Snapped Y1 (m)",
            "Snapped X2 (m)",
            "Snapped Y2 (m)",
        ]
    )
    _skip_sec = {2, 3, 4, 5, 6, 8}
    xs: list[float] = []
    ys: list[float] = []
    for row in rows:
        filtered = [v for i, v in enumerate(row) if i not in _skip_sec]
        if str(row[7] or "").strip() == "Plinth" and (row[11] == 0 or row[11] == 0.0):
            filtered[5] = 90
        ws.append(filtered)
        xs.extend([float(row[12]), float(row[14])])
        ys.extend([float(row[13]), float(row[15])])
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    autosize = {
        "A": 7, "B": 8, "C": 18, "D": 10, "E": 10, "F": 10, "G": 12, "H": 12, "I": 12, "J": 12,
    }
    for column_letter, width in autosize.items():
        ws.column_dimensions[column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return xs, ys


def append_primary_beam_present_column(ws_beams) -> None:
    header_cells = list(ws_beams.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_map = {normalize_header(value): idx + 1 for idx, value in enumerate(header_cells)}
    present_header = "Present"
    present_key = normalize_header(present_header)
    if present_key in header_map:
        present_col = header_map[present_key]
    else:
        present_col = ws_beams.max_column + 1
        ws_beams.cell(row=1, column=present_col).value = present_header
    for row_idx in range(2, ws_beams.max_row + 1):
        ws_beams.cell(row=row_idx, column=present_col).value = "YES"
    ws_beams.column_dimensions[get_column_letter(present_col)].width = 10


def prune_main_workbook(wb) -> None:
    keep_order = ["Columns", "Primary Beams", FINAL_SECONDARY_SHEET, "Extra column", SHEAR_WALL_TEMPLATE_SHEET, COLUMN_LANDSCAPE_TEMPLATE_SHEET]
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in keep_order:
            del wb[sheet_name]
    for target_index, sheet_name in enumerate(keep_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(wb[sheet_name], offset=target_index - wb.index(wb[sheet_name]))


def unique_sorted(values: list[float]) -> list[float]:
    ordered = []
    for value in sorted(values):
        if not ordered or value != ordered[-1]:
            ordered.append(value)
    return ordered


def align_new_headers(*worksheets) -> None:
    for ws in worksheets:
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append leveled node coordinates to the current column/primary-beam workbook and rewrite beam endpoints to those nodes."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help="Input column-beam-pair workbook. Defaults to the single discovered workbook in the project folders.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="Output workbook path. Defaults to node_coordinates_<timestamp>.xlsx beside the input workbook.",
    )
    parser.add_argument(
        "--walls",
        type=Path,
        default=DEFAULT_WALL_INPUT_PATH,
        help="Optional walls workbook. If omitted, the script discovers the single wall workbook in the project folders.",
    )
    parser.add_argument(
        "--geometry-input",
        type=Path,
        default=DEFAULT_GEOMETRY_INPUT_PATH,
        help="Optional floor/nonplinth geometry workbook containing Rectangle/Balcony/Staircase sheets.",
    )
    parser.add_argument(
        "--dxf-columns",
        type=Path,
        default=DEFAULT_DXF_COLUMNS_INPUT_PATH,
        help="DXF column rectangles workbook (*_col_rectangles_m_v2_wall_assisted.xlsx). Used to read column bounding-box geometry for anchor computation.",
    )
    parser.add_argument(
        "--level-tolerance-m",
        type=float,
        default=DEFAULT_NODE_LEVEL_TOLERANCE_M,
        help="Tolerance in meters used to level/merge close node X/Y values.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = resolve_input_workbook(args.input)
    wall_path = resolve_wall_workbook(args.walls)
    geometry_input_path = resolve_geometry_workbook(args.geometry_input)
    dxf_columns_path = resolve_dxf_columns_workbook(args.dxf_columns)
    output_path = args.output.resolve() if args.output else input_path.parent / f"node_coordinates_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    other_output_path = build_other_output_path(output_path)
    unfiltered = load_unfiltered_module()
    legacy_node = load_legacy_node_module()

    dxf_geometry = load_dxf_column_geometry(dxf_columns_path)

    wb = load_workbook(input_path)
    if "Columns" not in wb.sheetnames or "Primary Beams" not in wb.sheetnames:
        raise SystemExit("Input workbook must contain 'Columns' and 'Primary Beams' sheets.")

    ws_cols = wb["Columns"]
    ws_beams = wb["Primary Beams"]

    columns, _ = load_columns(ws_cols, dxf_geometry)
    raw_secondary_rows = load_raw_secondary_rows(wb)
    x_beam_widths_by_column, y_beam_widths_by_column, beam_header_map = collect_beam_widths(ws_beams)
    primary_beam_rows = load_primary_beam_rows(ws_beams, beam_header_map)
    raw_nodes, leveled_nodes = compute_leveled_nodes(
        columns,
        x_beam_widths_by_column=x_beam_widths_by_column,
        y_beam_widths_by_column=y_beam_widths_by_column,
        tolerance_m=args.level_tolerance_m,
    )

    append_node_columns(ws_cols, columns, leveled_nodes)
    write_shear_wall_template_sheet(wb)
    write_column_landscape_sheet(wb, ws_cols)
    rewrite_primary_beams(ws_beams, leveled_nodes, beam_header_map)
    append_primary_beam_present_column(ws_beams)
    node_x_vals = unique_sorted([coords[0] for coords in leveled_nodes.values()])
    node_y_vals = unique_sorted([coords[1] for coords in leveled_nodes.values()])

    if legacy_node is not None:
        rect_rows, balcony_rows, staircase_rows = build_harmonized_other_rows(
            legacy_node,
            geometry_input_path,
            node_x_vals,
            node_y_vals,
            sec_beam_rows=None,
        )
        write_other_coordinates_workbook(other_output_path, rect_rows, balcony_rows, staircase_rows)
        rectangle_x, rectangle_y = rectangle_axes_from_rows(rect_rows)

        generated_secondary_rows = build_generated_secondary_raw_rows(
            unfiltered,
            other_output_path,
            wall_path,
            columns,
            leveled_nodes,
            primary_beam_rows,
            raw_secondary_rows,
        )
        final_secondary_rows = build_final_secondary_rows(generated_secondary_rows, leveled_nodes, rectangle_x, rectangle_y)

        rect_rows, balcony_rows, staircase_rows = build_harmonized_other_rows(
            legacy_node,
            geometry_input_path,
            node_x_vals,
            node_y_vals,
            sec_beam_rows=final_secondary_rows,
        )
        write_other_coordinates_workbook(other_output_path, rect_rows, balcony_rows, staircase_rows)
        rectangle_x, rectangle_y = rectangle_axes_from_rows(rect_rows)

        generated_secondary_rows = build_generated_secondary_raw_rows(
            unfiltered,
            other_output_path,
            wall_path,
            columns,
            leveled_nodes,
            primary_beam_rows,
            raw_secondary_rows,
        )
        final_secondary_rows = build_final_secondary_rows(generated_secondary_rows, leveled_nodes, rectangle_x, rectangle_y)

        rect_rows, balcony_rows, staircase_rows = build_harmonized_other_rows(
            legacy_node,
            geometry_input_path,
            node_x_vals,
            node_y_vals,
            sec_beam_rows=final_secondary_rows,
        )
        ew_rows, ew_support_segs = build_extra_wall_rows(
            unfiltered,
            legacy_node,
            wall_path,
            node_x_vals,
            node_y_vals,
            primary_beam_rows,
            leveled_nodes,
            final_secondary_rows,
        )
        write_other_coordinates_workbook(other_output_path, rect_rows, balcony_rows, staircase_rows, ew_rows=ew_rows, ew_support_segments=ew_support_segs)
    else:
        final_secondary_rows = build_final_secondary_rows(raw_secondary_rows, leveled_nodes, [], [])

    write_final_secondary_sheet(wb, final_secondary_rows)
    if "Node spacing review" in wb.sheetnames:
        del wb["Node spacing review"]
    if "Extra column" not in wb.sheetnames:
        ws_extra = wb.create_sheet("Extra column")
        ws_extra.append([
            "Extra Column",
            "extra_col_x",
            "extra_col_z",
            "extra_col_yd",
            "extra_col_zd",
            "extra_col_orientation",
            "extra_beam_x_loc",
            "extra_beam_y_loc",
            "extra_beam_width_x_mm",
            "extra_beam_width_y_mm",
        ])
        for cell in ws_extra[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    prune_main_workbook(wb)
    align_new_headers(ws_cols, ws_beams, wb[SHEAR_WALL_TEMPLATE_SHEET], wb[COLUMN_LANDSCAPE_TEMPLATE_SHEET])
    wb.save(output_path)

    print(f"Input workbook  : {input_path}")
    print(f"Wall workbook   : {wall_path}")
    print(f"Geometry source : {geometry_input_path}")
    print(f"DXF columns     : {dxf_columns_path}")
    print(f"Output workbook : {output_path}")
    if legacy_node is None:
        print("Other workbook  : skipped (legacy node geometry helper missing)")
        print("Secondary mode  : fallback from existing raw secondary sheets")
    else:
        print(f"Other workbook  : {other_output_path}")
    print(f"Columns updated : {len(columns)}")
    print(f"Secondary rows  : {len(final_secondary_rows)}")
    print(f"Node floor used : {NODE_SOURCE_FLOOR}")
    print(f"Level tolerance : {args.level_tolerance_m:.3f} m")
    for key in sorted(raw_nodes, key=lambda name: int(name[1:]))[:5]:
        raw_node = raw_nodes[key]
        leveled_node = leveled_nodes[key]
        print(f"{key}: raw={raw_node} leveled={leveled_node}")


if __name__ == "__main__":
    main()
